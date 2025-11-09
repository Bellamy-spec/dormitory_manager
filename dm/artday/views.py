from django.shortcuts import render
from .forms import CostumeForm, ProgramForm, PerformerForm, DesignerForm, SetNumForm
from django.http import HttpResponseRedirect, Http404, HttpResponse
from .models import Costume, Program, Performer, Designer
from django.urls import reverse
from .tools import DataTool
import os
from openpyxl.styles import Side, Border, Font, Alignment
from io import BytesIO
import datetime
from openpyxl import Workbook
from django.conf import settings


# 根路径
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 实例化数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}艺术节报名系统统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def get_max_num(year):
    """取出当前最大服装编号"""
    # 初始化最大编号值
    max_num = 0

    for costume in Costume.objects.filter(year=year):
        if costume.num > max_num:
            max_num = costume.num

    return max_num


def sorted_costumes(year):
    """对所有服装对象按一定规则排序，以元组形式返回排序后对象"""
    # 对所有服装对象按编号进行分类
    costume1, costume2 = [], []
    for costume in Costume.objects.filter(year=year):
        if costume.num == 0:
            costume1.append(costume)
        else:
            costume2.append(costume)

    # 对编号不为0的服装对象进行排序
    costume2.sort(key=lambda x: x.num)

    # 合并返回
    return tuple(costume2 + costume1)


def all_program_sec(year):
    """以列表形式返回某年所有节目的口令"""
    sec_list = []
    for program in Program.objects.filter(year=year):
        sec_list.append(program.sec)
    print(sec_list)
    return sec_list


def all_costume_sec(year):
    """以列表形式返回某年所有服装的口令"""
    sec_list = []
    for costume in Costume.objects.filter(year=year):
        sec_list.append(costume.sec)
    return sec_list


# Create your views here.
def index(request):
    """主页"""
    context = {'title': '{}艺术节报名系统'.format(settings.USER_NAME)}
    return render_ht(request, 'artday/index.html', context)


def add_program(request):
    """报名节目"""
    # raise Http404
    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ProgramForm()
    else:
        # 对POST提交的数据作出处理
        form = ProgramForm(request.POST)
        if form.is_valid():
            new_program = form.save(commit=False)

            # 设置年份属性
            new_program.year = str(datetime.datetime.now().year)

            # 保存之前，载入所有已保存对象口令
            secs = all_program_sec(new_program.year)

            # 保存
            new_program.save()

            # 更新口令，保证不重复
            while new_program.sec == '000000' or new_program.sec in secs:
                new_program.update_sec()
                new_program.save()

            context = {
                'msg': '你的节目口令是{}，请牢记！将作为你以后修改节目信息的重要依据'.format(new_program.sec),
                'year': new_program.year,
            }
            return render_ht(request, 'artday/tip_sec.html', context)

    tip = '报名须知：节目内容要积极向上，要符合社会主义核心价值观，着装要得体，不得出现日韩系内容，' \
          '否则将一票否决，不予通过！'
    return render_ht(request, 'artday/add_program.html', {'form': form, 'tip': tip})


def program_list(request, year):
    """节目列表"""
    # 取出所有节目
    programs = Program.objects.filter(year=year)

    return render_ht(request, 'artday/program_list.html', context={
        'programs': programs,
        'is_manager': request.user.username in DT.managers,
        'year': year,
    })


def show_program(request, program_id):
    """显示节目"""
    # 取出要显示的节目对象
    program = Program.objects.get(id=program_id)

    return render_ht(request, 'artday/show_program.html', {
        'program': program, 'is_manager': request.user.username in DT.managers})


def add_costume(request):
    """报名服装"""
    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = CostumeForm()
    else:
        # 对POST提交的数据作出处理
        form = CostumeForm(request.POST, request.FILES)
        if form.is_valid():
            new_costume = form.save(commit=False)

            # # 设计图纸为必填项
            # if not new_costume.drawing:
            #     return render(request, 'artday/add_costume.html', {
            #         'form': form, 'err': '请上传服装设计图纸'})

            # 设置年份属性
            new_costume.year = str(datetime.datetime.now().year)

            # 保存之前，载入所有已保存对象口令
            secs = all_costume_sec(new_costume.year)

            # 保存
            new_costume.save()

            # 更新口令，保证不重复
            while new_costume.sec == '000000' or new_costume.sec in secs:
                new_costume.update_sec()
                new_costume.save()

            context = {
                'msg': '你的服装口令是{}，请牢记！将作为你以后修改服装信息的重要依据'.format(new_costume.sec),
                'year': new_costume.year,
            }
            return render_ht(request, 'artday/tip_sec.html', context)

    return render_ht(request, 'artday/add_costume.html', {'form': form, 'err': ''})


def costume_list(request, year):
    """服装列表"""
    return render_ht(request, 'artday/costume_list.html', context={
        'costumes': sorted_costumes(year),
        'is_manager': request.user.username in DT.managers,
        'year': year,
    })


def show_costume(request, costume_id):
    """显示服装"""
    # 取出要显示的服装对象
    costume = Costume.objects.get(id=costume_id)

    # 是否含有设计图纸
    if costume.drawing:
        have_pic = True
    else:
        have_pic = False

    return render_ht(request, 'artday/show_costume.html', context={
        'costume': costume,
        'is_manager': request.user.username in DT.managers,
        'have_pic': have_pic,
    })


def delete_program(request, program_id, load_to):
    """删除节目"""
    # 仅管理员有此权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的节目对象
    program = Program.objects.get(id=program_id)

    # 执行删除操作
    program.delete()

    if load_to != 'a':
        # 重定向至口令查看页
        return HttpResponseRedirect(reverse('artday:see_program_sec', args=[program.year]))
    else:
        # 重定向到节目列表页面
        return HttpResponseRedirect(reverse('artday:program_list', args=[program.year]))


def delete_costume(request, costume_id, load_to):
    """删除服装"""
    # 仅管理员有此权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的服装对象
    costume = Costume.objects.get(id=costume_id)

    # 删除图片文件
    try:
        pic = BASE_DIR + '/media/' + str(costume.drawing)
        os.remove(pic)
    except IsADirectoryError:
        # 不存在文件，忽略此步
        pass

    # 执行删除操作
    costume.delete()

    if load_to != 'a':
        # 重定向至口令查看页
        return HttpResponseRedirect(reverse('artday:see_costume_sec', args=[costume.year]))
    else:
        # 重定向到服装列表页面
        return HttpResponseRedirect(reverse('artday:costume_list', args=[costume.year]))


def see_performer(request, program_id):
    """查看某个节目的表演者"""
    # 取出相对应的节目对象
    program = Program.objects.get(id=program_id)

    # 取出该节目的所有表演者
    performers = Performer.objects.filter(program_belong=program)

    context = {
        'program': program,
        'performers': performers,
        'is_manager': request.user.username in DT.managers,
    }
    return render_ht(request, 'artday/see_performer.html', context)


def see_designer(request, costume_id):
    """查看某个服装的设计师"""
    # 取出相应的服装对象
    costume = Costume.objects.get(id=costume_id)

    # 取出该服装的所有设计师
    designers = Designer.objects.filter(costume_belong=costume)

    context = {
        'costume': costume,
        'designers': designers,
        'is_manager': request.user.username in DT.managers,
    }
    return render_ht(request, 'artday/see_designer.html', context)


def edit_program(request, program_id, program_sec):
    """编辑节目信息"""
    # 取出要编辑的节目对象
    program = Program.objects.get(id=program_id)

    # 检验匹配
    match = program.sec == program_sec

    # 仅管理员或知口令者有此权限
    if not match:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，用对象原有属性填充表单
        form = ProgramForm(instance=program)
    else:
        # 对POST提交的数据作出处理
        form = ProgramForm(instance=program, data=request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('artday:show_program', args=[program.id]))

    return render_ht(request, 'artday/edit_program.html', {'form': form, 'program': program})


def edit_costume(request, costume_id, costume_sec):
    """编辑服装信息"""
    # 取出要编辑的服装对象
    costume = Costume.objects.get(id=costume_id)

    # 检验匹配
    match = costume.sec == costume_sec

    # 仅管理员或知道口令者有此权限
    if not match:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，用对象原有属性填充表单
        form = CostumeForm(instance=costume)
    else:
        # 对POST提交的数据作出处理
        form = CostumeForm(instance=costume, data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('artday:show_costume', args=[costume.id]))

    return render_ht(request, 'artday/edit_costume.html', {'form': form, 'costume': costume})


def add_performer(request, program_id):
    """为指定节目增加表演者"""
    # 取出需要增加表演者的节目对象
    program = Program.objects.get(id=program_id)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = PerformerForm()
    else:
        # 对POST提交的数据作出处理
        form = PerformerForm(request.POST)
        if form.is_valid():
            new_performer = form.save(commit=False)
            new_performer.program_belong = program
            new_performer.save()
            return HttpResponseRedirect(reverse('artday:see_performer', args=[program_id]))

    return render_ht(request, 'artday/add_performer.html', {'form': form, 'program': program})


def add_designer(request, costume_id):
    """为指定服装增加设计师"""
    # 取出要增加设计师的服装对象
    costume = Costume.objects.get(id=costume_id)

    if request.method != 'POST':
        # 提交数据，创建新的表单
        form = DesignerForm()
    else:
        # 对POST提交的数据作出处理
        form = DesignerForm(request.POST)
        if form.is_valid():
            new_designer = form.save(commit=False)
            new_designer.costume_belong = costume
            new_designer.save()
            return HttpResponseRedirect(reverse('artday:see_designer', args=[costume_id]))

    return render_ht(request, 'artday/add_designer.html', {'form': form, 'costume': costume})


def delete_performer(request, performer_id):
    """删除表演者"""
    # 取出要删除的表演者对象及其所属的节目对象
    performer = Performer.objects.get(id=performer_id)
    program = performer.program_belong

    # 执行删除操作
    performer.delete()

    # 重定向至查看表演者页面
    return HttpResponseRedirect(reverse('artday:see_performer', args=[program.id]))


def delete_designer(request, designer_id):
    """删除设计师"""
    # 取出要删除的设计师对象及其所属的服装对象
    designer = Designer.objects.get(id=designer_id)
    costume = designer.costume_belong

    # 执行删除操作
    designer.delete()

    # 重定向至查看设计师页面
    return HttpResponseRedirect(reverse('artday:see_designer', args=[costume.id]))


def export_program(request, year):
    """导出节目单"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 新建表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].alignment = Alignment(wrap_text=True)
    st['A1'].value = '序号'
    st['B1'].value = '节目名称'
    st['C1'].value = '节目类别'
    st['D1'].value = '节目负责人'
    st['D2'].value = '姓名'
    st['E2'].value = '所在班级'
    st['F1'].alignment = Alignment(wrap_text=True)
    st['F1'].value = '所需话筒个数'
    st['G1'].value = '节目简介'
    st['H1'].value = '备注'

    # 合并需要合并的单元格
    for c in ['A', 'B', 'C', 'F', 'G', 'H']:
        st.merge_cells(range_string='{}1:{}2'.format(c, c))
    st.merge_cells(range_string='D1:E1')

    # 设置表头字体格式
    for r in range(1, 3):
        for col in range(1, 9):
            st.cell(row=r, column=col).font = Font(bold=True)

    # 取出所有节目并排序
    programs = list(Program.objects.filter(year=year))
    programs.sort(key=lambda x: DT.get_cs_tup(x.owner_class))
    programs.sort(key=lambda x: DT.range[x.tp])

    # 初始行
    row = 3

    # 循环遍历写入节目信息
    for program in programs:
        # 写入序号
        st.cell(row=row, column=1).value = row - 2

        # 写入节目名称和表演形式
        st.cell(row=row, column=2).value = program.name
        st.cell(row=row, column=3).value = program.tp

        # 写入负责人姓名和所在班级
        st.cell(row=row, column=4).value = program.owner
        st.cell(row=row, column=5).value = program.owner_class

        # 写入所需话筒数量，节目简介和备注
        st.cell(row=row, column=6).value = program.mac_nums
        st.cell(row=row, column=7).value = program.desc
        st.cell(row=row, column=8).value = program.etc

        # G、H、B、D四列设置为自动换行
        st.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=8).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

        # 下一行
        row += 1

    # 调整列宽
    set_width_dict(st, {'A': 3, 'B': 15, 'C': 9, 'D': 7, 'E': 9, 'F': 8, 'G': 38, 'H': 38})

    # 添加边框
    add_border(st, start_row=1, end_row=row - 1, start_column=1, end_column=8)

    # 输出
    return write_out(wb)


def export_costume(request, year):
    """导出服装"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 新建表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].alignment = Alignment(wrap_text=True)
    st['A1'].value = '序号'
    st['B1'].value = '服装名称'
    st['C1'].value = '服装模特'
    st['C2'].value = '姓名'
    st['D2'].value = '所在班级'
    st['E1'].value = '服装联络人'
    st['E2'].value = '姓名'
    st['F2'].value = '所在班级'
    st['G1'].value = '服装简介'
    st['H1'].value = '设计图纸链接'

    # 设置表头字体格式
    for r in range(1, 3):
        for col in range(1, 9):
            st.cell(row=r, column=col).font = Font(bold=True)

    # 合并需要合并的单元格
    for c in ['A', 'B', 'G', 'H']:
        st.merge_cells(range_string='{}1:{}2'.format(c, c))
    st.merge_cells(range_string='C1:D1')
    st.merge_cells(range_string='E1:F1')

    # 初始行
    row = 3

    # 循环遍历所有服装，写入信息
    for costume in sorted_costumes(year):
        # 写入基本信息
        st.cell(row=row, column=1).value = costume.num
        st.cell(row=row, column=2).value = costume.name
        st.cell(row=row, column=3).value = costume.mt
        st.cell(row=row, column=4).value = costume.mt_class
        st.cell(row=row, column=5).value = costume.owner
        st.cell(row=row, column=6).value = costume.owner_class

        # 以可换行的方式写入服装简介
        st.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=7).value = costume.desc

        # 如果存在设计图纸，则插入其链接
        if costume.drawing:
            url = 'http://39.106.137.22/media/' + str(costume.drawing)
            st.cell(row=row, column=8).alignment = Alignment(wrap_text=True)
            st.cell(row=row, column=8).value = url

        # 下一行
        row += 1

    # 调整列宽
    set_width_dict(st, {'A': 3, 'B': 16, 'C': 9, 'D': 9, 'E': 9, 'F': 9, 'G': 35, 'H': 35})

    # 添加边框
    add_border(st, start_row=1, end_row=row - 1, start_column=1, end_column=8)

    # 输出
    return write_out(wb)


def export_sec_costume(request, year):
    """导出含有口令的服装信息"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 新建表格
    wb = Workbook()
    st = wb.active

    # 初始行
    row = 1

    for costume in sorted_costumes(year):
        st.cell(row=row, column=1).value = '服装编号'
        st.cell(row=row, column=2).value = '服装名称'
        st.cell(row=row, column=3).value = '服装口令'
        row += 1
        st.cell(row=row, column=1).value = costume.num
        st.cell(row=row, column=2).value = costume.name
        st.cell(row=row, column=3).value = DT.secl[costume.id - 1]
        row += 2

    # 设置字体格式并且居中
    for r in range(1, row - 1):
        for col in range(1, 4):
            st.cell(row=r, column=col).font = Font(size=13)
            st.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # 设置页边距（单位：英寸）
    st.page_margins.top = 0.55
    st.page_margins.bottom = 0.55

    # 调整列宽
    set_width_dict(st, {'A': 22, 'B': 22, 'C': 22})

    # 输出
    return write_out(wb)


def set_num(request, costume_id):
    """给服装设置编号"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出相应的服装对象
    costume = Costume.objects.get(id=costume_id)

    # 给服装设置默认编号
    costume.num = get_max_num(costume.year) + 1

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = SetNumForm(instance=costume)
    else:
        # 对POST提交的数据作出处理
        form = SetNumForm(instance=costume, data=request.POST)
        if form.is_valid():
            form.save()

            # 重定向至服装列表页
            return HttpResponseRedirect(reverse('artday:see_costume_sec', args=[costume.year]))

    return render_ht(request, 'artday/set_num.html', {'costume': costume, 'form': form})


def change_costume(request, costume_sec):
    """编辑服装信息"""
    # 取出相应的服装对象
    try:
        costume_id = DT.sec[costume_sec]
        costume = Costume.objects.get(id=costume_id)
    except:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，用对象原有属性填充表单
        form = CostumeForm(instance=costume)
    else:
        # 对POST提交的数据作出处理
        form = CostumeForm(instance=costume, data=request.POST, files=request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('artday:show_costume', args=[costume.id]))

    return render_ht(request, 'artday/change_costume.html', context={
        'form': form, 'costume': costume, 'sec': costume_sec})


def sec_input(request):
    """输入口令页面"""
    if request.method == 'POST':
        # 获取POST提交的口令
        sec = request.POST.get('sec', '')

        # 重定向至修改页
        return HttpResponseRedirect(reverse('artday:change_costume', args=[sec]))
    else:
        return render_ht(request, 'artday/sec_input.html', {})


def years(request):
    """年份列表"""
    # 起始年份
    start_year = 2023

    # 当前年份
    end_year = datetime.datetime.now().year

    # 生成可遍历的年份集
    year_list = []
    for y in range(start_year, end_year + 1):
        year_list.append(str(y))

    context = {'years': tuple(year_list)}
    return render_ht(request, 'artday/years.html', context)


def year_main(request, year):
    """年份主页"""
    # 生成标题
    title = '{}年艺术节报名情况查看'.format(year)

    context = {'title': title, 'year': year, 'is_manager': request.user.username in DT.managers}
    return render_ht(request, 'artday/year_main.html', context)


def program_sec_input(request, year):
    """输入节目口令"""
    if request.method == 'POST':
        # 对POST提交的数据做出处理
        sec = request.POST.get('sec', '')

        # 尝试取出节目对象
        try:
            program = Program.objects.filter(year=year, sec=sec)[0]
        except IndexError:
            context = {'year': year, 'err': '未查询到相关节目'}
            return render_ht(request, 'artday/program_sec_input.html', context)
        else:
            # 重定位到节目编辑页面
            return HttpResponseRedirect(reverse('artday:edit_program', args=[
                program.id, program.sec]))

    else:
        # 加载新的输入页面
        context = {'year': year, 'err': ''}
        return render_ht(request, 'artday/program_sec_input.html', context)


def costume_sec_input(request, year):
    """输入服装口令"""
    if request.method == 'POST':
        # 对POST提交的数据做出处理
        sec = request.POST.get('sec', '')

        # 尝试取出服装对象
        try:
            costume = Costume.objects.filter(year=year, sec=sec)[0]
        except IndexError:
            context = {'year': year, 'err': '未查询到相关服装'}
            return render_ht(request, 'artday/costume_sec_input.html', context)
        else:
            # 重定位到节目编辑页面
            return HttpResponseRedirect(reverse('artday:edit_costume', args=[
                costume.id, costume.sec]))

    else:
        # 加载新的输入页面
        context = {'year': year, 'err': ''}
        return render_ht(request, 'artday/costume_sec_input.html', context)


def see_program_sec(request, year):
    """查看节目口令"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出所有节目对象
    programs = Program.objects.filter(year=year)

    context = {'year': year, 'programs': programs}
    return render_ht(request, 'artday/see_program_sec.html', context)


def see_costume_sec(request, year):
    """查看服装口令"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出所有服装对象
    costumes = Costume.objects.filter(year=year).order_by('num')

    context = {'year': year, 'costumes': costumes}
    return render_ht(request, 'artday/see_costume_sec.html', context)


def update_program_sec(request, program_id):
    """更新节目口令"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出节目对象
    program = Program.objects.get(id=program_id)

    # 更新口令，保存
    program.update_sec()
    program.save()

    # 重定向至口令查看页
    return HttpResponseRedirect(reverse('artday:see_program_sec', args=[program.year]))


def update_costume_sec(request, costume_id):
    """更新服装口令"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 取出节目对象
    costume = Costume.objects.get(id=costume_id)

    # 更新口令，保存
    costume.update_sec()
    costume.save()

    # 重定向至口令查看页
    return HttpResponseRedirect(reverse('artday:see_costume_sec', args=[costume.year]))
