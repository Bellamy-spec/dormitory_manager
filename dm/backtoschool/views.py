from django.shortcuts import render
from .models import Task, Class, Student
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from .tools import DataTool
from .forms import ClassForm, StudentForm
from datetime import datetime
from django.contrib.auth.views import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment
from io import BytesIO
from dm.scores import models, data
from django.conf import settings


# 实例化工具类
DT = DataTool()
DT_OUT = data.Data()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.


def index(request):
    """返校情况统计程序主页"""
    # 取出所有任务
    tasks = Task.objects.all()

    # 已完成，进行中两个列表
    doing_list, done_list = [], []

    # 循环遍历所有任务进行分类
    for task in tasks:
        # 循环遍历该任务下所有班级是否都已完成，并依此分类
        for cs in Class.objects.filter(task=task):
            if not cs.done:
                doing_list.append(task)
                break

    for task in tasks:
        if task not in doing_list:
            done_list.append(task)

    # 哈希
    doing_list = tuple(doing_list)
    done_list = tuple(done_list)

    context = {
        'doing': doing_list,
        'done': done_list,
        'title': '{}返校情况统计系统'.format(settings.USER_NAME),
        'can_public': request.user.username in DT.super_users,
    }
    return render_ht(request, 'backtoschool/index.html', context)


def one_task(request, task_id):
    """具体任务页"""
    # 取出具体任务
    task = Task.objects.get(id=task_id)

    # 制定标题
    title = '{}返校情况统计'.format(task.date_str)

    # 已完成，未完成两个班级列表
    done_class, doing_class = [], []

    # 默认高三13班不在
    can_add = True

    # 分类
    task_done = True
    for cs in Class.objects.filter(task=task):
        if cs.done:
            done_class.append(cs)
        else:
            doing_class.append(cs)
            task_done = False

        # 已有高三13班
        if cs.name == '高三13班':
            can_add = False

    # 排序，哈希
    done_class.sort(key=lambda x: DT.get_class_num(x.name))
    doing_class.sort(key=lambda x: DT.get_class_num(x.name))
    done_class = tuple(done_class)
    doing_class = tuple(doing_class)

    context = {
        'title': title,
        'done_class': done_class,
        'doing_class': doing_class,
        'task': task,
        'can_delete': request.user.username in DT.super_users,
        'can_mark': not task_done,
        'can_add': can_add,
    }
    return render_ht(request, 'backtoschool/task.html', context)


def see_class(request, class_id):
    """查看班级返校情况"""
    # 取得指定的班级
    cs = Class.objects.get(id=class_id)

    # 未提交班级无此页面
    if not cs.done:
        raise Http404

    # 制定标题
    title = '{}{}返校情况'.format(cs.task.date_str, cs.name)

    # 获取未返校学生
    abst = Student.objects.filter(_class=cs)

    context = {'title': title, 'cs': cs, 'absents': abst}
    return render_ht(request, 'backtoschool/see.html', context)


@ login_required()
def edit_class(request, class_id, err):
    """班级情况编辑"""
    if err == 'a':
        err = ''
    elif err == 'b':
        err = '首次上报，无上一次数据'

    # 权限设置
    if request.user.username not in DT.operators:
        raise Http404

    # 取出对应的班级
    cs = Class.objects.get(id=class_id)

    # 获取逻辑班级字符串
    grade = cs.name[:2]
    lg = DT_OUT.logic_grade[grade]
    ct = cs.name[2:]
    if len(ct) < 3:
        ct = '0' + ct
    lgc = lg + ct

    # 获取学生列表
    st_list = models.get_students(lgc)

    # 制定标题
    title = '{}{}返校情况编辑'.format(cs.task.date_str, cs.name)

    # 获取未返校学生
    abst = Student.objects.filter(_class=cs)

    if request.method != 'POST':
        # 未提交数据，创建一个新表单
        form = ClassForm(instance=cs)
    else:
        # 对提交对数据进行处理
        form = ClassForm(request.POST)
        if form.is_valid():
            new = form.save(commit=False)
            cs.total = new.total

            # 班级总人数必须大于0
            if cs.total < 1:
                return render_ht(request, 'backtoschool/edit.html', context={
                    'form': form,
                    'form2': StudentForm(),
                    'cs': cs,
                    'title': title,
                    'absents': abst,
                    'err': '班级总人数必须大于0',
                    'st_list': tuple(st_list),
                })

            # 计算未到人数
            if cs.absent_students != '无':
                absent_students = cs.absent_students.split(',')
            else:
                absent_students = []
            cs.come = cs.total - len(absent_students)

            # 已做
            cs.done = True

            # 保存
            cs.save()
            return HttpResponseRedirect(reverse('backtoschool:see', args=[class_id]))

    context = {
        'form': form,
        'form2': StudentForm(),
        'cs': cs,
        'title': title,
        'absents': abst,
        'err': err,
        'st_list': tuple(st_list)
    }
    return render_ht(request, 'backtoschool/edit.html', context)


@ login_required()
def add_absent_student(request, class_id):
    """新增请假学生"""
    # 权限设置
    if request.user.username not in DT.operators:
        raise Http404

    # 取出对应班级
    cs = Class.objects.get(id=class_id)

    # 获取逻辑班级字符串
    grade = cs.name[:2]
    lg = DT_OUT.logic_grade[grade]
    ct = cs.name[2:]
    if len(ct) < 3:
        ct = '0' + ct
    lgc = lg + ct

    # 获取学生列表
    st_list = models.get_students(lgc)

    # 制定标题
    title = '新增{}{}请假学生'.format(cs.task.date_str, cs.name)

    if request.method != 'POST':
        # 未提交数据，创建一个新表单
        form = StudentForm()
    else:
        # 对提交的数据进行处理
        form = StudentForm(request.POST)
        if form.is_valid():
            new_student = form.save(commit=False)
            new_student._class = cs

            # 姓名当中不能含有数字和英文逗号字符
            if DT.num_regex.findall(new_student.name) or ',' in new_student.name:
                return render_ht(request, 'backtoschool/add.html', context={
                    'form': form,
                    'cs': cs,
                    'title': title,
                    'err': '姓名格式错误',
                    'st_list': tuple(st_list),
                })

            # 宿舍号只能含有数字且必须为三位
            if not (new_student.dorm.isdigit() and len(new_student.dorm) == 3):
                return render_ht(request, 'backtoschool/add.html', context={
                    'form': form,
                    'cs': cs,
                    'title': title,
                    'err': '宿舍号格式错误',
                    'st_list': tuple(st_list),
                })

            new_student.save()

            # 班级数据也要更新
            if cs.absent_students != '无':
                absents = cs.absent_students.split(',')
            else:
                absents = []
            absents.append(new_student.name)
            cs.absent_students = ','.join(absents)
            cs.come = cs.total - len(absents)
            cs.save()

            return HttpResponseRedirect(reverse('backtoschool:edit', args=[class_id, 'a']))

    context = {'form': form, 'cs': cs, 'title': title, 'err': '', 'st_list': tuple(st_list)}
    return render_ht(request, 'backtoschool/add.html', context)


@ login_required()
def delete_student(request, student_id):
    """删除请假学生"""
    # 权限设置
    if request.user.username not in DT.operators:
        raise Http404

    # 取出要删除的学生及对应班级
    st = Student.objects.get(id=student_id)
    cs = st._class

    # 更新班级信息
    absents = cs.absent_students.split(',')
    absents.remove(st.name)
    if absents:
        absents_str = ','.join(absents)
    else:
        absents_str = '无'
    cs.absent_students = absents_str
    cs.come = cs.total - len(absents)
    cs.save()

    # 执行删除
    st.delete()
    return HttpResponseRedirect(reverse('backtoschool:edit', args=[cs.id, 'a']))


@ login_required()
def public_task(request):
    """发布任务"""
    # 限制管理员以外的用户访问该页面
    if request.user.username not in DT.super_users:
        raise Http404

    if request.method == 'POST':
        # 对POST提交的数据做出处理
        grade1 = request.POST.get('grade1', '')
        grade2 = request.POST.get('grade2', '')
        grade3 = request.POST.get('grade3', '')
        dt = request.POST.get('dt', '')

        # 未选择日期时给出提示
        if not dt:
            return render_ht(request, 'backtoschool/public.html', {'err': '请选择返校日期'})

        # 生成年级列表
        if grade1 or grade2 or grade3:
            # 至少有一个年级被选中时才执行此项操作
            grades = []
            if grade1:
                grades.append('高一')
            if grade2:
                grades.append('高二')
            if grade3:
                grades.append('高三')
        else:
            return render_ht(request, 'backtoschool/public.html', {'err': '请至少选择一个年级'})

        # TODO:创建任务对象
        task = Task()
        task.date_str = dt
        task.grade_include = ','.join(grades)
        task.save()

        # TODO:创建班级对象
        for grade in grades:
            for i in range(1, 13):
                class_name = grade + str(i) + '班'

                # TODO:取得该班级上次保存的总人数
                classes = Class.objects.filter(name=class_name).order_by('task')
                classes = list(classes)
                if classes:
                    cs_last = classes[-1]
                else:
                    cs_last = None

                cs = Class()
                cs.name = class_name
                cs.task = task
                cs.grade = grade

                # 记录之前保存的班级总人数
                if cs_last:
                    cs.total = cs_last.total

                cs.save()

        return HttpResponseRedirect(reverse('backtoschool:index'))

    # 未提交数据，创建新的发布表单
    return render_ht(request, 'backtoschool/public.html', {'err': ''})


@ login_required()
def delete_task(request, task_id):
    """删除任务"""
    # 限制管理员以外的用户访问此链接
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的任务
    task = Task.objects.get(id=task_id)

    # 执行删除
    task.delete()

    # 重定向到index页
    return HttpResponseRedirect(reverse('backtoschool:index'))


def export(request, task_id):
    """导出一次的返校情况"""
    # 取出相应的任务
    task = Task.objects.get(id=task_id)

    # 新建Excel文件和工作表
    wb = Workbook()
    st = wb.active

    # 大标题样式
    ft1 = Font(size=20, bold=True)

    # 表头样式
    ft2 = Font(size=14, bold=True)

    # 正文样式
    ft3 = Font(size=14)

    # 设置列宽
    column_width = {
        'A': 11,
        'B': 11,
        'C': 11,
        'D': 9,
        'E': 6,
        'F': 6,
        'G': 50,
    }
    for k in column_width.keys():
        st.column_dimensions[k].width = column_width[k]

    # 写入大标题
    st['A1'].value = '{}返校情况统计'.format(task.date_str)
    st['A1'].font = ft1
    st.merge_cells(range_string='A1:G1')

    # 写入表头
    st['A2'].value = '班级'
    st['B2'].value = '应到人数'
    st['C2'].value = '实到人数'
    st['D2'].value = '请假学生情况'
    st['D3'].value = '姓名'
    st['E3'].value = '性别'
    st['F3'].value = '宿舍'
    st['G3'].value = '请假原因及预计返校时间'
    for c in range(1, 8):
        st.cell(row=2, column=c).font = ft2
        st.cell(row=3, column=c).font = ft2

    # 合并需要合并的单元格
    for col in ['A', 'B', 'C']:
        st.merge_cells(range_string='{}2:{}3'.format(col, col))
    st.merge_cells(range_string='D2:G2')

    # 写入主体部分之前先取出班级并进行排序
    class_list = []
    for cs in Class.objects.filter(task=task):
        class_list.append(cs)
    class_list.sort(key=lambda x: DT.get_class_num(x.name))

    # 当前行
    row = 4

    # 开始写入主体
    for cs in class_list:
        st.cell(row=row, column=1).value = cs.name
        if cs.done:
            st.cell(row=row, column=2).value = cs.total
            st.cell(row=row, column=3).value = cs.come
        else:
            st.cell(row=row, column=2).value = '未提交'

        # 写入请假学生
        row_right = row
        for student in Student.objects.filter(_class=cs):
            st.cell(row=row_right, column=4).value = student.name
            st.cell(row=row_right, column=5).value = student.gender
            st.cell(row=row_right, column=6).value = student.dorm
            st.cell(row=row_right, column=7).value = student.reason
            row_right += 1

        # 必要时对左半部分合并单元格
        if row_right - row >= 2:
            for col in ['A', 'B', 'C']:
                # 合成合并字符串
                merge_string = col + str(row) + ':' + col + str(row_right - 1)
                st.merge_cells(range_string=merge_string)

        # 下一行
        if row_right > row:
            row = row_right
        else:
            row += 1

    # 主体部分设置字体
    for r in range(4, row):
        for c in range(1, 8):
            st.cell(row=r, column=c).font = ft3

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=7)

    # 垂直居中
    alignment = Alignment(vertical='center')
    for r in range(2, row):
        for c in range(1, 8):
            st.cell(row=r, column=c).alignment = alignment

    return write_out(wb)


def mark_done(request, task_id):
    """标记某个任务为已完成"""
    # 限制管理员以外的用户访问此链接
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要更改的任务对象
    task = Task.objects.get(id=task_id)

    # 遍历该任务下的班级对象，设置已完成
    for cs in Class.objects.filter(task=task):
        cs.done = True
        cs.save()

    # 重定向至首页
    return HttpResponseRedirect(reverse('backtoschool:index'))


def add_last(request, cs_id):
    """载入上次请假学生"""
    # 取出相应的班级对象和班级名
    cs = Class.objects.get(id=cs_id)
    cs_name = cs.name

    # TODO:取出该班上一次返校任务对象
    # 收集所有id小于本次的对象
    css = []
    for ob in Class.objects.filter(name=cs_name):
        if ob.id < cs_id:
            css.append(ob)

    # 若为空，则为第一次，给出提示
    if not css:
        return HttpResponseRedirect(reverse('backtoschool:edit', kwargs={
            'class_id': cs_id, 'err': 'b'}))

    # 否则取出id最大的对象
    css.sort(key=lambda x: x.id, reverse=True)
    cs_last = css[0]

    # TODO:取出上次对象的所有学生对象添加到此次
    for student in Student.objects.filter(_class=cs_last):
        student_copy = student.deep_clone()
        student_copy._class = cs
        student_copy.save()

        # 班级数据也要更新
        if cs.absent_students != '无':
            absents = cs.absent_students.split(',')
        else:
            absents = []
        absents.append(student_copy.name)
        cs.absent_students = ','.join(absents)
        cs.save()

    return HttpResponseRedirect(reverse('backtoschool:edit', kwargs={
        'class_id': cs_id, 'err': 'a'}))


def add_13(request, task_id):
    """某个任务增加高三13班"""
    # 限制管理员以外的用户访问此链接
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要更改的任务对象
    task = Task.objects.get(id=task_id)

    # 取得高三13班上次保存的总人数
    classes = Class.objects.filter(name='高三13班').order_by('task')
    classes = list(classes)
    if classes:
        cs_last = classes[-1]
    else:
        cs_last = None

    # 实例化高三13班对象
    cs = Class()
    cs.name = '高三13班'
    cs.task = task
    cs.grade = '高三'

    # 记录之前保存的班级总人数
    if cs_last:
        cs.total = cs_last.total

    # 保存实例
    cs.save()

    # 重定向至任务页
    return HttpResponseRedirect(reverse('backtoschool:task', args=[task_id]))
