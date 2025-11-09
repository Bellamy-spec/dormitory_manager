from django.shortcuts import render
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.views import login_required
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from .tools import DataTool
from .models import BKRecord
from .forms import BKRecordForm
from openpyxl.styles import Side, Border
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from django.conf import settings


# 实例化数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}中招监考信息上报系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def correct_id(id_number):
    """检查所给的身份证号格式是否合法"""
    # 位数须为18
    if len(id_number) != 18:
        return False

    # 前17位须为数字
    if not id_number[:17].isdigit():
        return False

    # 最后一位须为数字或X
    if not (id_number[17].isdigit() or id_number[17] == 'X'):
        return False

    # TODO:校验
    s = 0
    for i in range(18):
        # 取得每一位上的数字
        if id_number[i].isdigit():
            n = int(id_number[i])
        else:
            n = 10

        # 计算权
        w = 2 ** (17 - i) % 11

        # 累加
        s += n * w

    if s % 11 != 1:
        return False

    # 通过所有检查，格式正确
    return True


def all_id_numbers():
    """获取数据库中的所有身份证号"""
    # 初始化存放列表
    id_list = []

    # 循环遍历所有记录，加入身份证号
    for ob in BKRecord.objects.all():
        id_list.append(ob.id_number)

    return id_list


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """主页"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    return render_ht(request, 'bank_id/index.html', context={
        'is_manager': request.user.username in DT.managers,
        'title': '{}中招监考信息上报系统'.format(settings.USER_NAME),
    })


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到导航页
            return HttpResponseRedirect(reverse('bank_id:index'))
        else:
            return render_ht(request, 'bank_id/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'bank_id/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('bank_id:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'bank_id/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'bank_id/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'bank_id/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'bank_id/set_pwd.html', {'err': ''})


def review(request):
    """预览所有信息"""
    # 限制非管理员用户访问此页面
    if request.user.username not in DT.managers:
        raise Http404

    # 取出相对应的记录
    if request.user.username in DT.manager_points.keys():
        point = DT.manager_points[request.user.username]
        records = BKRecord.objects.filter(work_point=point)
    else:
        records = BKRecord.objects.all()

    context = {'records': records}
    return render_ht(request, 'bank_id/review.html', context)


def delete(request, record_id):
    """删除记录"""
    # 限制非管理员用户访问此页面
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的记录
    record = BKRecord.objects.get(id=record_id)

    # 执行删除操作
    record.delete()

    # 重定向至查看页
    return HttpResponseRedirect(reverse('bank_id:review'))


def send_up(request):
    """填报信息"""
    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = BKRecordForm()
    else:
        # 对POST提交的数据作出处理
        form = BKRecordForm(request.POST)
        if form.is_valid():
            new_record = form.save(commit=False)

            # 避免重复
            if new_record.id_number in all_id_numbers():
                # 生成错误提示
                msg = '已存在身份证号{}的用户，请勿重复填报，如果有误，请先联系管理员删除'\
                    .format(new_record.id_number)
                return render_ht(request, 'bank_id/send_up.html', context={
                    'form': form, 'err': msg, 'bk_dict': DT.bank_dict})

            # 验证身份证号是否符合规范
            if not correct_id(new_record.id_number):
                return render_ht(request, 'bank_id/send_up.html', context={
                    'form': form, 'err': '身份证号错误', 'bk_dict': DT.bank_dict})

            # 验证银行卡号是否符合规范
            if not (new_record.bank_id.isdigit() and len(new_record.bank_id) >= 16):
                return render_ht(request, 'bank_id/send_up.html', context={
                    'form': form, 'err': '银行卡号错误', 'bk_dict': DT.bank_dict})

            # 根据银行卡号补全银行信息
            try:
                new_record.bank_tp = DT.bank_dict[new_record.bank_id[:4]]
            except KeyError:
                new_record.bank_tp = '未知'

            # 保存
            new_record.save()
            return HttpResponse('提交成功，感谢您的支持！')

    return render_ht(request, 'bank_id/send_up.html', context={
        'form': form, 'err': '', 'bk_dict': DT.bank_dict})


def export(request):
    """导出信息"""
    # 限制非管理员用户访问此页面
    if request.user.username not in DT.managers:
        raise Http404

    # 取出相对应的记录
    if request.user.username in DT.manager_points.keys():
        point = DT.manager_points[request.user.username]
        records = BKRecord.objects.filter(work_point=point)
    else:
        records = BKRecord.objects.all()

    # 新建文件，定位表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '姓名'
    st['B1'].value = '监考考点'
    st['C1'].value = '身份证号'
    st['D1'].value = '银行卡号'
    st['E1'].value = '所属银行'
    st['F1'].value = '提交时间'

    # 初始行
    row = 2

    # 循环遍历写入主体
    for ob in records:
        st.cell(row=row, column=1).value = ob.name
        st.cell(row=row, column=2).value = ob.work_point
        st.cell(row=row, column=3).value = ob.id_number
        st.cell(row=row, column=4).value = ob.bank_id
        st.cell(row=row, column=5).value = ob.bank_tp

        # 转换时区
        dt = ob.datetime_added.astimezone(DT.tz)

        st.cell(row=row, column=6).value = datetime.strftime(dt, '%Y-%m-%d %H:%M:%S')

        # 下一行
        row += 1

    # 设置列宽
    set_width_dict(st, {'B': 12, 'C': 20, 'D': 20, 'F': 20})

    # 添加边框
    add_border(st, start_row=1, end_row=row - 1, start_column=1, end_column=5)

    # 输出
    return write_out(wb)
