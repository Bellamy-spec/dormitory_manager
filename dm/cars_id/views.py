from django.shortcuts import render
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.views import login_required
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from .tools import DataTool
from .models import CarRecord
from .forms import CarForm
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment
from io import BytesIO
from datetime import datetime, date
from django.conf import settings


# 实例化数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}车辆管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def check_char(char):
    """检查车牌号后几位字符是否符合规范"""
    if char.isdigit():
        # 数字一定符合规范
        return True
    elif char.isupper():
        # 大写字母有可能符合规范
        if char in ['O', 'I']:
            return False
        else:
            return True
    else:
        # 其他情况都不符合规范
        return False


def check_char2(char):
    """检查车牌号第二位字符是否符合规范"""
    if char.isupper():
        if char in ['O', 'I']:
            return False
        else:
            return True
    else:
        return False


def all_cars():
    """获取系统中已有的所有车牌号"""
    cars = []
    for car_obj in CarRecord.objects.all():
        cars.append(car_obj.car1)
        if car_obj.car2:
            cars.append(car_obj.car2)
    return cars


def all_phone_numbers():
    """获取系统中已有的所有手机号"""
    phone_numbers = []
    for car_obj in CarRecord.objects.all():
        phone_numbers.append(car_obj.phone_number)
    return phone_numbers


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """MainPage"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    # 制定标题
    title = '{}车牌号信息统计系统'.format(settings.USER_NAME)

    context = {'title': title, 'is_manager': request.user.username in DT.managers}
    return render_ht(request, 'cars_id/index.html', context)


def login1(request):
    """登录页"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到主页
            return HttpResponseRedirect(reverse('cars_id:index'))
        else:
            return render_ht(request, 'cars_id/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'cars_id/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('cars_id:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'cars_id/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'cars_id/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'cars_id/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'cars_id/set_pwd.html', {'err': ''})


def see(request):
    """查看车牌信息"""
    # 限制管理员以外的用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出所有车牌信息
    cars = CarRecord.objects.all()

    # 标题
    title = '{}车牌信息查看'.format(settings.USER_NAME)

    context = {'title': title, 'cars': cars}
    return render_ht(request, 'cars_id/see.html', context)


def add(request):
    """录入车牌信息"""
    if request.method == 'POST':
        # 对POST提交的数据做出处理
        form = CarForm(request.POST)
        if form.is_valid():
            # 生成新的车牌信息对象
            new_car = form.save(commit=False)

            # 检查手机号是否符合规范
            if not (new_car.phone_number.isdigit() and len(new_car.phone_number) == 11):
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '手机号格式错误',
                })

            # 检查手机号是否已在系统中
            if new_car.phone_number in all_phone_numbers():
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '手机号{}已在系统中，如有问题，请联系管理员'.format(new_car.phone_number),
                })

            # 检查车牌号1是否符合规范
            if len(new_car.car1) == 7:
                # 车牌号7位的为燃油车
                new_car.is_new_energy1 = False
            elif len(new_car.car1) == 8:
                # 车牌号8位的为新能源车
                new_car.is_new_energy1 = True
            else:
                # 其他长度的车牌号不符合规范
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '车牌号1长度错误',
                })

            # 首位为省份
            if new_car.car1[0] not in DT.provinces:
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '车牌号1省份错误',
                })

            # 第二位为大写字母且不包括O,I
            if not check_char2(new_car.car1[1]):
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '车牌号1第二位字符错误',
                })

            # 剩余位数必须为数字或大写字母且不为O或I
            for char in new_car.car1[2:]:
                if not check_char(char):
                    return render_ht(request, 'cars_id/add.html', context={
                        'form': CarForm(),
                        'err': '车牌号1格式错误，请确认其中是否有大写字母O,I，'
                               '小写字母或空格、标点符号等非法字符',
                    })

            # 检查车牌号1是否已在系统中
            if new_car.car1 in all_cars():
                return render_ht(request, 'cars_id/add.html', context={
                    'form': CarForm(),
                    'err': '车牌号{}已在系统中，如有问题，请联系管理员'.format(new_car.car1),
                })

            # 从前端读取是否输入了车牌号2
            if request.POST.get('car2', ''):
                new_car.car2 = request.POST.get('car2', '')

                # 检查车牌号2是否符合规范
                if len(new_car.car2) == 7:
                    # 车牌号7位的为燃油车
                    new_car.is_new_energy2 = False
                elif len(new_car.car2) == 8:
                    # 车牌号8位的为新能源车
                    new_car.is_new_energy2 = True
                else:
                    # 其他长度的车牌号不符合规范
                    return render_ht(request, 'cars_id/add.html', context={
                        'form': CarForm(),
                        'err': '车牌号2长度错误',
                    })

                # 首位为省份
                if new_car.car2[0] not in DT.provinces:
                    return render_ht(request, 'cars_id/add.html', context={
                        'form': CarForm(),
                        'err': '车牌号2省份错误',
                    })

                # 第二位为大写字母且不包括O,I
                if not check_char2(new_car.car2[1]):
                    return render_ht(request, 'cars_id/add.html', context={
                        'form': CarForm(),
                        'err': '车牌号2第二位字符错误',
                    })

                # 剩余位数必须为数字或大写字母且不为O或I
                for char in new_car.car2[2:]:
                    if not check_char(char):
                        return render_ht(request, 'cars_id/add.html', context={
                            'form': CarForm(),
                            'err': '车牌号2格式错误，请确认其中是否有大写字母O,I，'
                                   '小写字母或空格、标点符号等非法字符',
                        })

                # 检查车牌号2是否已在系统中
                if new_car.car2 in all_cars():
                    return render_ht(request, 'cars_id/add.html', context={
                        'form': CarForm(),
                        'err': '车牌号{}已在系统中，如有问题，请联系管理员'.format(new_car.car2),
                    })
            else:
                # 避免保存错误的数据类型
                new_car.is_new_energy2 = False

            # 检查无误，保存对象
            new_car.save()
            return HttpResponse('提交成功，感谢您的支持！')

    # 未提交数据，创建新的表单
    return render_ht(request, 'cars_id/add.html', {'form': CarForm(), 'err': ''})


def delete(request, record_id):
    """删除记录"""
    # 限制管理员以外的用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的记录
    c = CarRecord.objects.get(id=record_id)

    # 执行删除操作
    c.delete()

    # 重定向到当前页
    return HttpResponseRedirect(reverse('cars_id:see'))


def export(request):
    """以Excel表格形式导出系统数据"""
    # 限制管理员以外的用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 打开文件，定位表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '客户编号'
    st['B1'].value = '客户姓名'
    st['C1'].value = '通行证类型'
    st['D1'].value = '车类型'
    st['E1'].value = '车位类型'
    st['F1'].value = '注册车牌'
    st['G1'].value = '介质类型'
    st['H1'].value = 'OBU'
    st['I1'].value = '有效期至'

    # 初始行
    row = 2

    # 逐条写入
    for car in CarRecord.objects.all():
        # 判断车辆数
        if not car.car2:
            # 只有一辆车的情况
            st['A' + str(row)].value = row - 1
            st['B' + str(row)].value = car.name
            st['C' + str(row)].value = '固定车'
            st['D' + str(row)].value = '小车'
            st['E' + str(row)].value = '临时车位'
            st['F' + str(row)].value = car.car1
            st['G' + str(row)].value = '车牌'
            st['H' + str(row)].value = '是'
            st['I' + str(row)].value = date(2024, 12, 31)

            # 下一行
            row += 1

        else:
            # 有两辆车的情况，先写入第一辆车
            st['A' + str(row)].value = row - 1
            st['B' + str(row)].value = car.name + '1'
            st['C' + str(row)].value = '固定车'
            st['D' + str(row)].value = '小车'
            st['E' + str(row)].value = '临时车位'
            st['F' + str(row)].value = car.car1
            st['G' + str(row)].value = '车牌'
            st['H' + str(row)].value = '是'
            st['I' + str(row)].value = date(2024, 12, 31)

            # 下一行
            row += 1

            # 写入第二辆车
            st['A' + str(row)].value = row - 1
            st['B' + str(row)].value = car.name + '2'
            st['C' + str(row)].value = '固定车'
            st['D' + str(row)].value = '小车'
            st['E' + str(row)].value = '临时车位'
            st['F' + str(row)].value = car.car2
            st['G' + str(row)].value = '车牌'
            st['H' + str(row)].value = '是'
            st['I' + str(row)].value = date(2024, 12, 31)

            # 下一行
            row += 1

    # 设置表格列宽
    set_width_dict(st, {'A': 9, 'B': 9, 'C': 12, 'D': 8, 'E': 9, 'F': 10,
                        'G': 9, 'H': 5, 'I': 12})

    # 导出，返回
    return write_out(wb)


def download(request):
    """导出含手机号信息的数据格式"""
    # 限制管理员以外的用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 打开文件，定位表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '车主姓名'
    st['B1'].value = '人员类别'
    st['C1'].value = '联系电话'
    st['D1'].value = '车牌号'

    # 初始行
    row = 2

    # 遍历数据，逐项写入
    for car in CarRecord.objects.all():
        st.cell(row=row, column=1).value = car.name
        st.cell(row=row, column=2).value = car.tp
        st.cell(row=row, column=3).value = car.phone_number
        st.cell(row=row, column=4).value = car.car1

        if car.car2:
            # 如果有第二辆车，则另起一行写入
            row += 1
            st.cell(row=row, column=4).value = car.car2

            # 合并前三列单元格
            for c in range(1, 4):
                st.merge_cells(start_row=row - 1, end_row=row, start_column=c, end_column=c)

        # 下一行
        row += 1

    # 添加边框线
    add_border(st, start_row=1, start_column=1, end_row=row - 1, end_column=4)

    # 设置表格列宽
    set_width_dict(st, {'A': 9, 'B': 9, 'C': 15, 'D': 12})

    # 垂直居中
    alignment = Alignment(vertical='center')
    for r in range(1, row):
        for c in range(1, 5):
            st.cell(row=r, column=c).alignment = alignment

    # 导出，返回
    return write_out(wb)


def find_car_owner(car_id):
    """根据车牌号以元组形式返回车主相关信息"""
    if car_id != '':
        # 仅在获得非空字符串时执行查询操作
        for car in CarRecord.objects.all():
            if car.car1 == car_id or car.car2 == car_id:
                return car.name, car.phone_number


def que(request):
    """查询车牌信息"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        car_id = request.POST.get('car_id', '').replace(' ', '').replace('·', '').upper()
        car_msg = find_car_owner(car_id)

        # 生成提示语
        if not car_msg:
            tip_msg = '未查询到相关信息'
        else:
            tip_msg = '车牌号{}的车主信息如下：'.format(car_id)

        # 结果返回及显示
        return render_ht(request, 'cars_id/que.html', context={
            'tip_msg': tip_msg, 'car_msg': car_msg, 'has_msg': True})

    # 以GET方式首次访问该页
    return render_ht(request, 'cars_id/que.html', context={
        'tip_msg': None, 'car_msg': None, 'has_msg': False})
