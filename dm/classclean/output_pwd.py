"""导出卫生部成员密码对照表"""
from tools import DataTool
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


# 实例化静态数据类
DT = DataTool()


def main():
    """主函数"""
    # 打开新的表格文件
    wb = Workbook()
    st = wb.active

    # 写入首行
    st['A1'].value = '姓名'
    st['B1'].value = '密码'

    # 初始行
    row = 2

    # 写入主体
    for name, pwd in DT.owners_pwd.items():
        st.cell(row=row, column=1).value = name
        st.cell(row=row, column=2).value = pwd

        # 下一行
        row += 1

    # 设置行高
    for r in range(1, row):
        st.row_dimensions[r].height = 25

        # 设置字体
        st.cell(row=r, column=1).font = Font(size=14)
        st.cell(row=r, column=2).font = Font(size=14)

        # 设置居中
        st.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
        st.cell(row=r, column=2).alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    st.column_dimensions['A'].width = 15
    st.column_dimensions['B'].width = 15

    # 保存并关闭文件
    wb.save('output_pwd.xlsx')
    wb.close()


if __name__ == '__main__':
    main()
