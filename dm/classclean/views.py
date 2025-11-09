from django.shortcuts import render
from .tools import DataTool
from .models import ClassCleanRecord, ClassCleanScore, OutLookRecord
from datetime import datetime, timedelta
from django.contrib.auth.views import login_required
from .forms import RecordForm, ScoreForm, OutLookRecordForm
from django.http import Http404, HttpResponseRedirect, HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font, Alignment
from io import BytesIO
from su_manage.models import get_owners_pwd
from dm.scores import models, data
from django.conf import settings


# 实例化工具类
DT = DataTool()
DT_OUT = data.Data()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """主页"""
    context = {
        'title': '{}班级卫生评分系统'.format(settings.USER_NAME),
        'grades': tuple(DT.grades_num.items()),
        'is_manager': request.user.username in DT.super_users,
    }
    return render_ht(request, 'classclean/index.html', context)


def show_dates(request, grade_num):
    """显示日期列表"""
    # 取出相对应的记录
    records = ClassCleanRecord.objects.filter(grade_num=grade_num)

    # 无重复地取出所有日期
    dates = []
    for record in records:
        date_str = datetime.strftime(record.date_added, '%Y-%m-%d')
        dates.append(date_str)

    # 去重，排序，哈希
    dates = list(set(dates))
    dates.sort()
    dates = tuple(dates)

    # 今日日期字符串
    today_str = datetime.strftime(datetime.today(), '%Y-%m-%d')

    context = {
        'title': '请选择日期：',
        'dates': dates,
        'today': today_str,
        'grade': grade_num,
    }
    return render_ht(request, 'classclean/dates.html', context)


def show_records(request, grade_num, date_str):
    """查看一个年级一天的记录"""
    # 字符串转化日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 取出相对应的记录
    records = ClassCleanRecord.objects.filter(grade_num=grade_num, date_added=date)

    # 制定标题
    title = '{}年级{}班级卫生扣分记录'.format(DT.reverse_grade_num[grade_num], date_str)

    context = {
        'title': title,
        'records': records,
        'grade': grade_num,
        'user': request.user,
    }
    return render_ht(request, 'classclean/look.html', context)


@ login_required()
def add_record(request):
    """新增记录"""
    # 限制管理员以外的用户访问此页
    if request.user.username not in DT.super_users:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新表单
        form = RecordForm()
    else:
        # 对POST提交的数据作出处理
        form = RecordForm(request.POST)
        if form.is_valid():
            new_record = form.save(commit=False)
            new_record.grade_num = DT.grades_num[new_record.grade]
            new_record.class_and_grade = new_record.grade + str(new_record.cs) + '班'
            new_record.owner = request.user

            # 所属日期
            if request.POST.get('dt', ''):
                new_record.date_added = request.POST.get('dt', '')
            else:
                # 未选择日期时给出错误提示
                return render_ht(request, 'classclean/add.html', context={
                    'form': form, 'err': '请选择日期'})

            new_record.save()
            return HttpResponseRedirect(reverse('classclean:index'))

    return render_ht(request, 'classclean/add.html', {'form': form, 'err': ''})


@ login_required()
def delete_record(request, record_id):
    """删除记录"""
    # 取出要删除的记录
    record = ClassCleanRecord.objects.get(id=record_id)

    # 只有添加这条记录的用户才能够删除它
    if record.owner != request.user:
        raise Http404

    # 记录当前页面年级、日期，以便删除之后重定向
    grade = record.grade_num
    date = datetime.strftime(record.date_added, '%Y-%m-%d')

    record.delete()
    return HttpResponseRedirect(reverse('classclean:look', args=[grade, date]))


def count(grade, ym):
    """以字典形式返回某个月一个年级各班的卫生得分"""
    # 初始化字典
    mtd, su_dict = {}, {}
    for cs in DT.cs_dict.values():
        gc = grade + cs
        mtd[gc] = [5]
        su_dict[gc] = []

    # 记录德育处评分
    for ob in ClassCleanRecord.objects.filter(grade=grade):
        if ob.date_added.year == ym[0] and ob.date_added.month == ym[1]:
            mtd[ob.class_and_grade][0] -= ob.decrease / 10

    # 记录卫生部评分
    for score in ClassCleanScore.objects.filter(grade=grade):
        if score.date_added.year == ym[0] and score.date_added.month == ym[1]:
            su_dict[score.class_and_grade].append(float(score.score))

    # 合并字典，统一保留两位小数
    for gc in mtd:
        try:
            # 加入学生会卫生部评分
            mtd[gc].append(sum(su_dict[gc]) / len(su_dict[gc]))
        except ZeroDivisionError:
            # 默认满分
            mtd[gc].append(5)

        # 保留位数，记录总分
        mtd[gc][0] = round(mtd[gc][0], 2)
        mtd[gc][1] = round(mtd[gc][1], 2)
        mtd[gc].append(sum(mtd[gc]))

    return mtd


def mt_main(request):
    """月总结主页"""
    # TODO:先确定有哪些月份
    today = datetime.today()
    ym_last = (today.year, today.month)

    # 初始年月固定为2023年3月
    ym = [2023, 3]
    ym_list = []

    # 年份小于或当年份等于时月份小于
    while ym[0] < ym_last[0] or (ym[0] == ym_last[0] and ym[1] < ym_last[1]):
        ym_list.append(('{}年{}月'.format(*ym), '{}-{}'.format(*ym)))

        # 年月加1
        if ym[1] < 12:
            ym[1] += 1
        else:
            ym[0] += 1
            ym[1] = 1

    # 元组化
    ym_tuple = tuple(ym_list)
    ym_now = ('{}年{}月（不完全）'.format(*ym), '{}-{}'.format(*ym))

    context = {
        'ym_before': ym_tuple,
        'ym_now': ym_now,
        'title': '选择年级/月份',
        'grades': tuple(DT.grades_num.items())
    }
    return render_ht(request, 'classclean/mt_main.html', context)


def export_mt(request, grade_num, month_str):
    """导出一个年级一个月的月总结"""
    # 生成年级字符串及年月元组
    grade = DT.reverse_grade_num[grade_num]
    ym = datetime.strptime(month_str, '%Y-%m')
    ym_tup = (ym.year, ym.month)

    # 获取数据字典
    mtd = count(grade, ym_tup)

    # 打开文件，加载工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = '{}年{}月{}年级班级卫生得分'.format(ym_tup[0], ym_tup[1], grade)
    st['A1'].font = Font(size=14, bold=True)
    st.merge_cells(range_string='A1:D1')

    # 写入表头
    st['A2'].value = '班级'
    st['B2'].value = '德育处评分'
    st['C2'].value = '卫生部评分'
    st['D2'].value = '总分'
    st['E2'].value = '平均分'
    st['A2'].font = Font(bold=True)
    st['B2'].font = Font(bold=True)
    st['C2'].font = Font(bold=True)
    st['D2'].font = Font(bold=True)
    st['E2'].font = Font(bold=True)

    # 初始行
    row = 3

    for gc, score in mtd.items():
        st.cell(row=row, column=1).value = gc
        st.cell(row=row, column=2).value = score[0]
        st.cell(row=row, column=3).value = score[1]
        st.cell(row=row, column=4).value = score[2]
        st.cell(row=row, column=5).value = score[2] / 2

        # 下一行
        row += 1

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=5)

    # 设置列宽
    set_width_dict(st, {'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20})

    # 输出
    return write_out(wb)


def show_mt(request, grade_num, month_str):
    """查看一个年级一个月的总结"""
    # 生成年级字符串及年月元组
    grade = DT.reverse_grade_num[grade_num]
    ym = datetime.strptime(month_str, '%Y-%m')
    ym_tup = (ym.year, ym.month)

    # 制定标题
    title = '{}年{}月{}年级班级卫生得分'.format(ym_tup[0], ym_tup[1], grade)

    # 获取数据字典
    mtd = count(grade, ym_tup)

    context = {'title': title, 'data': tuple(mtd.items())}
    return render_ht(request, 'classclean/show_mt.html', context)


def up_load(request, gc):
    """学生会检查结果上报"""
    # print(gc)

    # 获取当前日期、班级
    date = datetime.today()
    date_str = '{}年{}月{}日'.format(date.year, date.month, date.day)
    grade = DT.reverse_grade_num[int(gc.split('-')[0])]
    cs = gc.split('-')[1]
    class_and_grade = grade + cs + '班'

    # 取得卫生部（'02'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('02'), **get_owners_pwd('01')}

    # 不存在的情况
    if int(gc.split('-')[0]) < 1 or int(gc.split('-')[0]) > 3:
        raise Http404
    if int(cs) < 1 or int(cs) > 12:
        raise Http404

    # TODO:Mark
    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ScoreForm()
    else:
        # 对POST提交的数据作出处理
        form = ScoreForm(data=request.POST)
        if form.is_valid():
            new_score = form.save(commit=False)

            # 设置年级、班级、日期（已默认为今天）
            new_score.grade = grade
            new_score.grade_num = int(gc.split('-')[0])
            new_score.cs = int(cs)
            new_score.class_and_grade = class_and_grade

            # 检验分数
            if new_score.score < 3 or new_score.score > 5:
                return render_ht(request, 'classclean/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '得分必须介于3.0~5.0之间',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检查代号
            if new_score.owner not in opd.keys():
                return render_ht(request, 'classclean/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '代号错误',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检验密码
            if new_score.pwd != opd[new_score.owner][0]:
                return render_ht(request, 'classclean/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '检查码错误',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检查人字符串显示
            new_score.owner_show = new_score.owner + opd[new_score.owner][1]

            # 记录问题
            desc_list = []
            for d in DT.desc.keys():
                if request.POST.get(d, ''):
                    desc_list.append(DT.desc[d])
            new_score.update_desc(desc_list)

            # 清除之前添加的同班级、同日期记录，自动只保留最新
            before_score = ClassCleanScore.objects.filter(
                date_added=date, class_and_grade=class_and_grade)
            if before_score:
                for ob in before_score:
                    ob.delete()

            # 保存，重定向查看
            new_score.save()
            return HttpResponseRedirect(reverse('classclean:see_score', args=[new_score.id]))

    context = {
        'cag': class_and_grade,
        'date': date_str,
        'form': form,
        'err': '',
        'gc': gc,
        'descs': tuple(DT.desc.items()),
    }
    return render_ht(request, 'classclean/up_load.html', context)


def see_score(request, score_id):
    """学生会检查评分查看"""
    # 取出评分对象
    score_ob = ClassCleanScore.objects.get(id=score_id)

    context = {'score': score_ob}
    return render_ht(request, 'classclean/see_score.html', context)


def su_dl(request):
    """学生会检查结果日期列表"""
    dates = []
    for ob in ClassCleanScore.objects.all():
        date_str = datetime.strftime(ob.date_added, '%Y-%m-%d')
        dates.append(date_str)

    # 去重、排序、可哈希
    dates = list(set(dates))
    dates.sort(reverse=True)
    dates = tuple(dates)

    context = {'dates': dates}
    return render_ht(request, 'classclean/su_dl.html', context)


def su_day(request, date_str):
    """学生会检查结果一日页面"""
    # 字符串转化日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 取出当天全部对象
    scores = list(ClassCleanScore.objects.filter(date_added=date))

    # 排序，可哈希
    scores.sort(key=lambda x: x.cs)
    scores.sort(key=lambda x: x.grade_num)
    scores = tuple(scores)

    context = {'scores': scores, 'is_manager': request.user.username in DT.super_users}
    return render_ht(request, 'classclean/su_day.html', context)


def del_score(request, score_id):
    """删除学生会评分对象"""
    # 限制这项操作权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的对象及所属日期
    score = ClassCleanScore.objects.get(id=score_id)
    date = score.date_added
    date_str = datetime.strftime(date, '%Y-%m-%d')

    # 执行删除操作
    score.delete()

    # 重定向至当前日期页面
    return HttpResponseRedirect(reverse('classclean:su_day', args=[date_str]))


def pub_grade_list(request):
    """结果公示年级列表页面"""
    # 获取今天日期
    today = datetime.today()

    # 往前连续7天
    date_list = []
    for i in range(7):
        date = today - timedelta(days=i)
        date_str = datetime.strftime(date, '%Y-%m-%d')
        date_list.append(date_str)

    return render_ht(request, 'classclean/pub_grade_list.html', {'dates': tuple(date_list)})


def format_score(grade_num, date_str):
    """结构化数据"""
    # 日期字符串转化为日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 获取班级列表
    gcs = DT.get_gc(grade_num)

    # 初始化德育处评分字典
    dyc_score = dict(zip(gcs, [5.0] * 12))

    # 记录德育处评分
    for ob in ClassCleanRecord.objects.filter(grade_num=grade_num, date_added=date):
        dyc_score[ob.class_and_grade] -= ob.decrease / 10

    # 记录卫生部评分
    su_score = {}
    for gc in gcs:
        try:
            sus = ClassCleanScore.objects.filter(class_and_grade=gc, date_added=date)[0]
        except IndexError:
            score = '未反馈'
        else:
            score = sus.score

        su_score[gc] = score

    # 生成可哈希前端显示结构化数据
    scores = []
    for gc in gcs:
        # 计算总分
        if su_score[gc] == '未反馈':
            # 未反馈，默认满分
            s = dyc_score[gc] + 5
        else:
            s = dyc_score[gc] + float(su_score[gc])
        scores.append((gc, dyc_score[gc], su_score[gc], s))
    scores = tuple(scores)

    # 制定标题
    title = '{}年{}月{}日{}年级卫生检查结果公示'.format(
        date.year, date.month, date.day, DT.reverse_grade_num[grade_num])

    return scores, title


def pub(request, grade_num, date_str):
    """某年级某天结果公示"""
    # 获取数据结构和标题
    scores, title = format_score(grade_num, date_str)

    context = {'scores': scores, 'title': title, 'grade_num': grade_num, 'date_str': date_str}
    return render_ht(request, 'classclean/pub.html', context)


def pub_export(request, grade_num, date_str):
    """某天某年级结果公示导出"""
    # 获取数据结构和标题
    scores, title = format_score(grade_num, date_str)

    # 打开新的文件表格
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title
    st['A1'].font = Font(size=28, bold=True)
    st['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 合并标题行单元格
    st.merge_cells(range_string='A1:D1')

    # 写入表头
    st['A2'].value = '班级'
    st['B2'].value = '德育处评分'
    st['C2'].value = '卫生部评分'
    st['D2'].value = '总分'

    # 设置表头格式
    for c in ['A', 'B', 'C', 'D']:
        st[c + '2'].font = Font(size=20, bold=True)
        st[c + '2'].alignment = Alignment(horizontal='center', vertical='center')

    # 初始行
    row = 3

    # 写入数据
    for s in scores:
        st.cell(row=row, column=1).value = s[0]
        st.cell(row=row, column=2).value = s[1]
        st.cell(row=row, column=3).value = s[2]
        st.cell(row=row, column=4).value = s[3]

        # 下一行
        row += 1

    # 设置主体格式
    for r in range(3, row):
        for col in range(1, 5):
            st.cell(row=r, column=col).font = Font(size=20)
            st.cell(row=r, column=col).alignment = Alignment(horizontal='center', vertical='center')

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=4)

    # 设置列宽
    set_width_dict(st, {'A': 20, 'B': 20, 'C': 20, 'D': 20})

    # 设置行高
    st.row_dimensions[1].height = 100
    for ro in range(2, row):
        st.row_dimensions[ro].height = 40

    # 生成文件名
    grade_str = DT.reverse_grade_num[grade_num]
    date = datetime.strptime(date_str, '%Y-%m-%d')
    fn = '{}年级{}月{}日卫生.xlsx'.format(grade_str, date.month, date.day)

    # 输出
    return write_out(wb)


def yryb_upload(request, gc):
    """仪容仪表检查填报"""
    # 获取当前日期、班级
    date = datetime.today()
    date_str = '{}年{}月{}日'.format(date.year, date.month, date.day)
    grade = DT.reverse_grade_num[int(gc.split('-')[0])]
    cs = gc.split('-')[1]
    class_and_grade = grade + cs + '班'

    # 取得卫生部（'02'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('02'), **get_owners_pwd('01')}

    # 获取逻辑班级字符串
    lg = DT_OUT.logic_grade[grade]
    lgc = lg + DT.str_two(int(cs)) + '班'

    # 获取学生字典
    std = {}
    i = 0
    for st in models.get_students(lgc):
        if i % 5 == 4:
            etr = True
        else:
            etr = False
        std[i] = (st, etr, 'hair' + str(i), 'jewelry' + str(i))

        # 索引加1
        i += 1

    # 不存在的情况
    if int(gc.split('-')[0]) < 1 or int(gc.split('-')[0]) > 3:
        raise Http404
    if int(cs) < 1 or int(cs) > 12:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = OutLookRecordForm()
    else:
        # 对POST提交的数据作出处理
        form = OutLookRecordForm(request.POST)
        if form.is_valid():
            new_score = form.save(commit=False)

            # 设置年级、班级、日期（已默认为今天）
            new_score.grade = grade
            new_score.grade_num = int(gc.split('-')[0])
            new_score.cs = int(cs)
            new_score.class_and_grade = class_and_grade

            # 检验代号
            if new_score.owner not in opd.keys():
                return render_ht(request, 'classclean/yryb_upload.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '代号错误！',
                    'gc': gc,
                    'st_list': tuple(std.items()),
                })

            # 检验密码
            if new_score.pwd != opd[new_score.owner][0]:
                return render_ht(request, 'classclean/yryb_upload.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '检查码错误！',
                    'gc': gc,
                    'st_list': tuple(std.items()),
                })

            # 检查人字符串显示
            new_score.owner_show = new_score.owner + opd[new_score.owner][1]

            # TODO:记录烫发、染发名单
            hair_record_list = []
            for j in range(i):
                if request.POST.get(std[j][2], ''):
                    hair_record_list.append(std[j][0])
            new_score.update_hair_record(hair_record_list)

            # TODO:记录佩戴饰品名单
            jewelry_record_list = []
            for k in range(i):
                if request.POST.get(std[k][3], ''):
                    jewelry_record_list.append(std[k][0])
            new_score.update_jewelry_record(jewelry_record_list)

            # 清除之前添加的同班级、同日期记录，自动只保留最新
            before_score = OutLookRecord.objects.filter(
                date_added=date, class_and_grade=class_and_grade)
            if before_score:
                for ob in before_score:
                    ob.delete()

            # TODO:保存、重定向
            new_score.save()
            return HttpResponseRedirect(reverse('classclean:see_yryb_record',
                                                args=[new_score.id]))

    context = {
        'cag': class_and_grade,
        'date': date_str,
        'form': form,
        'err': '',
        'gc': gc,
        'st_list': tuple(std.items()),
    }
    return render_ht(request, 'classclean/yryb_upload.html', context)


def see_yryb_record(request, yryb_id):
    """显示仪容仪表检查记录结果"""
    # 取出要显示的记录
    record_ob = OutLookRecord.objects.get(id=yryb_id)

    context = {'record': record_ob}
    return render_ht(request, 'classclean/see_yryb_record.html', context)


def yryb_main(request):
    """仪容仪表检查结果主页"""
    dates = []
    for ob in OutLookRecord.objects.all():
        date_str = datetime.strftime(ob.date_added, '%Y-%m-%d')
        dates.append(date_str)

    # 去重、排序、可哈希
    dates = list(set(dates))
    dates.sort(reverse=True)
    dates = tuple(dates)

    context = {'dates': dates}
    return render_ht(request, 'classclean/yryb_main.html', context)


def yryb_day(request, grade_num, date_str):
    """查看某年级某一天仪容仪表检查不合格名单"""
    # 字符串转化日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 制定标题
    title = '{}{}年级仪容仪表检查不合格名单'.format(DT.reverse_grade_num[grade_num], date_str)

    # 取出该年级当天全部对象
    records = list(OutLookRecord.objects.filter(grade_num=grade_num, date_added=date))

    # 排序
    records.sort(key=lambda x: x.cs)

    context = {
        'records': tuple(records),
        'is_manager': request.user.username in DT.super_users,
        'title': title,
        'grade_num': grade_num,
        'date_str': date_str,
    }
    return render_ht(request, 'classclean/yryb_day.html', context)


def delete_yryb_record(request, yryb_id):
    """删除仪容仪表检查记录"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的记录
    yryb_record = OutLookRecord.objects.get(id=yryb_id)

    # 记录年级和日期
    grade_num = yryb_record.grade_num
    date_str = datetime.strftime(yryb_record.date_added, '%Y-%m-%d')

    # 执行删除操作
    yryb_record.delete()

    # 重定向至查看页
    return HttpResponseRedirect(reverse('classclean:yryb_day', args=[grade_num, date_str]))


def yryb_export(request, grade_num, date_str):
    """导出一个年级一天的仪容仪表检查记录"""
    # 字符串转化日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 制定标题
    title = '{}{}年级仪容仪表检查不合格名单'.format(DT.reverse_grade_num[grade_num], date_str)

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 第1行写入标题
    st['A1'].value = title

    # 标题格式
    st['A1'].font = Font(size=18, bold=True)
    st['A1'].alignment = Alignment(horizontal='center')
    st.merge_cells(range_string='A1:C1')

    # 第2行写入表头
    st['A2'].value = '班级'
    st['B2'].value = '烫发、染发学生名单'
    st['C3'].value = '佩戴饰品学生名单'

    # 表头加粗
    st['A2'].font = Font(bold=True)
    st['B2'].font = Font(bold=True)
    st['C2'].font = Font(bold=True)

    # 设置合适的列宽
    set_width_dict(st, {'A': 9, 'B': 36, 'C': 36})

    # 从第3行开始写入数据
    row = 3
    for record in OutLookRecord.objects.filter(grade_num=grade_num, date_added=date):
        st.cell(row=row, column=1).value = record.class_and_grade
        st.cell(row=row, column=2).value = record.hair_record
        st.cell(row=row, column=3).value = record.jewelry_record

        # 下一行
        row += 1

    # 输出
    return write_out(wb)
