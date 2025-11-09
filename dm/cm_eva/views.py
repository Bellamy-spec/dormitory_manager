from django.shortcuts import render
from django.contrib.auth.views import login_required
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from django.contrib.auth.models import Group
from datetime import datetime
from .tools import DataTool
from su_manage.models import get_owners_pwd
from .forms import CMScoreForm
from .models import CMScore
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.conf import settings


# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}班会评价系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """主页"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    context = {'title': '{}班会评价系统'.format(settings.USER_NAME)}
    return render_ht(request, 'cm_eva/index.html', context)


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        un = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=un, password=password)

        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到主页
            return HttpResponseRedirect(reverse('cm_eva:index'))
        else:
            return render_ht(request, 'cm_eva/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'cm_eva/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('cm_eva:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'cm_eva/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'cm_eva/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'cm_eva/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'cm_eva/set_pwd.html', {'err': ''})


def up_load(request, gc):
    """检查结果上报"""
    # 获取当前日期、班级
    date = datetime.today()
    date_str = '{}年{}月{}日'.format(date.year, date.month, date.day)
    grade = DT.reverse_grade_num[int(gc.split('-')[0])]
    cs = gc.split('-')[1]
    class_and_grade = grade + cs + '班'

    # 取得宣传部（'03'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('03'), **get_owners_pwd('01')}

    # 不存在的情况
    if int(gc.split('-')[0]) < 1 or int(gc.split('-')[0]) > 3:
        raise Http404
    if int(cs) < 1 or int(cs) > 12:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = CMScoreForm()
    else:
        # 对POST提交的数据作出处理
        form = CMScoreForm(data=request.POST)
        if form.is_valid():
            new_score = form.save(commit=False)

            # 设置年级、班级、日期（已默认为今天）
            new_score.grade = grade
            new_score.grade_num = int(gc.split('-')[0])
            new_score.cs = int(cs)
            new_score.class_and_grade = class_and_grade

            # 检验分数
            if new_score.score < 0 or new_score.score > 5:
                return render_ht(request, 'cm_eva/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '得分必须介于0.0~5.0之间',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检查代号
            if new_score.owner not in opd.keys():
                return render_ht(request, 'cm_eva/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '代号错误',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检查密码
            if new_score.pwd != opd[new_score.owner][0]:
                return render_ht(request, 'cm_eva/up_load.html', context={
                    'cag': class_and_grade,
                    'date': date_str,
                    'form': form,
                    'err': '检查码错误',
                    'gc': gc,
                    'descs': tuple(DT.desc.items()),
                })

            # 检查人字符串显示
            new_score.make_show()

            # 记录问题
            desc_list = []
            for d in DT.desc.keys():
                if request.POST.get(d, ''):
                    desc_list.append(DT.desc[d])
            new_score.update_desc(desc_list)

            # 自动保留最新
            before_score = CMScore.objects.filter(
                date_added=date, class_and_grade=class_and_grade)
            if before_score:
                for ob in before_score:
                    ob.delete()

            # TODO:保存，重定向查看
            new_score.save()
            return HttpResponseRedirect(reverse('cm_eva:see_score', args=[new_score.id]))

    context = {
        'cag': class_and_grade,
        'date': date_str,
        'form': form,
        'err': '',
        'gc': gc,
        'descs': tuple(DT.desc.items()),
    }
    return render_ht(request, 'cm_eva/up_load.html', context)


def see_score(request, score_id):
    """显示单次评价"""
    # 取出评分对象
    score_ob = CMScore.objects.get(id=score_id)

    context = {'score': score_ob}
    return render_ht(request, 'cm_eva/see_score.html', context)


def date_list(request):
    """日期列表"""
    # 取得所有日期，去重
    dl = []
    for score in CMScore.objects.all():
        date_str = datetime.strftime(score.date_added, '%Y-%m-%d')
        dl.append(date_str)
    dl = list(set(dl))

    context = {'dl': sorted(dl)}
    return render_ht(request, 'cm_eva/date_list.html', context)


def day(request, date_str):
    """查看某一天所有班级班会评分情况"""
    # 字符串转化为日期
    date = datetime.strptime(date_str, '%Y-%m-%d')

    # 制定标题
    title = '{}班会检查结果'.format(date_str)

    # 取出该日期下所有评分对象
    scores = list(CMScore.objects.filter(date_added=date))

    # 按年级、班级排序
    scores.sort(key=lambda x: x.cs)
    scores.sort(key=lambda x: x.grade_num)

    context = {'scores': scores, 'is_manager': request.user.username in DT.super_users, 'title': title}
    return render_ht(request, 'cm_eva/day.html', context)


def mt_main(request):
    """月总结主页"""
    # 初始化月份存放列表
    ym_list = []

    # 记录包含的月份
    for score in CMScore.objects.all():
        date = score.date_added
        ym = '{}-{}'.format(date.year, DT.str_two(date.month))
        ym_list.append(ym)

    # 去重
    ym_list = list(set(ym_list))

    context = {'yml': sorted(ym_list)}
    return render_ht(request, 'cm_eva/mt_main.html', context)


def get_mtd(grade_num, ym):
    """取得某个年级、某个月的月总结大数据列表"""
    # 该年级所有班级
    grade = DT.reverse_grade_num[grade_num]
    gc_list = []
    for cs in range(1, 13):
        gc_list.append(grade + str(cs) + '班')

    # 建立班级-班会评分对应字典，可变类型须逐项创建
    ys_dict = {}
    for gc in gc_list:
        ys_dict[gc] = {}

    # 记录所有可能的日期
    dates_list = []

    # 循环遍历该年级所有班会评价记录
    for score in CMScore.objects.filter(grade_num=grade_num):
        date = score.date_added
        score_ym = '{}-{}'.format(date.year, DT.str_two(date.month))

        # 日期转化为字符串
        date_str = datetime.strftime(date, '%Y-%m-%d')

        # 判断是否添加
        if score_ym == ym:
            gc = score.class_and_grade
            ys_dict[gc][date_str] = score.score

            # 添加日期
            dates_list.append(date_str)

    # 日期去重、排序
    dates_list = list(set(dates_list))
    dates_list.sort()

    # 计算平均分
    for sd in ys_dict.values():
        if sd:
            ave = np.mean(list(sd.values()))
            sd['mean'] = ave

    # TODO:数据结构二次解析
    # 表头
    head = ['日期'] + dates_list + ['平均']
    large_data_list = [head]

    # 构建各班数据
    for gc in gc_list:
        gs = [gc]

        # 按照固定日期顺序写入数据
        for ds in dates_list:
            # 尝试匹配班级、日期数据
            try:
                score = ys_dict[gc][ds]
            except KeyError:
                gs.append('未反馈')
            else:
                gs.append(score)

        # 尝试匹配平均值
        try:
            ave_score = ys_dict[gc]['mean']
        except KeyError:
            gs.append('未反馈')
        else:
            gs.append(ave_score)

        # 班级数据加入大数据
        large_data_list.append(gs)

    # 返回计算的大数据列表
    return large_data_list


def mt(request, grade_num, ym):
    """月总结"""
    # 获取大数据列表
    ldl = get_mtd(grade_num, ym)

    # 制定标题
    grade = DT.reverse_grade_num[grade_num]
    title = '{}年级{}班会月总结'.format(grade, ym)

    context = {'ldl': ldl, 'title': title, 'grade_num': grade_num, 'ym': ym}
    return render_ht(request, 'cm_eva/mt.html', context)


def export_mt(request, grade_num, ym):
    """导出班会月总结"""
    # 获取大数据列表
    ldl = get_mtd(grade_num, ym)

    # 制定标题
    grade = DT.reverse_grade_num[grade_num]
    title = '{}年级{}班会月总结'.format(grade, ym)

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title

    # 写入数据
    row = 2
    for line in ldl:
        col = 1
        for data in line:
            st.cell(row=row, column=col).value = data
            col += 1
        row += 1

    # 标题行合并单元格
    rs = 'A1:{}1'.format(get_column_letter(len(ldl[0])))
    st.merge_cells(range_string=rs)

    # 输出
    return write_out(wb)


def delete_score(request, score_id):
    """删除班会评分记录"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的对象
    score = CMScore.objects.get(id=score_id)

    # 获取日期字符串
    date_str = datetime.strftime(score.date_added, '%Y-%m-%d')

    # 执行删除操作
    score.delete()

    # 重定向至当日页面
    return HttpResponseRedirect(reverse('cm_eva:day', args=[date_str]))
