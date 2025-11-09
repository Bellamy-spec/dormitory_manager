from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.shortcuts import render
from django.urls import reverse
from .scores import models, froms, data
import datetime
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.views import login_required
from django.contrib.auth.models import Group, User
from openpyxl import Workbook, load_workbook
from io import BytesIO
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
import calendar
import numpy as np
import os
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
from matplotlib.font_manager import FontProperties
import time
import threading
import copy
import json
from django.conf import settings
from selenium.webdriver.common.by import By
from selenium.common.exceptions import ElementClickInterceptedException, NoSuchElementException
from urllib3.exceptions import MaxRetryError
import base64


# 实例化数据类
DT = data.Data()
# print(DT.dorm)


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def correct_id(id_number):
    """检查所给的身份证号格式是否合法"""
    # 特殊身份证号直接予以通过
    if id_number in DT.special_id:
        return True

    # 位数须为18
    if len(id_number) != 18:
        return False

    # 前17位须为数字
    if not id_number[:17].isdigit():
        return False

    # 最后一位须为数字或X
    if not (id_number[17].isdigit() or id_number[17] == 'X'):
        return False

    # 校验
    s = 0
    for i in range(18):
        # 取得每一位上的数字
        if id_number[i].isdigit():
            n = int(id_number[i])
        else:
            n = 10

        # 计算权
        w = 2 ** (17 - i) % 11

        # 累加
        s += n * w

    if s % 11 != 1:
        return False

    # 通过所有检查，格式正确
    return True


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def all_sep_id():
    """以列表形式返回所有新生身份证号"""
    id_list = []
    for ob in models.NewStudent.objects.all():
        if ob.id_number:
            id_list.append(ob.id_number)
    return id_list


def all_sep_gc(year):
    """以列表形式返回一年所有新生班级"""
    gc_list = []
    for ob in models.NewStudent.objects.filter(grade_year=year):
        gc_list.append(ob.gc)

    # 去重、排序、返回
    gc_list = list(set(gc_list))
    gc_list.sort()
    return gc_list


def all_sep_dorm(year):
    """以列表形式返回一年所有新生宿舍号"""
    dorm_list = []
    for ob in models.NewStudent.objects.filter(grade_year=year):
        dorm_list.append(ob.dorm)

    # 去重、排序、返回
    dorm_list = list(set(dorm_list))
    dorm_list.sort()
    return dorm_list


def remove_lay(filepath, t):
    """延时t秒删除文件"""
    time.sleep(t)
    try:
        os.remove(filepath)
    except FileNotFoundError:
        # 不用你来亲自动手删除了！
        pass


def get_gender(id_number):
    """根据身份证号返回性别"""
    # 取出倒数第二位数字值
    gender_key = int(id_number[-2])

    # 判断
    if gender_key // 2:
        gender = '男'
    else:
        gender = '女'

    return gender


def float_to_str(num):
    """优化浮点数的字符串表示：最大保留两位，自动去零"""
    if len(str(num).split('.')[1]) > 2:
        return str(round(num, 2))
    elif num % 1 == 0:
        return str(int(num))
    else:
        return str(num)


def index(request):
    week = datetime.datetime.now().weekday()
    if week < 5:
        img_filename = 'images/{}.jpg'.format(week)
    else:
        img_filename = 'images/{}.png'.format(week)
    context = {'hello': '{}宿舍管理系统'.format(settings.USER_NAME), 'weekday': img_filename,
               'is_manager': request.user.username in DT.manager}

    # # 测试
    # print(models.format_gc_students())

    return render_ht(request, 'index.html', context)


def index_out(request):
    """索引页"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    week = datetime.datetime.now().weekday()
    if week < 5:
        img_filename = 'images/{}.jpg'.format(week)
    else:
        img_filename = 'images/{}.png'.format(week)
    context = {'hello': '{}学生日常行为规范管理系统'.format(settings.USER_NAME), 'weekday': img_filename,
               'su': request.user.is_staff}
    return render_ht(request, 'index_out.html', context)


@login_required()
def all_records(request):
    """显示所有扣分记录"""
    records = models.Record.objects.all()
    context = {
        'records': records,
        'title': '所有记录',
        'user': request.user,
        'is_manager': request.user.username in DT.manager,
    }
    return render_ht(request, 'records.html', context)


@login_required()
def new_record(request):
    """新增记录"""
    # raise Http404
    if request.method != 'POST':
        # 未提交数据，创建一个新表单
        form = froms.RecordForm()
    else:
        # 对提交的数据进行处理
        form = froms.RecordForm(request.POST)
        if form.is_valid():
            new = form.save(commit=False)
            new.owner = request.user
            new.decrease = DT.decrease[new.tp]
            new.dormitory = request.POST.get('dormitory', '')

            try:
                new.gender = DT.teacher_gender[DT.get_manager(new.dormitory)]
                new.student, new.class_and_grade = models.get_st(new.dormitory, new.bed)
            except KeyError:
                raise Http404

            new.bed_area = dict(DT.ap)[new.bed]
            new.cs = DT.get_cs(new.class_and_grade)
            new.grade = DT.get_grade(new.class_and_grade)
            new.date_group = DT.get_date_group(datetime.datetime.now())
            y, m, d = new.date_group.year, new.date_group.month, new.date_group.day
            new.date_group_str = DT.get_date_group_str(new.date_group)
            new.save()
            return HttpResponseRedirect(reverse('records_by_date', args=[y, m, d]))

    # 当前登录的宿管老师负责的宿舍（需重新实例化）
    dorms = DT.dormitory_by_manager(request.user.username)

    context = {'form': form, 'dorms': dorms, 'empty_bed': models.empty_bed()}
    return render_ht(request, 'add.html', context)


@login_required()
def records_by_date(request, year, month, day):
    """显示某一天记录"""
    dt = datetime.date(year, month, day)
    records = models.Record.objects.filter(date_group=dt)

    # 排序
    records = list(records)
    records.sort(key=lambda x: x.cs, reverse=True)
    records.sort(key=lambda x: x.grade, reverse=True)

    # 化为可哈希数据类型
    records = tuple(records)

    # 制定标题
    title = '{}记录'.format(DT.get_date_group_str(dt))

    context = {
        'records': records,
        'title': title,
        'user': request.user,
        'is_manager': request.user.username in DT.manager,
    }
    return render_ht(request, 'records.html', context)


def show_dates(request, grade=None):
    """显示所有日期"""
    # 收集所有日期
    dates = []
    if grade is None:
        records = models.Record.objects.all()
    else:
        records = models.Record.objects.filter(grade=grade)
    for record in records:
        dt = record.date_group
        # 仅显示开始日期之后的日期
        if dt > DT.start_date:
            dates.append(dt)

    # 获取今天、昨天日期
    today = datetime.datetime.date(datetime.datetime.now(tz=DT.tz))
    yestoday = today - datetime.timedelta(days=1)

    # 去重，排序，可哈希
    dates = list(set(dates))
    dates.sort()
    date_delta = []
    for date in dates:
        date_str = datetime.datetime.strftime(date, '%Y-%m-%d')
        if date == today:
            date_to_show = '今日'
        elif date == yestoday:
            if datetime.datetime.now().hour >= 9:
                date_to_show = '昨日中午——今日早上'
            else:
                date_1 = date + datetime.timedelta(days=1)
                date_1_str = datetime.datetime.strftime(date_1, '%Y-%m-%d')
                date_to_show = '{}中午——{}早上'.format(date_str, date_1_str)
        else:
            date_1 = date + datetime.timedelta(days=1)
            date_1_str = datetime.datetime.strftime(date_1, '%Y-%m-%d')
            date_to_show = '{}中午——{}早上'.format(date_str, date_1_str)
        date_delta.append((date, date_to_show, date_str))
    date_delta = tuple(date_delta)

    context = {'dates': date_delta, 'grade': grade, 'has_grade': grade is not None}
    return render_ht(request, 'dates.html', context)


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到导航页
            return HttpResponseRedirect(reverse('index_out'))
        else:
            return render_ht(request, 'login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('index_out'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'set_pwd.html', {'err': ''})


@login_required()
def delete_record(request, record_id):
    """删除记录"""
    r = models.Record.objects.get(id=record_id)

    # 不能删除别人的记录（管理员除外）
    if r.owner != request.user and request.user.username not in DT.manager:
        raise Http404

    r.delete()
    return HttpResponseRedirect(reverse('records'))


def grades(request):
    """分年级"""
    return render_ht(request, 'grades.html', {'grades': DT.grade})


def show_records(request, grade, date_str, gender=None):
    """根据日期，年级查看"""
    # 取得指定年级、日期数据
    dt = datetime.datetime.date(datetime.datetime.strptime(date_str, '%Y-%m-%d'))
    if gender is None:
        records_all = models.Record.objects.filter(grade=grade, date_group=dt)
    else:
        records_all = models.Record.objects.filter(grade=grade, gender=gender,
                                                   date_group=dt)

    # 获取所有宿舍号
    records, records_dorm = [], []
    for obj in records_all:
        records.append(obj)

        # TODO:2025年六月上半月原宿舍号转新宿舍号
        if dt.year == 2025 and dt.month == 6:
            if dt.day < 17 and obj.dormitory in DT.get_old_new(2).keys():
                dor = DT.get_old_new(2)[obj.dormitory]
            elif dt.day < 19 and obj.dormitory in DT.get_old_new(1).keys():
                dor = DT.get_old_new(1)[obj.dormitory]
            else:
                dor = obj.dormitory
            records_dorm.append(dor)
        else:
            records_dorm.append(obj.dormitory)

    # 应在此处排序
    records.sort(key=lambda x: x.cs)
    records.sort(key=lambda x: x.tm)

    # 化为可哈希数据类型
    records = tuple(records)

    # 当天表扬宿舍
    praise_set = set(DT.get_dormitory_by_grade(grade, gender)) - set(records_dorm)
    praise_dict = {}
    for dorm in praise_set:
        praise_dict.setdefault(DT.get_class(dorm), [])
        praise_dict[DT.get_class(dorm)].append(dorm)

    # 表扬宿舍按班级排序
    praise_list = sorted(praise_dict.items(), key=lambda x: DT.get_cs(x[0]))

    # 转换数据类型
    praise = []
    for p in praise_list:
        praise.append((p[0], '，'.join(p[1])))
    praise = tuple(praise)      # 完成praise的生成

    # 制定标题
    date_1 = dt + datetime.timedelta(days=1)
    date_1_str = datetime.datetime.strftime(date_1, '%Y-%m-%d')
    grade_str = dict(DT.grade)[grade]
    if gender == 'boy':
        title = '{}年级男寝{}中午——{}早上记录'.format(grade_str, date_str, date_1_str)
    elif gender == 'girl':
        title = '{}年级女寝{}中午——{}早上记录'.format(grade_str, date_str, date_1_str)
    else:
        title = '{}年级{}中午——{}早上记录'.format(grade_str, date_str, date_1_str)

    # 已搬宿舍
    if DT.change_dorm:
        title += '(表扬宿舍以新宿舍号展示)'

    context = {
        'records': records,
        'title': title,
        'praise': praise,
        'grade': grade_str,
        'grade_num': grade,
        'date_str': date_str,
        'gender': gender,

        # 当天或开始日期之前不显示表扬宿舍
        'not_today': DT.start_date < dt < datetime.datetime.date(datetime.datetime.now(tz=DT.tz)),
    }
    return render_ht(request, 'show_records.html', context)


def export_excel(request, grade, date_str, gender=None):
    """导出一个年级一天的数据为Excel表格"""
    # 取得指定年级、日期数据
    dt = datetime.datetime.date(datetime.datetime.strptime(date_str, '%Y-%m-%d'))
    if gender is None:
        records_all = models.Record.objects.filter(grade=grade, date_group=dt)
    else:
        records_all = models.Record.objects.filter(grade=grade, gender=gender,
                                                   date_group=dt)

    # 筛选对应日期记录
    records, records_dorm = [], []
    for obj in records_all:
        records.append(obj)

        # TODO:2025年六月上半月原宿舍号转新宿舍号
        if dt.year == 2025 and dt.month == 6:
            if dt.day < 17 and obj.dormitory in DT.get_old_new(2).keys():
                dor = DT.get_old_new(2)[obj.dormitory]
            elif dt.day < 19 and obj.dormitory in DT.get_old_new(1).keys():
                dor = DT.get_old_new(1)[obj.dormitory]
            else:
                dor = obj.dormitory
            records_dorm.append(dor)
        else:
            records_dorm.append(obj.dormitory)

    # 应在此处排序
    records.sort(key=lambda x: x.cs)
    records.sort(key=lambda x: x.tm)

    # 创建表格
    wb = Workbook()
    st = wb.active

    # 设置列宽
    width_dict = {'A': 12, 'B': 16, 'C': 8, 'D': 11, 'E': 10, 'F': 29}
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]

    # 写入第一行
    ft1 = Font(size=18, bold=True)
    grade_str = dict(DT.grade)[grade]
    date_1 = dt + datetime.timedelta(days=1)
    date_1_str = datetime.datetime.strftime(date_1, '%Y-%m-%d')
    st['A1'].font = ft1
    if gender == 'boy':
        st['A1'] = '{}年级男寝{}中午——{}早上'.format(grade_str, date_str, date_1_str)
    elif gender == 'girl':
        st['A1'] = '{}年级女寝{}中午——{}早上'.format(grade_str, date_str, date_1_str)
    else:
        st['A1'] = '{}年级{}中午——{}早上'.format(grade_str, date_str, date_1_str)
    st.merge_cells(range_string='A1:F1')

    # 写入表头
    ft2 = Font(size=14, bold=True)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        st[col + '2'].font = ft2
    st['A2'] = '时间类型'
    st['B2'] = '所属班级'
    st['C2'] = '宿舍号'
    st['D2'] = '床铺'
    st['E2'] = '问题类型'
    st['F2'] = '具体情况'

    # # 设置主体部分字体
    # ft3 = Font(size=14)

    # 写入主体
    row = 3
    for record in records:
        # 设置字体大小
        for col in range(1, 7):
            st.cell(row=row, column=col).font = ft2

        st.cell(row=row, column=1).value = record.tm
        st.cell(row=row, column=2).value = record.class_and_grade
        st.cell(row=row, column=3).value = record.dormitory
        st.cell(row=row, column=4).value = record.bed_area
        st.cell(row=row, column=5).value = record.tp

        # E列设置自动换行
        st.cell(row=row, column=6).alignment = Alignment(wrapText=True)
        st.cell(row=row, column=6).value = record.reason
        row += 1

    # 非当天导出表扬部分
    if dt != datetime.datetime.date(datetime.datetime.now(tz=DT.tz)):
        # 写入表扬标题
        st.cell(row=row, column=1).font = ft2
        st.cell(row=row, column=1).value = '表扬'
        st.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

        # 当天表扬宿舍
        praise_set = set(DT.get_dormitory_by_grade(grade, gender)) - set(records_dorm)
        praise_dict = {}
        for dorm in praise_set:
            praise_dict.setdefault(DT.get_class(dorm), [])
            praise_dict[DT.get_class(dorm)].append(dorm)

        # 表扬宿舍按班级排序
        praise_list = sorted(praise_dict.items(), key=lambda x: DT.get_cs(x[0]))

        # 写入表扬主体
        for cd in praise_list:
            st.cell(row=row, column=1).font = ft2
            st.cell(row=row, column=1).value = cd[0]
            st.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            dr = '，'.join(cd[1])
            st.cell(row=row, column=3).font = ft2
            st.cell(row=row, column=3).value = dr
            st.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            row += 1

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=6)

    # 重设F列列宽
    st.column_dimensions['F'].width = 30

    # 设置页边距（单位：英寸）
    st.page_margins.top = 0.75
    st.page_margins.bottom = 0.75
    st.page_margins.left = 0.95
    st.page_margins.right = 0.1

    return write_out(wb)


def show_all_by_grade(request, grade):
    """显示一个年级全部记录"""
    records_all = models.Record.objects.filter(grade=grade).order_by('date_group')

    # 按班级、宿舍排序
    records = []
    for r in records_all:
        records.append(r)
    records.sort(key=lambda x: x.dormitory)
    records.sort(key=lambda x: x.cs)

    # 可哈希
    records = tuple(records)

    title = '{}年级全部记录'.format(dict(DT.grade)[grade])
    context = {'records': records, 'title': title, 'grade': grade}
    return render_ht(request, 'grade_all.html', context)


def export_grade_all(request, grade):
    """导出一个年级全部数据"""
    records_all = models.Record.objects.filter(grade=grade).order_by('date_group')

    # 按班级排序
    records = []
    for r in records_all:
        records.append(r)
    records.sort(key=lambda x: x.dormitory)
    records.sort(key=lambda x: x.cs)

    title = '{}年级全部记录'.format(dict(DT.grade)[grade])

    # 创建表格
    wb = Workbook()
    st = wb.active

    # 设置列宽
    width_dict = {'A': 14, 'B': 6, 'C': 32, 'D': 10, 'E': 8, 'F': 30}
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]

    # 写入第一行
    ft1 = Font(size=20, bold=True)
    st['A1'].font = ft1
    st['A1'] = title
    st.merge_cells(range_string='A1:F1')

    # 写入表头
    ft2 = Font(bold=True)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        st[col + '2'].font = ft2
    st['A2'] = '所属班级'
    st['B2'] = '宿舍号'
    st['C2'] = '日期'
    st['D2'] = '时间类型'
    st['E2'] = '问题类型'
    st['F2'] = '具体情况'

    # 写入主体
    row = 3
    for record in records:
        st.cell(row=row, column=1).value = record.class_and_grade
        st.cell(row=row, column=2).value = record.dormitory
        st.cell(row=row, column=3).value = record.date_group_str
        st.cell(row=row, column=4).value = record.tm
        st.cell(row=row, column=5).value = record.tp

        # F列设置自动换行
        st.cell(row=row, column=6).alignment = Alignment(wrapText=True)
        st.cell(row=row, column=6).value = record.reason
        row += 1

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=6)

    return write_out(wb)


@login_required()
def change_record(request, record_id):
    """修改记录"""
    # 取出对应记录
    r = models.Record.objects.get(id=record_id)

    # 不能修改别人的记录（管理员除外）
    if r.owner != request.user and request.user.username not in DT.manager:
        raise Http404

    if request.method != 'POST':
        # 初次请求，使用当前记录填充表单
        form = froms.RecordForm(instance=r)
    else:
        # 对于POST提交的数据作出处理
        form = froms.RecordForm(instance=r, data=request.POST)
        if form.is_valid():
            new = form.save(commit=False)

            # 修改原始数据
            r.tm = new.tm
            r.tp = new.tp
            r.reason = new.reason
            r.dormitory = request.POST.get('dormitory', '')
            r.date_group_str = request.POST.get('date_group', '')
            r.date_group = DT.date_group_from_str(r.date_group_str)

            # 修改关联数据
            r.decrease = DT.decrease[r.tp]
            r.gender = DT.teacher_gender[DT.get_manager(r.dormitory)]
            r.student, r.class_and_grade = models.get_st(r.dormitory, r.bed)
            r.bed_area = dict(DT.ap)[r.bed]
            r.cs = DT.get_cs(r.class_and_grade)
            r.grade = DT.get_grade(r.class_and_grade)
            r.save()
            return HttpResponseRedirect(reverse('records_by_date', args=[
                r.date_group.year, r.date_group.month, r.date_group.day]))

    # 当前登录的宿管老师负责的宿舍
    dorms = DT.dormitory_by_manager(request.user.username)

    # 获取七天之内日期列表
    dt = DT.get_date_group(r.date_added)
    dates = []
    for i in range(7):
        dtn = dt - datetime.timedelta(days=i)
        dates.append(DT.get_date_group_str(dtn))

    context = {'form': form, 'dorms': dorms, 'dates': dates, 'record': r}
    return render_ht(request, 'change.html', context)


def month_together(request):
    """月总结主页"""
    # 确定有哪些月份，即应截止到哪个月份，先取得当前西一区时间
    nd = datetime.datetime.now(tz=DT.tz)

    # 取得当前一天后西一区日期
    ndt = datetime.datetime.date(nd) + datetime.timedelta(days=1)

    # 做到此月即可（不超过此年月）
    ym_last = (ndt.year, ndt.month)

    # 初始年月固定为2024年8月
    ym = [2024, 8]
    ym_list = []

    # 年份小于或当年份等于时月份小于
    while ym[0] < ym_last[0] or (ym[0] == ym_last[0] and ym[1] < ym_last[1]):
        ym_list.append(('{}年{}月'.format(*ym), '{}-{}'.format(*ym)))

        # 年月加1
        if ym[1] < 12:
            ym[1] += 1
        else:
            ym[0] += 1
            ym[1] = 1

    # 元组化
    ym_tuple = tuple(ym_list)
    ym_now = ('{}年{}月（不完全）'.format(*ym), '{}-{}'.format(*ym))

    context = {
        'ym_before': ym_tuple,
        'ym_now': ym_now,
        'title': '选择年级/月份',
        'grades': DT.grade,
    }
    return render_ht(request, 'month_together.html', context)


def fs(grade, start_date, total_days):
    """以字典形式返回一段时间的总结"""
    # 初始化字典
    key_list = DT.get_dormitory_by_grade1(grade)
    key_list.sort(key=lambda x: DT.get_cs(DT.get_class(x)))

    # 列表在python中为可变类型，此处必须逐项创建列表副本
    value_dict = {}
    score_list = [0] * 9
    for t in DT.tp:
        value_dict[t[0]] = score_list[:]
    value_dict['表扬'] = [0]

    # 字典在python中为可变类型，此处必须逐项创建字典副本
    monthtogether_dict = {}
    for d in key_list:
        monthtogether_dict[d] = copy.deepcopy(value_dict)

    # 开始循环遍历这个月的每一天
    for i in range(total_days):
        # 计算出每天的西一区日期
        date = start_date + datetime.timedelta(days=i)

        # 取出当天该年级所有记录
        records = models.Record.objects.filter(grade=grade, date_group=date)

        # 初始化表扬宿舍名单
        praise_list = DT.get_dormitory_by_grade1(grade)

        # 循环遍历每条记录，开始统计
        for record in records:
            # # 118宿舍即为原253宿舍
            # if record.dormitory == '253':
            #     dor = '118'
            # else:
            #     dor = record.dormitory
            dor, bed = record.dormitory, record.bed

            # TODO:原宿舍号转新宿舍号（必须限时间）
            # if record.date_group < datetime.date(2024, 6, 25):
            #     if dor in DT.change_dorm_2.keys():
            #         dor = DT.change_dorm_2[dor]
            #     elif dor in DT.change_dorm_1.keys():
            #         dor = DT.change_dorm_1[dor]
            if record.date_group < datetime.date(2025, 6, 17) and grade == 2:
                dor = DT.get_old_new(2)[dor]
            elif record.date_group < datetime.date(2025, 6, 19) and grade == 1:
                dor = DT.get_old_new(1)[dor]

            monthtogether_dict[dor][record.tp][bed] += 1
            if dor in praise_list:
                praise_list.remove(dor)

        # 仅在当天记录非空的前提下记录表扬
        if records:
            for dorm in praise_list:
                monthtogether_dict[dorm]['表扬'][0] += 1

    return monthtogether_dict


def config_dates(month_str):
    """返回月总结开始日期和天数"""
    # 对日期的初步处理，得出这个月的总天数
    month_str_list = month_str.split('-')
    year = int(month_str_list[0])
    month = int(month_str_list[1])
    total_days = calendar.monthrange(year, month)[1]

    # 起始日期，为上个月最后一天
    start_date = datetime.date(year, month, 1) - datetime.timedelta(days=1)

    return start_date, total_days


def class_summary(mtd, simple_count=False):
    """根据宿舍月总结字典返回班级卫生总分、纪律总分及未关灯次数，以列表形式"""
    # 初始化三个列表
    if simple_count:
        ws = [0] * 12
        jl = [0] * 12
    else:
        ws = [5] * 12
        jl = [5] * 12
    jn = [0] * 12

    # 循环遍历宿舍月总结字典
    for dorm, scores in mtd.items():
        for tp, sl in scores.items():
            # 不处理表扬
            if tp != '表扬':
                for j in range(len(sl)):
                    # 获取班级名称
                    class_and_grade = models.get_st(dorm, j)[1]

                    # 获取班级数字列表及班级个数
                    cs = DT.num_regex.findall(class_and_grade)
                    total = len(cs)
                    for i in range(total):
                        cs[i] = int(cs[i])

                    for c in cs:
                        if simple_count:
                            # 简单计数模式
                            if tp == '卫生':
                                # 卫生扣分
                                ws[c - 1] += sl[j] / total
                            elif tp == '未关灯':
                                # 未关灯次数
                                jn[c - 1] += sl[j] / total
                            else:
                                # 纪律扣分
                                jl[c - 1] += sl[j] / total
                        else:
                            if tp == '卫生':
                                # 卫生扣分
                                ws[c - 1] -= sl[j] / 30 / total
                            elif tp == '未关灯':
                                # 未关灯次数
                                jn[c - 1] += sl[j] / total
                            else:
                                # 纪律扣分
                                jl[c - 1] -= sl[j] / 20 / total

    # 依次返回三个列表
    return ws, jl, jn


def dorm_summary(mtd, cs):
    """根据月总结字典返回一个班级各宿舍扣分、表扬次数字典"""
    # 初始化两个字典
    kf, by = {}, {}

    # 循环遍历宿舍月总结字典
    for dorm, scores in mtd.items():
        # 初步确定可能的班级
        class_and_grade = models.get_st(dorm, 0)[1]
        css = DT.num_regex.findall(class_and_grade)
        for i in range(len(css)):
            css[i] = int(css[i])
        if cs not in css:
            # 省去不必要的操作
            continue

        for tp, sl in scores.items():
            if tp == '表扬':
                # 先处理表扬
                by.setdefault(dorm, 0)
                by[dorm] += sl[0] / len(css)
            else:
                # 重新确定班级
                for j in range(len(sl)):
                    # 获取班级名称
                    cag = models.get_st(dorm, j)[1]
                    cl = DT.num_regex.findall(cag)
                    for k in range(len(cl)):
                        cl[k] = int(cl[k])

                    # 班级个数
                    total = len(cl)

                    # 记录次数
                    if cs in cl:
                        kf.setdefault(dorm, 0)
                        kf[dorm] += sl[j] / total

    # 返回两个字典
    return kf, by


def student_summary(mtd, cs):
    """根据月总结字典返回学生记录次数字典"""
    # 初始化字典
    st_dict = {}

    # 循环遍历宿舍月总结字典
    for dorm, scores in mtd.items():
        # 初步确定可能的班级
        class_and_grade = models.get_st(dorm, 0)[1]
        css = DT.num_regex.findall(class_and_grade)
        for i in range(len(css)):
            css[i] = int(css[i])
        if cs not in css:
            # 省去不必要的操作
            continue

        for tp, sl in scores.items():
            if tp != '表扬':
                for j in range(len(sl)):
                    # 仅记录非零
                    if sl[j]:
                        # 确定学生
                        if j == 0:
                            student = '公共区域'
                        else:
                            student = models.get_st(dorm, j)[0]
                            if not student:
                                student = '公共区域'

                        # 重新确定班级
                        cag = models.get_st(dorm, j)[1]
                        cl = DT.num_regex.findall(cag)
                        for k in range(len(cl)):
                            cl[k] = int(cl[k])

                        # 班级个数
                        total = len(cl)

                        # 记录次数
                        if cs in cl:
                            st_dict.setdefault(student, 0)
                            st_dict[student] += sl[j] / total

    # 排序返回
    return dict(sorted(st_dict.items(), key=lambda x: x[1]))


def make_title(grade, month_str=None, date_info=None):
    """判断总结类型并制定总结标题"""
    if month_str is not None:
        title = '{}年级{}宿舍月总结'.format(DT.grade[grade - 1][1], month_str)
    elif date_info is not None:
        start_date = date_info[0]
        end_date = date_info[0] + datetime.timedelta(days=date_info[1] - 1)
        title = '{}年级{}至{}宿舍总结'.format(DT.grade[grade - 1][1], start_date, end_date)
    else:
        title = ''

    # 已搬宿舍
    if DT.change_dorm:
        title += '(已表示为新宿舍号)'

    return title


def export_fs(request, grade, start_date, total_days, title=''):
    """导出一个年级自定义时间段总结"""
    # # 毕业提示
    # if DT.change_dorm and grade == 3:
    #     return HttpResponse('高三年级已毕业')

    # 无标题参数，为直接外部访问情况
    if not title:
        # 字符串转日期
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

        # 获取标题
        title = make_title(grade, date_info=(start_date, total_days))

    # 取得总结情况字典
    mtd = fs(grade, start_date, total_days)

    # 新建工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    ft1 = Font(size=20, bold=True)
    st['A1'].font = ft1
    st['A1'].value = title

    # 写入首行
    ft2 = Font(size=14, bold=True)
    st['A2'].value = '班级'
    st['B2'].value = '宿舍号'
    st['A2'].font = ft2
    st['B2'].font = ft2
    c = 3
    for ite in list(mtd.values())[0].keys():
        st.cell(row=2, column=c).value = ite
        st.cell(row=2, column=c).font = ft2
        c += 1

    # 标题行合并单元格
    st.merge_cells(start_row=1, start_column=1, end_row=1, end_column=c - 1)

    # 设置班级列列宽
    st.column_dimensions['A'].width = 20

    # 设置主题部分字体
    ft3 = Font(size=14)

    # 写入每一项
    row = 3
    for dorm, score_dict in mtd.items():
        # 写入班级
        st.cell(row=row, column=1).value = DT.get_class(dorm)

        # 写入宿舍号
        st.cell(row=row, column=2).value = dorm

        # 写入各项统计结果
        c = 3
        for score in score_dict.values():
            st.cell(row=row, column=c).value = sum(score)
            c += 1

        # 设置字体大小
        for col in range(1, c):
            st.cell(row=row, column=col).font = ft3

        row += 1

    # 添加边框
    add_border(st, start_row=1, start_column=1, end_row=row - 1, end_column=c - 1)

    # 写入班级总分
    row += 1
    down_head = row
    st['A' + str(row)].value = '班级汇总结果'
    st.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    st['A' + str(row)].value = '班级'
    st['B' + str(row)].value = '卫生总分'
    st['C' + str(row)].value = '纪律总分'
    st['D' + str(row)].value = '未关灯次数'
    row += 1

    # 取得班级数据
    ws, jl, jn = class_summary(mtd)

    # 循环遍历每个班级并写入
    for c_num in range(12):
        st['A' + str(row)].value = str(c_num + 1) + '班'
        st['B' + str(row)].value = ws[c_num]
        st['C' + str(row)].value = jl[c_num]
        st['D' + str(row)].value = jn[c_num]
        row += 1

    # 再次添加边框
    add_border(st, start_row=down_head, start_column=1, end_row=row - 1, end_column=4)

    # 再次设置字体
    for ro in range(down_head, row):
        for co in range(1, 5):
            st.cell(row=ro, column=co).font = ft3

    return write_out(wb)


def export_grade_month(request, grade, month_str):
    """导出一个年级一个月的月总结"""
    # if grade == 3:
    #     # 暂不显示已毕业年级的月总结
    #     return HttpResponse('已毕业')

    # 制定标题
    title = make_title(grade, month_str=month_str)

    # 解析日期
    date_info = config_dates(month_str)

    return export_fs(request, grade, date_info[0], date_info[1], title=title)


def show_fs(request, grade, start_date, total_days, title=''):
    """查看一个年级自定义时间段总结"""
    # # 毕业提示
    # if DT.change_dorm and grade == 3:
    #     return HttpResponse('高三年级已毕业')

    # 无标题参数，为直接外部访问情况
    if not title:
        # 字符串转日期
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

        # 获取标题
        title = make_title(grade, date_info=(start_date, total_days))

    # 取得总结情况字典
    mtd = fs(grade, start_date, total_days)

    # 构造表头
    table_head = ['班级', '宿舍号']
    for ite in list(mtd.values())[0].keys():
        table_head.append(ite)
    table_head = tuple(table_head)

    # 构造便于网页模版调用的可哈希数据类型
    mt_list = []
    for dorm, scores in mtd.items():
        mt_item = [DT.get_class(dorm), dorm]
        for score in scores.values():
            mt_item.append(sum(score))
        mt_list.append(tuple(mt_item))
    mt_tuple = tuple(mt_list)

    context = {'title': title, 'table_head': table_head, 'mt': mt_tuple}
    return render_ht(request, 'month_summary.html', context)


def show_grade_month(request, grade, month_str):
    """查看一个年级一个月的月总结"""
    # if grade == 3:
    #     # 暂不显示已毕业年级的月总结
    #     return HttpResponse('已毕业')

    # 解析日期
    date_info = config_dates(month_str)

    # 制定标题
    title = make_title(grade, month_str=month_str)

    return show_fs(request, grade, date_info[0], date_info[1], title=title)


def make_visual_title(grade_str, month_str=None, date_info=None):
    """制定统计图表标题"""
    if month_str is not None:
        # 处理月份
        ym = month_str.split('-')
        ym_str = '{}年{}月'.format(*ym)

        # 制定标题
        title = '{}{}年级宿舍月总结统计图表'.format(ym_str, grade_str)
    elif date_info is not None:
        # 开始日期和结束日期
        start_date = date_info[0]
        end_date = date_info[0] + datetime.timedelta(days=date_info[1] - 1)

        # 制定标题
        title = '{}至{}{}年级宿舍总结统计图表'.format(start_date, end_date, grade_str)
    else:
        title = ''

    # 已搬宿舍
    if DT.change_dorm:
        title += '(已表示为新宿舍号)'

    return title


def visual_fs(request, grade, start_date, total_days, month_info=''):
    """一个年级自定义时间段总结图表"""
    # # 毕业提示
    # if DT.change_dorm and grade == 3:
    #     return HttpResponse('高三年级已毕业')

    # 处理年级
    grade_str = DT.grade_dict[grade]

    # 无月份信息参数，为直接外部访问情况
    if not month_info:
        # 字符串转日期
        start_date_str = start_date
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

        # 获取标题
        title = make_visual_title(grade_str, date_info=(start_date, total_days))

        # 可变链接格式
        month_info = '{}/{}'.format(start_date_str, total_days)
    else:
        # 获取标题
        title = make_visual_title(grade_str, month_str=month_info)

    # 取得总结情况字典
    mtd = fs(grade, start_date, total_days)
    ws, jl, jn = class_summary(mtd, simple_count=True)

    # 制作数据标签
    labels, cv_list = [], []
    for cs in range(1, 13):
        labels.append(grade_str + str(cs) + '班')
        cv = '{}-{}'.format(grade, cs)
        cv_list.append(cv)

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 设置可显示的中文字体
    font = FontProperties(fname='media/STKAITI.TTF', size=11)
    title_font = FontProperties(fname='media/STKAITI.TTF', size=16)

    # 设置柱子的宽度
    bar_width = 0.8

    # 设置x轴的位置
    x_1 = [x * 3 for x in list(range(12))]
    x_2 = [x + bar_width for x in x_1]
    x_3 = [x + bar_width for x in x_2]

    # 清空画布
    plt.cla()

    # 设置画布尺寸
    plt.figure(figsize=(12, 6))

    # 调整图形在画布中的位置
    plt.subplots_adjust(left=0.1, bottom=0.1, top=0.9, right=0.82)

    # 依次绘制三组数据
    plt.bar(x_1, ws, width=bar_width, color='blue', label='卫生扣分次数', alpha=0.7)
    plt.bar(x_2, jl, width=bar_width, color='yellow', label='纪律扣分次数', alpha=0.7)
    plt.bar(x_3, jn, width=bar_width, color='green', label='节能扣分次数', alpha=0.7)

    # 显示数值
    for j in range(12):
        # 优化整数与小数显示
        t1 = float_to_str(ws[j])
        t2 = float_to_str(jl[j])
        t3 = float_to_str(jn[j])

        # 数值文本显示
        plt.text(x_1[j], ws[j], t1, ha='center', va='bottom', fontsize=6)
        plt.text(x_2[j], jl[j], t2, ha='center', va='bottom', fontsize=6)
        plt.text(x_3[j], jn[j], t3, ha='center', va='bottom', fontsize=6)

    # 设置x轴的刻度标签
    plt.xticks(x_2, labels, fontproperties=font)

    # 设置标题
    plt.title(title, fontproperties=title_font)

    # 添加图例
    plt.legend(prop=font, bbox_to_anchor=(1.2, 1))

    # 保存图像
    fn = 'dorm_mtd_{}_{}_{}.png'.format(grade, start_date, total_days)
    plt.savefig(os.path.join('media', 'dorm_mtd', fn))

    # 生成班级选项
    cs_options = tuple(zip(cv_list, labels))

    # 前端显示保存的图像
    context = {'fn': fn, 'cs_options': cs_options, 'month': month_info}
    return render_ht(request, 'visual_grade_month.html', context)


def visual_grade_month(request, grade, month_str):
    """一个年级一个月的月总结图表"""
    # 解析日期
    date_info = config_dates(month_str)

    return visual_fs(request, grade, date_info[0], date_info[1], month_info=month_str)


def make_vc_title(gc, month_str=None, date_info=None):
    """制定班级统计图表标题"""
    if month_str is not None:
        ym = '{}年{}月'.format(*month_str.split('-'))
        title = '{}{}宿舍统计图表'.format(gc, ym)
    elif date_info is not None:
        start_date = date_info[0]
        end_date = date_info[0] + datetime.timedelta(days=date_info[1] - 1)
        title = '{}{}至{}宿舍统计图表'.format(gc, start_date, end_date)
    else:
        title = ''

    # 已搬宿舍
    if DT.change_dorm:
        title += '(已表示为新宿舍号)'

    return title


def visual_class_fs(request, gc_str, start_date, total_days, month_info=''):
    """查看自定义时间段班级大数据统计图表"""
    # 处理班级、年级信息
    grade = int(gc_str.split('-')[0])
    cs = int(gc_str.split('-')[1])
    gc = DT.grade_dict[grade] + str(cs) + '班'

    # 无月份信息参数，为直接外部访问情况
    if not month_info:
        # 字符串转日期
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()

        # 获取标题
        title = make_vc_title(gc, date_info=(start_date, total_days))

        # 时间标题
        end_date = start_date + datetime.timedelta(days=total_days - 1)
        time_title = '{}至{}'.format(start_date, end_date)
    else:
        # 月总结标题
        title = make_vc_title(gc, month_str=month_info)

        # 时间标题
        time_title = '{}年{}月'.format(*month_info.split('-'))

    # 取得总结字典
    mtd = fs(grade, start_date, total_days)

    # 处理总结字典
    kf, by = dorm_summary(mtd, cs)

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 字体设置
    font = FontProperties(fname='media/STKAITI.TTF', size=14)
    title_font = FontProperties(fname='media/STKAITI.TTF', size=18)

    # 设置柱子的宽度
    bar_width = 0.5

    # 清空画布
    plt.cla()

    # 调整图形在画布中的位置
    plt.subplots_adjust(left=0.05, bottom=0.1, top=0.9, right=0.95)

    # 画扣分次数统计图
    plt.bar(list(kf.keys()), list(kf.values()), width=bar_width, color='red', alpha=0.7)
    plt.title('{}{}宿舍扣分次数统计'.format(gc, time_title), fontproperties=title_font)

    # 设置x轴的刻度标签和名称
    plt.xticks(list(kf.keys()), list(kf.keys()), fontproperties=font)
    plt.xlabel('宿舍', fontproperties=font)

    # 显示数值
    for dorm1, num1 in kf.items():
        # 优化整数与小数显示
        t1 = float_to_str(num1)
        plt.text(dorm1, num1, t1, ha='center', va='bottom', fontsize=10)

    # 保存图像
    fn1 = 'kf_{}_{}_{}.png'.format(gc_str, start_date, total_days)
    plt.savefig(os.path.join('media', 'dorm_mtd', fn1))

    # 清空画布
    plt.cla()

    # 调整图形在画布中的位置
    plt.subplots_adjust(left=0.05, bottom=0.1, top=0.9, right=0.95)

    # 画表扬次数统计图
    plt.bar(list(by.keys()), list(by.values()), width=bar_width, color='green', alpha=0.7)
    plt.title('{}{}宿舍表扬次数统计'.format(gc, time_title), fontproperties=title_font)

    # 设置x轴的刻度标签和名称
    plt.xticks(list(by.keys()), list(by.keys()), fontproperties=font)
    plt.xlabel('宿舍', fontproperties=font)

    # 显示数值
    for dorm2, num2 in by.items():
        # 优化整数与小数显示
        t2 = float_to_str(num2)
        plt.text(dorm2, num2, t2, ha='center', va='bottom', fontsize=10)

    # 保存图象
    fn2 = 'by_{}_{}_{}.png'.format(gc_str, start_date, total_days)
    plt.savefig(os.path.join('media', 'dorm_mtd', fn2))

    if grade in DT.nst_grade:
        # 具体到学生的图表
        st_dict = student_summary(mtd, cs)

        # 清空画布
        plt.cla()

        # 设置画布大小
        fig_height = 2.5 + 0.25 * len(st_dict)
        plt.figure(figsize=(12, fig_height))

        # 调整图形在画布中的位置
        plt.subplots_adjust(left=0.15, bottom=0.1, top=0.9, right=0.95)

        # 绘制横向条形图
        plt.barh(list(st_dict.keys()), list(st_dict.values()), height=0.5)
        plt.title('学生扣分次数统计', fontproperties=title_font)
        plt.yticks(list(st_dict.keys()), list(st_dict.keys()), fontproperties=font)

        # 显示数值
        for st, num3 in st_dict.items():
            # 优化整数与小数显示
            t3 = float_to_str(num3)
            plt.text(num3, st, t3, va='center', ha='left', fontsize=12)

        # 保存图象
        fn3 = 'st_{}_{}_{}.png'.format(gc_str, start_date, total_days)
        plt.savefig(os.path.join('media', 'dorm_mtd', fn3))
    else:
        fn3 = ''

    context = {
        'title': title,
        'grade': grade,
        'fn1': fn1,
        'fn2': fn2,
        'fn3': fn3,
    }
    return render_ht(request, 'visual_class.html', context)


def visual_class(request, gc_str, month_str):
    """查看班级大数据统计情况图表"""
    # 解析日期信息
    date_info = config_dates(month_str)

    return visual_class_fs(request, gc_str, date_info[0], date_info[1], month_info=month_str)


def advice(request):
    """满意度调查主页"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    context = {'title': '{}宿舍管理满意度调查'.format(settings.USER_NAME)}
    return render_ht(request, 'advice.html', context)


def all_investigation():
    """返回所有已存在的调查学年"""
    school_years = []
    for ob in models.Investigation.objects.all():
        school_years.append(ob.school_year)
    return school_years


def add_investigation(request):
    """发布调查"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = froms.InvestigationForm()
    else:
        # 对POST提交的数据作出处理
        form = froms.InvestigationForm(request.POST)
        if form.is_valid():
            new_investigation = form.save(commit=False)

            # 防止重复发布
            sy = new_investigation.school_year
            if sy in all_investigation():
                return render_ht(request, 'add_investigation.html', context={
                    'form': form,
                    'err': '已存在{}的调查任务，请勿重复发布'.format(sy),
                })

            # 保存
            new_investigation.save()

            # 重定向至调查管理页
            return HttpResponseRedirect(reverse('inv_manage', args=[new_investigation.id]))

    context = {'form': form, 'err': ''}
    return render_ht(request, 'add_investigation.html', context)


def syl(request):
    """学年选择页面"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 获取所有调查对象
    all_invest = models.Investigation.objects.all()

    context = {'invests': all_invest}
    return render_ht(request, 'syl.html', context)


def inv_manage(request, investigation_id):
    """管理调查对象"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出调查对象
    inv = models.Investigation.objects.get(id=investigation_id)

    # 按钮名称
    if inv.active:
        btn_name = '停止任务'
    else:
        btn_name = '开始任务'

    # 制定标题
    title = '{}宿舍满意度调查'.format(inv.school_year)

    # 链接当中包含的年份字符串
    start_year = inv.school_year[:4]
    end_year = inv.school_year[5:9]
    link_year = '{}-{}'.format(start_year, end_year)

    # 生成问卷填写链接
    if os.name == 'nt':
        link = 'http://127.0.0.1:8000/do_paper/{}'.format(link_year)
    else:
        link = 'http://zz106gz.cn/do_paper/{}'.format(link_year)

    context = {'inv': inv, 'title': title, 'btn_name': btn_name, 'link': link}
    return render_ht(request, 'inv_manage.html', context)


def delete_inv(request, investigation_id):
    """删除调查对象"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出要删除的对象
    inv = models.Investigation.objects.get(id=investigation_id)

    # 执行删除操作
    inv.delete()

    # 重定向至年份选择页
    return HttpResponseRedirect(reverse('syl'))


def change_active(request, investigation_id):
    """更改调查状态"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出要更改的对象
    inv = models.Investigation.objects.get(id=investigation_id)

    # 更改状态
    inv.active = not inv.active
    inv.save()

    # 重定向至调查管理页
    return HttpResponseRedirect(reverse('inv_manage', args=[investigation_id]))


def get_paper_total(inv, dorm):
    """获取某个宿舍已收集的问卷数量"""
    return len(models.Paper.objects.filter(inv=inv, dorm=dorm))


def do_paper(request, school_year):
    """填写问卷"""
    # 生成学年
    sy = '{}~{}学年'.format(*school_year.split('-'))

    # 尝试匹配系统中已有的、活跃状态的调查对象
    try:
        inv = models.Investigation.objects.filter(school_year=sy, active=True)[0]
    except IndexError:
        # 未匹配到
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = froms.PaperForm()
    else:
        # 对POST提交的数据作出处理
        form = froms.PaperForm(request.POST)
        if form.is_valid():
            new_paper = form.save(commit=False)

            # 匹配宿管老师
            new_paper.teacher = DT.get_manager(new_paper.dorm)

            # 检测性别与宿舍号是否匹配
            if new_paper.gender != DT.gender_ec[DT.teacher_gender[new_paper.teacher]]:
                return render_ht(request, 'do_paper.html', context={
                    'inv': inv,
                    'inv_str': school_year,
                    'form': form,
                    'grade_class': DT.get_grade_class(),
                    'class_dorm': DT.get_class_dorm1(),
                    'class_dorm_boy': DT.get_class_dorm1(gender='boy'),
                    'class_dorm_girl': DT.get_class_dorm1(gender='girl'),
                    'err': '性别与宿舍号不匹配！',
                })

            # 检测是否超过宿舍上限
            if get_paper_total(inv, new_paper.dorm) >= 8:
                return render_ht(request, 'do_paper.html', context={
                    'inv': inv,
                    'inv_str': school_year,
                    'form': form,
                    'grade_class': DT.get_grade_class(),
                    'class_dorm': DT.get_class_dorm1(),
                    'class_dorm_boy': DT.get_class_dorm1(gender='boy'),
                    'class_dorm_girl': DT.get_class_dorm1(gender='girl'),
                    'err': '{}宿舍填写问卷数量已达上限'.format(new_paper.dorm),
                })

            # 关联调查对象
            new_paper.inv = inv
            new_paper.num = inv.total + 1
            inv.total += 1
            inv.save()

            # 生成问卷编号
            new_paper.make_num_name()

            # 保存
            new_paper.save()
            return HttpResponse('感谢您参与此次调查！')

    context = {
        'inv': inv,
        'inv_str': school_year,
        'form': form,
        'grade_class': DT.get_grade_class(),
        'class_dorm': DT.get_class_dorm1(),
        'class_dorm_boy': DT.get_class_dorm1(gender='boy'),
        'class_dorm_girl': DT.get_class_dorm1(gender='girl'),
        'err': '',
    }
    return render_ht(request, 'do_paper.html', context)


def paper_manage(request, inv_id):
    """管理问卷"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出调查对象及所有下属的问卷对象
    inv = models.Investigation.objects.get(id=inv_id)
    papers = models.Paper.objects.filter(inv=inv).order_by('num')

    # 制定标题
    title = '{}所有已收集问卷'.format(inv.school_year)

    context = {'inv': inv, 'papers': papers, 'title': title}
    return render_ht(request, 'paper_manage.html', context)


def delete_paper(request, paper_id):
    """删除问卷"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出要删除的问卷对象及对应的调查对象
    paper = models.Paper.objects.get(id=paper_id)
    inv = paper.inv

    # 执行删除操作
    paper.delete()

    # 重定向至问卷管理页
    return HttpResponseRedirect(reverse('paper_manage', args=[inv.id]))


def teacher_list(request, inv_id):
    """宿管老师列表"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出相应的调查对象
    inv = models.Investigation.objects.get(id=inv_id)

    # 所有相关的老师
    tl = []
    for ob in models.Paper.objects.filter(inv=inv):
        teacher = ob.teacher
        teacher_index = DT.teacher_list.index(teacher)
        tl.append((teacher, teacher_index))
    tl = list(set(tl))

    # 制定标题
    title = '{}宿舍管理满意度调查统计结果'.format(inv.school_year)

    context = {'inv': inv, 'title': title, 'tl': tuple(tl)}
    return render_ht(request, 'teacher_list.html', context)


def teacher_result(request, inv_id, teacher_id):
    """宿管老师统计结果"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出相应的调查对象
    inv = models.Investigation.objects.get(id=inv_id)

    # 获取宿管老师姓名
    teacher = DT.teacher_list[teacher_id]

    # 记录得分和建议
    rl, adv13, adv14 = [], [], []
    for ob in models.Paper.objects.filter(inv=inv, teacher=teacher):
        rl.append([ob.t1, ob.t2, ob.t3, ob.t4, ob.t5, ob.t6,
                   ob.t7, ob.t8, ob.t9, ob.t10, ob.t11, ob.t12])
        if ob.t13:
            adv13.append(ob.t13)
        if ob.t14:
            adv14.append(ob.t14)

    # 计算各项平均分
    rl = np.array(rl)
    rs = np.round(rl.mean(axis=0), decimals=3)

    # 拼接评分项目
    rs = np.hstack((np.array(DT.sel_items).reshape(-1, 1), rs.reshape(-1, 1)))

    context = {'inv': inv, 'teacher': teacher, 'rs': tuple(rs),
               'adv13': tuple(adv13), 'adv14': tuple(adv14)}
    return render_ht(request, 'teacher_result.html', context)


def visual(request, inv_id):
    """显示可视化统计图表"""
    # 禁止非管理员用户访问
    if request.user.username not in DT.manager:
        raise Http404

    # 取出相应的调查对象
    inv = models.Investigation.objects.get(id=inv_id)

    # 制定标题
    title = '{}宿舍管理满意度调查结果统计图表'.format(inv.school_year)

    scores, labels = [], []
    for teacher in DT.dorm().keys():
        rl = []
        for ob in models.Paper.objects.filter(inv=inv, teacher=teacher):
            rl.append([ob.t1, ob.t2, ob.t3, ob.t4, ob.t5, ob.t6,
                       ob.t7, ob.t8, ob.t9, ob.t10, ob.t11, ob.t12])

        # 没有评价结果，直接略过
        if not rl:
            continue

        # 计算各项平均分
        rl = np.array(rl)
        rs = list(np.round(rl.mean(axis=0), decimals=3))

        # 加入总结果
        scores.append(rs)
        labels.append(teacher)

    # 存放文件名
    fns = []

    # 无统计结果，直接返回
    if not scores:
        context = {'fns': tuple(fns), 'title': title, 'inv': inv, 'tip': '暂无统计结果'}
        return render_ht(request, 'visual.html', context)

    # 转化为数组并转置
    scores = np.array(scores).transpose()

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    font = FontProperties(fname='media/STKAITI.TTF', size=17)

    # 作出可视化图表
    for i in range(len(DT.sel_items)):
        # 获取数值
        values = scores[i]

        # 清空画布
        plt.cla()

        # 作条形图
        plt.bar(labels, values)
        plt.xticks(labels, fontproperties=font)
        plt.tick_params(axis='x', labelsize=14)

        # 设置图表标题
        plt.title(DT.sel_items[i], fontproperties=font)

        # 显示数值
        for j in range(len(labels)):
            plt.text(labels[j], values[j], str(values[j]), ha='center', va='bottom',
                     fontsize=14)

        # 保存图象
        fn = 't' + str(i + 1) + '.png'
        fns.append(fn)
        plt.savefig(os.path.join('media', 'advice', fn))

    # 前端显示保存的图象
    context = {'fns': tuple(fns), 'title': title, 'inv': inv, 'tip': ''}
    return render_ht(request, 'visual.html', context)


def sep_class(request):
    """分班分寝主页"""
    # 限制访问权限
    if request.user.username not in DT.manager:
        raise Http404

    return render_ht(request, 'sep_class.html', {})


def sep_add(request):
    """分班分寝操作页"""
    # 限制访问权限
    if request.user.username not in DT.manager:
        raise Http404

    return render_ht(request, 'sep_add.html', context={
        'err': '',
        'form': froms.FileUploadForm(),
        'have_button': False,
        'err_id': -1,
    })


def sep_temp(request):
    """下载模板"""
    # 限制访问权限
    if request.user.username not in DT.manager:
        raise Http404

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '身份证号'
    st['B1'].value = '姓名'
    st['C1'].value = '性别'
    st['D1'].value = '班级'
    st['E1'].value = '宿舍'
    st['F1'].value = '床铺号'
    st['G1'].value = '入学年份（默认为当前年份）：'
    st.column_dimensions['G'].width = 30

    # 导出
    return write_out(wb, fn='sep_temp.xlsx')


def add_new_student(request):
    """加入新生"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 获取表单提交内容
    form = froms.FileUploadForm(request.POST, request.FILES)

    if form.is_valid():
        # 临时保存文件
        upload_file = form.cleaned_data['file']
        file_path = 'media/temp/' + upload_file.name

        # 保存到指定路径
        with open(file_path, 'wb') as f:
            for chunk in upload_file.chunks():
                f.write(chunk)

        # 处理上传的文件
        try:
            wb = load_workbook(file_path)
            st = wb.active
        except InvalidFileException:
            # 删除临时文件
            os.remove(file_path)

            # 错误提示信息
            return render_ht(request, 'sep_add.html', context={
                'err': '文件格式必须为xlsx',
                'form': froms.FileUploadForm(),
                'have_button': False,
                'err_id': -1,
            })

        # 删除临时文件
        os.remove(file_path)

        # 成功数和失败数
        success, fail = 0, 0

        # 错误提示填充
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # 从第二行开始逐行读取
        for row in range(2, st.max_row + 1):
            # 读取信息
            name = st.cell(row=row, column=2).value
            id_number = st.cell(row=row, column=1).value
            gender = st.cell(row=row, column=3).value
            cs = int(st.cell(row=row, column=4).value)
            dorm = str(st.cell(row=row, column=5).value)
            bed = int(st.cell(row=row, column=6).value)

            # 处理数据，生成新生对象
            if id_number:
                if not correct_id(id_number):
                    # 身份证号格式错误
                    st.cell(row=row, column=7).value = '身份证号格式错误'
                    st.cell(row=row, column=7).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue
            else:
                id_number = ''

            if gender:
                # 处理性别
                if gender not in ['男', '女']:
                    st.cell(row=row, column=7).value = '性别只能是“男”或“女”'
                    st.cell(row=row, column=7).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue
            else:
                gender = '未知'

            if id_number and gender:
                # 处理身份证号与性别的关系
                if gender != models.get_gender(id_number):
                    st.cell(row=row, column=7).value = '身份证号与性别不匹配'
                    st.cell(row=row, column=7).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

            if id_number in all_sep_id():
                # 防止身份证号重复
                st.cell(row=row, column=7).value = '身份证号重复'
                st.cell(row=row, column=7).fill = yellow_fill

                # 失败数加1
                fail += 1
                continue

            if len(name) > 5:
                # 姓名长度不能超过5个字
                st.cell(row=row, column=7).value = '姓名长度不能超过5个字'
                st.cell(row=row, column=7).fill = yellow_fill

                # 失败数加1
                fail += 1
                continue

            if len(dorm) != 3:
                # 宿舍号须为三位
                st.cell(row=row, column=7).value = '宿舍号须为三位'
                st.cell(row=row, column=7).fill = yellow_fill

                # 失败数加1
                fail += 1
                continue

            if cs < 1 or cs > 12:
                # 班级范围应为1~12
                st.cell(row=row, column=7).value = '班级范围应为1~12'
                st.cell(row=row, column=7).fill = yellow_fill

                # 失败数加1
                fail += 1
                continue

            if bed < 1 or bed > 8:
                # 床铺号范围应为1~8
                st.cell(row=row, column=7).value = '床铺号范围应为1~8'
                st.cell(row=row, column=7).fill = yellow_fill

                # 失败数加1
                fail += 1
                continue

            # 通过所有检测，可以创建对象
            nst = models.NewStudent()
            nst.name = name
            nst.gender = gender
            nst.id_number = id_number
            nst.cs = cs
            nst.dorm = dorm
            nst.bed = bed

            # 入学年份
            if st['H1'].value:
                nst.grade_year = int(st['H1'].value)
            else:
                nst.grade_year = datetime.datetime.now().year

            # 补全属性，存入数据库
            nst.fill_blank()
            nst.save()

            # 成功数加1
            success += 1

        # 生成提示信息
        tip_msg = '共上传{}条数据，成功{}条，失败{}条'.format(success + fail, success, fail)

        if fail > 0:
            # 存在失败数据，保存失败文件
            st['G1'].value = '错误提示'
            st['G1'].fill = yellow_fill
            st['G1'].font = Font(bold=True)
            st.column_dimensions['G'].width = 50

            # TODO:错误信息文件提示
            # 设定错误文件名
            n = 0

            while True:
                # 生成文件名
                if n > 0:
                    filename = 'error({}).xlsx'.format(n)
                else:
                    filename = 'error.xlsx'

                # 判断是否已有文件名
                if filename in os.listdir(os.path.join('media', 'temp')):
                    # 下一个号
                    n += 1
                else:
                    # 可用文件名
                    wb.save(os.path.join('media', 'temp', filename))
                    break

            # 可下载错误文件
            have_button = True
            err_id = n

            # 创建并启动延时删除文件线程
            remove_thread = threading.Thread(target=remove_lay, args=(
                os.path.join('media', 'temp', filename), 60))
            remove_thread.start()
        else:
            # 无须下载错误文件
            have_button = False
            err_id = -1

        # 返回提示
        return render_ht(request, 'sep_add.html', context={
            'err': tip_msg,
            'form': froms.FileUploadForm(),
            'have_button': have_button,
            'err_id': err_id,
        })


def download_error(request, error_num):
    """下载错误文件并然后删除"""
    # # 限制权限
    # if request.user.username not in DT.manager:
    #     raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 连接文件名及文件路径
    if error_num > 0:
        filename = 'error({}).xlsx'.format(error_num)
    else:
        filename = 'error.xlsx'
    file_path = os.path.join('media', 'temp', filename)

    # 下载
    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f, content_type='application/vnd.ms-excel')

            # 设置文件名称
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
    except FileNotFoundError:
        return HttpResponse('没有错误文件可以下载！')

    # 输出
    return response


def sep_yl(request):
    """入学年份选择"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 获取所有已存在的入学年份
    years = []
    for student in models.NewStudent.objects.all():
        years.append(student.grade_year)
    years = tuple(set(years))

    context = {'years': years}
    return render_ht(request, 'sep_yl.html', context)


def sep_see(request, grade_year):
    """分年份查看"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 查询框赋值
    ipt = ''
    gc_val = 'all'
    dorm_val = 'all'

    # 获取所有新生对象
    students = []
    if request.method != 'POST':
        # 非POST方法访问，加入所有新生对象
        for ob in models.NewStudent.objects.filter(grade_year=grade_year):
            students.append(ob)
    else:
        # POST方法访问，处理查询信息
        # 获取前端输入信息
        que = request.POST.get('que', '')
        gc_val = request.POST.get('sep-gc', '')
        dorm_val = request.POST.get('sep-dorm', '')

        # 加入包含查询信息的新生对象
        for ob in models.NewStudent.objects.filter(grade_year=grade_year):
            if que in ob.name or que in ob.id_number:
                if ob.gc == gc_val or gc_val == 'all':
                    if ob.dorm == dorm_val or dorm_val == 'all':
                        students.append(ob)

        # 查询框重新赋值
        ipt = que

    # 依次按照床铺号、宿舍号、班级排序
    students.sort(key=lambda x: x.bed)
    students.sort(key=lambda x: x.dorm)
    students.sort(key=lambda x: x.cs)

    context = {
        'students': tuple(students),
        'year': grade_year,
        'ipt': ipt,
        'all_gc': tuple(all_sep_gc(grade_year)),
        'all_dorm': tuple(all_sep_dorm(grade_year)),
        'gc_val': gc_val,
        'dorm_val': dorm_val,
        'n': len(students),
    }
    return render_ht(request, 'sep_see.html', context)


def sep_del(request, student_id):
    """删除新生对象"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 取出新生对象，记录所属年份
    student = models.NewStudent.objects.get(id=student_id)
    year = student.grade_year

    # 执行删除操作
    student.delete()

    # 重定向至查看页
    return HttpResponseRedirect(reverse('sep_see', args=[year]))


def sep_del_year(request, year):
    """一键清除一年的所有新生"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 获取要删除的新生对象
    students = models.NewStudent.objects.filter(grade_year=year)

    # 依次执行删除
    for ob in students:
        ob.delete()

    # 重定向至年份选择页面
    return HttpResponseRedirect(reverse('sep_yl'))


def sep_change_info(request, student_id):
    """修改新生信息"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 取出新生对象，记录所属年份
    student = models.NewStudent.objects.get(id=student_id)
    year = student.grade_year
    title = '修改新生信息'

    if request.method != 'POST':
        # 未提交数据，用当前新生对象填充表单
        form = froms.ChangeInfoForm(instance=student)
    else:
        # 对POST提交的数据作出处理
        form = froms.ChangeInfoForm(instance=student, data=request.POST)
        if form.is_valid():
            change = form.save(commit=False)

            # 验证身份证号
            if change.id_number:
                if not correct_id(change.id_number):
                    return render_ht(request, 'sep_change_info.html', context={
                        'student': student,
                        'year': year,
                        'form': form,
                        'err': '身份证号格式错误',
                        'title': title,
                    })

            # 判断身份证号与性别是否匹配
            if change.gender != '未知' and change.id_number:
                if change.gender != models.get_gender(change.id_number):
                    return render_ht(request, 'sep_change_info.html', context={
                        'student': student,
                        'year': year,
                        'form': form,
                        'err': '身份证号与性别不匹配',
                        'title': title,
                    })

            # 防止身份证号重复
            if change.id_number in filter(lambda x: x != student.id_number, all_sep_id()):
                return render_ht(request, 'sep_change_info.html', context={
                    'student': student,
                    'year': year,
                    'form': form,
                    'err': '身份证号重复',
                    'title': title,
                })

            # 验证班级
            if change.cs < 1 or change.cs > 12:
                return render_ht(request, 'sep_change_info.html', context={
                    'student': student,
                    'year': year,
                    'form': form,
                    'err': '班级必须在1~12之间',
                    'title': title,
                })

            # 验证床号
            if change.bed < 1 or change.bed > 8:
                return render_ht(request, 'sep_change_info.html', context={
                    'student': student,
                    'year': year,
                    'form': form,
                    'err': '床铺号必须在1~8之间',
                    'title': title,
                })

            # 修改新生信息
            student.name = change.name
            student.id_number = change.id_number
            student.gender = change.gender
            student.cs = change.cs
            student.dorm = change.dorm
            student.bed = change.bed
            student.fill_blank()
            student.save()

            # 重定向至查看页
            return HttpResponseRedirect(reverse('sep_see', args=[year]))

    context = {'student': student, 'year': year, 'form': form, 'err': '', 'title': title}
    return render_ht(request, 'sep_change_info.html', context)


def sep_add_one(request, year):
    """单个添加某一年入学的新生"""
    # 限制权限
    if request.user.username not in DT.manager:
        raise Http404

    # 制定标题
    title = '添加{}年入学新生'.format(year)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = froms.ChangeInfoForm()
    else:
        # 对POST提交的数据作出处理
        form = froms.ChangeInfoForm(data=request.POST)
        if form.is_valid():
            new_st = form.save(commit=False)

            # 验证身份证号
            if new_st.id_number:
                if not correct_id(new_st.id_number):
                    return render_ht(request, 'sep_add_one.html', context={
                        'year': year,
                        'form': form,
                        'err': '身份证号格式错误',
                        'title': title,
                    })

            if new_st.gender != '未知' and new_st.id_number:
                if new_st.gender != models.get_gender(new_st.id_number):
                    return render_ht(request, 'sep_add_one.html', context={
                        'year': year,
                        'form': form,
                        'err': '身份证号与性别不匹配',
                        'title': title,
                    })

            # 防止身份证号重复
            if new_st.id_number in all_sep_id():
                return render_ht(request, 'sep_add_one.html', context={
                    'year': year,
                    'form': form,
                    'err': '身份证号重复',
                    'title': title,
                })

            # 验证班级
            if new_st.cs < 1 or new_st.cs > 12:
                return render_ht(request, 'sep_add_one.html', context={
                    'year': year,
                    'form': form,
                    'err': '班级必须在1~12之间',
                    'title': title,
                })

            # 验证床铺号
            if new_st.bed < 1 or new_st.bed > 8:
                return render_ht(request, 'sep_add_one.html', context={
                    'year': year,
                    'form': form,
                    'err': '床铺号必须在1~8之间',
                    'title': title,
                })

            # 设置入学年份为给定年份
            new_st.grade_year = year

            # 补充完整属性，保存
            new_st.fill_blank()
            new_st.save()

            # 重定向至查看页
            return HttpResponseRedirect(reverse('sep_see', args=[year]))

    context = {'year': year, 'form': form, 'err': '', 'title': title}
    return render_ht(request, 'sep_add_one.html', context)


def sep_que(request):
    """按身份证号查询"""
    # 未开放查询
    if datetime.datetime.now() < DT.que_start_time and request.user.username != 'zz106dyc':
        return HttpResponse('尚未开放查询，敬请期待！')

    if request.method != 'POST':
        # 未提交数据，加载新的查询页面
        return render_ht(request, 'sep_que.html', {'err': ''})
    else:
        # 对POST提交的数据作出处理
        id_number = request.POST.get('id-number', '')

        # 身份证号不能为空
        if not id_number:
            return render_ht(request, 'sep_que.html', context={
                'err': '身份证号不能为空',
            })

        try:
            # 匹配新生对象
            st = models.NewStudent.objects.filter(id_number=id_number)[0]
        except IndexError:
            # 未匹配到
            return render_ht(request, 'sep_que.html', context={
                'err': '不存在身份证号为{}的学生'.format(id_number),
            })
        else:
            # 获取原始班级字符串
            rgc = DT.get_original_gc(st.gc)

            # 家长群二维码链接
            if st.grade_year >= 2025:
                code_link = '{}-{}'.format(st.grade_year, DT.str_two(st.cs))
            else:
                code_link = ''

            # 返回结果显示页面
            return render_ht(request, 'sep_que_result.html', context={
                'st': st,
                'rgc': rgc,
                'code_link': code_link,
                'title': '{}分班分寝查询结果'.format(settings.USER_NAME),
            })


def sep_export(request, year):
    """导出一年的新生数据"""
    # 打开新的表格文件
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '姓名'
    st['B1'].value = '性别'
    st['C1'].value = '身份证号'
    st['D1'].value = '班级'
    st['E1'].value = '宿舍'
    st['F1'].value = '床铺号'

    # 初始行
    row = 2

    for ob in models.NewStudent.objects.filter(grade_year=year):
        st.cell(row=row, column=1).value = ob.name
        st.cell(row=row, column=2).value = ob.gender
        st.cell(row=row, column=3).value = ob.id_number
        st.cell(row=row, column=4).value = ob.gc
        st.cell(row=row, column=5).value = ob.dorm
        st.cell(row=row, column=6).value = ob.bed

        # 下一行
        row += 1

    # 输出
    return write_out(wb)


def settable_summary(request):
    """自定义时间段总结"""
    # 获取今天的日期字符串
    today_str = datetime.datetime.strftime(datetime.date.today(), '%Y-%m-%d')

    if request.method != 'POST':
        # 未提交数据，加载设置页
        context = {'today': today_str, 'err': ''}
        return render_ht(request, 'settable_summary.html', context)
    else:
        # 对POST提交的数据做出处理
        grade_num = int(request.POST.get('grade', ''))
        start_date = request.POST.get('start', '')
        end_date = request.POST.get('end', '')

        # 检验日期合法性
        if start_date > end_date:
            return render_ht(request, 'settable_summary.html', context={
                'today': today_str,
                'err': '开始日期不能晚于结束日期',
            })

        # 连结日期范围
        date_area = '{}_{}'.format(start_date, end_date)

        # 重定向至查看页
        return HttpResponseRedirect(reverse('free_summary', args=[grade_num, date_area]))


def free_summary(request, grade_num, date_area):
    """自定义时间段总结查看"""
    # 解析开始日期和天数
    se = date_area.split('_')
    start_date = datetime.datetime.strptime(se[0], '%Y-%m-%d').date()
    end_date = datetime.datetime.strptime(se[1], '%Y-%m-%d').date()
    total_days = (end_date - start_date).days + 1

    # 生成标题
    title = make_title(grade_num, date_info=(start_date, total_days))

    context = {
        'title': title,
        'grade_num': grade_num,
        'start_date_str': se[0],
        'total_days': total_days,
    }
    return render_ht(request, 'free_summary.html', context)


def user_manage(request):
    """用户管理"""
    # 限制访问权限
    if not request.user.is_staff:
        raise Http404

    # 获取格式化存储的所有用户对象
    user_list = []
    sg = Group.objects.get(name='Student')
    for user in User.objects.all():
        if sg in user.groups.all():
            is_student = True
        else:
            is_student = False
        user_list.append((user, is_student))

    context = {'users': tuple(user_list)}
    return render_ht(request, 'user_manage.html', context)


def reset_pwd(request, user_id):
    """重置用户密码"""
    # 限制访问权限
    if not request.user.is_staff:
        raise Http404

    # 取得用户对象
    user = User.objects.get(id=user_id)

    # 根据用户类型不同，重置为不同的初始密码
    sg = Group.objects.get(name='Student')
    if sg in user.groups.all():
        user.set_password('106gzxsh')
        tip_msg = '账号{}{}的密码已重置为106gzxsh'.format(user.username, user.last_name)
    else:
        user.set_password('106dycdyc')
        tip_msg = '账号{}的密码已重置为106dycdyc'.format(user.username)
    user.save()

    # 跳转至提示页
    context = {'tip_msg': tip_msg}
    return render_ht(request, 'tip_reset.html', context)


def change_graduated(request, grade_year):
    """标记/取消标记毕业"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出当前年级宿舍列表
    dorms = models.get_dorms(grade_year)

    # 逐个修改
    for student in models.NewStudent.objects.filter(grade_year=grade_year):
        student.graduated = not student.graduated
        student.save()

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 找到并且删除目标宿舍
    for dorm in dorms:
        for dl in dm_dict.values():
            if dorm in dl:
                dl.remove(dorm)

                # 已完成删除，不必进行后面的循环，节省算力
                break

    # 写入新宿管信息
    with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
        fo.write(json.dumps(dm_dict, ensure_ascii=False))

    # 重定向至查看页
    return HttpResponseRedirect(reverse('sep_see', args=[grade_year]))


def sep_change_dorm(request, grade_year):
    """搬宿舍"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 制定标题
    title = '给{}级搬宿舍'.format(grade_year)

    context = {'title': title, 'year': grade_year, 'form': froms.SepChangeDormForm(), 'errs': (),
               'have_button': False, 'err_id': -1}
    return render_ht(request, 'sep_change_dorm.html', context)


def sep_cd_temp(request, grade_year):
    """下载搬宿舍模板"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '原宿舍'
    st['B1'].value = '新宿舍'

    # 表头加粗
    st['A1'].font = Font(bold=True)
    st['B1'].font = Font(bold=True)

    # 依次写入原宿舍
    row = 2
    for dorm in models.get_dorms(grade_year):
        st.cell(row=row, column=1).value = dorm
        row += 1

    # 输出
    return write_out(wb, fn='sep_cd.xlsx')


def sep_cd(request, grade_year):
    """模板上传搬宿舍"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 制定标题
    title = '给{}级搬宿舍'.format(grade_year)

    # 获取表单提交内容
    form = froms.SepChangeDormForm(request.POST, request.FILES)

    # 获取年级索引
    lgd = str(grade_year) + '级'
    grade = DT.logic_grade_reverse[lgd]
    grade_index = DT.grade_dict_reverse[grade] - 1

    if form.is_valid():
        # 临时保存文件
        upload_file = form.cleaned_data['file']
        file_path = 'media/temp/' + upload_file.name

        # 保存到指定路径
        with open(file_path, 'wb') as f:
            for chunk in upload_file.chunks():
                f.write(chunk)

        # 处理上传的文件
        try:
            wb = load_workbook(file_path)
            st = wb.active
        except InvalidFileException:
            # 删除临时文件
            os.remove(file_path)

            # 错误提示信息
            return render_ht(request, 'sep_change_dorm.html', context={
                'title': title,
                'year': grade_year,
                'form': froms.SepChangeDormForm(),
                'errs': ('文件格式必须为xlsx',),
            })

        # 删除临时文件
        os.remove(file_path)

        # 读取信息存入原宿舍-新宿舍字典
        old_new = {}
        for row in range(2, st.max_row + 1):
            old_dorm = str(st.cell(row=row, column=1).value)
            new_dorm = str(st.cell(row=row, column=2).value)
            if len(new_dorm) == 3 and new_dorm.isdigit():
                # 读入行信息
                old_new[old_dorm] = (new_dorm, row)

        # 第三列为信息提示列
        st['C1'].value = '提示'

        # 防止宿舍管理权限重复
        exist_dm_dorm, exist_st_dorm = {}, {}
        # 遍历时要注意创建字典副本
        old_new_copy = old_new.copy()
        for odm, ndm in old_new_copy.items():
            if ndm[0] in DT.all_dormitory() or ndm[0] in DT.get_all_dormitory():
                # 宿管权限冲突
                if ndm[0] in DT.all_dormitory():
                    exist_dm_dorm[odm] = ndm

                # 学生住宿冲突
                if ndm[0] in DT.get_all_dormitory():
                    exist_st_dorm[odm] = ndm

                # 原字典中先删除新宿舍有重复的键值对
                del old_new[odm]

        # 用于记录未在字典的宿舍列表
        nd = []

        # 对该年级学生逐个操作
        done = {}
        for student in models.NewStudent.objects.filter(grade_year=grade_year, graduated=False):
            try:
                student.dorm = old_new[student.dorm][0]
            except KeyError:
                # 该提示仅对不在原字典中的提出
                if student.dorm not in old_new_copy.keys():
                    if student.dorm not in nd:
                        nd.append(student.dorm)
            else:
                student.save()

                # 记录实际搬迁情况
                if student.dorm not in done.keys():
                    done[student.dorm] = old_new[student.dorm][0]

                    # 输出文件中搬迁成功提示
                    st.cell(row=old_new[student.dorm][1], column=3).value = '搬迁成功'

        # 读取原宿管信息
        with open('media/dorm_manager.json', encoding='gbk') as fi:
            dm_dict = json.loads(fi.read())

        # 宿管信息变更
        for dl in dm_dict.values():
            # 此处必须用列表副本进行遍历，因为列表本身会发生变化
            for dorm in dl[:]:
                if dorm in done.keys():
                    dl.remove(dorm)
                    dl.append(done[dorm])

        # 写入新宿管信息
        with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
            fo.write(json.dumps(dm_dict, ensure_ascii=False))

        # 记入当年新旧宿舍搬迁对应字典（实际搬迁为准）
        # 获取年份及文件路径
        year = datetime.datetime.now().year
        filepath = 'media/old_new_{}.json'.format(year)

        # 打开文件之前确定文件存在，否则写入空字典列表
        if not os.path.exists(filepath):
            with open(file_path, 'w') as f:
                f.write(json.dumps([{}, {}]))

        old_new_d_l = DT.get_old_new()

        # 拼接新搬迁部分
        old_new_d_l[grade_index].update(done)

        # 输出拼接后的字典到json
        with open(filepath, 'w') as fo_1:
            fo_1.write(json.dumps(old_new_d_l))

        # 生成提示信息
        err_list = []

        # 宿管权限冲突提示
        for odor, ndor in exist_dm_dorm.items():
            # 获取宿管老师
            tm = DT.get_manager(ndor[0])

            # 生成提示信息
            err_msg = '新宿舍{}为{}老师管理的宿舍，原宿舍{}搬入失败'.format(ndor[0], tm, odor)
            err_list.append(err_msg)

            # 输出文件中提示信息
            st.cell(row=ndor[1], column=3).value = err_msg

        # 学生入住冲突提示
        for odor, ndor in exist_st_dorm.items():
            # 获取入住班级
            gc_str = DT.get_class(ndor[0])

            # 生成提示信息
            err_msg = '新宿舍{}为{}学生入住的宿舍，原宿舍{}搬入失败'.format(ndor[0], gc_str, odor)
            err_list.append(err_msg)

            # 输出文件中提示信息
            st.cell(row=ndor[1], column=3).value = err_msg

        # 原宿舍未搬迁提示
        err_row = st.max_row + 1
        for dor in nd:
            err_msg = '{}宿舍未搬迁，如不需要搬迁，请忽略'.format(dor)
            err_list.append(err_msg)

            # 输出文件中异常提示
            st.cell(row=err_row, column=3).value = err_msg

            # 下一行
            err_row += 1

        # TODO:提示信息文件输出
        # 设定错误文件名
        n = 0

        while True:
            # 生成文件名
            if n > 0:
                filename = 'error({}).xlsx'.format(n)
            else:
                filename = 'error.xlsx'

            # 判断是否已有文件名
            if filename in os.listdir(os.path.join('media', 'temp')):
                # 下一个号
                n += 1
            else:
                # 可用文件名
                wb.save(os.path.join('media', 'temp', filename))
                break

        # 创建并启动延时删除文件线程
        remove_thread = threading.Thread(target=remove_lay, args=(
            os.path.join('media', 'temp', filename), 60))
        remove_thread.start()

        return render_ht(request, 'sep_change_dorm.html', context={
            'title': title,
            'year': grade_year,
            'form': froms.SepChangeDormForm(),
            'errs': tuple(err_list),
            'have_button': True,
            'err_id': n,
        })


def sep_three(li):
    """将给定列表分为三个一组"""
    # 输出列表
    output_list, small_list = [], []

    # 原列表先排序
    li.sort()

    # 保证遍历次数为3的倍数且能覆盖所有元素
    if len(li) % 3:
        total = (len(li) // 3 + 1) * 3
    else:
        total = len(li)

    for i in range(total):
        if i % 3 == 0:
            # 重置小列表
            small_list = []

        # 加入元素
        try:
            small_list.append(li[i])
        except IndexError:
            # 末尾空白元素
            small_list.append('')

        if i % 3 == 2:
            # 小列表已满，加入大列表
            output_list.append(small_list)

    return output_list


def get_sep_dict(ori_dict):
    """取得分裂之后用于传给前端的字典"""
    # 分裂为三个一组
    sep_dict = {}

    # 计数变量
    count = 0
    for k, v in ori_dict.items():
        sep_dict[k] = (sep_three(v), count)
        count += 1

    return sep_dict


def dm_power(request):
    """宿舍管理权限分配"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 前端显示格式
    sep_dict = get_sep_dict(dm_dict)

    # 宿管权限与入住学生信息不一致提示
    dm_more, st_more = DT.compare_all_dorms()
    tip_list = []
    for dor in dm_more:
        # 获取宿舍管理老师
        tm = DT.get_manager(dor)

        # 生成提示信息
        tip_msg = '{}老师管理的{}宿舍没有学生入住，如无学生入住，建议删除管理权限'.format(tm, dor)
        tip_list.append(tip_msg)

    for dor in st_more:
        # 获取宿舍年级
        grade = DT.grade_dict[DT.get_grade_by_dorm(dor)]

        # 生成提示信息
        tip_msg = '{}年级宿舍{}没有宿管老师，请尽快分配'.format(grade, dor)
        tip_list.append(tip_msg)

    context = {'dm_dict': tuple(sep_dict.items()), 'err': '', 'tips': tuple(tip_list)}
    return render_ht(request, 'dm_power.html', context)


def dmp_add(request):
    """宿舍管理权限添加"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 前端读取老师和要添加的宿舍
    teacher = request.POST.get('teacher', '')
    dorm = request.POST.get('ipt', '')

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 检验输入的信息是否符合规范
    if not dorm.isdigit() or len(dorm) != 3:
        # 前端显示格式
        sep_dict = get_sep_dict(dm_dict)

        return render_ht(request, 'dm_power.html', context={
            'dm_dict': tuple(sep_dict.items()),
            'err': '宿舍号格式错误',
        })

    # 检验信息的重复性
    for tm, dl in dm_dict.items():
        if dorm in dl:
            # 前端显示格式
            sep_dict = get_sep_dict(dm_dict)

            # 生成错误提示信息
            err_msg = '{}宿舍为{}老师管理的宿舍，如需更改，请先删除再添加'.format(dorm, tm)

            return render_ht(request, 'dm_power.html', context={
                'dm_dict': tuple(sep_dict.items()),
                'err': err_msg,
            })

    # 宿管信息变更
    dm_dict[teacher].append(dorm)

    # 写入新宿管信息
    with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
        fo.write(json.dumps(dm_dict, ensure_ascii=False))

    # 重定向至宿管权限页
    return HttpResponseRedirect(reverse('dm_power'))


def dmp_del(request, dorm):
    """删除宿舍管理权限"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 找到并且删除目标宿舍
    for dl in dm_dict.values():
        if dorm in dl:
            dl.remove(dorm)

            # 已完成删除，不必进行后面的循环，节省算力
            break

    # 写入新宿管信息
    with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
        fo.write(json.dumps(dm_dict, ensure_ascii=False))

    # 重定向至宿管权限页
    return HttpResponseRedirect(reverse('dm_power'))


def add_teacher(request):
    """新增宿管老师"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 前端读取宿管老师姓名
    teacher_name = request.POST.get('ipt', '')

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 加入新键
    dm_dict.setdefault(teacher_name, [])

    # 写入新宿管信息
    with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
        fo.write(json.dumps(dm_dict, ensure_ascii=False))

    # 重定向至宿管权限页
    return HttpResponseRedirect(reverse('dm_power'))


def del_teacher(request):
    """删除宿管老师"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 前端读取宿管老师姓名
    teacher = request.POST.get('teacher', '')

    # 读取原宿管信息
    with open('media/dorm_manager.json', encoding='gbk') as fi:
        dm_dict = json.loads(fi.read())

    # 删除键值对
    del dm_dict[teacher]

    # 写入新宿管信息
    with open('media/dorm_manager.json', 'w', encoding='gbk') as fo:
        fo.write(json.dumps(dm_dict, ensure_ascii=False))

    # 重定向至宿管权限页
    return HttpResponseRedirect(reverse('dm_power'))


def dm_lines_main(request):
    """搬宿舍轨迹主页"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确定起止年份和可遍历年份对象
    start_year = 2025
    end_year = datetime.datetime.now().year
    years = tuple(range(start_year, end_year + 1))

    context = {'years': years}
    return render_ht(request, 'dm_lines_main.html', context)


def dm_lines(request, year):
    """年份搬宿舍轨迹"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 读取宿舍搬迁轨迹字典
    try:
        old_new = DT.get_old_new(year=year)
    except FileNotFoundError:
        # 未找到文件，提示该年份尚未搬迁宿舍
        return HttpResponse('{}年还没有搬迁宿舍'.format(year))

    context = {
        'year': year,
        'old_new_1': tuple(sorted(old_new[0].items(), key=lambda x: x[0])),
        'old_new_2': tuple(sorted(old_new[1].items(), key=lambda x: x[0])),
    }
    return render_ht(request, 'dm_lines.html', context)


def del_dm_line(request, year, grade_index, od_key):
    """删除宿舍搬迁轨迹"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 生成文件路径
    filepath = 'media/old_new_{}.json'.format(year)

    # 读取宿舍搬迁轨迹字典
    old_new = DT.get_old_new(year=year)

    # 执行删除操作
    del old_new[grade_index][od_key]

    # 重新写入
    with open(filepath, 'w') as fo:
        fo.write(json.dumps(old_new))

    # 重定向至当年轨迹查看与管理页面
    return HttpResponseRedirect(reverse('dm_lines', args=[year]))


def add_dm_line(request, year):
    """新增宿舍搬迁轨迹"""
    # 限制用户权限
    if request.user.username not in DT.manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = froms.AddLineForm()
    else:
        # 对POST提交的数据作出处理
        form = froms.AddLineForm(request.POST)

        if form.is_valid():
            # 读取输入的数据
            grade_index = form.cleaned_data['grade_index']
            old_dorm = form.cleaned_data['old_dorm']
            new_dorm = form.cleaned_data['new_dorm']

            # 生成文件路径
            filepath = 'media/old_new_{}.json'.format(year)

            # 读取宿舍搬迁轨迹字典
            old_new = DT.get_old_new(year=year)

            # 验证原宿舍、新宿舍是否已存在
            for ond in old_new:
                if old_dorm in ond.keys():
                    return render_ht(request, 'add_dm_line.html', context={
                        'year': year,
                        'form': form,
                        'err': '原宿舍{}已存在'.format(old_dorm),
                    })
                if new_dorm in ond.values():
                    return render_ht(request, 'add_dm_line.html', context={
                        'year': year,
                        'form': form,
                        'err': '新宿舍{}已存在'.format(new_dorm)
                    })

            # 通过验证，执行添加操作
            old_new[int(grade_index)][old_dorm] = new_dorm

            # 重新写入
            with open(filepath, 'w') as fo:
                fo.write(json.dumps(old_new))

            # 重定向至轨迹查看页面
            return HttpResponseRedirect(reverse('dm_lines', args=[year]))

    context = {'year': year, 'form': form, 'err': ''}
    return render_ht(request, 'add_dm_line.html', context)


def see_leave(request, err):
    """
        查看请假离校学生
        err为0：打开新浏览器
        err为1：使用旧浏览器，并给出验证码错误提示
        err为2：打开新浏览器，并提示再次尝试
    """
    # 制定标题
    title = '查看当前请假离校学生'

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')
        download_path = '/tmp'
    else:
        download_path = 'C:\\Users\\45920\\Downloads'

    if request.method != 'POST':
        # 登录链接
        login_url = 'https://www.12kcool.com/#/102064'

        # 打开新浏览器
        if err in (0, 2):
            DT.refresh_browser()

        # 进入登录页面
        DT.b.get(login_url)
        # time.sleep(1)

        # 获取验证码图像
        captcha_ele = DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]/div[4]/img')
        captcha_binary = captcha_ele.screenshot_as_png
        captcha_binary = base64.b64encode(captcha_binary)
        captcha_str = captcha_binary.decode(encoding='utf-8')
        DT.current_captcha_str = captcha_str

        # 提示信息
        if err == 1:
            err_msg = '验证码错误！'
        elif err == 2:
            err_msg = '获取数据失败，请再次尝试'
        else:
            err_msg = ''

        context = {'title': title, 'err': err_msg, 'image': captcha_str}
        return render_ht(request, 'see_leave.html', context)
    else:
        # 用户名、密码、验证码
        un = 'admin'
        pwd = 'Ffkj-102064'
        captcha = request.POST.get('captcha', '')

        try:
            # 依次输入
            DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                 '/div[1]/input').send_keys(un)
            DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                 '/div[2]/input').send_keys(pwd)
            DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                 '/div[3]/input').send_keys(captcha)

            # 点击登录
            DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]/button').click()
        except AttributeError:
            # 重试
            return HttpResponseRedirect(reverse('see_leave', args=[2]))
        except MaxRetryError:
            # 也重试
            return HttpResponseRedirect(reverse('see_leave', args=[2]))
        else:
            time.sleep(1)

        # 进入目标链接
        target_url = 'https://www.12kcool.com/approve/#/approve/studentManage/approve'
        DT.b.get(target_url)
        time.sleep(3)

        # 先删除下载目录中的其他同名文件
        if os.name != 'nt':
            # 服务器上需要在一个范围内查找文件
            for root, dirs, files in os.walk(download_path):
                for fn in files:
                    if fn in ('学生请假记录.xls', '学生请假记录.xlsx'):
                        fp = os.path.join(root, fn)
                        os.remove(fp)
        else:
            for fn in os.listdir(download_path):
                if fn in ('学生请假记录.xls', '学生请假记录.xlsx'):
                    fp = os.path.join(download_path, fn)
                    os.remove(fp)

        # 点击导出按钮
        try:
            DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/section/main/div/div/div[3]/div[2]/'
                                                 'div/div/div/div[2]/div[1]/div/div[2]/button[3]').click()
        except ElementClickInterceptedException:
            # 登录失败，重新加载页面
            return HttpResponseRedirect(reverse('see_leave', args=[1]))
        except NoSuchElementException:
            # 登录失败，重新加载页面
            return HttpResponseRedirect(reverse('see_leave', args=[1]))
        else:
            # TODO:记录正确的验证码
            captcha_str = DT.current_captcha_str
            bank_path = os.path.join('media', 'captcha_bank.json')

            # print(json.loads(json.dumps({})))

            # 首次使用确定文件存在
            if not os.path.exists(bank_path):
                # print('首次创建！')
                with open(bank_path, 'w', encoding='utf-8') as ff:
                    ff.write(json.dumps({}))

            with open(bank_path, encoding='utf-8') as cfi:
                # print(captcha_binary)
                # print('here')
                # print(captcha_str)
                captcha_bank_ds = cfi.read()

            try:
                captcha_bank_dict = json.loads(captcha_bank_ds)
            except json.JSONDecodeError:
                # 出现未知原因错误，跳过记录验证码，直接显示获取结果
                pass
            else:
                if captcha_str not in captcha_bank_dict.keys():
                    captcha_bank_dict[captcha_str] = captcha
                    with open(bank_path, 'w', encoding='utf-8') as cfo:
                        cfo.write(json.dumps(captcha_bank_dict))

            # 退出浏览器
            time.sleep(2)
            DT.b.quit()

            # 处理已成功导出的数据
            if os.name != 'nt':
                # 服务器上需要在一个范围内查找文件
                found = False
                for root, dirs, files in os.walk(download_path):
                    for fn in files:
                        if fn == '学生请假记录.xls':
                            data_fp_old = os.path.join(root, fn)
                            data_fp_new = os.path.join(root, '学生请假记录.xlsx')

                            # 找到即终止
                            found = True
                            break

                    if found:
                        break

                # 走完循环仍未找到，提示获取数据失败
                if not found:
                    return HttpResponseRedirect(reverse('see_leave', args=[2]))

            else:
                # 本地简单操作
                data_fp_old = os.path.join(download_path, '学生请假记录.xls')
                if not os.path.exists(data_fp_old):
                    return HttpResponseRedirect(reverse('see_leave', args=[2]))
                data_fp_new = os.path.join(download_path, '学生请假记录.xlsx')

            # 强制更名为xlsx文件后再打开
            os.rename(data_fp_old, data_fp_new)
            with open(data_fp_new, 'rb') as f:
                file_data = f.read()
            in_mem_file = BytesIO(file_data)
            wb = load_workbook(in_mem_file, read_only=True)
            st = wb.active

            # 记录当前请假学生
            leave_students = []
            for row in range(4, st.max_row + 1):
                # 读取请假开始、结束时间
                stt_str = st.cell(row=row, column=5).value
                edt_str = st.cell(row=row, column=6).value
                stt = datetime.datetime.strptime(stt_str, '%Y-%m-%d %H:%M')
                edt = datetime.datetime.strptime(edt_str, '%Y-%m-%d %H:%M')

                # 读取审批状态
                perm = st.cell(row=row, column=10).value

                # 符合条件者予以记录
                if stt <= datetime.datetime.now() <= edt and perm == '审批通过':
                    # 请假者班级
                    gc = st.cell(row=row, column=3).value

                    # # 判断逻辑年级
                    # grade_year = gc[:4]
                    # logic_grade = str(grade_year - 3) + '级'
                    #
                    # # 取得班级数字
                    # cs_num = int(gc[-2:])
                    #
                    # # 重新赋值班级
                    # grade = DT.logic_grade_reverse[logic_grade]
                    # gc = grade + str(cs_num) + '班'

                    # 请假者姓名、请假类型
                    name = st.cell(row=row, column=2).value
                    leave_type = st.cell(row=row, column=4).value

                    # 记录三个数据
                    leave_students.append((gc, name, leave_type))

            # 成功读取文件内容之后即可删除
            wb.close()
            os.remove(data_fp_new)

            # 显示获取到的结果
            context = {'leave_students': tuple(leave_students)}
            return render_ht(request, 'leave_result.html', context)


def captcha_manage(request):
    """验证码管理页面"""
    # 权限严格控制
    if not request.user.is_staff:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 读取验证码字典
    bank_path = os.path.join('media', 'captcha_bank.json')
    with open(bank_path, encoding='utf-8') as f:
        captcha_bank_ds = f.read()
    captcha_bank_dict = json.loads(captcha_bank_ds)
    cb_list = list(captcha_bank_dict.items())

    # 分左中右三栏
    if len(cb_list) % 3 == 0:
        n_left = len(cb_list) // 3
    else:
        n_left = len(cb_list) // 3 + 1
    cb_left, cb_mid, cb_right = [], [], []
    for ob_left in cb_list[:n_left]:
        cb_left.append(ob_left)
    for ob_mid in cb_list[n_left:n_left * 2]:
        cb_mid.append(ob_mid)
    for ob_right in cb_list[n_left * 2:]:
        cb_right.append(ob_right)

    # 三栏合
    left_right = []
    for i in range(len(cb_left)):
        otb_left = cb_left[i]
        otb_mid = cb_mid[i]
        try:
            otb_right = cb_right[i]
        except IndexError:
            otb_right = (None, None)
        left_right.append((otb_left, otb_mid, otb_right))

    context = {'cb': tuple(left_right), 'n': len(cb_list)}
    return render_ht(request, 'captcha_manage.html', context)


def my_ai(request):
    """常用AI工具导航页"""
    return render_ht(request, 'my_ai.html', {})
