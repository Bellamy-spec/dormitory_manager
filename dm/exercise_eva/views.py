# -*- coding: utf-8 -*-
from django.shortcuts import render
from django.contrib.auth import logout, login, authenticate, update_session_auth_hash
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from django.contrib.auth.views import login_required
from .tools import DataTool
import datetime
from .models import ShortAbst, ExerciseScore, ECOScore
from .forms import ShortAbstForm, ExerciseScoreForm, ECOScoreForm
from su_manage.models import get_owners_pwd, Member, SchoolYear
from dm.scores.models import format_gc_students, get_students
import json
from long_leave.views import consult
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.conf import settings


# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}课间操评价系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """程序主页"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    return render_ht(request, 'exercise_eva/index.html', context={
        'is_manager': request.user.username in DT.super_users,
        'title': '{}课间操评价系统'.format(settings.USER_NAME),
    })


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        un = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=un, password=password)

        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到主页
            return HttpResponseRedirect(reverse('exercise_eva:index'))
        else:
            return render_ht(request, 'exercise_eva/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'exercise_eva/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('exercise_eva:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'exercise_eva/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'exercise_eva/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'exercise_eva/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'exercise_eva/set_pwd.html', {'err': ''})


def short_abst_manage(request, date_str=None):
    """短假学生管理"""
    # 先清除会话中的登记人代号字段
    if 'owner_code' in request.session.keys() and request.user.username in DT.super_users:
        del request.session['owner_code']

    # 取得日期和日期字符串（默认今天）
    if date_str:
        date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        cur = False
    else:
        date = datetime.date.today()
        date_str = datetime.datetime.strftime(date, '%Y-%m-%d')
        cur = True

    # 取得当日所有短假学生
    students = list(ShortAbst.objects.filter(date_added=date))

    # 按班级排序
    students.sort(key=lambda x: x.cs)
    students.sort(key=lambda x: x.grade_num)

    context = {'dt': date_str, 'students': tuple(students), 'cur': cur}
    if request.user.username in DT.super_users:
        # 管理员用户导向管理页面
        return render_ht(request, 'exercise_eva/short_abst_manage.html', context)
    else:
        # 普通用户导向基础版页面
        return render_ht(request, 'exercise_eva/short_abst.html', context)


def load_short_in(request):
    """短假登记进入"""
    # 先清除会话中的登记人代号字段
    if 'owner_code' in request.session.keys():
        del request.session['owner_code']

    # 取得今天的日期和日期字符串
    date = datetime.date.today()
    date_str = datetime.datetime.strftime(date, '%Y-%m-%d')

    # 取得文体部（'04'）、纪律部（'05'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('04'), **get_owners_pwd('05'), **get_owners_pwd('01')}

    if request.method != 'POST':
        # 未提交数据，加载登录页
        context = {'dt': date_str, 'err': ''}
        return render_ht(request, 'exercise_eva/load_short_in.html', context)
    else:
        # 对POST提交的数据作出处理
        code = request.POST.get('code', '')
        pwd = request.POST.get('pwd', '')

        # 验证代号
        if code not in opd.keys():
            return render_ht(request, 'exercise_eva/load_short_in.html', context={
                'dt': date_str,
                'err': '代号错误！',
            })

        # 验证密码
        if pwd != opd[code][0]:
            return render_ht(request, 'exercise_eva/load_short_in.html', context={
                'dt': date_str,
                'err': '密码错误！',
            })

        # 记录代号
        request.session['owner_code'] = code

        # 通过验证，可重定向至登记页
        return HttpResponseRedirect(reverse('exercise_eva:load_short_abst'))


def load_short_abst(request):
    """短假学生登记"""
    # 取得检查人代号
    if request.user.username in DT.super_users:
        owner_code = ''
        if 'owner_code' in request.session.keys():
            del request.session['owner_code']
    else:
        try:
            owner_code = request.session['owner_code']
        except KeyError:
            raise Http404

    # 取得今天的日期和日期字符串
    date = datetime.date.today()
    date_str = datetime.datetime.strftime(date, '%Y-%m-%d')

    # 班级学生对应字典及生成json字符串
    gs_dict = format_gc_students(grades=('高一', '高二'), logic=False)
    st_house = json.dumps(gs_dict, ensure_ascii=False)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ShortAbstForm()
    else:
        # 对POST提交的数据作出处理
        form = ShortAbstForm(request.POST)
        if form.is_valid():
            new_abst = form.save(commit=False)

            # 补全对象属性
            new_abst.owner = owner_code
            new_abst.gc_fill()
            new_abst.make_show()

            # 更新相关的评分属性
            try:
                # 尝试匹配评分对象
                score_ob = ExerciseScore.objects.filter(
                    date_added=date, class_and_grade=new_abst.class_and_grade)[0]
            except IndexError:
                # 未匹配到，可以直接忽略
                # print('不存在{}{}课间操评价记录'.format(date_str, new_abst.class_and_grade))
                pass
            else:
                score_ob.short_abst += 1
                score_ob.calculate_score()
                score_ob.save()

            # 保存对象并重定向至短假学生管理页
            new_abst.save()
            return HttpResponseRedirect(reverse('exercise_eva:short_abst_manage'))

    context = {'dt': date_str, 'form': form, 'st_house': st_house, 'code': owner_code}
    if request.user.username in DT.super_users:
        return render_ht(request, 'exercise_eva/load_short_abst.html', context)
    else:
        # 电子班牌使用情景
        return render_ht(request, 'exercise_eva/load_short_abst_1.html', context)


def del_short_abst(request, student_id):
    """删除短假学生"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的学生对象
    student = ShortAbst.objects.get(id=student_id)

    # 根据所属日期，判断要返回的页面
    date = student.date_added
    date_str = datetime.datetime.strftime(date, '%Y-%m-%d')
    if date == datetime.date.today():
        cur = True
    else:
        cur = False

    # 执行删除操作
    student.delete()

    # 重定向至查看页
    if cur:
        return HttpResponseRedirect(reverse('exercise_eva:short_abst_manage'))
    else:
        return HttpResponseRedirect(reverse('exercise_eva:date_short_abst', args=[date_str]))


def short_abst_dates(request):
    """短假学生日期列表"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取得日期列表
    date_list = []
    for student in ShortAbst.objects.all():
        if student.date_added != datetime.date.today():
            # 转为日期字符串再添加
            date_str = datetime.datetime.strftime(student.date_added, '%Y-%m-%d')
            date_list.append(date_str)

    # 去重、排序（倒序）
    date_list = list(set(date_list))
    date_list.sort(reverse=True)

    context = {'dates': tuple(date_list)}
    return render_ht(request, 'exercise_eva/short_abst_dates.html', context)


def up_load(request):
    """检查结果上报"""
    # 取得今天日期和日期字符串
    date = datetime.date.today()
    date_str = datetime.datetime.strftime(date, '%Y-%m-%d')

    # 取得文体部（'04'）、纪律部（'05'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('04'), **get_owners_pwd('05'), **get_owners_pwd('01')}

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ExerciseScoreForm()
    else:
        # 对POST提交的数据作出处理
        form = ExerciseScoreForm(request.POST)
        if form.is_valid():
            new_score = form.save(commit=False)

            # 检验分数
            if new_score.quality_score < 3 or new_score.quality_score > 5:
                return render_ht(request, 'exercise_eva/up_load.html', context={
                    'form': form,
                    'dt': date_str,
                    'err': '得分必须介于3.0~5.0之间',
                })

            # 检查代号
            if new_score.owner not in opd.keys():
                return render_ht(request, 'exercise_eva/up_load.html', context={
                    'form': form,
                    'dt': date_str,
                    'err': '代号错误！',
                })

            # 检查密码
            if new_score.pwd != opd[new_score.owner][0]:
                return render_ht(request, 'exercise_eva/up_load.html', context={
                    'form': form,
                    'dt': date_str,
                    'err': '密码错误！',
                })

            # 拉取应跑操人数
            normal = consult(new_score.class_and_grade)[2]
            if normal < 0:
                # 班主任未上报，默认为本班全体住宿人数减去长假人数
                normal = len(get_students(new_score.class_and_grade, logic=False))
                normal -= len(consult(new_score.class_and_grade)[0])
            new_score.to_come = normal

            # 拉取短假人数
            new_score.short_abst = len(ShortAbst.objects.filter(
                class_and_grade=new_score.class_and_grade, date_added=date))

            # 拉取学生会工作人员人数
            new_score.work_abst = len(Member.objects.filter(
                class_and_grade=new_score.class_and_grade,
                work_abst2=True,
                school_year=SchoolYear.objects.get(current=True),
            ))

            # 计算总分
            new_score.calculate_score()

            # 补全属性
            new_score.gc_fill()
            new_score.make_show()

            # 自动保留最新，覆盖重复
            before_score = ExerciseScore.objects.filter(
                class_and_grade=new_score.class_and_grade, date_added=date)
            if before_score:
                for ob in before_score:
                    ob.delete()

            # 保存，跳转至提示页
            new_score.save()
            return render_ht(request, 'exercise_eva/tip_success.html', context={
                'msg': '{}{}课间操评价成功！'.format(date_str, new_score.class_and_grade),
                'href': '/exercise_eva/up_load/',
            })

    context = {'form': form, 'dt': date_str, 'err': ''}
    return render_ht(request, 'exercise_eva/up_load.html', context)


def score_dates(request):
    """评价结果日期列表"""
    # 获取所有日期，以字符串形式存入列表
    date_list = []
    for score in ExerciseScore.objects.all():
        date_str = datetime.datetime.strftime(score.date_added, '%Y-%m-%d')
        if date_str not in date_list:
            date_list.append(date_str)

    # 日期排序（倒序）
    date_list.sort(reverse=True)

    context = {'dates': tuple(date_list)}
    return render_ht(request, 'exercise_eva/score_dates.html', context)


def see_score(request, date_str):
    """评价结果查看页面"""
    # 根据日期字符串取得日期对象
    date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

    # 取得所有评分对象
    scores = list(ExerciseScore.objects.filter(date_added=date))

    # 评分对象按年级-班级排序
    scores.sort(key=lambda x: x.cs)
    scores.sort(key=lambda x: x.grade_num)

    context = {'scores': tuple(scores), 'dt': date_str,
               'is_manager': request.user.username in DT.super_users}
    return render_ht(request, 'exercise_eva/see_score.html', context)


def del_score(request, score_id):
    """删除课间操评价记录"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的对象
    score = ExerciseScore.objects.get(id=score_id)

    # 获取所属日期字符串
    date_str = datetime.datetime.strftime(score.date_added, '%Y-%m-%d')

    # 执行删除操作
    score.delete()

    # 重定向至日期查看页面
    return HttpResponseRedirect(reverse('exercise_eva:see_score', args=[date_str]))


def mt_main(request):
    """月总计主页"""
    # 初始化月份存放列表
    ym_list = []

    # 记录包含的月份
    for score in ExerciseScore.objects.all():
        date = score.date_added
        ym = '{}-{}'.format(date.year, DT.str_two(date.month))
        ym_list.append(ym)

    # 去重
    ym_list = list(set(ym_list))

    context = {'yml': sorted(ym_list)}
    return render_ht(request, 'exercise_eva/mt_main.html', context)


def get_mtd(grade_num, ym):
    """取得某个年级、某个月的月总结大数据列表"""
    # 该年级所有班级
    grade = DT.reverse_grade_num[grade_num]
    gc_list = []
    for cs in range(1, 13):
        gc_list.append(grade + str(cs) + '班')

    # 建立班级-课间操评分对应字典，可变类型须逐项创建
    ys_dict = {}
    for gc in gc_list:
        ys_dict[gc] = {}

    # 记录所有可能的日期
    dates_list = []

    # 循环遍历该年级所有课间操评价记录
    for score in ExerciseScore.objects.filter(grade_num=grade_num):
        date = score.date_added
        score_ym = '{}-{}'.format(date.year, DT.str_two(date.month))

        # 日期转化为字符串
        date_str = datetime.datetime.strftime(date, '%Y-%m-%d')

        # 判断是否添加
        if score_ym == ym:
            gc = score.class_and_grade
            ys_dict[gc][date_str] = score.total_score

            # 添加日期
            dates_list.append(date_str)

    # 日期去重、排序
    dates_list = list(set(dates_list))
    dates_list.sort()

    # 计算平均分
    for sd in ys_dict.values():
        if sd:
            ave = np.mean(list(sd.values()))
            sd['mean'] = ave

    # 数据结构二次解析
    # 表头
    head = ['日期'] + dates_list + ['平均', '节能记录次数']
    large_data_list = [head]

    # 构建各班数据
    for gc in gc_list:
        gs = [gc]

        # 按照固定日期顺序写入数据
        for ds in dates_list:
            # 尝试匹配班级、日期数据
            try:
                score = ys_dict[gc][ds]
            except KeyError:
                gs.append('未反馈')
            else:
                gs.append(score)

        # 尝试匹配平均值
        try:
            ave_score = ys_dict[gc]['mean']
        except KeyError:
            gs.append('未反馈')
        else:
            gs.append(ave_score)

        # TODO:获取节能记录次数
        eco_count = 0
        for eco_score in ECOScore.objects.filter(class_and_grade=gc):
            # 记录年月
            eco_ym = '{}-{}'.format(eco_score.date_added.year, DT.str_two(eco_score.date_added.month))
            if eco_ym == ym and eco_score.desc != '':
                n = len(eco_score.desc.split('；'))
                eco_count += n
        gs.append(eco_count)

        # 班级数据加入大数据
        large_data_list.append(gs)

    # 返回计算的大数据列表
    return large_data_list


def mt(request, grade_num, ym):
    """月总结"""
    # 获取大数据列表
    ldl = get_mtd(grade_num, ym)

    # 制定标题
    grade = DT.reverse_grade_num[grade_num]
    title = '{}年级{}课间操月总结'.format(grade, ym)

    context = {'ldl': ldl, 'title': title, 'grade_num': grade_num, 'ym': ym}
    return render_ht(request, 'exercise_eva/mt.html', context)


def export_mt(request, grade_num, ym):
    """导出课间操月总结"""
    # 获取大数据列表
    ldl = get_mtd(grade_num, ym)

    # 制定标题
    grade = DT.reverse_grade_num[grade_num]
    title = '{}年级{}班会月总结'.format(grade, ym)

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title

    # 写入数据
    row = 2
    for line in ldl:
        col = 1
        for data in line:
            st.cell(row=row, column=col).value = data
            col += 1
        row += 1

    # 标题行合并单元格
    rs = 'A1:{}1'.format(get_column_letter(len(ldl[0])))
    st.merge_cells(range_string=rs)

    # 输出
    return write_out(wb)


def eco_upload(request):
    """节能检查结果上报"""
    # 取得今天日期和日期字符串
    date = datetime.date.today()
    date_str = datetime.datetime.strftime(date, '%Y-%m-%d')

    # 取得文体部（'04'）、纪律部（'05'）和主席团（'01'）成员代号与检查码对应字典
    opd = {**get_owners_pwd('04'), **get_owners_pwd('05'), **get_owners_pwd('01')}

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ECOScoreForm()
    else:
        # 对POST提交的数据作出处理
        form = ECOScoreForm(request.POST)
        if form.is_valid():
            new_eco = form.save(commit=False)

            # 检查代号
            if new_eco.owner not in opd.keys():
                return render_ht(request, 'exercise_eva/eco_upload.html', context={
                    'form': form,
                    'dt': date_str,
                    'err': '代号错误！',
                    'descs': tuple(DT.desc.items()),
                })

            # 检查密码
            if new_eco.pwd != opd[new_eco.owner][0]:
                return render_ht(request, 'exercise_eva/up_load.html', context={
                    'form': form,
                    'dt': date_str,
                    'err': '密码错误！',
                    'descs': tuple(DT.desc.items()),
                })

            # 完善问题描述
            desc_list = []
            for d in DT.desc.keys():
                if request.POST.get(d, ''):
                    desc_list.append(DT.desc[d])
            new_eco.update_desc(desc_list)

            # 补全对象属性
            new_eco.make_show()
            new_eco.gc_fill()

            # 自动保留最新
            before_score = ECOScore.objects.filter(
                class_and_grade=new_eco.class_and_grade, date_added=date)
            if before_score:
                for ob in before_score:
                    ob.delete()

            # 保存，跳转至提示页
            new_eco.save()
            return render_ht(request, 'exercise_eva/tip_success.html', context={
                'msg': '{}{}节能评价成功！'.format(date_str, new_eco.class_and_grade),
                'href': '/exercise_eva/eco_upload/',
            })

    context = {'form': form, 'dt': date_str, 'err': '', 'descs': tuple(DT.desc.items())}
    return render_ht(request, 'exercise_eva/eco_upload.html', context)


def eco_dates(request):
    """节能记录日期列表"""
    # 获取所有日期，以字符串形式存入列表
    date_list = []
    for score in ECOScore.objects.all():
        date_str = datetime.datetime.strftime(score.date_added, '%Y-%m-%d')
        if date_str not in date_list:
            date_list.append(date_str)

    # 日期排序（倒序）
    date_list.sort(reverse=True)

    context = {'dates': tuple(date_list)}
    return render_ht(request, 'exercise_eva/eco_dates.html', context)


def see_eco(request, date_str):
    """评价结果查看页面"""
    # 根据日期字符串取得日期对象
    date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

    # 取得所有评分对象
    scores = list(ECOScore.objects.filter(date_added=date))

    # 评分对象按年级-班级排序
    scores.sort(key=lambda x: x.cs)
    scores.sort(key=lambda x: x.grade_num)

    context = {'scores': tuple(scores), 'dt': date_str,
               'is_manager': request.user.username in DT.super_users}
    return render_ht(request, 'exercise_eva/see_eco.html', context)


def del_eco(request, score_id):
    """删除节能记录"""
    # 验证权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的对象
    score = ECOScore.objects.get(id=score_id)

    # 获取所属日期字符串
    date_str = datetime.datetime.strftime(score.date_added, '%Y-%m-%d')

    # 执行删除操作
    score.delete()

    # 重定向至日期查看页面
    return HttpResponseRedirect(reverse('exercise_eva:see_eco', args=[date_str]))
