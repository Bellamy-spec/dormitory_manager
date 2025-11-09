"""
    作者：郭雨健
    功能：解析日志为更易于人类理解的表格形式
    版本：1.0
    日期：2024.12.20
"""
from openpyxl import Workbook
import sys


# http请求方法类型大全
METHOD_LIST = ['GET', 'POST', 'HEAD', 'PUT', 'DELETE', 'PATCH', 'CONNECT', 'OPTIONS', 'TRACE']


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def get_ip_log(st, row):
    """获取IP访问情况写入工作表"""
    # 生成文件名
    try:
        d = sys.argv[1]
    except IndexError:
        fp = '/var/log/apache2/ip-access.log'
    else:
        fp = '/var/log/apache2/ip-access.log.{}'.format(d)

    # 加载日志文件
    try:
        with open(fp) as f:
            log_text = f.read()
    except FileNotFoundError:
        # 提示文件不存在并退出
        print('文件{}不存在'.format(fp))
        sys.exit()

    # 逐行写入
    for record in log_text.split('\n')[:-1]:
        rl = record.split(' ')
        st.cell(row=row, column=1).value = 'IP访问'
        st.cell(row=row, column=2).value = rl[0]
        st.cell(row=row, column=3).value = rl[3][1:]

        # 格式化处理
        method = rl[5][1:]
        url = rl[6]
        code = rl[8]

        if method not in METHOD_LIST:
            code = url
            url = method
            method = '--'

        st.cell(row=row, column=4).value = method
        st.cell(row=row, column=5).value = url
        st.cell(row=row, column=6).value = code

        # 下一行
        row += 1

    return row


def get_cn_log(st, row):
    """获取域名访问情况写入工作表"""
    # 生成文件名
    try:
        d = sys.argv[1]
    except IndexError:
        fp = '/var/log/apache2/zz106gz-access.log'
    else:
        fp = '/var/log/apache2/zz106gz-access.log.{}'.format(d)

    # 加载日志文件
    try:
        with open(fp) as f:
            log_text = f.read()
    except FileNotFoundError:
        # 提示文件不存在并退出
        print('文件{}不存在'.format(fp))
        sys.exit()

    # # 调试输出
    # fo = open('log_out.txt', 'w')

    # 逐行写入
    for record in log_text.split('\n')[:-1]:
        rl = record.split(' ')
        st.cell(row=row, column=1).value = '域名访问'
        st.cell(row=row, column=2).value = rl[0]
        st.cell(row=row, column=3).value = rl[3][1:]

        # 格式化处理
        method = rl[5][1:]
        url = rl[6]
        code = rl[8]

        if method not in METHOD_LIST:
            code = url
            url = method
            method = '--'

        st.cell(row=row, column=4).value = method
        st.cell(row=row, column=5).value = url
        st.cell(row=row, column=6).value = code

        # 下一行
        row += 1

        # # 调试输出
        # fo.write(str(rl) + '\n')

    # # 调试文件关闭
    # fo.close()

    return row


def main():
    """主函数"""
    # 打开输出文件
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '访问方式'
    st['B1'].value = '访问者IP'
    st['C1'].value = '访问时间'
    st['D1'].value = '请求方法'
    st['E1'].value = '请求链接'
    st['F1'].value = '状态码'

    # 写入主体
    get_ip_log(st, get_cn_log(st, 2))

    # 设置列宽
    set_width_dict(st, {'A': 10, 'B': 19, 'C': 22, 'D': 9, 'E': 75, 'F': 8})

    # 保存输出，关闭文件
    wb.save('access_log.xlsx')
    wb.close()


if __name__ == '__main__':
    main()
