from django.shortcuts import render
from .tools import DataTool
from .models import LongLeaveRecord, AbsentStudents, ClassInfo, StayTask, StayRecord
import datetime
from django.http import Http404, HttpResponseRedirect, HttpResponse
from .forms import LongLeaveForm, ClassForm, ChangeTotalForm, AddAbsentForm, StayTaskForm
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border, Alignment
from io import BytesIO
from django.contrib.auth.views import login_required
from dm.scores.models import format_gc_students, all_students, st_gc, NewStudent
from dm.scores.data import Data
from django.conf import settings
from exercise_eva.models import ExerciseScore
import json
from selenium.webdriver.common.by import By
import base64
import time
from urllib3.exceptions import MaxRetryError
import os


# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def all_cs_date():
    """返回所有班级，日期二元组列表"""
    cs_date_list = []
    for cs in ClassInfo.objects.all():
        cs_date_list.append((cs.class_and_grade, cs.date))
    return cs_date_list


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """主页"""
    # 自动删除过期人员
    for record in LongLeaveRecord.objects.all():
        if datetime.datetime.date(datetime.datetime.today()) > record.end_date:
            record.delete()

    context = {
        'is_manager': request.user.username in DT.managers,
        'title': '{}课间操长期假人员管理系统'.format(settings.USER_NAME),
    }
    return render_ht(request, 'long_leave/index.html', context)


def see(request):
    """查看所有长期假人员"""
    # 此页为管理页，非管理员无权限访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出所有记录
    rs = LongLeaveRecord.objects.all().order_by('cs')

    # 列表化
    rs = list(rs)

    # 排序
    rs.sort(key=lambda x: DT.grades_dict[x.grade])

    # 元组化
    rs = tuple(rs)

    # 制定标题
    title = '课间操长期假人员管理'

    context = {'title': title, 'rs': rs}
    return render_ht(request, 'long_leave/see.html', context)


def add(request):
    """新增"""
    # 非管理员无权限访问
    if request.user.username not in DT.managers:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = LongLeaveForm()
    else:
        # 对POST提交的数据作出处理
        form = LongLeaveForm(request.POST)
        if form.is_valid():
            new = form.save(commit=False)

            # 截止日期不能早于今天
            if new.end_date < datetime.date.today():
                return render_ht(request, 'long_leave/add.html', context={
                    'form': form,
                    'err': '截止日期不能早于今天',
                    'st_house': format_gc_students(grades=('高一', '高二')),
                    'all_students': all_students(grades=('高一', '高二')),
                    'lg': Data().logic_grade,
                    'st_gc': st_gc(grades=('高一', '高二')),
                })

            # 年级班级信息的进一步完善
            new.grade = DT.get_grade(new.class_and_grade)
            new.cs = DT.get_cs(new.class_and_grade)

            # 保存对象，重定向至管理页
            new.save()
            return HttpResponseRedirect(reverse('long_leave:manage'))

    return render_ht(request, 'long_leave/add.html', context={
        'form': form,
        'err': '',
        'st_house': format_gc_students(grades=('高一', '高二')),
        'all_students': all_students(grades=('高一', '高二')),
        'lg': Data().logic_grade,
        'st_gc': st_gc(grades=('高一', '高二')),
    })


def delete(request, r_id):
    """删除记录"""
    # 非管理员无权限访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的记录
    r = LongLeaveRecord.objects.get(id=r_id)

    # 执行删除操作
    r.delete()

    # 重定向到管理页
    return HttpResponseRedirect(reverse('long_leave:manage'))


def change_date(request, r_id):
    """改变日期和请假类型"""
    # 非管理员无权限访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要改变的记录
    r = LongLeaveRecord.objects.get(id=r_id)

    # 制定标题
    title = '更改{}的请假期限和请假类型'.format(r.name)

    if request.method == 'POST':
        # 对POST提交的数据作出处理
        new_date_str = request.POST.get('ed', '')
        if new_date_str:
            # 字符串转换日期
            new_date = datetime.datetime.strptime(new_date_str, '%Y-%m-%d')
            new_date = datetime.datetime.date(new_date)

            # 不能早于今天
            if new_date < datetime.date.today():
                return render_ht(request, 'long_leave/change.html', context={
                    'title': title,
                    'err': '截止日期不能早于当前日期',
                    'r': r,
                })

            # 更改截止日期
            r.end_date = new_date

        # 读取并更改请假类型，保存
        if request.POST.get('tp', ''):
            r.tp = True
        else:
            r.tp = False
        r.save()

        # 重定向到管理页
        return HttpResponseRedirect(reverse('long_leave:manage'))

    return render_ht(request, 'long_leave/change.html', {'title': title, 'r': r, 'err': ''})


def consult(gc):
    """根据班级查询长假人员，以列表形式返回"""
    # 初始化人员列表
    leavers, absts = [], []

    # 逐个添加长期假人员
    # TODO:修改显示
    for r in LongLeaveRecord.objects.filter(class_and_grade=gc):
        if not r.desc:
            if r.tp:
                # 备注请假类型
                leavers.append('{}({})'.format(r.name, '在班'))
            else:
                # 无备注，直接添加姓名
                leavers.append(r.name)
        else:
            if r.tp:
                # 双备注
                leavers.append('{}({})({})'.format(r.name, '在班', r.desc))
            else:
                # 括号内显示备注
                leavers.append('{}({})'.format(r.name, r.desc))

    # 逐个添加请假不在校人员
    cs = ClassInfo.objects.filter(class_and_grade=gc, date=datetime.date.today())
    if cs:
        for abst in AbsentStudents.objects.filter(class_belong=cs[0]):
            absts.append(abst.name)

        # 计算正常应跑操人数
        normal = cs[0].total - len(set(leavers + absts))

        return leavers, absts, normal, cs[0].total
    else:
        return leavers, absts, -1, -1


def consult_all():
    """以字典形式返回班级与长假人员关系"""
    # 初始化字典
    leaver_dict = {}

    # 循环遍历12个班级，写入字典
    for gc in DT.get_all_classes():
        leaver_list, abst_list, normal, total = consult(gc[0])
        leaver_str = ', '.join(leaver_list)
        abst_str = ', '.join(abst_list)
        leaver_dict[gc[0]] = (total, leaver_str, abst_str, normal)

    return leaver_dict


def look(request):
    """查看"""
    # 制定标题
    try:
        today_str = datetime.datetime.strftime(datetime.date.today(), '%Y年%m月%d日')
        title = '{}两操请假人员情况'.format(today_str)
    except UnicodeEncodeError:
        # 另一种生成等效标题的方式
        today = datetime.date.today()
        title = '{}年{}月{}日两操请假人员情况'.format(today.year, today.month, today.day)

    context = {'title': title, 'data': tuple(consult_all().items())}
    return render_ht(request, 'long_leave/look.html', context)


def export(request):
    """导出两操长期假人员为excel表格"""
    # 打开文件，定位表格
    wb = Workbook()
    st = wb.active

    # 写入表头，合并单元格
    try:
        today_str = datetime.datetime.strftime(datetime.date.today(), '%Y年%m月%d日')
        st['A1'].value = '{}两操请假名单'.format(today_str)
    except UnicodeEncodeError:
        # 另一种生成等效标题的方式
        today = datetime.date.today()
        st['A1'].value = '{}年{}月{}日两操请假人员情况'.format(today.year, today.month, today.day)
    st.merge_cells(range_string='A1:E1')

    # 设置表头格式
    st['A1'].font = Font(size=28, bold=True)
    st['A1'].alignment = Alignment(horizontal='center')

    # 主体部分格式
    ft_main = Font(size=11)

    # 写入首行
    st['A2'].value = '班级'
    st['B2'].value = '应到人数'
    st['C2'].value = '长期假学生'
    st['D2'].value = '请假不在校学生'
    st['E2'].value = '正常应跑操人数'
    # st['F2'].value = '实到'
    # st['G2'].value = '体育委员签字确认'

    # 设置字体格式
    for c in range(1, 7):
        st.cell(row=2, column=c).font = ft_main

    # 初始行
    row = 3

    # 开始循环遍历写入数据
    for gc, data in consult_all().items():
        # 写入班级
        st.cell(row=row, column=1).value = gc

        # 写入应到人数
        if data[0] < 0:
            st.cell(row=row, column=2).value = '未上报'
        else:
            st.cell(row=row, column=2).value = data[0]

        # 写入长期假人员
        st.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=3).value = data[1]

        # 写入请假不在校人员
        st.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
        st.cell(row=row, column=4).value = data[2]

        # 写入正常跑操人数
        if data[3] < 0:
            st.cell(row=row, column=5).value = '未上报'
        else:
            st.cell(row=row, column=5).value = data[3]

        # 设置字体格式
        for c in range(1, 6):
            st.cell(row=row, column=c).font = ft_main

        # 下一行
        row += 1

    # 添加边框
    add_border(st, start_row=2, end_row=row - 1, start_column=1, end_column=5)

    # TODO:调整列宽
    set_width_dict(st, {'A': 10, 'B': 10, 'C': 33, 'D': 33, 'E': 15})

    # TODO:调整行高
    for r in range(2, row):
        st.row_dimensions[r].height = 26

    # 设置页边距（单位：英寸）
    st.page_margins.left = 0.25
    st.page_margins.right = 0.25

    return write_out(wb)


def index_out(request):
    """上级页"""
    context = {'title': '{}请假与在校人员管理系统'.format(settings.USER_NAME)}
    return render_ht(request, 'long_leave/index_out.html', context)


def daily(request):
    """请假上报系统首页"""
    return render_ht(request, 'long_leave/daily.html', {})


@ login_required()
def send_up(request):
    """上报"""
    # 非操作员无权限访问
    if request.user.username not in DT.operators:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ClassForm()
    else:
        # 对POST提交的数据作出处理
        form = ClassForm(request.POST)
        if form.is_valid():
            new_class = form.save(commit=False)

            # 不能晚于当前日期
            if new_class.date > datetime.date.today():
                return render_ht(request, 'long_leave/send_up.html', context={
                    'form': form,
                    'have_link': False,
                    'err': '填报日期不能晚于今天',
                    'cs': None,
                    'cs_house': DT.get_cs_dict(),
                    'cs_house_reverse': DT.get_reverse_cs_dict(),
                    'cs_str': '0-0',
                    'absts': (),
                    'ct': 0,
                    'total': 0,
                    'st_house': format_gc_students(),
                    'lg': Data().logic_grade,
                })

            # TODO:补充完整实例属性
            new_class.grade_num = DT.grades_dict[new_class.grade]
            new_class.class_and_grade = new_class.grade + new_class.cs
            new_class.cs_num = DT.get_cs(new_class.class_and_grade)
            new_class.come = new_class.total

            # TODO:验证是否已存在
            cs_date = (new_class.class_and_grade, new_class.date)
            if cs_date in all_cs_date():
                # 生成班级，日期字符串
                gc = new_class.class_and_grade
                date_str = datetime.datetime.strftime(new_class.date, '%Y-%m-%d')
                return render_ht(request, 'long_leave/send_up.html', context={
                    'form': form,
                    'have_link': True,
                    'err': '已存在{}{}的记录，请勿重复提交！'.format(gc, date_str),
                    'cs': ClassInfo.objects.filter(class_and_grade=gc, date=new_class.date)[0],
                    'cs_house': DT.get_cs_dict(),
                    'cs_house_reverse': DT.get_reverse_cs_dict(),
                    'cs_str': '0-0',
                    'absts': (),
                    'ct': 0,
                    'total': 0,
                    'st_house': format_gc_students(),
                    'lg': Data().logic_grade,
                })

            # 须先保存才能绑定请假学生
            new_class.save()

            # TODO:绑定请假学生
            for i in range(int(request.POST.get('count', ''))):
                # 获取姓名字符串
                name_index = 'name' + str(i)
                st_name = request.POST.get(name_index, '')

                # 检验姓名是否为空，若为空则不执行任何操作
                if st_name == '':
                    continue

                # 获取请假原因字符串
                reason_index = 'reason' + str(i)
                st_reason = request.POST.get(reason_index, '')

                # 请假原因不能为空
                if st_reason == '':
                    continue

                # 创建新的学生对象
                abst = AbsentStudents()
                abst.name = st_name
                abst.reason = st_reason
                abst.class_belong = new_class

                # 所属班级对象属性改变
                new_class.come -= 1

                # 保存学生对象
                abst.save()

            # TODO:更新课间操评分属性
            try:
                # 尝试匹配评分对象
                score_ob = ExerciseScore.objects.filter(
                    date_added=new_class.date, class_and_grade=new_class.class_and_grade)[0]
            except IndexError:
                # 未匹配到，直接忽略
                pass
            else:
                score_ob.to_come = new_class.come - len(consult(new_class.class_and_grade)[0])
                score_ob.calculate_score()
                score_ob.save()

            # 提交班级对象
            new_class.save()
            return HttpResponseRedirect(reverse('long_leave:show_cs', args=[new_class.id]))

    return render_ht(request, 'long_leave/send_up.html', context={
        'form': form,
        'have_link': False,
        'err': '',
        'cs': None,
        'cs_house': DT.get_cs_dict(),
        'cs_house_reverse': DT.get_reverse_cs_dict(),
        'cs_str': '0-0',
        'absts': (),
        'ct': 0,
        'total': 0,
        'st_house': format_gc_students(),
        'lg': Data().logic_grade,
    })


def dates(request, grade_num):
    """获取日期列表"""
    dts = []
    for cs in ClassInfo.objects.filter(grade_num=grade_num):
        dts.append(cs.date)

    # 去重，排序，可哈希
    dts = list(set(dts))
    dts.sort(reverse=True)
    dts = tuple(dts)

    context = {'dates': dts, 'grade': grade_num}
    return render_ht(request, 'long_leave/dates.html', context)


def grades(request):
    """一天的年级列表"""
    return render_ht(request, 'long_leave/grades.html', {'grades': tuple(DT.grades_dict.items())})


def one_day(request, grade_num, date_str):
    """查看一个年级一天的情况"""
    # 初始化所有班级列表
    grade_str = DT.grades_dict_reverse[grade_num]
    cs_list = []
    for i in range(1, 13):
        cs_list.append(grade_str + str(i) + '班')

    # 如果年级为高三，增加13班
    if grade_str == '高三':
        cs_list.append('高三13班')

    # 根据日期字符串生成日期
    dt = datetime.datetime.strptime(date_str, '%Y-%m-%d')

    # 遍历对应记录，作出操作
    cs_info = []
    for cs in ClassInfo.objects.filter(grade_num=grade_num, date=dt):
        # 生成请假学生姓名字符串
        abst_list = []
        for abst in AbsentStudents.objects.filter(class_belong=cs):
            abst_list.append(abst.name)
        abst_str = ', '.join(abst_list)

        # 去除已完成
        if cs.class_and_grade in cs_list:
            cs_list.remove(cs.class_and_grade)

        # 加入信息
        cs_info.append((cs, abst_str))

    # 制定标题
    title = '{}{}年级各班在校与请假人员情况'.format(date_str, grade_str)

    # 排序
    cs_info.sort(key=lambda x: x[0].cs_num)

    context = {
        'title': title,
        'css': tuple(cs_info),
        'not_done': tuple(cs_list),
        'grade': grade_num,
    }
    return render_ht(request, 'long_leave/records.html', context)


def show_cs(request, cs_id):
    """查看某个班级具体记录"""
    # 取出要查看的班级对象
    cs = ClassInfo.objects.get(id=cs_id)

    # 取出该班级所有请假学生对象
    absts = AbsentStudents.objects.filter(class_belong=cs)

    # 制定标题
    title = '{}{}在校与请假人员情况'.format(cs.date, cs.class_and_grade)

    context = {
        'title': title,
        'cs': cs,
        'absts': absts,
        'is_operator': request.user.username in DT.operators,
    }
    return render_ht(request, 'long_leave/csinfo.html', context)


def delete_st(request, st_id):
    """删除请假学生"""
    # 非操作员无权限访问
    if request.user.username not in DT.operators:
        raise Http404

    # 取出要删除的学生及其所属班级对象
    st = AbsentStudents.objects.get(id=st_id)
    cs = st.class_belong

    # 执行删除操作
    st.delete()

    # 所属班级对象的属性改变，保存
    cs.come += 1
    cs.save()

    # 重定向到班级详情页
    return HttpResponseRedirect(reverse('long_leave:show_cs', args=[cs.id]))


def change_total(request, cs_id):
    """更改已提交班级的应到人数"""
    # 非操作员无权限访问
    if request.user.username not in DT.operators:
        raise Http404

    # 取出要更改的班级对象
    cs = ClassInfo.objects.get(id=cs_id)

    # 生成日期字符串
    date_str = datetime.datetime.strftime(cs.date, '%Y-%m-%d')

    # 制定标题
    title = '更改{}{}应到人数'.format(cs.class_and_grade, date_str)

    # 计算请假人数
    n = cs.total - cs.come

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ChangeTotalForm(instance=cs)
    else:
        # 根据POST提交的数据作出相应的修改
        form = ChangeTotalForm(instance=cs, data=request.POST)
        if form.is_valid():
            cs = form.save(commit=False)

            # 重新计算实到人数
            cs.come = cs.total - n

            # 保存对象，重定向至班级详情页
            cs.save()
            return HttpResponseRedirect(reverse('long_leave:show_cs', args=[cs.id]))

    context = {'title': title, 'form': form, 'cs': cs}
    return render_ht(request, 'long_leave/change_total.html', context)


def add_st(request, cs_id):
    """已提交班级新增请假学生"""
    # 非操作员无权限访问
    if request.user.username not in DT.operators:
        raise Http404

    # 取出需要新增请假学生的班级对象
    cs = ClassInfo.objects.get(id=cs_id)

    # 取得逻辑班级字符串
    lg = Data().logic_grade[cs.grade]
    if len(cs.cs) < 3:
        lcs = '0' + cs.cs
    else:
        lcs = cs.cs
    lgc = lg + lcs

    # 生成日期字符串
    date_str = datetime.datetime.strftime(cs.date, '%Y-%m-%d')

    # 制定标题
    title = '新增{}{}请假学生'.format(cs.class_and_grade, date_str)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = AddAbsentForm()
    else:
        # 对POST提交的数据作出处理
        form = AddAbsentForm(request.POST)
        if form.is_valid():
            new_st = form.save(commit=False)

            # 绑定所属班级，保存
            new_st.class_belong = cs
            new_st.save()

            # 所属班级对象属性改变
            cs.come -= 1
            cs.save()

            # 重定向至班级详情页
            return HttpResponseRedirect(reverse('long_leave:show_cs', args=[cs.id]))

    context = {
        'form': form,
        'title': title,
        'cs': cs,
        'st_list': tuple(format_gc_students()[lgc]),
    }
    return render_ht(request, 'long_leave/add_st.html', context)


def get_last(request, cs_str, total):
    """获取某个班级最近一次请假学生"""
    # 解析班级名称
    cs_name = DT.get_reverse_cs_dict()[cs_str]

    # 获取确定对象
    try:
        cs_ob = list(ClassInfo.objects.filter(class_and_grade=cs_name).order_by('id'))[-1]
    except IndexError:
        # 系统内还没有该班级数据，给出提示
        return render_ht(request, 'long_leave/send_up.html', context={
            'form': ClassForm(),
            'have_link': False,
            'err': '{}是第一次上报，无上一次数据'.format(cs_name),
            'cs': None,
            'cs_house': DT.get_cs_dict(),
            'cs_house_reverse': DT.get_reverse_cs_dict(),
            'cs_str': cs_str,
            'absts': (),
            'ct': 0,
            'total': total,
            'st_house': format_gc_students(),
            'lg': Data().logic_grade,
        })
    else:
        # 获取所有请假学生信息
        absts = []
        i = 0
        for ob in AbsentStudents.objects.filter(class_belong=cs_ob):
            absts.append((ob.name, ob.reason, 'name' + str(i), 'reason' + str(i),
                          'd' + str(i), 'del({})'.format(i), 'del' + str(i)))
            i += 1

        # 查到的信息传回前端
        return render_ht(request, 'long_leave/send_up.html', context={
            'form': ClassForm(),
            'have_link': False,
            'err': '',
            'cs': None,
            'cs_house': DT.get_cs_dict(),
            'cs_house_reverse': DT.get_reverse_cs_dict(),
            'cs_str': cs_str,
            'absts': tuple(absts),
            'ct': len(absts),
            'total': total,
            'st_house': format_gc_students(),
            'lg': Data().logic_grade,
        })


def stay_dorm_main(request):
    """留宿上报主页"""
    # 获取所有已完成和未完成任务对象
    done_tasks = StayTask.objects.filter(done=True)
    not_done_tasks = StayTask.objects.filter(done=False)

    context = {'is_manager': request.user.username in DT.managers, 'dt': done_tasks, 'ndt': not_done_tasks}
    return render_ht(request, 'long_leave/stay_dorm_main.html', context)


def public_task(request):
    """发布新的留宿上报任务"""
    # 限制非管理员用户访问
    if request.user.username not in DT.managers:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = StayTaskForm()
    else:
        # 对POST提交的数据作出处理
        form = StayTaskForm(request.POST)
        if form.is_valid():
            new_task = form.save(commit=False)

            # TODO:设置包含年级
            grade1 = request.POST.get('grade1', '')
            grade2 = request.POST.get('grade2', '')
            grade3 = request.POST.get('grade3', '')

            if grade1 or grade2 or grade3:
                # 须至少勾选一个年级
                grade_list = []
                if grade1:
                    grade_list.append('高一')
                if grade2:
                    grade_list.append('高二')
                if grade3:
                    grade_list.append('高三')
            else:
                # 给出错误提示
                return render_ht(request, 'long_leave/public.html', context={
                    'form': form,
                    'err': '请至少勾选一个年级',
                })

            # 补全属性
            new_task.set_date_str()
            new_task.set_grade_str(grade_list)

            # 保存，重定向
            new_task.save()
            return HttpResponseRedirect(reverse('long_leave:task_main', args=[new_task.id]))

    context = {'form': form, 'err': ''}
    return render_ht(request, 'long_leave/public.html', context)


def task_main(request, task_id):
    """任务主页"""
    # 取出相应的任务对象
    task = StayTask.objects.get(id=task_id)

    # 取出该任务下所有已上报的留宿信息记录
    records = list(StayRecord.objects.filter(task_belong=task))

    # 记录排序
    records.sort(key=lambda x: x.bed)
    records.sort(key=lambda x: x.dorm)
    records.sort(key=lambda x: x.cs)
    records.sort(key=lambda x: x.grade_num)

    # 制定标题
    if task.done:
        title = '{}留宿学生-已完成上报'.format(task.date_str)
        btn_name = '取消标记已完成'
    else:
        title = '{}留宿学生-上报中'.format(task.date_str)
        btn_name = '标记为已完成'

    context = {
        'task': task,
        'is_manager': request.user.username in DT.managers,
        'is_operator': request.user.username in DT.operators,
        'title': title,
        'btn_name': btn_name,
        'records': records,
    }
    return render_ht(request, 'long_leave/task_main.html', context)


def change_done(request, task_id):
    """变更任务状态"""
    # 限制非管理员用户访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要操作的任务对象
    task = StayTask.objects.get(id=task_id)

    # 执行变更操作
    task.done = not task.done

    # 保存，重定向
    task.save()
    return HttpResponseRedirect(reverse('long_leave:task_main', args=[task_id]))


def delete_task(request, task_id):
    """删除任务"""
    # 限制非管理员用户访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的任务对象
    task = StayTask.objects.get(id=task_id)

    # 执行删除操作
    task.delete()

    # 重定向至留宿上报主页（上级页）
    return HttpResponseRedirect(reverse('long_leave:stay_dorm_main'))


def delete_stay_record(request, record_id):
    """删除留宿学生记录"""
    # 限制非管理员用户访问
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的记录对象和所属任务对象
    record = StayRecord.objects.get(id=record_id)
    task = record.task_belong

    # 执行删除操作
    record.delete()

    # 重定向至任务主页（当前页）
    return HttpResponseRedirect(reverse('long_leave:task_main', args=[task.id]))


@ login_required()
def stay_send_up(request, task_id, err):
    """上报留宿信息"""
    # 非操作员无权限访问
    if request.user.username not in DT.operators:
        raise Http404

    # 取出相应的任务对象
    task = StayTask.objects.get(id=task_id)

    # 已完成任务不可上报
    if task.done:
        raise Http404

    # 生成班级选项
    grade_list = task.get_grade_list()
    gc_options = DT.get_all_classes(grades=grade_list)

    # 班级学生对应字典
    st_dict = format_gc_students(grades=tuple(grade_list), logic=False, include_id=True)
    st_house = json.dumps(st_dict, ensure_ascii=False)

    if request.method == 'POST':
        # 对POST提交的数据作出处理
        gc = request.POST.get('gc', '')

        # 班级不允许为空
        if not gc:
            return render_ht(request, 'long_leave/stay_send_up.html', context={
                'task': task,
                'gc_options': gc_options,
                'st_house': st_house,
                'err': '请选择班级',
                'gc_selected': '',
                'image': DT.current_captcha_str,
            })

        # 读取学生对象id存为列表
        st_list = st_dict[gc]
        checked_id_list = []
        for student in st_list:
            sid = 's{}'.format(student[1])
            if request.POST.get(sid, ''):
                checked_id_list.append(student[1])

        # 至少选择一个学生
        if not checked_id_list:
            return render_ht(request, 'long_leave/stay_send_up.html', context={
                'task': task,
                'gc_options': gc_options,
                'st_house': st_house,
                'err': '请至少勾选一个学生',
                'gc_selected': gc,
                'image': DT.current_captcha_str,
            })

        if DT.tr < DT.max_tr:
            # 试验次数加1
            DT.tr += 1

            # 用户名、密码、验证码
            un = 'admin'
            pwd = 'Ffkj-102064'
            captcha = request.POST.get('captcha', '').lower()

            try:
                # 依次输入
                DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                     '/div[1]/input').send_keys(un)
                DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                     '/div[2]/input').send_keys(pwd)
                DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]'
                                                     '/div[3]/input').send_keys(captcha)

                # 点击登录
                DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]/button').click()
            except AttributeError:
                # 重试
                return HttpResponseRedirect(reverse('long_leave:stay_send_up', args=[task_id, 2]))
            except MaxRetryError:
                # 也重试
                return HttpResponseRedirect(reverse('long_leave:stay_send_up', args=[task_id, 2]))
            else:
                time.sleep(2)

            # 判断是否成功登入
            if '登录' in DT.b.page_source:
                # 未成功登录，返回验证码错误提示
                return HttpResponseRedirect(reverse('long_leave:stay_send_up', args=[task_id, 1]))
            else:
                # 成功登录，记录正确的验证码
                captcha_str = DT.current_captcha_str
                bank_path = os.path.join('media', 'captcha_bank.json')

                # 首次使用确定文件存在
                if not os.path.exists(bank_path):
                    # print('首次创建！')
                    with open(bank_path, 'w', encoding='utf-8') as ff:
                        ff.write(json.dumps({}))

                with open(bank_path, encoding='utf-8') as cfi:
                    captcha_bank_ds = cfi.read()

                try:
                    captcha_bank_dict = json.loads(captcha_bank_ds)
                except json.JSONDecodeError:
                    # 出现未知原因错误，跳过记录验证码，直接显示获取结果
                    pass
                else:
                    if captcha_str not in captcha_bank_dict.keys():
                        captcha_bank_dict[captcha_str] = captcha
                        with open(bank_path, 'w', encoding='utf-8') as cfo:
                            cfo.write(json.dumps(captcha_bank_dict))

                # 退出浏览器
                time.sleep(2)
                DT.b.quit()
        else:
            # 重置试验次数之后直接放行
            DT.tr = 0

        # 逐个创建留宿学生记录对象
        for id_num in checked_id_list:
            new_stay_record = StayRecord()

            # 关联任务对象
            new_stay_record.task_belong = task

            # 设置年级、班级
            new_stay_record.class_and_grade = gc
            new_stay_record.grade = gc[:2]
            new_stay_record.grade_num = DT.grades_dict[gc[:2]]
            new_stay_record.cs = DT.get_cs(gc)

            # 匹配学生对象
            student_ob = NewStudent.objects.get(id=id_num)
            new_stay_record.student_related = student_ob
            new_stay_record.name = student_ob.name
            new_stay_record.dorm = student_ob.dorm
            new_stay_record.bed = student_ob.bed

            # 保存
            new_stay_record.save()

        # 重定向至任务主页
        return HttpResponseRedirect(reverse('long_leave:task_main', args=[task_id]))
    else:
        # 打开新的无头浏览器
        DT.refresh_browser()

        # 进入登录页面
        login_url = 'https://www.12kcool.com/#/102064'
        DT.b.get(login_url)

        # 获取验证码图像
        captcha_ele = DT.b.find_element(by=By.XPATH, value='//*[@id="app"]/section/main/div/div[2]/div[4]/img')
        captcha_binary = captcha_ele.screenshot_as_png
        captcha_binary = base64.b64encode(captcha_binary)
        captcha_str = captcha_binary.decode(encoding='utf-8')
        DT.current_captcha_str = captcha_str

        # 提示信息
        if err == 1:
            err_msg = '验证码错误！'
        elif err == 2:
            err_msg = '获取数据失败，请再次尝试'
        else:
            err_msg = ''

    context = {
        'task': task,
        'gc_options': gc_options,
        'st_house': st_house,
        'err': err_msg,
        'gc_selected': '',
        'image': captcha_str
    }
    return render_ht(request, 'long_leave/stay_send_up.html', context)
