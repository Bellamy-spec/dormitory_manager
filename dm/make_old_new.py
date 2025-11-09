# -*- coding: utf-8 -*-
"""
    作者：郭雨健
    功能：取得原宿舍与新宿舍对应关系字典，并存入json
    日期：2025.6.18
"""
import sys
import openpyxl
import datetime
import json


def main():
    """主函数"""
    # 确定年份
    try:
        year = sys.argv[1]
    except IndexError:
        year = str(datetime.datetime.now().year)

    # 打开数据源文件
    wb = openpyxl.load_workbook('old_new.xlsx')
    st = wb.active

    # 存入字典
    old_new_dict_1, old_new_dict_2 = {}, {}
    for row in range(2, st.max_row + 1):
        old_dorm = str(st.cell(row=row, column=1).value)
        new_dorm = str(st.cell(row=row, column=2).value)
        grade = st.cell(row=row, column=3).value
        if grade == 1:
            old_new_dict_1[old_dorm] = new_dorm
        elif grade == 2:
            old_new_dict_2[old_dorm] = new_dorm

    # 存入json文件
    filename = 'old_new_{}.json'.format(year)
    with open(filename, 'w') as f:
        f.write(json.dumps([old_new_dict_1, old_new_dict_2]))


if __name__ == '__main__':
    main()
