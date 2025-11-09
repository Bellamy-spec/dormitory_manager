# coding=utf-8
from django.shortcuts import render
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from django.contrib.auth.views import login_required
from .tools import DataTool
from .forms import TaskForm, ClassSubmitForm, ChangeDateForm, TempAddForm, AddStudentForm
from .models import Task, ClassSubmit, StudentSubmit
import os
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
import threading
import time
import json
from django.utils.encoding import escape_uri_path
import zipfile
from PIL import Image
import shutil
import fitz
from django.conf import settings


# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}期末评优评先上报系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def all_term():
    """获取系统已存在所有任务学期，以列表形式返回"""
    term_list = []
    for task in Task.objects.all():
        term_list.append(task.term)
    return term_list


def all_gc(task):
    """获取某任务所有已提交班级，以列表形式返回"""
    gc_list = []
    for class_submit in ClassSubmit.objects.filter(task_belong=task):
        gc_list.append(class_submit.gc)
    return gc_list


def all_student(class_submit):
    """获取某班级所有获奖学生"""
    student_list = []
    for student_submit in StudentSubmit.objects.filter(class_belong=class_submit):
        student_list.append(student_submit.name)
    return student_list


def get_praise_total(class_submit, praise_name):
    """获取班级奖项已获得人数"""
    return len(StudentSubmit.objects.filter(class_belong=class_submit, praise_name=praise_name))


def get_empty_xlsx(title):
    """获取空的模板表格"""
    # 打开新的文件和工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title
    st['A1'].font = Font(bold=True)
    st.merge_cells(range_string='A1:N1')

    # 写入表头
    st['A2'].value = '奖项'
    st['B2'].value = '姓名'
    st['C2'].value = '语文'
    st['D2'].value = '数学'
    st['E2'].value = '外语'
    st['F2'].value = '物理'
    st['G2'].value = '化学'
    st['H2'].value = '生物'
    st['I2'].value = '政治'
    st['J2'].value = '地理'
    st['K2'].value = '历史'
    st['L2'].value = '体育'
    st['M2'].value = '美术班名次'
    st['N2'].value = '文化班名次'
    for col in range(1, 15):
        st.cell(row=2, column=col).font = Font(bold=True)

    # 设置列宽
    wd = {'A': 11, 'B': 9, 'C': 6, 'D': 6, 'E': 6, 'F': 6, 'G': 6,
          'H': 6, 'I': 6, 'J': 6, 'K': 6, 'L': 6, 'M': 11, 'N': 11}
    set_width_dict(st, wd)

    return wb, st


def copy_row(source_st, target_st, source_row):
    """将source_st的第source_row行整行复制到target_st末尾"""
    # 获取末尾行索引
    target_row = target_st.max_row + 1

    # 逐格复制
    for col in range(1, source_st.max_column + 1):
        target_st.cell(row=target_row, column=col).value = source_st.cell(row=source_row, column=col).value


def process_xlsx(class_submit, wb, wb_save=None):
    """处理xlsx文件，生成获奖学生对象，记录成功失败，并且保存成功数据"""
    if wb_save is None:
        wb_save, st_save = get_empty_xlsx(class_submit.task_belong.title)
    else:
        st_save = wb_save.active

    st = wb.active

    # 人数限制
    class_imd = class_submit.get_items_max()

    # 错误提示填充
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # 生成获奖学生对象
    success, fail = 0, 0
    for row in range(3, st.max_row + 1):
        # 读取奖项和姓名
        praise_name = st.cell(row=row, column=1).value
        student_name = st.cell(row=row, column=2).value

        # 学生不存在
        if student_name not in class_submit.get_student_list():
            # 给出错误提示
            st.cell(row=row, column=15).value = '{}不是{}学生'.format(student_name, class_submit.gc)
            st.cell(row=row, column=15).fill = yellow_fill

            # 记录错误，下一个
            fail += 1
            # fail_lines.append(row)
            continue

        # 学生重复
        if student_name in all_student(class_submit):
            st.cell(row=row, column=15).value = '学生{}已得奖，不可重复得奖'.format(student_name)
            st.cell(row=row, column=15).fill = yellow_fill

            # 记录错误，下一个
            fail += 1
            # fail_lines.append(row)
            continue

        try:
            max_up = class_imd[praise_name]
        except KeyError:
            # 奖项不存在
            st.cell(row=row, column=15).value = '奖项不存在'
            st.cell(row=row, column=15).fill = yellow_fill

            # 记录错误，下一个
            fail += 1
            # fail_lines.append(row)
            continue

        # 人数已达上限
        if get_praise_total(class_submit, praise_name) >= max_up:
            # 奖项不存在
            st.cell(row=row, column=15).value = '{}获奖人数已达上限'.format(praise_name)
            st.cell(row=row, column=15).fill = yellow_fill

            # 记录错误，下一个
            fail += 1
            # fail_lines.append(row)
            continue

        # 通过检测，成功数加1
        success += 1

        # 复制成功行到保存文件
        copy_row(st, st_save, row)

        # 创建新的获奖学生对象
        new_st = StudentSubmit(name=student_name, praise_name=praise_name, class_belong=class_submit)

        # 获取班内编号
        new_st.get_num()

        # 生成获奖证书及证书模板文件
        new_st.make_cert()
        new_st.make_cert(simple=False)

        # 保存学生对象
        new_st.save()

    # 生成提示信息
    tip_msg = '共上传{}条数据，成功{}条，失败{}条，可点击上方的已提交班级查看详情'.format(
        success + fail, success, fail)

    if fail > 0:
        # 存在失败数据，保存失败文件
        st['O2'].value = '错误提示'
        st['O2'].fill = yellow_fill
        st['O2'].font = Font(bold=True)
        st.column_dimensions['O'].width = 30

        # 错误信息文件提示，首先设定错误文件名
        n = 0

        while True:
            # 生成文件名
            if n > 0:
                filename = 'error({}).xlsx'.format(n)
            else:
                filename = 'error.xlsx'

            # 判断是否已有文件名
            if filename in os.listdir(os.path.join('media', 'temp')):
                # 下一个号
                n += 1
            else:
                # 可用文件名
                wb.save(os.path.join('media', 'temp', filename))
                break

        # 可下载错误文件
        have_button = True
        err_id = n

        # 创建并启动延时删除文件线程
        remove_thread = threading.Thread(target=remove_lay, args=(
            os.path.join('media', 'temp', filename), 60))
        remove_thread.start()
    else:
        # 无须下载错误文件
        have_button = False
        err_id = -1

    # 文件更名，重新保存
    save_fn = '{}_{}_{}.xlsx'.format(
        class_submit.task_belong.term, class_submit.grade_num, class_submit.cs)
    save_path = os.path.join(DT.base_dir, 'media', 'praises', 'class_submit', save_fn)
    wb_save.save(save_path)
    class_submit.xlsx_file = save_path

    # 返回前端需要加载的参数值
    return tip_msg, have_button, err_id


def get_row_index(st, msg, col):
    """找到某信息所在行"""
    for row in range(3, st.max_row + 1):
        if msg == st.cell(row=row, column=col).value:
            return row


def im_out(im_fp, fn=None):
    """下载图片文件"""
    # 设置文件名
    if fn is None:
        fn = os.path.basename(im_fp)
    else:
        fn += '.png'

    with open(im_fp, 'rb') as f:
        response = HttpResponse(f, content_type='image/png')

        # 设置文件名称
        # print(fn)
        # print(escape_uri_path(fn))
        response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(escape_uri_path(fn))

    # 输出
    return response


def zip_out(zip_path, fn=None):
    """下载压缩包并从服务器上删除"""
    # 设置文件名
    if fn is None:
        fn = os.path.basename(zip_path)
    else:
        fn += '.zip'

    with open(zip_path, 'rb') as f:
        response = HttpResponse(f, content_type='application/zip')
        response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(escape_uri_path(fn))

    # 删除并输出
    os.remove(zip_path)
    return response


def merge_to_pdf(path, fn=None):
    """遍历一个目录中的所有png图象，合成一个pdf文件在原目录中"""
    # 处理文件名
    if fn is None:
        fn = path
    fn += '.pdf'

    # 打开一个新的PDF文件
    doc = fitz.open()

    # 遍历目录
    for file in sorted(os.listdir(path)):
        if file.endswith('png'):
            img_doc = fitz.open(os.path.join(path, file))
            pdf_bytes = img_doc.convert_to_pdf()
            pdf = fitz.open('pdf', pdf_bytes)
            doc.insert_pdf(pdf)

    # 保存，关闭
    doc.save(os.path.join(path, fn))
    doc.close()


def pdf_out(pdf_path, fn=None):
    """下载pdf文件并从服务器上删除其所在目录"""
    # 设置文件名
    if fn is None:
        fn = os.path.basename(pdf_path)
    else:
        fn += '.pdf'

    with open(pdf_path, 'rb') as f:
        response = HttpResponse(f, content_type='application/pdf')
        response['Content-Disposition'] = "attachment; filename*=UTF-8''{}".format(escape_uri_path(fn))

    # 删除所在目录并输出
    shutil.rmtree(os.path.dirname(pdf_path))
    return response


def remove_lay(filepath, t):
    """延时t秒删除文件"""
    time.sleep(t)
    try:
        os.remove(filepath)
    except FileNotFoundError:
        # 不用你来亲自动手删除了！
        pass


# Create your views here.
def index(request):
    """程序主页"""
    # 列出所有任务
    tasks = Task.objects.all().order_by('term')

    context = {
        'tasks': tasks,
        'is_manager': request.user.username in DT.super_users,
        'title': '{}期末评优评先上报系统'.format(settings.USER_NAME),
    }
    return render_ht(request, 'praises/index.html', context)


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        un = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=un, password=password)

        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到主页
            return HttpResponseRedirect(reverse('praises:index'))
        else:
            return render_ht(request, 'praises/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'praises/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('praises:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'praises/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'praises/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'praises/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'praises/set_pwd.html', {'err': ''})


def public_task(request):
    """发布新任务"""
    # 仅管理员用户可访问
    if request.user.username not in DT.super_users:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = TaskForm()
    else:
        # 对POST提交的数据作出处理
        form = TaskForm(request.POST)
        if form.is_valid():
            new_task = form.save(commit=False)

            # 生成标题
            new_task.make_title()

            # 防止学期重复
            if new_task.term in all_term():
                return render_ht(request, 'praises/public_task.html', context={
                    'form': form,
                    'err': '已存在{}任务，请勿重复发布！'.format(new_task.title),
                })

            # 设置奖项及人数限制
            im_dict = {}
            i = 0
            praise_name_list = []
            while True:
                # 前端读取奖项名称
                praise_name = request.POST.get('r{}_1'.format(i), '')
                if not praise_name:
                    # 已无新奖项，可打破循环
                    break

                # 检验奖项名称是否重复
                if praise_name in praise_name_list:
                    return render_ht(request, 'praises/public_task.html', context={
                        'form': form,
                        'err': '奖项名称不能重复！',
                    })
                praise_name_list.append(praise_name)

                # 前端读取限制方式和限制值
                if int(request.POST.get('r{}_2'.format(i), '')):
                    is_rate = True
                else:
                    is_rate = False
                max_val_str = request.POST.get('r{}_3'.format(i), '')

                # 检验输入的值是否合法
                if is_rate:
                    # 比例必须为0~1之间的数
                    try:
                        max_val = float(max_val_str)
                    except ValueError:
                        return render_ht(request, 'praises/public_task.html', context={
                            'form': form,
                            'err': '数值格式错误，比例须为0~1之间带小数点的数',
                        })
                    else:
                        if max_val <= 0 or max_val >= 1:
                            return render_ht(request, 'praises/public_task.html', context={
                                'form': form,
                                'err': '数值格式错误，比例须为0~1之间带小数点的数',
                            })
                else:
                    # 人数须为正整数
                    try:
                        max_val = int(max_val_str)
                    except ValueError:
                        return render_ht(request, 'praises/public_task.html', context={
                            'form': form,
                            'err': '数值格式错误，人数须为正整数',
                        })
                    else:
                        if max_val <= 0:
                            return render_ht(request, 'praises/public_task.html', context={
                                'form': form,
                                'err': '数值格式错误，人数须为正整数',
                            })

                # 生成人数限制提示信息
                if is_rate:
                    max_msg = '班级人数的{}%'.format(max_val * 100)
                else:
                    max_msg = '{}人'.format(max_val)

                # 加入字典
                im_dict[praise_name] = (is_rate, max_val, max_msg)

                # 索引号加1
                i += 1

            # 字典转为字段
            new_task.load_items_max(im_dict)

            # 保存
            new_task.save()

            # 重定向至任务查看页面
            return HttpResponseRedirect(reverse('praises:see_task', args=[new_task.id]))

    context = {'form': form, 'err': ''}
    return render_ht(request, 'praises/public_task.html', context)


def see_task(request, task_id):
    """任务查看页面"""
    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出要查看的任务对象
    task = Task.objects.get(id=task_id)

    # 取出奖项及人数限制字典
    imd = task.get_items_max()

    # 更改活跃状态按钮名称和制定标题
    if task.active:
        btn_name = '结束任务'
        title = task.title
    else:
        btn_name = '恢复任务'
        title = task.title + '（已结束）'

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ClassSubmitForm()
    else:
        # 任务已结束，不允许以POST方法提交数据
        if not task.active:
            return HttpResponse('任务已结束，不可再上报！')

        # 非操作员不允许提交数据
        DT.check_operator(request.user)

        # 对POST提交的数据作出处理
        form = ClassSubmitForm(request.POST, request.FILES)
        if form.is_valid():
            new_cs = form.save(commit=False)

            # 完善班级年级属性和学生信息
            new_cs.fill_grade_cs()
            new_cs.fill_students()

            # 关联任务对象
            new_cs.task_belong = task

            # 检验重复
            if new_cs.gc in all_gc(task):
                return render_ht(request, 'praises/see_task.html', context={
                    'task': task,
                    'imd': imd.items(),
                    'is_manager': request.user.username in DT.super_users,
                    'btn_name': btn_name,
                    'title': title,
                    'form': form,
                    'err': '已存在{}{}，请勿重复提交，详情可点击上方的已提交班级查看'.format(new_cs.gc, task.title),
                    'have_button': False,
                    'err_id': -1,
                    'is_operator': request.user.username in DT.operators,
                })

            # 第一次保存
            new_cs.save()

            # 对文件再处理
            load_path = os.path.join(DT.base_dir, 'media', str(new_cs.xlsx_file))
            # print(load_path)

            try:
                wb = load_workbook(load_path)
            except InvalidFileException:
                # 文件格式错误，提交失败，删除临时对象及文件
                new_cs.delete()
                os.remove(load_path)

                # 给出错误提示
                return render_ht(request, 'praises/see_task.html', context={
                    'task': task,
                    'imd': imd.items(),
                    'is_manager': request.user.username in DT.super_users,
                    'btn_name': btn_name,
                    'title': title,
                    'form': form,
                    'err': '文件格式必须为xlsx',
                    'have_button': False,
                    'err_id': -1,
                    'is_operator': request.user.username in DT.operators,
                })

            # 删除临时文件
            os.remove(load_path)

            # 处理表格，生成获奖学生对象
            tip_msg, have_button, err_id = process_xlsx(new_cs, wb)

            # 第二次保存，加载当前页
            new_cs.save()
            return render_ht(request, 'praises/see_task.html', context={
                'task': task,
                'imd': imd.items(),
                'is_manager': request.user.username in DT.super_users,
                'btn_name': btn_name,
                'title': title,
                'form': form,
                'err': tip_msg,
                'have_button': have_button,
                'err_id': err_id,
                'is_operator': request.user.username in DT.operators,
            })

    context = {
        'task': task,
        'imd': imd.items(),
        'is_manager': request.user.username in DT.super_users,
        'btn_name': btn_name,
        'title': title,
        'form': form,
        'err': '',
        'have_button': False,
        'err_id': -1,
        'is_operator': request.user.username in DT.operators,
    }
    return render_ht(request, 'praises/see_task.html', context)


def change_active(request, task_id):
    """更改任务活跃状态"""
    # 仅管理员用户可访问
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要更改的任务对象
    task = Task.objects.get(id=task_id)

    # 更改活跃状态并保存
    task.active = not task.active
    task.save()

    # 重定向至任务查看页
    return HttpResponseRedirect(reverse('praises:see_task', args=[task_id]))


def delete_task(request, task_id):
    """删除任务"""
    # 仅管理员用户可访问
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要更改的任务对象
    task = Task.objects.get(id=task_id)

    # 删除相关联的文件（递归删除）
    for class_submit in ClassSubmit.objects.filter(task_belong=task):
        for student_submit in StudentSubmit.objects.filter(class_belong=class_submit):
            # 删除学生证书（cert字段）
            try:
                filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert))
                os.remove(filepath)
            except IsADirectoryError:
                # 文件已不存在，可以忽略
                pass
            except FileNotFoundError:
                # 文件已不存在，可以忽略
                pass
            except PermissionError:
                # 忽略
                pass

            # 删除学生证书模板（cert_simple字段）
            try:
                filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert_simple))
                os.remove(filepath)
            except IsADirectoryError:
                # 文件已不存在，可以忽略
                pass
            except FileNotFoundError:
                # 文件已不存在，可以忽略
                pass
            except PermissionError:
                # 忽略
                pass

        # 删除班级提交Excel表格文件
        try:
            filepath = os.path.join(DT.base_dir, 'media', str(class_submit.xlsx_file))
            os.remove(filepath)
        except IsADirectoryError:
            # 文件已不存在，可以忽略
            pass
        except FileNotFoundError:
            # 文件已不存在，可以忽略
            pass
        except PermissionError:
            # 忽略
            pass

    # 删除任务
    task.delete()

    # 重定向至应用程序主页
    return HttpResponseRedirect(reverse('praises:index'))


def see_class(request, class_id):
    """班级提交查看"""
    # 非操作员不允许访问
    DT.check_operator(request.user)

    # 取出要查看的对象
    class_submit = ClassSubmit.objects.get(id=class_id)
    students = StudentSubmit.objects.filter(class_belong=class_submit)

    # 制定标题
    if class_submit.task_belong.active:
        title = '{}{}'.format(class_submit.gc, class_submit.task_belong.title)
    else:
        title = '{}{}（已结束）'.format(class_submit.gc, class_submit.task_belong.title)

    if request.method == 'POST':
        if not class_submit.task_belong.active:
            return HttpResponse('任务已结束，不可再上报！')

        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')

        # 对POST提交的数据作出处理
        form = TempAddForm(request.POST, request.FILES)
        if form.is_valid():
            # 临时保存文件
            upload_file = form.cleaned_data['file']
            file_path = 'media/temp/' + upload_file.name

            # 保存到指定路径
            with open(file_path, 'wb') as f:
                for chunk in upload_file.chunks():
                    f.write(chunk)

            # 处理上传的文件
            try:
                wb = load_workbook(file_path)
            except InvalidFileException:
                # 删除临时文件
                os.remove(file_path)

                # 错误提示信息
                return render_ht(request, 'praises/see_class.html', context={
                    'class_submit': class_submit,
                    'title': title,
                    'students': students,
                    'form': TempAddForm(),
                    'err': '文件格式必须为xlsx',
                    'have_button': False,
                    'err_id': -1,
                    'active': class_submit.task_belong.active,
                })

            # 删除临时文件
            os.remove(file_path)

            # 打开要保存的文件
            wb_save = load_workbook(class_submit.xlsx_file)

            # 处理表格，生成获奖学生对象
            tip_msg, have_button, err_id = process_xlsx(class_submit, wb, wb_save)

            # 再次保存班级提交对象
            class_submit.save()

            # 重新获取获奖学生对象
            students = StudentSubmit.objects.filter(class_belong=class_submit)

            # 重新加载当前页
            return render_ht(request, 'praises/see_class.html', context={
                'class_submit': class_submit,
                'title': title,
                'students': students,
                'form': TempAddForm(),
                'err': tip_msg[:-16],
                'have_button': have_button,
                'err_id': err_id,
                'active': class_submit.task_belong.active,
            })

    context = {
        'class_submit': class_submit,
        'title': title,
        'students': students,
        'form': TempAddForm(),
        'err': '',
        'have_button': False,
        'err_id': -1,
        'active': class_submit.task_belong.active,
    }
    return render_ht(request, 'praises/see_class.html', context)


def download_temp(request, task_id):
    """下载获奖学生名单模板"""
    # 非操作员不允许下载
    DT.check_operator(request.user)

    # 获取对应的任务对象
    task = Task.objects.get(id=task_id)

    # 获取奖项列表
    praise_list = list(task.get_items_max().keys())

    # 获取空表格
    wb, st = get_empty_xlsx(task.title)

    # 为A列添加奖项下拉选项
    praise_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(praise_list) + '"',
        showErrorMessage=True,
        errorTitle='无效数据',
        error='请从下拉列表当中选择',
    )
    praise_dv.add('A3:A22')
    st.add_data_validation(praise_dv)

    # 输出
    return write_out(wb, fn=task.term + '.xlsx')


def class_list(request, task_id):
    """班级列表"""
    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 奖项列表
    praise_list = task.get_items_max().keys()

    # 生成结构化数据
    data_list = []
    for class_submit in ClassSubmit.objects.filter(task_belong=task):
        # 获奖学生列表转字符串
        praise_str_list = []
        for prais_student_list in class_submit.get_praise_students().values():
            praise_str_list.append('，'.join(prais_student_list))

        # 加入数据
        data_list.append((class_submit, tuple(praise_str_list)))

    # 结构化数据排序
    data_list.sort(key=lambda x: x[0].cs)
    data_list.sort(key=lambda x: x[0].grade_num)
    # print(data_list)

    context = {
        'praise_list': praise_list,
        'data_list': tuple(data_list), 'task': task,
        'is_manager': request.user.username in DT.super_users,
        'is_operator': request.user.username in DT.operators,
    }
    return render_ht(request, 'praises/class_list.html', context)


def delete_class(request, class_id):
    """删除班级提交对象"""
    # 限制访问权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要删除的班级提交对象和所属任务对象
    class_submit = ClassSubmit.objects.get(id=class_id)
    task = class_submit.task_belong

    # 删除相关联的学生证书文件
    for student_submit in StudentSubmit.objects.filter(class_belong=class_submit):
        # 删除学生证书（cert字段）
        try:
            filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert))
            os.remove(filepath)
        except IsADirectoryError:
            # 文件已不存在，可以忽略
            pass
        except FileNotFoundError:
            # 文件已不存在，可以忽略
            pass
        except PermissionError:
            # 忽略
            pass

        # 删除学生证书模板（cert_simple字段）
        try:
            filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert_simple))
            os.remove(filepath)
        except IsADirectoryError:
            # 文件已不存在，可以忽略
            pass
        except FileNotFoundError:
            # 文件已不存在，可以忽略
            pass
        except PermissionError:
            # 忽略
            pass

    # 删除班级提交Excel表格文件
    try:
        filepath = os.path.join(DT.base_dir, 'media', str(class_submit.xlsx_file))
        # print(filepath)
        os.remove(filepath)
    except IsADirectoryError:
        # 文件已不存在，可以忽略
        pass
    except FileNotFoundError:
        # 文件已不存在，可以忽略
        pass
    except PermissionError:
        # 忽略
        pass

    # 执行删除操作
    class_submit.delete()

    # 重定向至班级列表页面
    return HttpResponseRedirect(reverse('praises:class_list', args=[task.id]))


def change_cert_date(request, task_id):
    """修改证书日期"""
    # 限制访问权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出要修改的任务对象
    task = Task.objects.get(id=task_id)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ChangeDateForm(instance=task)
    else:
        # 对POST提交的数据作出处理
        form = ChangeDateForm(instance=task, data=request.POST)
        if form.is_valid():
            form.save()

            # 重定向至任务查看页
            return HttpResponseRedirect(reverse('praises:see_task', args=[task_id]))

    context = {'task': task, 'form': form}
    return render_ht(request, 'praises/change_cert_date.html', context)


def delete_student(request, student_id):
    """删除获奖学生对象"""
    # 非操作员不允许访问
    DT.check_operator(request.user)

    # 取出要删除的学生对象和所属的班级对象
    student_submit = StudentSubmit.objects.get(id=student_id)
    class_submit = student_submit.class_belong

    if not class_submit.task_belong.active:
        return HttpResponse('任务已结束，不可再更改！')

    # 删除获奖证书（cert字段）
    try:
        filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert))
        os.remove(filepath)
    except IsADirectoryError:
        # 文件已不存在，可以忽略
        pass
    except FileNotFoundError:
        # 文件已不存在，可以忽略
        pass
    except PermissionError:
        # 忽略
        pass

    # 删除证书模板（cert_simple字段）
    try:
        filepath = os.path.join(DT.base_dir, 'media', str(student_submit.cert_simple))
        os.remove(filepath)
    except IsADirectoryError:
        # 文件已不存在，可以忽略
        pass
    except FileNotFoundError:
        # 文件已不存在，可以忽略
        pass
    except PermissionError:
        # 忽略
        pass

    # 班级提交名单文件（xlsx_file字段）删除相关行
    wb = load_workbook(class_submit.xlsx_file)
    st = wb.active

    # 获取要删除的行号
    row_index = get_row_index(st, student_submit.name, col=2)

    # 删除整行
    if row_index is not None:
        st.delete_rows(row_index)

        # 重新保存
        wb.save(str(class_submit.xlsx_file))

    # 执行删除操作
    student_submit.delete()

    # 重定向至班级查看页
    return HttpResponseRedirect(reverse('praises:see_class', args=[class_submit.id]))


def add_student(request, class_id):
    """单个添加获奖学生"""
    # 非操作员不允许提交数据
    DT.check_operator(request.user)

    # 取出班级提交对象
    class_submit = ClassSubmit.objects.get(id=class_id)

    if not class_submit.task_belong.active:
        return HttpResponse('任务已结束，不可再上报！')

    # 人数限制
    class_imd = class_submit.get_items_max()

    # 获取学生列表和奖项列表json字符串（用以填充下拉选项）
    student_list_json = class_submit.student_list
    praise_list_json = json.dumps(list(class_imd.keys()))

    # 制定标题
    title = '单个添加{}{}'.format(class_submit.gc, class_submit.task_belong.title)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = AddStudentForm()
    else:
        # 对POST提交的数据作出处理
        form = AddStudentForm(request.POST)
        if form.is_valid():
            new_st = form.save(commit=False)

            # 检验学生是否重复得奖
            if new_st.name in all_student(class_submit):
                return render_ht(request, 'praises/add_student.html', context={
                    'class_submit': class_submit,
                    'title': title,
                    'form': form,
                    'student_house': student_list_json,
                    'praise_house': praise_list_json,
                    'err': '学生{}已得奖，不可重复得奖'.format(new_st.name),
                    'sub_idx': DT.subject_idx.items(),
                })

            # 检验奖项人数是否已满
            if get_praise_total(class_submit, new_st.praise_name) >= class_imd[new_st.praise_name]:
                return render_ht(request, 'praises/add_student.html', context={
                    'class_submit': class_submit,
                    'title': title,
                    'form': form,
                    'student_house': student_list_json,
                    'praise_house': praise_list_json,
                    'err': '{}获奖人数已达上限'.format(new_st.praise_name),
                    'sub_idx': DT.subject_idx.items(),
                })

            # 关联班级提交对象
            new_st.class_belong = class_submit

            # 获取班内编号
            new_st.get_num()

            # 生成获奖证书及证书模板文件
            new_st.make_cert()
            new_st.make_cert(simple=False)

            # TODO:更新班级提交名单文件（xlsx_file字段）
            wb = load_workbook(class_submit.xlsx_file)
            st = wb.active

            # 确定行（末尾）
            row = st.max_row + 1

            # 写入奖项和姓名
            st.cell(row=row, column=1).value = new_st.praise_name
            st.cell(row=row, column=2).value = new_st.name

            # 依次写入各科成绩
            for si in DT.subject_idx.values():
                sub_name = 'sub' + str(si)
                st.cell(row=row, column=si).value = request.POST.get(sub_name, '')

            # 班级提交名单文件重新保存
            wb.save(str(class_submit.xlsx_file))

            # 保存，重定向至班级查看页面
            new_st.save()
            return HttpResponseRedirect(reverse('praises:see_class', args=[class_id]))

    context = {
        'class_submit': class_submit,
        'title': title,
        'form': form,
        'err': '',
        'student_house': student_list_json,
        'praise_house': praise_list_json,
        'sub_idx': DT.subject_idx.items(),
    }
    return render_ht(request, 'praises/add_student.html', context)


def download_cert(request, student_id):
    """下载学生的电子获奖证书"""
    # 非操作员用户不允许访问
    DT.check_operator(request.user)

    # 取出学生对象
    student = StudentSubmit.objects.get(id=student_id)

    # 输出
    return im_out(str(student.cert), fn=student.name)


def export_class_cert(request, class_id):
    """以压缩包形式导出班级学生获奖电子证书"""
    # 非操作员用户不允许访问
    DT.check_operator(request.user)

    # 取出对应的班级提交对象
    class_submit = ClassSubmit.objects.get(id=class_id)

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 进入工作目录
    os.chdir(os.path.join('media', 'praises'))

    # 创建新的zip文件和用以临时存放电子证书的目录
    temp_dir = '{}_{}'.format(class_submit.task_belong.title[:-8], class_submit.gc)
    os.mkdir(temp_dir)
    zip_path = temp_dir + '.zip'
    zipf = zipfile.ZipFile(zip_path, 'a')

    # 逐个添加获奖学生电子证书
    for student in StudentSubmit.objects.filter(class_belong=class_submit):
        # 先将证书放入临时目录
        im = Image.open(student.cert)
        temp_path = os.path.join(temp_dir, student.name + '.png')
        im.save(temp_path)
        im.close()

        # 加入压缩文件
        zipf.write(temp_path)

    # 关闭打开的zip文件
    zipf.close()

    # 删除临时目录
    shutil.rmtree(temp_dir)

    # 输出
    return zip_out(zip_path)


def export_task_xlsx(request, task_id):
    """以压缩包形式导出班级获奖名单"""
    # 限制访问权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 进入工作目录
    os.chdir(os.path.join('media', 'praises'))

    # 创建新的zip文件和用以临时存放xlsx表格的目录
    temp_dir = task.title[:-8]
    os.mkdir(temp_dir)
    zip_path = temp_dir + '.zip'
    zipf = zipfile.ZipFile(zip_path, 'a')

    # 逐个添加获奖名单电子表格
    for class_submit in ClassSubmit.objects.filter(task_belong=task):
        # 先将表格放入临时目录
        wb = load_workbook(class_submit.xlsx_file)
        temp_path = os.path.join(temp_dir, class_submit.gc + '.xlsx')
        wb.save(temp_path)
        wb.close()

        # 加入压缩文件
        zipf.write(temp_path)

    # 关闭打开的zip文件
    zipf.close()

    # 删除临时目录
    shutil.rmtree(temp_dir)

    # 输出
    return zip_out(zip_path)


def export_print_temp(request, task_id):
    """导出学生获奖证书打印模板"""
    # 限制访问权限
    if request.user.username not in DT.super_users:
        raise Http404

    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 进入工作目录
    os.chdir(os.path.join('media', 'praises'))

    # 创建临时目录
    temp_dir = task.term
    try:
        os.mkdir(temp_dir)
    except FileExistsError:
        # 临时目录不正确地已存在，须先删除
        shutil.rmtree(temp_dir)
        os.mkdir(temp_dir)

    # 循环遍历所有学生对象，获奖证书打印模板加入临时目录
    for class_submit in ClassSubmit.objects.filter(task_belong=task):
        for student in StudentSubmit.objects.filter(class_belong=class_submit):
            im = Image.open(student.cert_simple)
            temp_path = os.path.join(temp_dir, os.path.basename(str(student.cert_simple)))
            im.save(temp_path)
            im.close()

    # 图片合成pdf
    merge_to_pdf(temp_dir)
    pdf_path = os.path.join(temp_dir, temp_dir + '.pdf')

    # 输出
    return pdf_out(pdf_path)
