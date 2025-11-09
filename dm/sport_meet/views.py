from django.shortcuts import render
from .tools import DataTool
from .forms import PutNameForm
from .models import Athletes, PutName
from openpyxl.utils import get_column_letter
from django.http import Http404, HttpResponse
from django.contrib.auth.views import login_required
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font
from io import BytesIO
from datetime import datetime
from django.conf import settings
from dm.scores.models import format_gc_students
import json


# 实例化数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def all_year():
    """获取系统内已存在的所有年份"""
    year_list = []
    for athlete in Athletes.objects.all():
        year_list.append(athlete.year)

    # 去重
    year_list = list(set(year_list))

    return year_list


def all_num(year, class_and_grade):
    """取得一个班级的所有序号"""
    num_list = []
    for athlete in Athletes.objects.filter(year=year, class_and_grade=class_and_grade):
        num_list.append(athlete.num[2:])
    return num_list


def give_num(year, class_and_grade):
    """分配一个未被占用的序号"""
    n = 1
    while True:
        # 生成两位序号字符串
        num = str(n)
        if n < 10:
            num = '0' + num

        # 分配从小到大第一个未被占用的号
        if num in all_num(year, class_and_grade):
            n += 1
        else:
            return num


def all_athletes(year):
    """以班级姓名二元组的方式返回某个年份所有运动员"""
    athletes_list = []
    for a in Athletes.objects.filter(year=year):
        athletes_list.append((a.class_and_grade, a.name))
    return athletes_list


def get_aths(year, class_and_grade=None):
    """返回某个年份某个班级所有运动员及项目格式化元组"""
    # 初始化运动员对象列表
    aths = []

    if class_and_grade is None:
        all_ath = Athletes.objects.filter(year=year).order_by('num')
    else:
        all_ath = Athletes.objects.filter(year=year, class_and_grade=class_and_grade).order_by('num')

    # 取出所有运动员对象及其项目四元组
    for ath in all_ath:
        # 格式化项目为三元列表
        items = list(ath.format_items())
        items += [('', -1)] * (3 - ath.n)

        # 加入
        aths.append(tuple([ath] + items))

    return aths


def sep(nums, m):
    """从列表li中分出n个元素，前两位尽可能不重复，放在一个新列表中返回"""
    # 按照前两位出现次数由多到少排序
    cht = count_head_two(nums)
    nums.sort(key=lambda x: cht[x[:2]], reverse=True)

    # 记录已分得的编号前两位
    head_two = []

    # 初始化第一组编号列表
    nums1 = []

    # 记录元素遍历次数
    count = dict(zip(nums, [0] * len(nums)))
    c = 0

    # 开始分组
    while m > 0:
        # 从头弹出一个编号
        num = nums.pop(0)

        # 判断遍历次数
        if count[num] > c:
            c += 1

            # 清空已记录的前两位
            head_two = []

        # 尽可能使前两位错开
        if num[:2] not in head_two:
            nums1.append(num)
            head_two.append(num[:2])

            # 已存入一个元素
            m -= 1
        else:
            # 重新放回队尾
            nums.append(num)

            # 已遍历一次
            count[num] += 1

    # 返回
    return nums1


def count_head_two(li):
    """统计前两位出现的次数，以键值对形式返回"""
    # 初始化结果字典
    result = {}

    for ele in li:
        result.setdefault(ele[:2], 0)
        result[ele[:2]] += 1

    return result


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


# Create your views here.
def index(request):
    """主页"""
    context = {
        'is_manager': request.user.username in DT.manager,
        'is_su': request.user.username in DT.su,
        'title': '{}运动会个人项目报名系统'.format(settings.USER_NAME),
    }
    return render_ht(request, 'sport_meet/index.html', context)


def put_name(request):
    """报名参加项目"""
    # # 报名截止
    # raise Http404

    # 取得当前年份字符串
    year = str(datetime.now().year)

    # 班级学生对应字典及生成json字符串
    gs_dict = format_gc_students(grades=('高一', '高二'), logic=False)
    st_house = json.dumps(gs_dict, ensure_ascii=False)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = PutNameForm()
    else:
        # 对POST提交的数据作出处理
        form = PutNameForm(request.POST)
        if form.is_valid():
            new_put = form.save(commit=False)

            # 检验是否重复报名
            for put in PutName.objects.filter(year=year):
                if new_put.name == put.name and new_put.class_and_grade\
                        == put.class_and_grade and new_put.item == put.item:
                    err_msg = '{}{}已报名{}项目，请勿重复报名'.format(
                        put.class_and_grade, put.name, put.item)
                    return render_ht(request, 'sport_meet/put_name.html', {
                        'form': form, 'st_house': st_house, 'err': err_msg})

            # TODO:检验同班级同项目已报人数
            if len(PutName.objects.filter(class_and_grade=new_put.class_and_grade, gender=new_put.gender,
                                          item=new_put.item, year=year)) >= 3:
                return render_ht(request, 'sport_meet/put_name.html', context={
                    'form': form,
                    'st_house': st_house,
                    'err': '{}{}{}项目已报够3人，不可再报'.format(new_put.class_and_grade, new_put.gender, new_put.item),
                })

            # 是否首次报名
            first = False

            # 判断是否为系统已有运动员
            if (new_put.class_and_grade, new_put.name) not in all_athletes(year=year):
                # 确定为首次报名
                first = True

                # TODO:创建新的运动员对象
                new_ath = Athletes()

                # 完善新对象的属性
                new_ath.name = new_put.name
                new_ath.class_and_grade = new_put.class_and_grade
                new_ath.gender = new_put.gender
                new_ath.complete()

                # TODO:生成运动员编号
                # 年级、班级编号
                grade_num_str = DT.grade_num[new_ath.grade_str]
                class_num = get_column_letter(new_ath.cs)

                # 分配班级内序号
                inner_num = give_num(year=year, class_and_grade=new_ath.class_and_grade)

                # 拼接成完整的运动员编号
                new_ath.num = grade_num_str + class_num + inner_num

                # 更新运动员口令
                new_ath.update_pwd()

                # 设置运动员年份
                new_ath.year = year

                # 保存新的运动员对象
                new_ath.save()

            # TODO:给运动员添加项目
            # 取出相应的运动员对象
            athlete = Athletes.objects.filter(
                class_and_grade=new_put.class_and_grade, name=new_put.name, year=year)[0]

            if athlete.n < 3:
                # 已报项目未达到3个，可以增加项目
                athlete.add_item(new_put.item)

                # 保存运动员对象
                athlete.save()

                # 报名对象关联运动员对象
                new_put.athlete_belong = athlete

                # 设置年份、保存报名对象
                new_put.year = year
                new_put.save()

                # 跳转至运动员页面
                context = {
                    'athlete': athlete,
                    'first': first,
                    'items': athlete.format_items(),
                }
                return render_ht(request, 'sport_meet/show_athlete.html', context)

            else:
                return render_ht(request, 'sport_meet/put_name.html', context={
                    'form': form, 'st_house': st_house, 'err': '你已报名3个个人项目，不可再报'})

    return render_ht(request, 'sport_meet/put_name.html', {'form': form, 'st_house': st_house, 'err': ''})


def que(request):
    """查询及查看运动员"""
    if request.method != 'POST':
        # 未提交数据，加载查询页面
        return render_ht(request, 'sport_meet/que.html', {'err': '', 'yl': tuple(all_year())})
    else:
        # 对POST提交的数据作出处理
        # 前端获取用户输入的运动员编号及口令
        num = request.POST.get('num', '')
        pwd = request.POST.get('pwd', '')
        year = request.POST.get('year', '')

        # 年份、编号及口令均非空时执行操作
        if num and pwd and year:
            # 获取运动员对象
            try:
                athlete = Athletes.objects.filter(year=year, num=num)[0]
            except IndexError:
                # 查找不到
                return render_ht(request, 'sport_meet/que.html', context={
                    'err': '编号不存在', 'yl': tuple(all_year())})
            else:
                # 验证口令是否正确
                if pwd == athlete.pwd:
                    return render_ht(request, 'sport_meet/show_athlete.html', context={
                        'athlete': athlete,
                        'first': False,
                        'items': athlete.format_items(),
                    })
                else:
                    return render_ht(request, 'sport_meet/que.html', context={
                        'err': '运动员口令不正确', 'yl': tuple(all_year())})
        else:
            return render_ht(request, 'sport_meet/que.html', context={
                'err': '请输入运动员编号及口令', 'yl': tuple(all_year())})


def delete_item(request, athlete_id, athlete_pwd, item_idx, code, year):
    """删除项目"""
    # 取出运动员对象
    athlete = Athletes.objects.get(id=athlete_id)

    # 匹配口令
    if athlete.pwd != athlete_pwd:
        raise Http404

    # 获取要删除的项目
    item = DT.items[item_idx]

    if item in athlete.str_to_list():
        # 删除相应的报名记录
        try:
            put_name_record = PutName.objects.filter(
                year=year,
                name=athlete.name,
                class_and_grade=athlete.class_and_grade,
                item=item)[0]
        except IndexError:
            raise Http404
        else:
            # 执行删除
            put_name_record.delete()
            athlete.delete_item(item)

            # 保存运动员对象
            athlete.save()

            # 刷新显示页面
            if code == 0:
                # 以方式0删除，加载运动员个人页
                return render_ht(request, 'sport_meet/show_athlete.html', context={
                    'athlete': athlete,
                    'first': False,
                    'items': athlete.format_items(),
                })
            elif code == 1:
                # 以方式1删除
                if request.user.username in DT.manager:
                    # 若为管理员登录状态，返回所有运动员列表
                    aths = get_aths(year=year)
                else:
                    # 否则仅返回该运动员所在班级的运动员列表
                    aths = get_aths(year=year, class_and_grade=athlete.class_and_grade)

                return render_ht(request, 'sport_meet/athletes_list.html', context={
                    'aths': tuple(aths),
                    'is_manager': request.user.username in DT.manager,
                })
            else:
                # 收到其他code一律返回404
                raise Http404
    else:
        # 尚未报名此项目，返回404
        raise Http404


@ login_required()
def yl(request):
    """年份列表"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.manager:
        raise Http404

    return render_ht(request, 'sport_meet/yl.html', {'yl': tuple(all_year())})


@ login_required()
def admin(request, year):
    """管理端年份主页"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.manager:
        raise Http404

    return render_ht(request, 'sport_meet/admin.html', {'year': year})


@login_required()
def athletes(request, year):
    """运动员列表"""
    # 取得运动员对象及项目格式化列表
    aths = get_aths(year=year)

    context = {'aths': tuple(aths), 'is_manager': request.user.username in DT.manager, 'year': year}
    return render_ht(request, 'sport_meet/athletes_list.html', context)


def class_login(request):
    """班级登录页面"""
    # 加载班级元组
    css = DT.get_class()

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        return render_ht(request, 'sport_meet/class_login.html', context={
            'css': css, 'err': '', 'yl': tuple(all_year())})
    else:
        # 对POST提交的数据作出处理
        # 前端获取班级、年份和密码
        year = request.POST.get('year', '')
        cs = request.POST.get('cs', '')
        pwd = request.POST.get('pwd', '')

        # 验证密码是否正确
        if pwd != DT.class_pwd[year][cs]:
            return render_ht(request, 'sport_meet/class_login.html', context={
                'css': css, 'err': '班级口令不正确', 'yl': tuple(all_year())})

        # 取得运动员对象及项目格式化列表
        aths = get_aths(year=year, class_and_grade=cs)

        context = {'aths': tuple(aths), 'is_manager': request.user.username in DT.manager, 'year': year}
        return render_ht(request, 'sport_meet/athletes_list.html', context)


def delete_athlete(request, athlete_id, athlete_pwd, code):
    """删除运动员或更新运动员口令"""
    # 取出要删除的运动员对象及班级、年份
    athlete = Athletes.objects.get(id=athlete_id)
    class_and_grade = athlete.class_and_grade
    year = athlete.year

    # 验证口令
    if athlete_pwd != athlete.pwd:
        raise Http404

    if code == 0:
        # 操作代号为0，执行删除操作
        athlete.delete()
    elif code == 1:
        # 操作代号为1，执行更新口令操作
        athlete.update_pwd()
        athlete.save()
    else:
        # 其他操作代号一律返回404
        raise Http404

    if request.user.username in DT.manager:
        # 管理员登录状态下，返回所有运动员列表
        aths = get_aths(year=year)
    else:
        # 否则返回所在班级运动员列表
        aths = get_aths(year=year, class_and_grade=class_and_grade)

    return render_ht(request, 'sport_meet/athletes_list.html', context={
        'aths': tuple(aths),
        'is_manager': request.user.username in DT.manager,
        'year': year,
    })


def make_group(request, year):
    """项目分组"""
    # 非管理员登录状态下不允许执行此操作
    if request.user.username not in DT.manager:
        raise Http404

    # 打开文件，加载工作表
    wb = Workbook()
    st = wb.active

    # 初始行
    row = 1

    # 按照年级、性别进行分组
    for gr in [(1, '男'), (1, '女'), (2, '男'), (2, '女')]:
        for item, method in DT.item_group.items():
            # 记录初始行号
            row_min = row

            # 取出范围内所有运动员对象
            aths = []
            for ath in Athletes.objects.filter(grade_num=gr[0], gender=gr[1], year=year):
                if item in ath.str_to_list():
                    aths.append(ath)

            # 获取项目人数
            n = len(aths)

            # 项目无人报名
            if n == 0:
                # 写入标题并加粗
                st.cell(row=row, column=1).value = '{}{}子{}无人报名'.format(DT.grades[gr[0]], gr[1], item)
                st.cell(row=row, column=1).font = Font(bold=True)
                row += 1

                continue

            # 生成标题
            title = '{}{}子{} {}人'.format(DT.grades[gr[0]], gr[1], item, n)

            # 初始化分组大字典
            groups = {}

            if method == 3:
                # 方式3：分一个大组
                groups['第1组'] = []
                for ath in aths:
                    groups['第1组'].append(ath.num)

            elif method == 2:
                # 方式2：分两个大组
                # 记录所有编号
                nums = []
                for ath in aths:
                    nums.append(ath.num)

                # 计算第一组应分得的人数
                m = n // 2 + 1

                # 按照特定算法进行分组
                nums1 = sep(nums, m)

                # 已分好组，分别存入字典的两个值即可
                groups['第1组'] = nums1
                groups['第2组'] = nums

            elif method == 1:
                # 方式3：6个一组
                # 记录所有编号
                nums = []
                for ath in aths:
                    nums.append(ath.num)

                # 组号
                h = 1

                while len(nums) > 6:
                    # 按照特定算法进行分组
                    nums1 = sep(nums, 6)

                    # 组名
                    group_name = '第{}组'.format(h)

                    # 存入字典值
                    groups[group_name] = nums1

                    # 组号加一
                    h += 1

                # 最后一组存入字典值
                group_name = '第{}组'.format(h)
                groups[group_name] = nums

            # 分组情况写入表格
            # 写入标题并加粗
            st.cell(row=row, column=1).value = title
            st.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # 最大列号
            c_max = 1

            # 循环遍历写入编号
            for group, group_nums in groups.items():
                # 写入组名
                st.cell(row=row, column=1).value = group

                # 初始化列号
                c = 2

                # 写入编号
                for num in group_nums:
                    if c > DT.max_cols + 1:
                        # 重置行列号
                        c = 2
                        row += 1

                    st.cell(row=row, column=c).value = num

                    # 记下最大列号
                    if c > c_max:
                        c_max = c

                    # 下一列
                    c += 1

                # 下一行
                row += 1

            # TODO:合并单元格，添加边框
            st.merge_cells(start_row=row_min, end_row=row_min, start_column=1,
                           end_column=c_max)
            add_border(st, start_row=row_min, end_row=row - 1, start_column=1,
                       end_column=c_max)

            # 另起一行准备写下一个项目
            row += 1

    # 输出
    return write_out(wb)


def num_name(request, year):
    """导出编号与姓名对照表"""
    # 禁止非管理员访问
    if request.user.username not in DT.manager:
        raise Http404

    # 打开文件，加载工作表
    wb = Workbook()
    st = wb.active

    # 初始行
    row = 1

    # 循环遍历所有班级
    for css in DT.get_class():
        # 写入班名
        st.cell(row=row, column=1).value = css[0]
        st.cell(row=row, column=1).font = Font(bold=True)
        st.merge_cells(start_row=row, end_row=row, start_column=1, end_column=6)
        row += 1

        # 记下第一行
        row_min = row

        # 初始列
        c = 1

        # 循环遍历本班运动员并写入
        for athlete in Athletes.objects.filter(year=year, class_and_grade=css[0]).order_by('num'):
            # 生成信息
            msg = '{} {}'.format(athlete.num, athlete.name)

            if c > 6:
                # 重置行列
                c = 1
                row += 1

            # 写入信息
            st.cell(row=row, column=c).value = msg

            # 下一列
            c += 1

        # 添加边框
        add_border(st, start_row=row_min, end_row=row, start_column=1, end_column=6)

        # 下一行
        row += 1

    # 调整列宽
    width_dict = dict(zip(['A', 'B', 'C', 'D', 'E', 'F'], [14] * 6))
    set_width_dict(st, width_dict)

    # 输出
    return write_out(wb)


def show_put_name(request):
    """查看报名记录"""
    # 只有超级用户可访问此页
    if request.user.username not in DT.su:
        raise Http404

    # 取出所有报名对象
    put_names = PutName.objects.all()

    context = {'put_names': put_names}
    return render_ht(request, 'sport_meet/show_put_name.html', context)
