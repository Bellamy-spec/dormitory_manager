from django.shortcuts import render
from django.contrib.auth.views import login_required
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from .tools import DataTool
from .forms import SchoolYearForm, AddDepartmentForm, FileUploadForm, FbdpForm, MemberForm, ConfigWorkerForm
from .models import SchoolYear, Department, Member, StudentUser
import os
from io import BytesIO
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter
import threading
import time
import json
from django.core.exceptions import ObjectDoesNotExist
from django.conf import settings
from dm.scores.models import get_students


# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context=None):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    if context is None:
        context = {}
    context.update({'head_title': '{}学生会管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def all_department(exc=None, active_only=False):
    """返回所有部门代号和名称"""
    if exc is None:
        # 没有要去除的对象
        exc = []

    # 初始化代号和名称列表
    code_list, name_list = [], []

    # 取得所有部门
    if active_only:
        departments = Department.objects.filter(active=True)
    else:
        departments = Department.objects.all()

    # 循环遍历所有部门，加入
    for department in departments:
        code_list.append(department.code)
        name_list.append(department.name)

    # 去除
    for ob in exc:
        try:
            code_list.remove(ob.code)
        except ValueError:
            pass
        try:
            name_list.remove(ob.name)
        except ValueError:
            pass

    return code_list, name_list


def get_management(user):
    """获取用户的管理权限"""
    management = []

    if not user.username:
        # 未登录状态，直接返回空列表
        return management

    if user.username in DT.super_manager:
        # 教师管理员用户
        su = True
    else:
        try:
            # 学生用户须先取得用户对象
            st_user = StudentUser.objects.get(username=user.username)
        except ObjectDoesNotExist:
            # 其他情况，权限为空
            return management
        else:
            su = False

    for department in Department.objects.all():
        if su:
            # 教师管理员用户，直接添加权限
            management.append(department)
            continue

        if st_user in department.master_list():
            management.append(department)

    # print(management)
    return management


def get_stop_department(management):
    """获取所有停用及无权限部门"""
    department_ids = []
    for department in Department.objects.filter(active=False):
        department_ids.append(department.id)
    for department in Department.objects.filter(active=True):
        if department not in management:
            department_ids.append(department.id)
    return json.dumps(department_ids)


def get_department_wa():
    """以JSON形式返回所有部门承担课间操工作情况"""
    dw_dict = {}
    for department in Department.objects.filter(active=True):
        dw_dict[department.id] = department.work_abst2
    return json.dumps(dw_dict)


def get_empty_min(school_year, department):
    """获取同学年同部门最小可用编号"""
    # 取得所有已存在的成员编号列表
    num_list = []
    for member in Member.objects.filter(school_year=school_year, department=department):
        num_list.append(member.num)

    # 1~99号从小到大逐个判断
    for n in range(1, 99):
        if n not in num_list:
            return n


def remove_lay(filepath, t):
    """延时t秒删除文件"""
    time.sleep(t)
    try:
        os.remove(filepath)
    except FileNotFoundError:
        # 不用你来亲自动手删除了！
        pass


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


# Create your views here.
def index(request):
    """主页"""
    # 更新本届学年状态
    for i in range(2):
        for sy in SchoolYear.objects.all():
            sy.config_current()
            sy.save()

    return render_ht(request, 'su_manage/index.html', context={
        'is_super': request.user.username in DT.super_manager,
        'is_manager': get_management(request.user),
        'title': '{}学生会管理系统-首页'.format(settings.USER_NAME),
    })


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        un = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=un, password=password)

        # # 获取前端勾选角色
        # char = request.POST.get('char', '')

        if user:
            login(request, user)

            # 登录成功之后重定向到导航页
            return HttpResponseRedirect(reverse('su_manage:index'))
        else:
            return render_ht(request, 'su_manage/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'su_manage/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('su_manage:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'su_manage/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'su_manage/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'su_manage/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'su_manage/set_pwd.html', {'err': ''})


def change_step1(request, school_year_id):
    """第一步：设置学年"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        if school_year_id == 0:
            form = SchoolYearForm()
        else:
            school_year = SchoolYear.objects.get(id=school_year_id)
            form = SchoolYearForm(instance=school_year)
    else:
        # 对POST提交的数据作出处理
        if school_year_id == 0:
            form = SchoolYearForm(request.POST)
        else:
            school_year = SchoolYear.objects.get(id=school_year_id)
            form = SchoolYearForm(data=request.POST, instance=school_year)
        if form.is_valid():
            # 生成新学年
            nsy = form.save(commit=False)
            nsy.fill_blank()

            # 保存对象
            nsy.save()

            # 删除重名的其他学年对象
            for sy in SchoolYear.objects.filter(name=nsy.name):
                if sy != nsy:
                    sy.delete()

            # 所有部门负责人重置为空
            for department in Department.objects.all():
                department.master = ''
                department.master_str = ''
                department.save()

            # 重定向至第二步
            return HttpResponseRedirect(reverse('su_manage:change_step2', args=[nsy.id]))

    context = {'school_year_id': school_year_id, 'form': form}
    return render_ht(request, 'su_manage/change_step1.html', context)


def change_step2(request, school_year_id):
    """第二步：设置部门"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取学年对象
    school_year = SchoolYear.objects.get(id=school_year_id)

    # 生成标题
    title = '第二步：设置{}部门'.format(school_year.name)

    # 获取所有部门对象
    departments = Department.objects.filter(active=True)
    departments_1 = Department.objects.filter(active=False)

    context = {'school_year': school_year, 'title': title, 'departments': departments,
               'departments_1': departments_1}
    return render_ht(request, 'su_manage/change_step2.html', context)


def deactivate_department(request, department_id, school_year_id):
    """停用部门"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出部门对象
    department = Department.objects.get(id=department_id)

    # 执行停用操作
    department.active = False
    department.save()

    # 重定向至第二步或部门管理页
    if school_year_id > 0:
        return HttpResponseRedirect(reverse('su_manage:change_step2', args=[school_year_id]))
    else:
        return HttpResponseRedirect(reverse('su_manage:dp_manage'))


def activate_department(request, department_id, school_year_id):
    """启用部门"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出部门对象
    department = Department.objects.get(id=department_id)

    # 执行启用操作
    department.active = True
    department.save()

    # 重定向至第二步或部门管理页
    if school_year_id > 0:
        return HttpResponseRedirect(reverse('su_manage:change_step2', args=[school_year_id]))
    else:
        return HttpResponseRedirect(reverse('su_manage:dp_manage'))


def change_department(request, department_id, school_year_id):
    """修改部门"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取学生用户
    student_users = []
    for student_user in StudentUser.objects.all():
        if student_user.mem.school_year.current:
            student_users.append(student_user)
    student_users = tuple(student_users)

    if department_id > 0:
        # 修改模式
        department = Department.objects.get(id=department_id)
        title = '修改部门：{}'.format(department.name)
        masters = department.master

        if request.method != 'POST':
            # 未提交数据，创建新的表单
            form = AddDepartmentForm(instance=department)
        else:
            # 对POST提交的数据作出处理
            form = AddDepartmentForm(data=request.POST, instance=department)
            if form.is_valid():
                ndp = form.save(commit=False)
                ndp.make_code()

                # 防止部门代号重复（去掉自己）
                if ndp.code in all_department(exc=[ndp])[0]:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门代号{}已经存在'.format(ndp.code),
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 防止部门名称重复（去掉自己）
                if ndp.name in all_department(exc=[ndp])[1]:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门名称“{}”已经存在'.format(ndp.name),
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 设置部门负责人
                ul = []
                for user in student_users:
                    if request.POST.get('u' + str(user.id), ''):
                        ul.append(user)

                # 部门负责人不能超过五个
                if len(ul) > 5:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门名称负责人不可超过5个',
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 赋值字段
                ndp.config_master(ul)

                # 保存，重定向
                ndp.save()
                if school_year_id > 0:
                    return HttpResponseRedirect(reverse('su_manage:change_step2', args=[school_year_id]))
                else:
                    return HttpResponseRedirect(reverse('su_manage:dp_manage'))
    else:
        # 新增模式
        title = '新增部门'
        masters = ''

        if request.method != 'POST':
            # 未提交数据，创建新的表单
            form = AddDepartmentForm()
        else:
            # 对POST提交的数据作出处理
            form = AddDepartmentForm(request.POST)
            if form.is_valid():
                ndp = form.save(commit=False)
                ndp.make_code()

                # 防止部门代号重复
                if ndp.code in all_department()[0]:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门代号{}已经存在'.format(ndp.code),
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 防止部门名称重复
                if ndp.name in all_department()[1]:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门名称“{}”已经存在'.format(ndp.name),
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 设置部门负责人
                ul = []
                for user in student_users:
                    if request.POST.get('u' + str(user.id), ''):
                        ul.append(user)

                # 部门负责人不能超过五个
                if len(ul) > 5:
                    return render_ht(request, 'su_manage/change_department.html', context={
                        'title': title,
                        'form': form,
                        'department_id': department_id,
                        'school_year_id': school_year_id,
                        'err': '部门名称负责人不可超过5个',
                        'student_users': student_users,
                        'masters': masters,
                    })

                # 赋值字段
                ndp.config_master(ul)

                # 保存，重定向
                ndp.save()
                if school_year_id > 0:
                    return HttpResponseRedirect(reverse('su_manage:change_step2', args=[school_year_id]))
                else:
                    return HttpResponseRedirect(reverse('su:manage:dp_manage'))

    return render_ht(request, 'su_manage/change_department.html', context={
        'title': title,
        'form': form,
        'department_id': department_id,
        'school_year_id': school_year_id,
        'err': '',
        'student_users': student_users,
        'masters': masters,
    })


def change_step3(request, school_year_id):
    """第三步：加入成员"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出学年对象
    school_year = SchoolYear.objects.get(id=school_year_id)

    if request.method != 'POST':
        # 未提交数据，加载新的模板上传页面
        return render_ht(request, 'su_manage/change_step3.html', context={
            'school_year': school_year,
            'form': FileUploadForm(),
            'err': '',
            'have_button': False,
            'err_id': -1,
        })
    else:
        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')

        # 获取表单提交的文件
        form = FileUploadForm(request.POST, request.FILES)

        if form.is_valid():
            # 临时保存文件
            upload_file = form.cleaned_data['file']
            file_path = 'media/temp/' + upload_file.name

            # 保存到指定路径
            with open(file_path, 'wb') as f:
                for chunk in upload_file.chunks():
                    f.write(chunk)

            # 处理上传的文件
            try:
                wb = load_workbook(file_path)
                st = wb.active
            except InvalidFileException:
                # 删除临时文件
                os.remove(file_path)

                # 错误提示信息
                return render_ht(request, 'su_manage/change_step3.html', context={
                    'school_year': school_year,
                    'form': FileUploadForm(),
                    'err': '文件格式必须为xlsx',
                    'have_button': False,
                    'err_id': -1,
                })

            # 删除临时文件
            os.remove(file_path)

            # 成功数和失败数
            success, fail = 0, 0

            # 错误提示填充
            error_tip_column = 6
            error_tip_letter = get_column_letter(error_tip_column)
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 从第3行开始逐行读取
            for row in range(3, st.max_row + 1):
                # 读取信息
                department = st.cell(row=row, column=1).value
                gc = st.cell(row=row, column=2).value
                name = st.cell(row=row, column=3).value
                level_name = st.cell(row=row, column=4).value

                # 获取部门列表、班级列表
                department_list = all_department(active_only=True)[1]
                gc_list = DT.all_gc_options(simple=True)

                # 验证部门是否合法
                if department not in department_list:
                    st.cell(row=row, column=error_tip_column).value = '部门格式错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 获取部门对象
                department_ob = Department.objects.get(name=department)

                # 验证班级是否合法
                if gc not in gc_list:
                    st.cell(row=row, column=error_tip_column).value = '班级格式错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 姓名长度限制
                if len(name) > 5:
                    st.cell(row=row, column=error_tip_column).value = '姓名不能超过5个字符'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 验证姓名班级是否匹配
                if name not in get_students(gc, logic=False):
                    st.cell(row=row, column=error_tip_column).value = '{}不是{}学生'.format(name, gc)
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # # 编号格式限制
                # try:
                #     num = int(num_val)
                # except ValueError:
                #     st.cell(row=row, column=error_tip_column).value = '编号格式错误'
                #     st.cell(row=row, column=error_tip_column).fill = yellow_fill
                #
                #     # 失败数加1
                #     fail += 1
                #     continue
                # else:
                #     if num < 1 or num > 99:
                #         st.cell(row=row, column=error_tip_column).value = '编号格式错误'
                #         st.cell(row=row, column=error_tip_column).fill = yellow_fill
                #
                #         # 失败数加1
                #         fail += 1
                #         continue

                # 获取级别
                try:
                    level = DT.level_dict_reverse[level_name]
                except KeyError:
                    st.cell(row=row, column=error_tip_column).value = '职级格式错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 通过所有检测，创建对象
                nmb = Member()

                # TODO:分配编号
                new_num = get_empty_min(school_year, department_ob)
                if new_num:
                    nmb.num = new_num
                else:
                    st.cell(row=row, column=error_tip_column).value = '编号已满'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                nmb.name = name
                nmb.department = department_ob
                nmb.class_and_grade = gc
                nmb.level = level
                nmb.school_year = school_year
                nmb.fill_blank()
                nmb.update_pwd()

                # 保存对象，成功数加1
                nmb.save()
                success += 1

                # 当读取到“是”并且成员级别大于0时创建账号
                if st.cell(row=row, column=5).value == '是' and nmb.level > 0:
                    create_user(nmb)

            # 生成提示信息
            tip_msg = '共上传{}条数据，成功{}条，失败{}条'.format(success + fail, success, fail)

            if fail > 0:
                # 存在失败数据，保存失败文件
                st.cell(row=1, column=error_tip_column).value = '错误提示'
                st.cell(row=1, column=error_tip_column).fill = yellow_fill
                st.cell(row=1, column=error_tip_column).font = Font(bold=True)
                st.column_dimensions[error_tip_letter].width = 50

                # 设定错误文件名
                n = 0

                while True:
                    # 生成文件名
                    if n > 0:
                        filename = 'error({}).xlsx'.format(n)
                    else:
                        filename = 'error.xlsx'

                    # 判断是否已有文件名
                    if filename in os.listdir(os.path.join('media', 'temp')):
                        # 下一个号
                        n += 1
                    else:
                        # 可用文件名
                        wb.save(os.path.join('media', 'temp', filename))
                        break

                # 可下载错误文件
                have_button = True
                err_id = n

                # 创建并启动延时删除文件线程
                remove_thread = threading.Thread(target=remove_lay, args=(
                    os.path.join('media', 'temp', filename), 60))
                remove_thread.start()
            else:
                # 无须下载错误文件
                have_button = False
                err_id = -1

            # 返回提示
            return render_ht(request, 'su_manage/change_step3.html', context={
                'school_year': school_year,
                'form': FileUploadForm(),
                'err': tip_msg,
                'have_button': have_button,
                'err_id': err_id,
            })


def download_error(request, error_num):
    """下载错误文件并然后删除"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 连接文件名及文件路径
    if error_num > 0:
        filename = 'error({}).xlsx'.format(error_num)
    else:
        filename = 'error.xlsx'
    file_path = os.path.join('media', 'temp', filename)

    # 下载
    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f, content_type='application/vnd.ms-excel')

            # 设置文件名称
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
    except FileNotFoundError:
        return HttpResponse('没有错误文件可以下载！')

    # 输出
    return response


def download_temp(request):
    """下载批量导入成员模板"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 打开新的文件和工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '部门'
    st['B1'].value = '班级'
    st['C1'].value = '姓名'
    st['D1'].value = '职级'
    st['E1'].value = '是否创建账号'

    # 表头样式
    for ro in st['A1:E1']:
        for ce in ro:
            ce.font = Font(size=14, bold=True)
            ce.alignment = Alignment(horizontal='center')

    # 写入说明
    st['A2'].value = '请从下拉列表当中选择'
    st['B2'].value = '请从下拉列表当中选择'
    st['C2'].value = '请输入姓名'
    st['D2'].value = '请从下拉列表当中选择'
    st['E2'].value = '仅干部可创建账号，此操作对干事无效。干部创建账号后，自动成为部门负责人'

    # 说明样式
    for ro in st['A2:E2']:
        for ce in ro:
            ce.font = Font(name='楷体')
            ce.alignment = Alignment(wrap_text=True, vertical='center')
    st.row_dimensions[2].height = 60

    # 获取部门列表、班级列表和职级列表
    department_list = all_department(active_only=True)[1]
    gc_list = DT.all_gc_options(simple=True)
    level_list = list(DT.level_dict_reverse.keys())

    # 设置部门下拉选项
    department_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(department_list) + '"',
        showErrorMessage=True,
        errorTitle='无效数据',
        error='请从下拉列表当中选择',
    )
    department_dv.add('A3:A500')

    # 设置班级下拉选项
    gc_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(gc_list) + '"',
        showErrorMessage=True,
        errorTitle='无效数据',
        error='请从下拉列表当中选择',
    )
    gc_dv.add('B3:B500')

    # 设置职级下拉选项
    level_dv = DataValidation(
        type='list',
        formula1='"' + ','.join(level_list) + '"',
        showErrorMessage=True,
        errorTitle='无效数据',
        error='请从下拉列表当中选择',
    )
    level_dv.add('D3:D500')

    # 是或否
    bdv = DataValidation(
        type='list',
        formula1='"' + ','.join(['是', '否']) + '"',
        showErrorMessage=True,
        errorTitle='无效数据',
        error='请从下拉列表当中选择',
    )
    bdv.add('E3:E500')

    # 数据限制应用于表格
    st.add_data_validation(department_dv)
    st.add_data_validation(gc_dv)
    st.add_data_validation(level_dv)
    st.add_data_validation(bdv)

    # 设置列宽
    set_width_dict(st, {'A': 12, 'B': 12, 'C': 12, 'D': 12, 'E': 20})

    # 输出
    return write_out(wb, fn='su_temp.xlsx')


def member_list(request, school_year_id, member_id):
    """某一届学生会的成员列表"""
    # 取出学年对象及所属的成员对象（按部门排序）
    school_year = SchoolYear.objects.get(id=school_year_id)
    members = list(Member.objects.filter(school_year=school_year))

    # 成员排序（部门>职级>编号），可哈希
    members.sort(key=lambda x: x.num)
    members.sort(key=lambda x: x.level, reverse=True)
    members.sort(key=lambda x: x.department.id)
    members = tuple(members)

    # # 确认账号信息
    # for member in members:
    #     member.config_user()
    #     member.save()

    # 权限部门
    management = get_management(request.user)
    management = tuple(management)

    context = {
        'school_year': school_year,
        'members': members,
        'management': management,
        'is_super': request.user.username in DT.super_manager,
        'member_id': member_id,
    }
    return render_ht(request, 'su_manage/member_list.html', context)


def get_current(request):
    """获取当前学年"""
    try:
        csy = SchoolYear.objects.filter(current=True)[0]
    except IndexError:
        # 未匹配到当前学年，给出提示
        return render_ht(request, 'su_manage/tip_none.html')

    # 匹配到当前学年，进行重定向
    return HttpResponseRedirect(reverse('su_manage:member_list', args=[csy.id, 0]))


def school_year_list(request):
    """获取往届学年列表"""
    # 获取所有往届学年
    school_years_nc = SchoolYear.objects.filter(current=False)

    context = {'ncsys': school_years_nc}
    return render_ht(request, 'su_manage/school_year_list.html', context)


def fbdp(request):
    """完善部门介绍"""
    # 获取管理权限
    management = get_management(request.user)

    # 无管理权限者禁止访问此页
    if not management:
        raise Http404

    # 格式化表单数据
    form_data = []
    for department in management:
        form = FbdpForm(instance=department)
        form_data.append((department, form))

    return render_ht(request, 'su_manage/fbdp.html', context={'fd': form_data})


def submit_fbdp(request, department_id):
    """提交部门介绍完善数据"""
    # 获取部门对象
    department = Department.objects.get(id=department_id)

    # 验证权限
    if department not in get_management(request.user):
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 对POST提交的数据作出处理
    form = FbdpForm(instance=department, data=request.POST)
    if form.is_valid():
        form.save()

        # 重定向至部门介绍完善页
        return HttpResponseRedirect(reverse('su_manage:fbdp'))


def edit_member(request, school_year_id, member_id):
    """单个新增或修改成员"""
    # 获取管理权限
    management = get_management(request.user)

    # 无权限者禁止访问
    if not management:
        raise Http404

    # 获取学年对象
    school_year = SchoolYear.objects.get(id=school_year_id)

    # 获取停用及无权限部门列表、部门承担两操工作情况字典
    di = get_stop_department(management)
    dw = get_department_wa()

    if member_id > 0:
        # 修改模式
        member = Member.objects.get(id=member_id)
        title = '修改 成员:{} 的信息'.format(member.name)
        department = member.department
        to_show = json.dumps([member.work_abst2])

        # 记录原代号
        ori_code = member.code

        # 验证权限
        if department not in management:
            raise Http404

        if request.method != 'POST':
            # 未提交数据，创建新的表单
            form = MemberForm(instance=member)
        else:
            # 对POST提交的数据作出处理
            form = MemberForm(instance=member, data=request.POST)
            if form.is_valid():
                nmb = form.save(commit=False)

                # # 不上早操情况
                # # print(request.POST.get('work_abst1', ''))
                # if request.POST.get('work_abst1', ''):
                #     nmb.work_abst1 = True
                # else:
                #     nmb.work_abst1 = False

                # 不上课间操情况
                # print(request.POST.get('work_abst2', ''))
                if request.POST.get('work_abst2', ''):
                    nmb.work_abst2 = True
                else:
                    nmb.work_abst2 = False

                # 如调换部门则重新分配编号
                if nmb.department != department:
                    new_num = get_empty_min(school_year, nmb.department)
                    if new_num:
                        nmb.num = new_num
                    else:
                        return render_ht(request, 'su_manage/edit_member.html', context={
                            'school_year': school_year,
                            'title': title,
                            'member_id': member_id,
                            'form': form,
                            'err': '编号已满！',
                            'di': di,
                            'dw': dw,
                            'to_show': to_show,
                        })

                # 验证班级、姓名是否匹配
                if nmb.name not in get_students(nmb.class_and_grade, logic=False):
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '{}不是{}学生'.format(nmb.name, nmb.class_and_grade),
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # 验证部门是否正确
                if nmb.department.code not in all_department(active_only=True)[0] or \
                        nmb.department not in management:
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '部门错误！',
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # # 验证早操工作是否符合规范
                # if not nmb.department.work_abst1 and nmb.work_abst1 and nmb.code not in DT.special_member1:
                #     return render(request, 'su_manage/edit_member.html', context={
                #         'school_year': school_year,
                #         'title': title,
                #         'member_id': member_id,
                #         'form': form,
                #         'err': '部门不承担早操工作，成员不可不上早操！',
                #         'di': di,
                #         'dw': dw,
                #         'to_show': to_show,
                #     })

                # 验证课间操工作是否符合规范
                if not nmb.department.work_abst2 and nmb.work_abst2 and nmb.code not in DT.special_member2:
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '部门不承担课间操工作，成员不可不上课间操！',
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # 补充完整属性
                nmb.fill_blank()

                # 对账号的操作
                try:
                    user = StudentUser.objects.get(mem=member)
                except ObjectDoesNotExist:
                    # 不存在账号，可跳过此步
                    pass
                else:
                    if nmb.level > 0:
                        # 新级别为干部
                        if nmb.code != ori_code:
                            # 原部门负责人信息变更，权限取消
                            department.del_master(user)
                            department.save()

                            # 代号改变，用户名跟着变
                            user.username = nmb.code
                            user.save()

                            # 更新成员信息显示
                            nmb.config_user()
                    else:
                        # 新级别为干事，直接注销账号
                        delete_user(member, department=department)

                # 保存，重定向至成员列表页面
                nmb.save()
                return HttpResponseRedirect(reverse('su_manage:member_list', args=[school_year_id, member_id]))
    else:
        # 新增模式
        title = '新增成员'
        to_show = [False]

        if request.method != 'POST':
            # 未提交数据，创建新的表单
            form = MemberForm()
        else:
            # 对POST提交的数据作出处理
            form = MemberForm(request.POST)
            if form.is_valid():
                nmb = form.save(commit=False)

                # # 不上早操情况
                # if request.POST.get('work_abst1', ''):
                #     nmb.work_abst1 = True
                # else:
                #     nmb.work_abst1 = False

                # 不上课间操情况
                if request.POST.get('work_abst2', ''):
                    nmb.work_abst2 = True
                else:
                    nmb.work_abst2 = False

                # 分配编号
                new_num = get_empty_min(school_year, nmb.department)
                if new_num:
                    nmb.num = new_num
                else:
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '编号已满！',
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # 验证班级、姓名是否匹配
                if nmb.name not in get_students(nmb.class_and_grade, logic=False):
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '{}不是{}学生'.format(nmb.name, nmb.class_and_grade),
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # 验证部门是否正确
                if nmb.department.code not in all_department(active_only=True)[0] \
                        or nmb.department not in management:
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '部门错误！',
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # # 验证早操工作是否符合规范
                # if not nmb.department.work_abst1 and nmb.work_abst1:
                #     return render(request, 'su_manage/edit_member.html', context={
                #         'school_year': school_year,
                #         'title': title,
                #         'member_id': member_id,
                #         'form': form,
                #         'err': '部门不承担早操工作，成员不可不上早操！',
                #         'di': di,
                #         'dw': dw,
                #         'to_show': to_show,
                #     })

                # 验证课间操工作是否符合规范
                if not nmb.department.work_abst2 and nmb.work_abst2:
                    return render_ht(request, 'su_manage/edit_member.html', context={
                        'school_year': school_year,
                        'title': title,
                        'member_id': member_id,
                        'form': form,
                        'err': '部门不承担课间操工作，成员不可不上课间操！',
                        'di': di,
                        'dw': dw,
                        'to_show': to_show,
                    })

                # 补充完整属性
                nmb.school_year = school_year
                nmb.fill_blank()

                # 保存，重定向至成员列表页面
                nmb.save()
                return HttpResponseRedirect(reverse('su_manage:member_list', args=[school_year_id, member_id]))

    # 加载新页面
    context = {'school_year': school_year, 'title': title, 'member_id': member_id, 'form': form,
               'err': '', 'di': di, 'dw': dw, 'to_show': to_show}
    return render_ht(request, 'su_manage/edit_member.html', context)


def up_member(request, member_id):
    """更新成员检查码"""
    # 取出成员对象
    member = Member.objects.get(id=member_id)

    # 验证权限
    if member.department not in get_management(request.user):
        raise Http404

    # 执行更新口令操作
    member.update_pwd()
    member.save()

    # 重定向至成员列表页面
    return HttpResponseRedirect(reverse('su_manage:member_list', args=[member.school_year.id]))


def delete_member(request, member_id):
    """删除成员"""
    # 取出成员对象并记录学年
    member = Member.objects.get(id=member_id)
    school_year = member.school_year

    # 验证权限
    if member.department not in get_management(request.user):
        raise Http404

    # 执行删除操作
    member.delete()

    # 重定向至成员列表页面
    return HttpResponseRedirect(reverse('su_manage:member_list', args=[school_year.id]))


def see_department(request):
    """查看部门"""
    # 获取所有启用状态的部门
    departments = Department.objects.filter(active=True)

    context = {'departments': departments}
    return render_ht(request, 'su_manage/see_department.html', context)


def dp_manage(request):
    """部门管理"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取所有（启用、停用）状态的部门
    departments = Department.objects.all()

    context = {
        'departments': departments,
        'del_tip': '删除部门，该部门所有本届、往届成员将一并删除，请谨慎操作！如不用该部门，停用即可。你确定要删除吗？',
    }
    return render_ht(request, 'su_manage/dp_manage.html', context)


def delete_department(request, department_id):
    """删除部门"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出要删除的部门对象
    department = Department.objects.get(id=department_id)

    # 执行删除操作
    department.delete()

    # 重定向至部门管理页
    return HttpResponseRedirect(reverse('su_manage:dp_manage'))


def create_user(member):
    """为成员创建账号"""
    # 创建用户
    user = StudentUser(username=member.code)

    # 关联成员对象
    user.mem = member

    # 设置初始密码，先保存
    user.set_password('106gzxsh')
    user.save()

    # 准初始化行为
    user.add_student_group()
    user.set_last_name()

    # 自动设置负责人
    member.department.add_master(user)
    member.department.save()

    # 成员账号信息更新
    member.config_user()
    member.save()

    # 再保存
    user.save()


def add_user(request, member_id):
    """
        创建账号须满足以下三个条件：
        1.须为超级管理员创建；
        2.没有已存在账号；
        3.是干部，即职级大于0；
    """
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取成员对象
    member = Member.objects.get(id=member_id)

    # 已存在账号或级别不够
    if member.user_info or member.level == 0:
        raise Http404

    # 执行创建账号操作
    create_user(member)

    # 重定向至成员列表页面
    return HttpResponseRedirect(reverse('su_manage:member_list', args=[member.school_year.id]))


def delete_user(member, department=None):
    """删除与member相关联的用户对象"""
    # 取得用户对象
    user = StudentUser.objects.get(mem=member)

    if department is None:
        # 默认情况下，部门为成员所属部门
        department = member.department

    # 部门属性更改
    department.del_master(user)
    department.save()

    # 执行删除操作
    user.delete()

    # 成员属性更改
    member.config_user()
    member.save()


def del_user(request, member_id):
    """
        账号注销，须满足以下两个条件：
        1.须为超级管理员操作；
        2.有已存在账号；
    """
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取成员对象
    member = Member.objects.get(id=member_id)

    # 未存在账号
    if not member.user_info:
        raise Http404

    # 执行账号删除操作
    delete_user(member)

    # 重定向至成员列表页面
    return HttpResponseRedirect(reverse('su_manage:member_list', args=[member.school_year.id]))


def sp(request, member_id):
    """
        给用户重置密码，须满足两个条件：
        1.须为超级管理员操作；
        2.有已存在账号；
    """
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取成员对象
    member = Member.objects.get(id=member_id)

    # 未存在账号
    if not member.user_info:
        raise Http404

    # 获取用户对象
    user = StudentUser.objects.get(mem=member)

    # 设置为初始密码
    user.set_password('106gzxsh')
    user.save()

    # 提示语
    tip_msg = '账号{}{}的密码已重置为106gzxsh'.format(user.username, user.last_name)

    # 跳转至提示页
    context = {'tip_msg': tip_msg, 'school_year': member.school_year}
    return render_ht(request, 'su_manage/tip_reset.html', context)


def output_pwd(request, school_year_id):
    """导出某届所有学生会成员的检查码"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取学年对象
    school_year = SchoolYear.objects.get(id=school_year_id)

    # 打开新的表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '部门'
    st['B1'].value = '代号'
    st['C1'].value = '姓名'
    st['D1'].value = '检查码'

    # 表头加粗
    for c in range(1, 5):
        st.cell(row=1, column=c).font = Font(bold=True)

    # 从第2行开始逐项写入
    row = 2
    for member in Member.objects.filter(school_year=school_year):
        st.cell(row=row, column=1).value = member.department.name
        st.cell(row=row, column=2).value = member.code
        st.cell(row=row, column=3).value = member.name
        st.cell(row=row, column=4).value = member.pwd

        # 下一行
        row += 1

    # 行高统一设置为20
    for r in range(1, row):
        st.row_dimensions[r].height = 20

    # 列宽统一设置为10
    for cl in ['A', 'B', 'C', 'D']:
        st.column_dimensions[cl].width = 10

    # 输出
    return write_out(wb)


def school_year_manage(request):
    """学年管理"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出所有学年对象
    school_years = SchoolYear.objects.all()

    context = {'school_years': school_years}
    return render_ht(request, 'su_manage/school_year_manage.html', context)


def edit_sy(request, sy_id):
    """修改学年起止日期"""
    # 限制非超级管理员用户访问此页
    if request.user.username not in DT.super_manager:
        raise Http404

    # 取出要修改的学年对象
    sy = SchoolYear.objects.get(id=sy_id)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = SchoolYearForm(instance=sy)
    else:
        # 对POST提交的数据作出处理
        form = SchoolYearForm(instance=sy, data=request.POST)
        if form.is_valid():
            form.save()

            # 重定向至学年管理页
            return HttpResponseRedirect(reverse('su_manage:school_year_manage'))

    context = {'sy': sy, 'form': form}
    return render_ht(request, 'su_manage/edit_sy.html', context)


def config_worker(request, sy_id):
    """更改课间操检查工作人员"""
    # 限制权限
    if request.user.username not in DT.super_manager:
        raise Http404

    # 获取学年对象
    school_year = SchoolYear.objects.get(id=sy_id)

    if request.method != 'POST':
        # 未提交数据，加载新的上传页面
        return render_ht(request, 'su_manage/config_worker.html', context={
            'school_year': school_year,
            'form': ConfigWorkerForm(),
            'err': '',
            'have_button': False,
            'err_id': -1,
        })
    else:
        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')      # to 535

        # 获取表单提交的文件
        form = FileUploadForm(request.POST, request.FILES)

        if form.is_valid():
            # 临时保存文件
            upload_file = form.cleaned_data['file']
            file_path = 'media/temp/' + upload_file.name

            # 保存到指定路径
            with open(file_path, 'wb') as f:
                for chunk in upload_file.chunks():
                    f.write(chunk)

            # 处理上传的文件
            with open(file_path, 'rb') as fd:
                file_data = fd.read()
            in_mem_file = BytesIO(file_data)
            try:
                wb = load_workbook(in_mem_file)
                st = wb.active
            except InvalidFileException:
                # 删除临时文件
                os.remove(file_path)

                # 错误提示信息
                return render_ht(request, 'su_manage/config_worker.html', context={
                    'school_year': school_year,
                    'form': ConfigWorkerForm(),
                    'err': '文件格式必须为xlsx',
                    'have_button': False,
                    'err_id': -1,
                })

            # 删除临时文件
            os.remove(file_path)

            # 成功数和失败数
            success, fail = 0, 0

            # 错误提示填充
            error_tip_column = 6
            error_tip_letter = get_column_letter(error_tip_column)
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 记录工作人员代号
            worker_codes = []

            # 从第2行开始逐行读取
            for row in range(2, st.max_row + 1):
                # 读取信息
                department = st.cell(row=row, column=1).value
                code = st.cell(row=row, column=2).value
                name = st.cell(row=row, column=3).value
                gc = st.cell(row=row, column=4).value
                sp_work = st.cell(row=row, column=5).value

                # 尝试根据代号获取对象
                try:
                    member = Member.objects.filter(school_year=school_year, code=code)[0]
                except IndexError:
                    # 未匹配到，给出失败提示
                    st.cell(row=row, column=error_tip_column).value = '代号不存在'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 检验各项信息的正确性
                if department != member.department.name:
                    st.cell(row=row, column=error_tip_column).value = '部门错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                if name != member.name:
                    st.cell(row=row, column=error_tip_column).value = '姓名错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                if gc != member.class_and_grade:
                    st.cell(row=row, column=error_tip_column).value = '班级错误'
                    st.cell(row=row, column=error_tip_column).fill = yellow_fill

                    # 失败数加1
                    fail += 1
                    continue

                # 确认无误，将成员的是否课间操工作人员标记为“是”
                member.work_abst2 = True

                # # 职责清空
                # member.main_work = ''

                # 更新成员工作职责
                if member.main_work:
                    work_list = member.main_work.split(',')
                else:
                    work_list = []
                if sp_work not in work_list:
                    work_list.append(sp_work)
                    member.main_work = ','.join(work_list)

                # 保存对象，添加已处理，成功数加1
                member.save()
                worker_codes.append(member.code)
                success += 1

                # print(type(code), type(member.code))

            # print(worker_codes)

            # 其余均非工作人员，全部标记“否”
            for ob in Member.objects.filter(school_year=school_year):
                if ob.code not in worker_codes:
                    ob.work_abst2 = False
                    ob.save()

            # 生成提示信息
            tip_msg = '共上传{}条数据，成功{}条，失败{}条'.format(success + fail, success, fail)

            if fail > 0:
                # 存在失败数据，保存失败文件
                st.cell(row=1, column=error_tip_column).value = '错误提示'
                st.cell(row=1, column=error_tip_column).fill = yellow_fill
                st.cell(row=1, column=error_tip_column).font = Font(bold=True)
                st.column_dimensions[error_tip_letter].width = 20

                # 设定错误文件名
                n = 0

                while True:
                    # 生成文件名
                    if n > 0:
                        filename = 'error({}).xlsx'.format(n)
                    else:
                        filename = 'error.xlsx'

                    # 判断是否已有文件名
                    if filename in os.listdir(os.path.join('media', 'temp')):
                        # 下一个号
                        n += 1
                    else:
                        # 可用文件名
                        wb.save(os.path.join('media', 'temp', filename))
                        break

                # 可下载错误文件
                have_button = True
                err_id = n

                # 创建并启动延时删除文件线程
                remove_thread = threading.Thread(target=remove_lay, args=(
                    os.path.join('media', 'temp', filename), 60))
                remove_thread.start()
            else:
                # 无须下载错误文件
                have_button = False
                err_id = -1

            # 返回提示
            return render_ht(request, 'su_manage/config_worker.html', context={
                'school_year': school_year,
                'form': ConfigWorkerForm(),
                'err': tip_msg,
                'have_button': have_button,
                'err_id': err_id,
            })


def cw_temp(request):
    """下载更改课间操工作人员模板"""
    # 打开新的表格
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '部门'
    st['B1'].value = '代号'
    st['C1'].value = '姓名'
    st['D1'].value = '班级'
    st['E1'].value = '具体工作'
    st['A1'].font = Font(bold=True)
    st['B1'].font = Font(bold=True)
    st['C1'].font = Font(bold=True)
    st['D1'].font = Font(bold=True)
    st['E1'].font = Font(bold=True)

    # 输出
    return write_out(wb, fn='cw_temp.xlsx')
