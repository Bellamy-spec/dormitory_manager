from django.shortcuts import render
from .models import Teacher
from .forms import TeacherForm
from django.http import HttpResponse, Http404, HttpResponseRedirect
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Font
from io import BytesIO
import datetime
from django.conf import settings


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}学生日常行为规范管理系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def all_id_numbers():
    """获取已在系统中的所有身份证号"""
    id_numbers = []
    for ob in Teacher.objects.all():
        id_numbers.append(ob.id_number)
    return id_numbers


# Create your views here.
def index(request):
    """主页"""
    context = {
        'title': '{}教师基础信息收集系统'.format(settings.USER_NAME),
        'is_manager': request.user.username == 'zz106dyc',
    }
    return render_ht(request, 'teachers/index.html', context)


def add(request):
    """录入"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        form = TeacherForm(request.POST)
        if form.is_valid():
            new_teacher = form.save(commit=False)

            # 检查位数是否为18
            if len(new_teacher.id_number) != 18:
                return render_ht(request, 'teachers/add.html', context={
                    'form': TeacherForm(), 'err': '身份证号长度须为18'})

            # 检查身份证格式是否合法，默认合法
            is_legal = True

            # 前17位必须为数字
            if not new_teacher.id_number[:17].isdigit():
                is_legal = False

            # 最后一位必须为数字或X
            tail = new_teacher.id_number[17]
            if not (tail.isdigit() or tail == 'X'):
                is_legal = False

            # 对非法身份证格式的提示
            if not is_legal:
                return render_ht(request, 'teachers/add.html', context={
                    'form': TeacherForm(), 'err': '身份证号格式非法！'})

            # 校验性别
            if int(new_teacher.id_number[16]) % 2:
                new_teacher.gender1 = '男'
            else:
                new_teacher.gender1 = '女'
            if new_teacher.gender != new_teacher.gender1:
                return render_ht(request, 'teachers/add.html', context={
                    'form': TeacherForm(), 'err': '身份证号与性别不符！'})

            # 已在系统中的身份证号不提交
            if new_teacher.id_number in all_id_numbers():
                err = '身份证号{}已在系统中，请勿重复提交，如有问题，' \
                      '请联系管理员删除'.format(new_teacher.id_number)
                return render_ht(request, 'teachers/add.html', context={
                    'form': TeacherForm(), 'err': err})

            new_teacher.save()
            return HttpResponse('提交成功，感谢您的支持！')

    # 未提交数据，创建新的表单
    return render_ht(request, 'teachers/add.html', {'form': TeacherForm(), 'err': ''})


def see(request):
    """查看"""
    # 除管理员外无人有此权限
    if request.user.username != 'zz106dyc':
        raise Http404

    # 获取所有教师信息
    teachers = Teacher.objects.all()

    context = {'title': '教师信息', 'teachers': teachers}
    return render_ht(request, 'teachers/see.html', context)


def delete(request, teacher_id):
    """删除教师信息"""
    # 除管理员外无人有此权限
    if request.user.username != 'zz106dyc':
        raise Http404

    # 取得要删除的教师信息
    t = Teacher.objects.get(id=teacher_id)

    # 执行删除操作
    t.delete()

    # 重定向到当前页
    return HttpResponseRedirect(reverse('teachers:see'))


def load(request):
    """导出"""
    # 创建新表格
    wb = Workbook()
    st = wb.active

    # 设置列宽
    set_width_dict(st, {'A': 5, 'B': 5, 'C': 10, 'D': 6, 'E': 10})

    # 写入表头
    st['A1'].value = '姓名'
    st['B1'].value = '性别'
    st['C1'].value = '登录账号'
    st['D1'].value = '证件类型'
    st['E1'].value = '证件号码'

    # 从第二行开始写入主体
    row = 2
    for teacher in Teacher.objects.all():
        st.cell(row=row, column=1).value = teacher.name
        st.cell(row=row, column=2).value = teacher.gender
        st.cell(row=row, column=3).value = teacher.id_number
        st.cell(row=row, column=4).value = teacher.card_type
        st.cell(row=row, column=5).value = teacher.id_number
        row += 1

    # 表头加粗
    ft_head = Font(bold=True)
    for c in ['A', 'B', 'C', 'D', 'E']:
        st[c + '1'].font = ft_head

    # 添加边框
    add_border(st, start_row=1, end_row=row - 1, start_column=1, end_column=5)

    return write_out(wb)
