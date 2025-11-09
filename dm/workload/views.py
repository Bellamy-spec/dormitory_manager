from django.shortcuts import render
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.urls import reverse
from django.contrib.auth.views import login_required
from django.contrib.auth.models import Group
from .models import Task, WorkloadRecord, SubLesson
from .tools import DataTool
from .forms import TaskForm, WorkloadForm, AddSubForm, ChangeWeeksForm
from openpyxl.styles import Side, Border, Font, Alignment
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.conf import settings

# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}教师工作量统计系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def ct(st, start_row, start_column, end_row, end_column):
    """区域水平、垂直居中"""
    alignment = Alignment(horizontal='center', vertical='center')
    for r in range(start_row, end_row + 1):
        for c in range(start_column, end_column + 1):
            st.cell(row=r, column=c).alignment = alignment


def write_out(wb):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def all_names(task, grade_num, subject):
    """获取某个月份、年级、学科所有已填报老师的姓名"""
    # 初始化姓名列表
    names = []

    # 循环遍历、写入
    for obj in WorkloadRecord.objects.filter(month=task, grade_num=grade_num,
                                             subject=subject):
        names.append(obj.name)

    return names


def all_months():
    """返回所有任务的月份"""
    months = []
    for task in Task.objects.all():
        months.append((task.month, task.grade))
    return months


# Create your views here.
def index(request):
    """主页"""
    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    # 取出所有任务对象
    tasks = Task.objects.all().order_by('month')

    # 任务对象按月份分组
    task_month = {}
    for task in tasks:
        task_month.setdefault(task.month, [])
        task_month[task.month].append(task)

    # 格式化
    for tm in task_month.keys():
        fl = []

        # 按照高一、高二、高三年级的顺序逐个添加
        for grade in range(1, 4):
            # 是否找到
            found = False
            for task in task_month[tm]:
                if task.grade == grade:
                    fl.append(task)

                    # 找到了
                    found = True
                    break

            # 未找到元素，添加空值
            if not found:
                fl.append(None)

        # 重写字典值
        task_month[tm] = tuple(fl)

    return render_ht(request, 'workload/index.html', context={
        'task_month': tuple(task_month.items()),
        'is_manager': request.user.username in DT.managers,
        'title': '{}教师工作量统计系统'.format(settings.USER_NAME),
    })


def login1(request):
    """登录页"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到主页
            return HttpResponseRedirect(reverse('workload:index'))
        else:
            return render_ht(request, 'workload/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'workload/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('workload:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'workload/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'workload/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'workload/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'workload/set_pwd.html', {'err': ''})


def task_main(request, task_id):
    """某个具体任务的主页"""
    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 取得年月和年级
    ym = task.month.split('-')
    ym.append(DT.grades[task.grade])

    # 制定标题
    title = '{}年{}月{}年级课时量统计'.format(*ym)

    # 是否该年级的管理员
    try:
        if task.grade in DT.manager_areas[request.user.username]:
            is_manager = True
        else:
            is_manager = False
    except KeyError:
        is_manager = False

    context = {'task': task, 'title': title, 'is_manager': is_manager}
    return render_ht(request, 'workload/task_main.html', context)


def public(request):
    """发布新的课时量统计任务"""
    # 需要管理员权限
    if request.user.username not in DT.managers:
        raise Http404

    # 生成年级权限
    cannot_public = [1, 2, 3]
    for gr in DT.manager_areas[request.user.username]:
        cannot_public.remove(gr)
    cannot_public.sort(reverse=True)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = TaskForm()
    else:
        # 对POST提交的数据作出处理
        form = TaskForm(request.POST)
        if form.is_valid():
            new_task = form.save(commit=False)

            # 月份，年级不能重复
            if (new_task.month, new_task.grade) in all_months():
                err = '已存在{}{}的课时量统计任务，请勿重复发布，若之前发布有误，' \
                      '须先删除后再重新发布'.format(new_task.month, new_task.grade)
                old_task = Task.objects.filter(month=new_task.month, grade=new_task.grade)[0]
                return render_ht(request, 'workload/public.html', context={
                    'form': form,
                    'err': err,
                    'have_link': True,
                    'old_task': old_task,
                    'cp': tuple(cannot_public),
                })

            new_task.grade_str = DT.grades[new_task.grade]
            new_task.save()

            # 重定向到任务主页
            return HttpResponseRedirect(reverse('workload:task_main', args=[new_task.id]))

    return render_ht(request, 'workload/public.html', {
        'form': form, 'err': '', 'have_link': False, 'cp': tuple(cannot_public)})


def send_up(request, task_id):
    """填报"""
    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 制定标题
    ym = task.month.split('-')
    ym.append(task.grade_str)
    title = '填报{}年{}月{}年级课时量'.format(*ym)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = WorkloadForm()
    else:
        # 对POST提交的数据作出处理
        form = WorkloadForm(request.POST)
        if form.is_valid():
            new_record = form.save(commit=False)

            # TODO:补全属性
            new_record.month = task
            new_record.grade_num = task.grade
            new_record.grade = DT.grades[new_record.grade_num]

            # TODO:防止重复
            if new_record.name in all_names(task, new_record.grade_num,
                                            new_record.subject):
                # 生成错误信息
                msg = '已存在{}{}{}{}的提交记录，请勿重复提交，如有错误，' \
                      '可先删除后再重新提交。'.format(
                       task.month,
                       new_record.grade,
                       new_record.subject,
                       new_record.name)

                # 链接信息
                link_id = WorkloadRecord.objects.filter(
                    month=task,
                    grade_num=new_record.grade_num,
                    subject=new_record.subject,
                    name=new_record.name)[0].id

                return render_ht(request, 'workload/send_up.html', context={
                    'form': form,
                    'title': title,
                    'err': msg,
                    'task': task,
                    'link_id': link_id,
                })

            # TODO:完善班级数字
            cs_list = []
            for i in range(1, DT.max_class + 1):
                # 前端获取勾选情况
                nm = 'c' + str(i)
                cs = request.POST.get(nm, '')
                if cs:
                    cs_list.append(cs)

            if cs_list:
                # 所教班级非空时执行的操作
                new_record.update_cs(cs_list)
            else:
                # 否则给出提示
                return render_ht(request, 'workload/send_up.html', context={
                    'form': form,
                    'title': title,
                    'err': '必须至少勾选1个所教班级',
                    'task': task,
                    'link_id': 0,
                })

            # 获取上课周数
            wks = task.weeks

            # 根据周教案数和周课时数计算月教案数、月课时数
            # new_record.week_lessons = new_record.week_plans * new_record.cs_n
            new_record.month_plans = new_record.week_plans * wks
            new_record.month_lessons = new_record.week_lessons * wks

            # 前端获取特殊身份勾选情况
            check1 = request.POST.get('check1', '')
            check2 = request.POST.get('check2', '')
            check3 = request.POST.get('check3', '')
            check4 = request.POST.get('check4', '')

            # 是否班主任
            if check1:
                n1 = request.POST.get('val1', '')
                if n1:
                    new_record.headteacher = True
                    new_record.headteacher_n = int(n1)
                else:
                    return render_ht(request, 'workload/send_up.html', context={
                        'form': form,
                        'title': title,
                        'err': '班主任须填写班级人数',
                        'task': task,
                        'link_id': 0,
                    })

            # 是否年级长
            if check2:
                n2 = request.POST.get('val2', '')
                if n2:
                    new_record.grade_master = True
                    new_record.grade_master_n = int(n2)
                else:
                    return render_ht(request, 'workload/send_up.html', context={
                        'form': form,
                        'title': title,
                        'err': '年级长须填写年级人数',
                        'task': task,
                        'link_id': 0,
                    })

            # 是否备课组长
            if check3:
                n3 = request.POST.get('val3', '')
                if n3:
                    new_record.small_master = True
                    new_record.small_master_n = int(n3)
                else:
                    return render_ht(request, 'workload/send_up.html', context={
                        'form': form,
                        'title': title,
                        'err': '备课组长须填写备课组人数',
                        'task': task,
                        'link_id': 0,
                    })

            # 是否教研组长
            if check4:
                n4 = request.POST.get('val4', '')
                if n4:
                    new_record.big_master = True
                    new_record.big_master_n = int(n4)
                else:
                    return render_ht(request, 'workload/send_up.html', context={
                        'form': form,
                        'title': title,
                        'err': '教研组长须填写教研组人数',
                        'task': task,
                        'link_id': 0,
                    })

            # 保存对象，重定向至记录详情页
            new_record.save()
            return HttpResponseRedirect(reverse('workload:show_record', args=[new_record.id]))

    return render_ht(request, 'workload/send_up.html', context={
        'form': form,
        'title': title,
        'err': '',
        'task': task,
        'link_id': 0,
    })


def teacher_list(request, task_id):
    """显示某个任务已填报老师的页面"""
    # 取出所有相关记录
    task = Task.objects.get(id=task_id)
    records = WorkloadRecord.objects.filter(month=task)

    context = {'task': task, 'records': records}
    return render_ht(request, 'workload/teacher_list.html', context)


def show_record(request, record_id):
    """显示记录详情"""
    # 取出相应记录
    record = WorkloadRecord.objects.get(id=record_id)

    # 取出该记录的所有代课记录
    sbl = SubLesson.objects.filter(record_belong=record)

    # 制定标题
    title = '{}{}课时量填写情况'.format(record.name, record.month.month)

    context = {'record': record, 'title': title, 'sbl': sbl}
    return render_ht(request, 'workload/show_record.html', context)


def delete_task(request, task_id):
    """删除任务"""
    # 禁止非管理员执行此操作
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的任务对象
    task = Task.objects.get(id=task_id)

    # 不可以删除其他年级的任务
    if task.grade not in DT.manager_areas[request.user.username]:
        return HttpResponse('你无权删除其他年级发布的任务！')

    # 执行删除操作
    task.delete()

    # 重定向至主页
    return HttpResponseRedirect(reverse('workload:index'))


def delete_record(request, record_id):
    """删除记录"""
    # # 禁止非管理员执行此操作
    # if request.user.username not in DT.managers:
    #     raise Http404

    # 取出要删除的记录及所属任务
    r = WorkloadRecord.objects.get(id=record_id)
    task = r.month

    # 执行删除操作
    r.delete()

    # 重定向至列表页
    return HttpResponseRedirect(reverse('workload:teacher_list', args=[task.id]))


def export(request, task_id):
    """导出某个任务已填报的数据"""
    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 制定标题
    title = '{}{}教师课时量统计'.format(task.month, task.grade_str)

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title
    st['A1'].font = Font(size=14, bold=True)

    # 写入表头
    st['A2'].value = '教师姓名'
    st['B2'].value = '所教学科'
    st['C2'].value = '所教年级'
    st['D2'].value = '所教班级'
    st['E2'].value = '周教案数'
    st['F2'].value = '周课时数'
    st['G2'].value = '月教案数'
    st['H2'].value = '月课时数'
    st['I2'].value = '早辅导次数'
    st['J2'].value = '晚自习次数'
    st['K2'].value = '是否班主任'
    st['L2'].value = '班级人数'
    st['M2'].value = '是否年级长'
    st['N2'].value = '年级人数'
    st['O2'].value = '是否备课组长'
    st['P2'].value = '备课组人数'
    st['Q2'].value = '是否教研组长'
    st['R2'].value = '教研组人数'
    st['S2'].value = '代课总节数'

    # 初始行
    row = 3

    # 记录最大代课条数
    max_sub = 0

    # 逐条记录写入
    for record in WorkloadRecord.objects.filter(month=task):
        # 基础信息
        st.cell(row=row, column=1).value = record.name
        st.cell(row=row, column=2).value = record.subject
        st.cell(row=row, column=3).value = record.grade
        st.cell(row=row, column=4).value = record.css_format
        st.cell(row=row, column=5).value = record.week_plans
        st.cell(row=row, column=6).value = record.week_lessons
        st.cell(row=row, column=7).value = record.month_plans
        st.cell(row=row, column=8).value = record.month_lessons
        st.cell(row=row, column=9).value = record.morning_lessons
        st.cell(row=row, column=10).value = record.evening_lessons

        # 班主任信息
        if record.headteacher:
            st.cell(row=row, column=11).value = '是'
            st.cell(row=row, column=12).value = record.headteacher_n
        else:
            st.cell(row=row, column=11).value = '否'

        # 年级长信息
        if record.grade_master:
            st.cell(row=row, column=13).value = '是'
            st.cell(row=row, column=14).value = record.grade_master_n
        else:
            st.cell(row=row, column=13).value = '否'

        # 备课组长信息
        if record.small_master:
            st.cell(row=row, column=15).value = '是'
            st.cell(row=row, column=16).value = record.small_master_n
        else:
            st.cell(row=row, column=15).value = '否'

        # 教研组长信息
        if record.big_master:
            st.cell(row=row, column=17).value = '是'
            st.cell(row=row, column=18).value = record.big_master_n
        else:
            st.cell(row=row, column=17).value = '否'

        # 总代课节数
        total_sub, count = 0, 0
        for sb in SubLesson.objects.filter(record_belong=record):
            total_sub += sb.sub_lessons
            count += 1

            # 记录代课情况
            st.cell(row=2, column=18 + 2 * count).value = '被代课教师'
            st.cell(row=2, column=19 + 2 * count).value = '代课节数'
            st.cell(row=row, column=18 + 2 * count).value = sb.sub_teacher
            st.cell(row=row, column=19 + 2 * count).value = sb.sub_lessons

        st.cell(row=row, column=19).value = total_sub

        # 刷新最大代课节数
        if count > max_sub:
            max_sub = count

        # 下一行
        row += 1

    # 设置表头格式为加粗
    for c in range(1, 20 + 2 * max_sub):
        st.cell(row=2, column=c).font = Font(bold=True)

    # 标题行合并单元格
    st.merge_cells(start_row=1, start_column=1, end_row=1, end_column=19 + 2 * max_sub)

    # 单元格居中
    ct(st, start_row=1, start_column=1, end_row=row-1, end_column=19 + 2 * max_sub)

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row-1, end_column=19 + 2 * max_sub)

    # 列宽统一设置为12
    width_dict = {}
    for i in range(1, 20 + 2 * max_sub):
        width_dict[get_column_letter(i)] = 12
    set_width_dict(st, width_dict)

    # 输出
    return write_out(wb)


def add_sub(request, record_id):
    """添加代课情况"""
    # 取出相应记录
    record = WorkloadRecord.objects.get(id=record_id)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = AddSubForm()
    else:
        # 对POST提交的数据作出处理
        form = AddSubForm(request.POST)
        if form.is_valid():
            new_sb = form.save(commit=False)

            # 绑定记录
            new_sb.record_belong = record
            new_sb.save()

            # 重定向至记录详情页
            return HttpResponseRedirect(reverse('workload:show_record', args=[record_id]))

    context = {'record': record, 'form': form}
    return render_ht(request, 'workload/add_sub.html', context)


def delete_sub(request, sb_id):
    """删除代课记录"""
    # 取出要删除的记录
    sb = SubLesson.objects.get(id=sb_id)
    record = sb.record_belong

    # 执行删除操作
    sb.delete()

    # 重定向至记录详情页
    return HttpResponseRedirect(reverse('workload:show_record', args=[record.id]))


def change_weeks(request, task_id):
    """更改上课周数"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要更改的任务对象
    task = Task.objects.get(id=task_id)

    # 制定标题
    title = '更改{}年级{}上课周数'.format(task.grade_str, task.month)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = ChangeWeeksForm(instance=task)
    else:
        # 对POST提交的数据作出处理
        form = ChangeWeeksForm(instance=task, data=request.POST)
        if form.is_valid():
            form.save()

            # TODO:更改所有已填报老师的月教案数和月课时数
            for ob in WorkloadRecord.objects.filter(month=task):
                ob.month_plans = ob.week_plans * task.weeks
                ob.month_lessons = ob.week_lessons * task.weeks
                ob.save()

            # 重定向至任务主页
            return HttpResponseRedirect(reverse('workload:task_main', args=[task_id]))

    context = {'task': task, 'title': title, 'form': form}
    return render_ht(request, 'workload/change_weeks.html', context)
