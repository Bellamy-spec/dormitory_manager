"""一些方便调用的数据"""
import os
from datetime import datetime
import json
import openpyxl


class DataTool:
    """静态数据类"""
    def __init__(self):
        # 项目管理员
        self.managers = ['zz106dyc', '陈政', '张志云']

        # 选考科目
        self.subjects = (('素描或创意画', '素描或创意画'), ('书法或国画', '书法或国画'))

        # 考试时长（单位：小时）
        self.ter = 1.5

        # 考试成绩
        self.scores = {
            'A': 'A(优秀)',
            'B': 'B(良好)',
            'C': 'C(合格)',
            'D': 'D(不合格)',
            'O': 'O(未设置)',
        }
        self.scores_2025 = {
            'A': '合格',
            'B': '合格',
            'C': '合格',
            'D': '不合格',
            'O': '未设置',
        }

        # 成绩对应逻辑分值
        self.score_int = {
            'A': 50,
            'B': 40,
            'C': 30,
            'D': 0,
            'O': -1,
        }
        self.reverse_score_int = {
            50: 'A',
            40: 'B',
            30: 'C',
            0: 'D',
            -1: 'O',
        }

        # 操作员数量
        self.opes = 18

        # 项目操作员
        self.operators = self.managers + self.num_head()

        # 考生添加方式
        self.add_method = ((1, '自行报名'), (2, '模板导入'))

        # 考场位置
        self.pos = {
            '01': '启航楼204',
            '02': '启航楼206',
            '03': '启航楼208',
            '04': '启航楼210',
            '05': '启航楼304',
            '06': '启航楼306',
            '07': '启航楼308',
            '08': '启航楼310',
            '09': '启航楼404',
            '10': '启航楼406',
            '11': '启航楼408',
            '12': '启航楼410',
            '13': '求真楼204',
            '14': '求真楼206',
            '15': '求真楼208',
            '16': '求真楼210',
            '17': '求真楼304',
            '18': '求真楼306',
            '19': '求真楼308',
            '20': '求真楼310',
            '21': '求真楼404',
            '22': '求真楼406',
            '23': '求真楼408',
            '24': '求真楼410',
            '25': '尚美楼204',
            '26': '尚美楼206',
            '27': '尚美楼208',
            '28': '尚美楼210',
            '29': '尚美楼304',
            '30': '尚美楼306',
            '31': '尚美楼308',
            '32': '尚美楼310',
            '33': '尚美楼404',
            '34': '尚美楼406',
            '35': '尚美楼408',
            '36': '尚美楼410',
            '37': '启航楼504',
            '38': '启航楼506',
            '39': '求真楼504',
            '40': '求真楼506',
            '41': '尚美楼504',
            '42': '尚美楼506',
            '51': '启航楼一楼',
            '52': '启航楼一楼',
            '53': '启航楼一楼',
        }

        # # 准考证号第8位起始号码
        # self.num8 = 0

        # 特殊身份证号
        self.special_id = [
            'S242293(1)',
            '07771507',
            '128253785',
            'P782445QS',
        ]

        # 学校地址
        self.address = '郑州市郑东新区正光路36号'

        # 准考证号中间三位
        self.ei_mid = '106'

        # 存储验证码的字典
        self.captcha_dict = {}
        self.captcha_dict_str = ''

    @staticmethod
    def str_three(n):
        """整数n格式化为三位数字字符串"""
        if n < 10:
            return '00' + str(n)
        elif n < 100:
            return '0' + str(n)
        else:
            return str(n)

    @staticmethod
    def str_two(n):
        """整数n格式化为两位数字字符串"""
        if n < 10:
            return '0' + str(n)
        else:
            return str(n)

    def num_head(self):
        """返回项目操作员"""
        # 获取当前年份
        year = datetime.now().year

        head_list = []
        for i in range(1, self.opes + 1):
            head_list.append(str(year) + self.str_two(i))

        return head_list

    def get_subject_list(self):
        """获取考试科目列表"""
        subject_list = []
        for subject in self.subjects:
            subject_list.append(subject[0])
        return subject_list

    def show_sted(self, st, fst):
        """根据开始时间和时间长度生成时间段显示，接收时间对象"""
        # 生成考试时间显示格式
        dt_start = datetime.strptime(fst, '%Y-%m-%dT%H:%M')
        h = int(self.ter)
        m = int((self.ter - int(self.ter)) * 60)
        end_h = dt_start.hour + h
        end_m = dt_start.minute + m

        # 处理考试结束时间
        if end_m >= 60:
            end_m -= 60
            end_h += 1

        # TODO:生成格式化时间表示(须进一步修改)
        exam_time = '{} —— {}:{}'.format(st, end_h, self.str_two(end_m))
        return exam_time

    @ staticmethod
    def get_exam_time():
        """获取考试时间选项"""
        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')

        # 读取数据
        with open('exam_time.json') as f:
            exam_time_dict = json.loads(f.read())

        # 返回数据选项
        return tuple(exam_time_dict.items())

    @ staticmethod
    def get_middle_school_list(simple=False):
        """获取初中学校列表"""
        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')

        # 加载Excel表格
        wb = openpyxl.load_workbook('media/middle_school_list.xlsx')
        st = wb.active

        # 初始化存放列表
        middle_school_list = []

        # 逐行读取、写入
        for row in range(1, st.max_row + 1):
            middle_school_list.append(st.cell(row=row, column=1).value)

        # 随手关闭文件好习惯
        wb.close()

        if not simple:
            # 加入最后一个选项
            middle_school_list.append('零星返郑报考')

        return middle_school_list

    @staticmethod
    def get_middle_school_choices():
        """获取初中学校列表"""
        # 确保服务器上进入正确的目录，可正常运行
        if os.name != 'nt':
            os.chdir('/root/dormitory_manager/dm')

        # 加载Excel表格
        wb = openpyxl.load_workbook('media/middle_school_list.xlsx')
        st = wb.active

        # 初始化存放列表
        middle_school_list = []

        # 逐行读取、写入
        for row in range(1, st.max_row + 1):
            middle_school = st.cell(row=row, column=1).value
            middle_school_list.append((middle_school, middle_school))

        # 随手关闭文件好习惯
        wb.close()

        # 加入最后一个选项
        middle_school_list.append(('零星返郑报考', '零星返郑报考'))

        return tuple(middle_school_list)
