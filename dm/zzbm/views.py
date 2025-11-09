import random
from django.shortcuts import render
from django.contrib.auth.views import login_required
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.urls import reverse
from openpyxl.styles import Side, Border, Alignment, Font, PatternFill
from io import BytesIO
from datetime import datetime
from .models import Task, Student
from .tools import DataTool
from .forms import TaskForm, PutNameForm, FileUploadForm, StudentsUploadForm, ChangeInfoForm, ChangePutForm
import os
from PIL import Image, ImageDraw, ImageFont
import fitz
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.exceptions import InvalidFileException
import time
import threading
import smtplib
from email.mime.text import MIMEText
import json
from django.conf import settings


# 根路径
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# 实例化静态数据类
DT = DataTool()


def render_ht(request, template_name, context):
    """重新定义返回网页模板的方法"""
    # 加入选项卡标题
    context.update({'head_title': '{}中招美术测试报名系统'.format(settings.USER_NAME)})
    return render(request, template_name, context)


def correct_id(id_number):
    """检查所给的身份证号格式是否合法"""
    # 特殊身份证号直接予以通过
    if id_number in DT.special_id:
        return True

    # 位数须为18
    if len(id_number) != 18:
        return False

    # 前17位须为数字
    if not id_number[:17].isdigit():
        return False

    # 最后一位须为数字或X
    if not (id_number[17].isdigit() or id_number[17] == 'X'):
        return False

    # 校验
    s = 0
    for i in range(18):
        # 取得每一位上的数字
        if id_number[i].isdigit():
            n = int(id_number[i])
        else:
            n = 10

        # 计算权
        w = 2 ** (17 - i) % 11

        # 累加
        s += n * w

    if s % 11 != 1:
        return False

    # 通过所有检查，格式正确
    return True


def set_width_dict(st, width_dict):
    """设置表格列宽"""
    for k in width_dict.keys():
        st.column_dimensions[k].width = width_dict[k]


def add_border(st, start_row, start_column, end_row, end_column):
    """给表格的某部分添加完全的薄边框"""
    # 添加边框
    side = Side(style='thin')
    bd = Border(top=side, bottom=side, left=side, right=side)
    for r in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            st.cell(row=r, column=col).border = bd


def ct(st, start_row, start_column, end_row, end_column):
    """区域水平、垂直居中"""
    alignment = Alignment(horizontal='center', vertical='center')
    for r in range(start_row, end_row + 1):
        for c in range(start_column, end_column + 1):
            st.cell(row=r, column=c).alignment = alignment


def write_out(wb, fn=''):
    """黑科技：将编辑好的Excel表格文件通过浏览器下载到本地"""
    # 准备写入到IO中
    output = BytesIO()

    # 将Excel文件内容保存到IO中
    wb.save(output)

    # 重新定位到开始
    output.seek(0)

    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')

    # 依据日期时间设置默认文件名称
    ctime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    filename = '%s.xlsx' % ctime

    # 手动设置的文件名
    if fn:
        filename = fn

    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    # response.write(output.getvalue())
    wb.close()
    return response


def im_out(im, fn):
    """下载图片文件"""
    # 先临时保存文件到服务器
    fn += '.png'
    fp = os.path.join('media', 'card', fn)
    im.save(fp)

    with open(fp, 'rb') as f:
        response = HttpResponse(f, content_type='image/png')

        # 设置文件名称
        response['Content-Disposition'] = 'attachment; filename=%s' % fn

    # 删除临时存储的文件
    os.remove(fp)

    # 输出
    return response


def pdf_out(pdf_fn):
    """下载服务器上的pdf文件到本地，并删除其所在目录"""
    # 生成完整的pdf文件路径
    pdf_fp = os.path.join('media', 'seat_table', pdf_fn)

    with open(pdf_fp, 'rb') as f:
        response = HttpResponse(f, content_type='application/pdf')

        # 设置文件名称
        response['Content-Disposition'] = 'attachment; filename=%s' % pdf_fn

    # 删除上级目录
    shutil.rmtree(os.path.join('media', 'seat_table'))

    # 输出
    return response


def merge_to_pdf(path):
    """遍历一个目录中的所有png图象，合成一个pdf文件在原目录中"""
    # 打开一个新的PDF文件
    doc = fitz.open()

    # 遍历目录
    for file in sorted(os.listdir(path)):
        if file.endswith('png'):
            img_doc = fitz.open(os.path.join(path, file))
            pdf_bytes = img_doc.convert_to_pdf()
            pdf = fitz.open('pdf', pdf_bytes)
            doc.insert_pdf(pdf)

    # 保存，关闭
    doc.save(os.path.join(path, 'seat_table.pdf'))
    doc.close()


def give_exam_id(student):
    """给考生分配考号"""
    # 获取任务对象
    task = student.task_belong

    # 已分配标志
    give = False

    # 判断考生类型
    if student.subject == '素描或创意画':
        st = 1
        ed = task.mr1
        turn = task.turn1
        ed_max = 42
    else:
        st = 51
        ed = task.mr2
        turn = task.turn2
        ed_max = 53

    # 内层循环，遍历所有已存在考场
    for t in range(task.start_turn, turn + 1):
        if give:
            # 已完成分配，打破循环即可
            break

        if t == turn:
            e = ed + 1
        else:
            e = ed_max + 1

        for ri in range(st, e):
            # 生成考场号
            rm = str(t) + DT.str_two(ri)

            # 取得该考场已有考生人数
            try:
                rn = task.get_room_dict()[rm]
            except KeyError:
                # 无此考场，给考生分配为该考场01号
                student.room = rm
                student.seat = '01'
                student.save()
                student.make_exam_id()
                student.save()

                # 考场总数加1
                task.max_room += 1
                task.save()

                # 已分配，打破循环
                give = True
                break
            else:
                if rn < task.max_len:
                    # 考场未满，安排考场、考号
                    student.room = rm
                    student.seat = DT.str_two(rn + 1)
                    student.save()
                    student.make_exam_id()
                    student.save()

                    # 已分配，打破循环
                    give = True
                    break

    if not give:
        # 已存在考场均不可再分配，需要开新考场
        if ed < ed_max:
            new_room = str(turn) + DT.str_two(ed + 1)
            if student.subject == '素描或创意画':
                task.mr1 = ed + 1
            else:
                task.mr2 = ed + 1

            # 该考生分配为新考场01号
            student.room = new_room
            student.seat = '01'
            student.save()
            student.make_exam_id()
            student.save()

            # 最大考场号加1
            task.max_room += 1
            task.save()

        else:
            # if student.subject == '素描或创意画':
            #     task.turn1 += 1
            #     task.mr1 = task.get_mr(str(task.turn1))[0]
            #     ed = task.mr1
            # else:
            #     task.turn2 += 1
            #     task.mr2 = task.get_mr(str(task.turn2))[1]
            #     ed = task.mr2

            # TODO:本轮满，该考生暂不分配考场考号，等待手动处理场次之后手动分配，后续会有更高级的递归代码补充完善此部分
            pass


def remove_lay(filepath, t):
    """延时t秒删除文件"""
    time.sleep(t)
    try:
        os.remove(filepath)
    except FileNotFoundError:
        # 不用你来亲自动手删除了！
        pass


def all_id(task, add_method, exc=None):
    """返回某任务所有已报名非缺考考生的身份证号"""
    if exc is None:
        exc = []

    id_list = []
    for ob in Student.objects.filter(task_belong=task, add_method=add_method):
        if ob.id_number != '0' and not ob.miss:
            id_list.append(ob.id_number)

    # 去除
    for ob in exc:
        try:
            id_list.remove(ob.id_number)
        except ValueError:
            pass

    return id_list


def all_exam_id(task):
    """返回某任务所有已报名考生的准考证号"""
    ex_list = []
    for ob in Student.objects.filter(task_belong=task):
        ex_list.append(ob.exam_id)
    return ex_list


def all_pwd(task):
    """返回某任务所有已报名考生的报名序号"""
    pwd_list = []
    for ob in Student.objects.filter(task_belong=task):
        pwd_list.append(ob.pwd)
    return pwd_list


# Create your views here.
def index(request):
    """主页"""
    # 限制外界访问
    # if request.user.username not in DT.managers:
    #     raise Http404

    # 学生用户自动退出登录
    student_group = Group.objects.get(name='Student')
    if student_group in request.user.groups.all():
        logout(request)

    # 获取当前年份任务
    year = datetime.now().year
    tasks_now = Task.objects.filter(year=str(year))

    return render_ht(request, 'zzbm/index.html', context={
        'tasks_now': tasks_now,
        'is_manager': request.user.username in DT.managers,
        'title': '{}中招美术测试报名系统'.format(settings.USER_NAME),
    })


def login1(request):
    """用户登录"""
    if request.method == 'POST':
        # 对POST提交的数据作出处理
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, password=password)
        if user:
            # 限制学生用户登录
            student_group = Group.objects.get(name='Student')
            if student_group in user.groups.all():
                return render_ht(request, 'login.html', {'err': '你无权登录此系统！'})

            login(request, user)

            # 登录成功之后重定向到导航页
            return HttpResponseRedirect(reverse('zzbm:index'))
        else:
            return render_ht(request, 'zzbm/login.html', {'err': '用户名或密码错误！'})

    return render_ht(request, 'zzbm/login.html', {'err': ''})


def logout_view(request):
    """注销登录"""
    logout(request)
    return HttpResponseRedirect(reverse('zzbm:index'))


@login_required()
def set_pwd(request):
    """修改密码"""
    if request.method == 'POST':
        pwd = request.POST.get('old', '')
        if not authenticate(username=request.user.username, password=pwd):
            return render_ht(request, 'zzbm/set_pwd.html', {'err': '原密码不正确'})
        if request.POST.get('new', '') != request.POST.get('new_again', ''):
            return render_ht(request, 'zzbm/set_pwd.html', {'err': '两次输入的密码不一致'})

        # 修改密码
        new_pwd = request.POST.get('new', '')
        request.user.set_password(new_pwd)
        request.user.save()
        update_session_auth_hash(request, request.user)

        # 退出登录然后重新登录
        logout(request)
        return render_ht(request, 'zzbm/login.html', {'err': '密码已修改成功，请重新登录'})

    return render_ht(request, 'zzbm/set_pwd.html', {'err': ''})


def year_list(request):
    """显示年份列表"""
    # 非管理员禁止访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 获取往年任务
    tasks_before = []
    for task in Task.objects.all():
        if int(task.year) < datetime.now().year:
            tasks_before.append(task)

    return render_ht(request, 'zzbm/year_list.html', context={
        'tasks_before': tasks_before})


def task_main(request, task_id):
    """加载任务主页"""
    # 限制外界访问
    # if request.user.username not in DT.managers:
    #     raise Http404

    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    # 刷新活跃状态
    task.update_active()

    # 制定标题
    title = '{}{}年中招艺术后备生专业测试报名'.format(settings.USER_NAME, task.year)

    # 制定提示
    condition1 = not task.changed and task.active
    tip1 = '报名将于{}截止！'.format(task.end_date)
    condition2 = not task.active and task.changed
    tip2 = '报名已于{}截止！'.format(task.end_date)

    context = {
        'task': task,
        'title': title,
        'condition1': condition1,
        'tip1': tip1,
        'condition2': condition2,
        'tip2': tip2,
        'is_manager': request.user.username in DT.managers,
        'is_operator': request.user.username in DT.operators,
    }
    return render_ht(request, 'zzbm/task_main.html', context)


def public(request):
    """发布新的任务"""
    # 禁止非管理员访问
    if request.user.username not in DT.managers:
        raise Http404

    # 获取当前年份任务
    year = datetime.now().year
    tasks_now = Task.objects.filter(year=str(year))

    if tasks_now:
        # 当前年份已存在任务，不可再发布，给出提示
        return render_ht(request, 'zzbm/exist_tip.html', context={
            'task': tasks_now[0], 'year': year})
    else:
        # 不存在当前年份任务，可以发布
        if request.method != 'POST':
            # 未提交数据，创建新的表单
            form = TaskForm()
        else:
            # 对POST提交的数据作出处理
            form = TaskForm(request.POST)
            if form.is_valid():
                new_task = form.save(commit=False)

                # 补全任务属性(年份)
                new_task.give_year()

                # 保存任务，重定向至任务主页
                new_task.save()
                return HttpResponseRedirect(reverse('zzbm:task_main',
                                                    args=[new_task.id]))

        return render_ht(request, 'zzbm/public.html', {'form': form})


def delete_task(request, task_id):
    """删除任务"""
    # 禁止非管理员用户执行此项操作
    if request.user.username not in DT.managers:
        raise Http404

    # 取出要删除的任务对象
    task = Task.objects.get(id=task_id)

    # 删除所有报名考生的照片
    for student in Student.objects.filter(task_belong=task, add_method=1):
        try:
            pic = BASE_DIR + '/media/' + str(student.photo)
            os.remove(pic)
        except IsADirectoryError:
            # 不存在文件，忽略此步
            pass

    # 执行删除操作
    task.delete()

    # 重定向至主页
    return HttpResponseRedirect(reverse('zzbm:index'))


def task_manage(request, task_id):
    """任务管理页"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象
    task = Task.objects.get(id=task_id)

    # 第二个按钮名称
    if task.active:
        btn_name = '停止报名'
    else:
        btn_name = '恢复报名'

    # 第六个按钮名称
    if task.open_download:
        open_btn = '关闭准考证下载'
    else:
        open_btn = '开放准考证下载'

    # 第八个按钮名称
    if task.auto_give:
        give_btn = '关闭报名自动分配考号'
    else:
        give_btn = '开启报名自动分配考号'

    # 未分配考场考生数
    n = len(Student.objects.filter(task_belong=task, exam_id='未分配'))

    # 五个人数
    n1 = len(Student.objects.filter(task_belong=task, add_method=1))
    n2 = len(Student.objects.filter(task_belong=task, add_method=2))
    n3 = len(Student.objects.filter(task_belong=task, add_method=1, subject='素描或创意画'))
    n4 = len(Student.objects.filter(task_belong=task, add_method=1, subject='书法或国画'))
    n5 = len(Student.objects.filter(task_belong=task, miss=True))

    # 尾考场
    max_turn1 = str(max(task.turn1, task.start_turn)) + DT.str_two(task.mr1)
    max_turn2 = str(max(task.turn2, task.start_turn)) + DT.str_two(task.mr2)

    context = {'task': task, 'btn_name': btn_name, 'open_btn': open_btn, 'err': '',
               'n': n, 'give_btn': give_btn, 'n1': n1, 'n2': n2, 'n3': n3, 'n4': n4, 'n5': n5,
               'max_turn1': max_turn1, 'max_turn2': max_turn2}
    return render_ht(request, 'zzbm/task_manage.html', context)


def change_date(request, task_id):
    """更改任务截止日期"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象
    task = Task.objects.get(id=task_id)

    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = TaskForm(instance=task)
    else:
        # 对POST提交的数据作出处理
        form = TaskForm(instance=task, data=request.POST)
        if form.is_valid():
            form.save()

            # 任务状态作出相应的更改
            task.changed = False
            task.active = True
            task.save()

            # 重定向至任务管理页
            return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))

    context = {'task': task, 'form': form}
    return render_ht(request, 'zzbm/change_date.html', context)


def change_active(request, task_id):
    """停止或恢复报名"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象
    task = Task.objects.get(id=task_id)

    # 第二个按钮名称
    if task.active:
        btn_name = '停止报名'
    else:
        btn_name = '恢复报名'

    # 第六个按钮名称
    if task.open_download:
        open_btn = '关闭准考证下载'
    else:
        open_btn = '开放准考证下载'

    # 第八个按钮名称
    if task.auto_give:
        give_btn = '关闭报名自动分配考号'
    else:
        give_btn = '开启报名自动分配考号'

    # 未分配考场考生数
    n = len(Student.objects.filter(task_belong=task, exam_id='未分配'))

    # 准考证下载开放中，不可恢复报名，自动分配除外
    if not task.active and task.open_download and not task.auto_give:
        return render_ht(request, 'zzbm/task_manage.html', context={
            'task': task,
            'btn_name': btn_name,
            'open_btn': open_btn,
            'give_btn': give_btn,
            'err': '准考证下载开放中，不可恢复报名',
            'n': n,
        })

    # 更改状态
    task.active = not task.active
    task.save()

    # 重定向至任务管理页（当前页）
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def list_put(request, task_id, st_id):
    """列出已报名考生"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象和该任务下所有考生对象
    task = Task.objects.get(id=task_id)
    students = []
    for student in Student.objects.filter(task_belong=task).order_by('exam_id'):
        students.append(student)
    students.sort(key=lambda x: x.add_method)
    students = tuple(students)

    # 制定标题和提示内容
    title = '所有报名{}中招美术加试的考生（按报名顺序排列）'.format(task.year)
    msg = '暂无考生报名'
    ipt = ''

    if request.method == 'POST':
        # 方法为POST，处理查询信息
        # 前端读取查询内容
        que_msg = request.POST.get('que', '')

        # 选出有关对象
        st_list = []
        for ob in students:
            if que_msg in ob.name or que_msg in ob.id_number:
                st_list.append(ob)

        # 重新赋值变量
        students = tuple(st_list)
        title = '查询结果'
        msg = '未查询到相关结果'
        ipt = que_msg

    context = {'task': task, 'students': students, 'title': title, 'msg': msg,
               'ipt': ipt, 'st_id': str(st_id)}
    return render_ht(request, 'zzbm/list_put.html', context)


def put_name(request, task_id):
    """报名"""
    # 不允许报名
    # if request.user.username not in DT.managers:
    #     raise Http404

    # 开放报名时间
    if datetime.now() < datetime(2024, 5, 16, 18, 0, 0):
        return HttpResponse('暂未开放报名')

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 不可报名非活跃状态任务
    if not task.active:
        raise Http404

    # # 只可报名当年任务
    # if task.year != str(datetime.now().year):
    #     raise Http404

    # 生成标题
    title = '报名{}{}年中招美术加试'.format(settings.USER_NAME, task.year)

    # 判断请求方法，创建表单
    if request.method != 'POST':
        # 未提交数据，创建新的表单
        form = PutNameForm()
    else:
        # 对POST提交的数据作出处理
        form = PutNameForm(request.POST, request.FILES)
        if form.is_valid():
            new_put = form.save(commit=False)

            # 判断身份证号是否合法
            if not correct_id(new_put.id_number):
                return render_ht(request, 'zzbm/put_name.html', context={
                    'task': task,
                    'title': title,
                    'form': form,
                    'err': '身份证号格式非法！',
                })

            # 判断性别与身份证号是否匹配
            if len(new_put.id_number) == 18:
                if {0: '女', 1: '男'}[int(new_put.id_number[16]) % 2] != new_put.gender:
                    return render_ht(request, 'zzbm/put_name.html', context={
                        'task': task,
                        'title': title,
                        'form': form,
                        'err': '身份证号与性别不匹配，请检查是否输入有误！',
                    })

            # 防止重复报名
            if new_put.id_number in all_id(task, add_method=1):
                return render_ht(request, 'zzbm/put_name.html', context={
                    'task': task,
                    'title': title,
                    'form': form,
                    'err': '身份证号{}的考生已报名，请勿重复报名'.format(new_put.id_number),
                })

            # 判断手机号格式是否有误
            if len(new_put.phone_number) != 11 or new_put.phone_number[0] != '1':
                return render_ht(request, 'zzbm/put_name.html', context={
                    'task': task,
                    'title': title,
                    'form': form,
                    'err': '手机号格式有误！',
                })

            # 判断邮箱格式是否有误
            if new_put.email:
                if '@' not in new_put.email or '.com' not in new_put.email:
                    return render_ht(request, 'zzbm/put_name.html', context={
                        'task': task,
                        'title': title,
                        'form': form,
                        'err': '邮箱格式有误！',
                    })

            # 一寸照片必须上传
            if not new_put.photo:
                return render_ht(request, 'zzbm/put_name.html', context={
                    'task': task,
                    'title': title,
                    'form': form,
                    'err': '请上传个人一寸照片',
                })

            # 一寸照片大小不能超过1MB
            if new_put.photo.size > 1024 * 1024:
                return render_ht(request, 'zzbm/put_name.html', context={
                    'task': task,
                    'title': title,
                    'form': form,
                    'err': '上传一寸照片大小不能超过1MB',
                })

            # 记录报名序数
            new_put.task_belong = task
            new_put.num = task.added + 1
            task.added += 1
            task.save()

            # 保存，生成报名序号，再保存
            new_put.save()

            # 防止报名序号重复
            while new_put.pwd == '000000' or new_put.pwd in all_pwd(task):
                new_put.update_pwd()
            new_put.save()

            # 实时分配考场、考号
            if task.auto_give:
                give_exam_id(new_put)

            # # 加载提示页
            # return render(request, 'zzbm/tip_pwd.html', context={
            #     'pwd': new_put.pwd, 'task': task})

            # 重定向至信息查看页
            return HttpResponseRedirect(reverse('zzbm:student', args=[new_put.id, new_put.pwd]))

    context = {'task': task, 'title': title, 'form': form, 'err': ''}
    return render_ht(request, 'zzbm/put_name.html', context)


def que(request, task_id):
    """查询身份证号、姓名"""
    # 取出任务对象
    task = Task.objects.get(id=task_id)

    if request.method != 'POST':
        # 未提交数据，直接加载输入框
        return render_ht(request, 'zzbm/que.html', context={'task': task, 'err': ''})
    else:
        # 对POST提交的数据作出判断
        # 前端读取数据
        id_number = request.POST.get('id-number', '')
        name = request.POST.get('nm', '')

        try:
            # 锁定对象
            st_ob = Student.objects.filter(task_belong=task, id_number=id_number,
                                           miss=False, add_method=1)[0]
        except IndexError:
            # 未找到对象
            return render_ht(request, 'zzbm/que.html', context={
                'task': task, 'err': '身份证号{}的考生未报名'.format(id_number)})
        else:
            if name != st_ob.name:
                # 姓名不匹配
                return render_ht(request, 'zzbm/que.html', context={
                    'task': task, 'err': '身份证号与姓名不匹配'})

            # 重定向至考生查看页
            return HttpResponseRedirect(reverse('zzbm:student', args=[st_ob.id, st_ob.pwd]))


def stu(request, student_id, pwd):
    """查看考生"""
    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 检查口令是否匹配
    if pwd != student.pwd:
        raise Http404

    return render_ht(request, 'zzbm/student.html', context={'st': student})


def delete_st(request, student_id, delete_method):
    """删除已报名考生"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.operators:
        raise Http404

    # 取出要删除的考生对象及所属任务对象、序数
    student = Student.objects.get(id=student_id)
    task = student.task_belong
    num = student.num

    # 删除图片文件
    try:
        pic = BASE_DIR + '/media/' + str(student.photo)
        os.remove(pic)
    except IsADirectoryError:
        # 不存在文件，忽略此步
        pass

    # 删除考生对象
    student.delete()

    # 该任务已报名考生数减1
    task.added -= 1
    task.save()

    # 所有大于num的序数减1
    for ob in Student.objects.filter(task_belong=task):
        if ob.num > num:
            ob.num -= 1
            ob.save()

    if delete_method:
        # 重定向至考生列表页
        return HttpResponseRedirect(reverse('zzbm:list_put', args=[task.id, 0]))
    else:
        # 重定向至考生查看页
        return HttpResponseRedirect(reverse('zzbm:see_multi', args=[task.id]))


def set_len(request, task_id):
    """设置考场人数"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    if request.method == 'POST':
        # 前端获取设置数字
        ml = int(request.POST.get('ml', ''))

        # 考场人数不可减小
        if ml < task.max_len:
            return render_ht(request, 'zzbm/set_len.html', context={
                'task': task, 'err': '考场人数不可减小'})

        # 考场人数必须在30~60之间
        if ml > 60 or ml < 30:
            return render_ht(request, 'zzbm/set_len.html', context={
                'task': task, 'err': '考场人数必须在30~60之间'})

        # 更改任务属性
        task.max_len = ml
        task.save()

        # 外层循环，遍历所有未分配考号的考生
        for student in Student.objects.filter(
                task_belong=task, exam_id='未分配').order_by('subject'):

            # 给考生分配考场、考号
            give_exam_id(student)

        # 完成分配，重定向至任务管理页面
        return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))

    return render_ht(request, 'zzbm/set_len.html', context={'task': task, 'err': ''})


def set_time(request, task_id):
    """设置考试时间"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 制定标题
    title = '设置{}年中招美术加试考试时间：'.format(task.year)

    # 设置可选时间范围
    nt = datetime.now()
    dt_min = datetime.strftime(nt, '%Y-%m-%dT%H:%M')
    dt_max = '{}-12-31T23:59'.format(task.year)

    if request.method == 'POST':
        # 读取前端数据
        dt_str = request.POST.get('dt', '')
        dt_str_1 = request.POST.get('dt_1', '')
        dt_str_2 = request.POST.get('dt_2', '')
        dt_str_3 = request.POST.get('dt_3', '')
        dt_str_4 = request.POST.get('dt_4', '')
        dt_str_5 = request.POST.get('dt_5', '')

        if dt_str:
            dt = datetime.strptime(dt_str, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt.hour < 12:
                tp = '上午'
            else:
                tp = '下午'
            time_tuple = (dt.year, dt.month, dt.day, tp, dt.hour, DT.str_two(dt.minute))
            ch_time = '{}年{}月{}日 {}{}:{}'.format(*time_tuple)

            # 设置任务属性
            task.format_start_time = dt_str
            task.start_time = ch_time

        if dt_str_1:
            dt_1 = datetime.strptime(dt_str_1, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt_1.hour < 12:
                tp_1 = '上午'
            else:
                tp_1 = '下午'
            time_tuple_1 = (dt_1.year, dt_1.month, dt_1.day, tp_1, dt_1.hour, DT.str_two(dt_1.minute))
            ch_time_1 = '{}年{}月{}日 {}{}:{}'.format(*time_tuple_1)

            # 设置任务属性
            task.format_start_time_1 = dt_str_1
            task.start_time_1 = ch_time_1

        if dt_str_2:
            dt_2 = datetime.strptime(dt_str_2, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt_2.hour < 12:
                tp_2 = '上午'
            else:
                tp_2 = '下午'
            time_tuple_2 = (dt_2.year, dt_2.month, dt_2.day, tp_2, dt_2.hour, DT.str_two(dt_2.minute))
            ch_time_2 = '{}年{}月{}日 {}{}:{}'.format(*time_tuple_2)

            # 设置任务属性
            task.format_start_time_2 = dt_str_2
            task.start_time_2 = ch_time_2

        if dt_str_3:
            dt_3 = datetime.strptime(dt_str_3, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt_3.hour < 12:
                tp_3 = '上午'
            else:
                tp_3 = '下午'
            time_tuple_3 = (dt_3.year, dt_3.month, dt_3.day, tp_3, dt_3.hour, DT.str_two(dt_3.minute))
            ch_time_3 = '{}年{}月{}日 {}{}:{}'.format(*time_tuple_3)

            # 设置任务属性
            task.format_start_time_3 = dt_str_3
            task.start_time_3 = ch_time_3

        if dt_str_4:
            dt_4 = datetime.strptime(dt_str_4, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt_4.hour < 12:
                tp_4 = '上午'
            else:
                tp_4 = '下午'
            time_tuple_4 = (dt_4.year, dt_4.month, dt_4.day, tp_4, dt_4.hour, DT.str_two(dt_4.minute))
            ch_time_4 = '{}年{}月{}日 {}{}:{}'.format(*time_tuple_4)

            # 设置任务属性
            task.format_start_time_4 = dt_str_4
            task.start_time_4 = ch_time_4

        if dt_str_5:
            dt_5 = datetime.strptime(dt_str_5, '%Y-%m-%dT%H:%M')
            # 转换成易于中国人理解的日期字符串
            if dt_5.hour < 12:
                tp_5 = '上午'
            else:
                tp_5 = '下午'
            time_tuple_5 = (dt_5.year, dt_5.month, dt_5.day, tp_5, dt_5.hour, DT.str_two(dt_5.minute))
            ch_time_5 = '{}年{}月{}日 {}{}:{}'.format(*time_tuple_5)

            # 设置任务属性
            task.format_start_time_5 = dt_str_5
            task.start_time_5 = ch_time_5

        # 保存任务
        task.save()

        return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))

    context = {'task': task, 'title': title, 'min': dt_min, 'max': dt_max}
    return render_ht(request, 'zzbm/set_time.html', context)


def change_open(request, task_id):
    """开放/关闭准考证下载"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 第二个按钮名称
    if task.active:
        btn_name = '停止报名'
    else:
        btn_name = '恢复报名'

    # 第六个按钮名称
    if task.open_download:
        open_btn = '关闭准考证下载'
    else:
        open_btn = '开放准考证下载'

    # 第八个按钮名称
    if task.auto_give:
        give_btn = '关闭报名自动分配考号'
    else:
        give_btn = '开启报名自动分配考号'

    # 未分配考场考生数
    n = len(Student.objects.filter(task_belong=task, exam_id='未分配'))

    if not task.open_download:
        # 考试报名中，不可开放准考证下载，自动分配除外
        if task.active and not task.auto_give:
            return render_ht(request, 'zzbm/task_manage.html', context={
                'task': task,
                'btn_name': btn_name,
                'open_btn': open_btn,
                'give_btn': give_btn,
                'err': '考试报名中，不可开放准考证下载',
                'n': n,
            })

        # 未设置考试时间，不可开放准考证下载
        if not task.start_time:
            return render_ht(request, 'zzbm/task_manage.html', context={
                'task': task,
                'btn_name': btn_name,
                'open_btn': open_btn,
                'give_btn': give_btn,
                'err': '未设置考试时间，不可开放准考证下载',
                'n': n,
            })

        # 存在未分配考号的考生，不可开放准考证下载
        for student in Student.objects.filter(task_belong=task):
            if student.exam_id == '未分配':
                return render_ht(request, 'zzbm/task_manage.html', context={
                    'task': task,
                    'btn_name': btn_name,
                    'open_btn': open_btn,
                    'give_btn': give_btn,
                    'err': '存在未分配考号的考生，不可开放准考证下载',
                    'n': n,
                })

    # 更改任务属性
    task.open_download = not task.open_download
    task.save()

    # 重定向至任务管理页
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def download_card(request, student_id, pwd):
    """在线生成、下载准考证"""
    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 检验匹配
    if pwd != student.pwd:
        raise Http404

    # 两轮之外及未分配者暂不可下载准考证
    try:
        tm = int(student.room[0])
    except ValueError:
        return HttpResponse('暂不可下载准考证')
    else:
        if tm > 5 or student.exam_id == '未分配':
            return HttpResponse('暂不可下载准考证')

    # 未开放下载准考证的普通用户禁止下载
    if not student.task_belong.open_download and request.user.username != 'zz106dyc':
        raise Http404

    # 加载新画布
    image = Image.new('RGBA', (1240, 877), (255, 255, 255))
    draw = ImageDraw.Draw(image)

    # 添加标题
    title = '{}{}中招美术测试准考证'.format(settings.USER_NAME, student.task_belong.year)
    title_font = ImageFont.truetype('font/simhei.ttf', 48)
    title_width = draw.textsize(title, title_font)[0]
    title_x = (1240 - title_width) / 2
    draw.text((title_x, 75), title, font=title_font, fill='black')

    # 根据准考证号第8位获取考试时间
    st, fst = student.get_exam_time()

    # 该场考试时间尚未设置的情况
    if not st:
        return HttpResponse('暂不可下载准考证')

    # 生成考试时间显示格式
    dt_start = datetime.strptime(fst, '%Y-%m-%dT%H:%M')
    h = int(DT.ter)
    m = int((DT.ter - int(DT.ter)) * 60)
    end_h = dt_start.hour + h
    end_m = dt_start.minute + m

    # 处理考试结束时间
    if end_m >= 60:
        end_m -= 60
        end_h += 1

    # TODO:生成格式化时间表示(须进一步修改)
    exam_time = '{} —— {}:{}'.format(st, end_h, DT.str_two(end_m))

    # 写入考生信息
    msg_list = [
        student.exam_id,
        student.name,
        student.subject,
        exam_time,
        settings.USER_NAME,
        DT.address,
        student.room,
        DT.pos.get(student.room[1:]),
        student.seat,
    ]
    msg = '''\t\t准考证号：{}\n
    姓名：{}\n
    考试科目：{}\n
    考试时间：{}\n
    考点：{}（{}）\n
    考场号：{}（{}）\n
    座位号：{}\n'''.format(*msg_list)
    msg_font = ImageFont.truetype('font/simsun.ttc', 28)
    draw.text((35, 225), msg, font=msg_font, fill='black')

    # 粘贴考生照片
    photo = Image.open(os.path.join('media', str(student.photo)))
    image.paste(photo, (855, 225))

    # 盖章
    zhang = Image.open(os.path.join('media', 'zhang.png'))
    zhang = zhang.resize((int(zhang.width / 2), int(zhang.height / 2)))
    image.paste(zhang, (825, 550), mask=zhang)

    # 粘贴公众号二维码
    code = Image.open(os.path.join('media', '640.webp'))
    code = code.resize((int(code.width / 2), int(code.height / 2)))
    image.paste(code, (80, 600))

    # 输出png格式图片
    fn = student.task_belong.year + '_' + student.pwd
    return im_out(image, fn)


def get_seat_table(request, task_id):
    """在线生成、下载考场座次表"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 建立名为seat_table用来临时存放座次表图片的目录
    try:
        os.mkdir(os.path.join('media', 'seat_table'))
    except FileExistsError:
        # 目录已存在，则先删再建，确保目录开始为空
        shutil.rmtree(os.path.join('media', 'seat_table'))
        os.mkdir(os.path.join('media', 'seat_table'))

    # 大标题
    big_title = '{}{}中考美术加试座次表'.format(settings.USER_NAME, task.year)

    # 标题样式
    title_font = ImageFont.truetype('font/simhei.ttf', 36)

    # 设置字体
    font = ImageFont.truetype('font/simsun.ttc', 20)

    # 循环遍历所有考场
    for rm in task.all_room(e8='3'):
        # # 分段下载
        # if int(rm[1:]) <= 15:
        #     continue

        # 判断考场科目
        sub = task.judge_type(rm)

        # 小标题
        if sub:
            small_title = '第{}考场A（{}）'.format(rm, sub)
        else:
            small_title = '第{}考场A'.format(rm)

        # 打开新的图像
        image = Image.new('RGB', (1240, 1754), (255, 255, 255))
        draw = ImageDraw.Draw(image)

        # 标题居中
        big_title_width = draw.textsize(big_title, title_font)[0]
        big_title_x = (1240 - big_title_width) / 2
        draw.text((big_title_x, 70), big_title, font=title_font, fill='black')
        small_title_width = draw.textsize(small_title, title_font)[0]
        small_title_x = (1240 - small_title_width) / 2
        draw.text((small_title_x, 130), small_title, font=title_font, fill='black')

        # 依次写入考生信息
        # 初始位置
        x, y = 50, 190

        # 记录换页
        change_page = False

        for st in Student.objects.filter(task_belong=task, room=rm).order_by('seat'):
            # 粘贴照片
            photo = Image.open(os.path.join('media', str(st.photo)))
            image.paste(photo.resize((148, 207)), (x, y))

            # 写入信息
            msg = '{}\n{}'.format(st.exam_id, st.name)
            draw.text((x, y + 220), msg, font=font, fill='black')

            # 更改位置
            if x < 900:
                x += 185
            else:
                if y < 1100:
                    x = 50
                    y += 290
                else:
                    # 换页
                    # 临时保存
                    fn = rm + 'A.png'
                    fp = os.path.join('media', 'seat_table', fn)
                    image.save(fp)

                    # 小标题
                    if sub:
                        small_title = '第{}考场B（{}）'.format(rm, sub)
                    else:
                        small_title = '第{}考场B'.format(rm)

                    # 打开新的图像
                    image = Image.new('RGB', (1240, 1754), (255, 255, 255))
                    draw = ImageDraw.Draw(image)

                    # 标题居中
                    big_title_width = draw.textsize(big_title, title_font)[0]
                    big_title_x = (1240 - big_title_width) / 2
                    draw.text((big_title_x, 70), big_title, font=title_font, fill='black')
                    small_title_width = draw.textsize(small_title, title_font)[0]
                    small_title_x = (1240 - small_title_width) / 2
                    draw.text((small_title_x, 130), small_title, font=title_font, fill='black')

                    # 依次写入考生信息
                    # 初始位置
                    x, y = 50, 190

                    # 已换页
                    change_page = True

            # 调试位置
            # image.paste(photo.resize((118, 165)), (x, y))

        # 临时保存
        if change_page:
            if x != 50 or y != 190:
                fn = rm + 'B.png'
                fp = os.path.join('media', 'seat_table', fn)
                image.save(fp)
        else:
            fn = rm + 'A.png'
            fp = os.path.join('media', 'seat_table', fn)
            image.save(fp)

    # 合成pdf，输出
    merge_to_pdf(os.path.join('media', 'seat_table'))
    return pdf_out('seat_table.pdf')


def export_students(request, task_id):
    """导出考生信息为Excel表格"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象，生成标题
    task = Task.objects.get(id=task_id)
    title = '{}{}中考美术加试报名信息'.format(settings.USER_NAME, task.year)

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入标题
    st['A1'].value = title
    st.merge_cells(range_string='A1:N1')

    # 写入表头
    st['A2'].value = '序号'
    st['B2'].value = '姓名'
    st['C2'].value = '性别'
    st['D2'].value = '身份证号'
    st['E2'].value = '手机号'
    st['F2'].value = '初中毕业学校'
    st['G2'].value = '选考科目'
    st['H2'].value = '报名序号'
    st['I2'].value = '考场号'
    st['J2'].value = '座位号'
    st['K2'].value = '准考证号'
    st['L2'].value = '成绩'
    st['M2'].value = '是否缺考'
    st['N2'].value = '报名时间'

    # 标题和表头字体格式
    st['A1'].font = Font(size=14, bold=True)
    for c in range(1, 13):
        st.cell(row=2, column=c).font = Font(size=12, bold=True)

    # 初始行
    row = 3

    # 循环遍历每个已报名考生，写入表格
    for student in Student.objects.filter(task_belong=task).order_by('num'):
        st.cell(row=row, column=1).value = student.num
        st.cell(row=row, column=2).value = student.name
        st.cell(row=row, column=3).value = student.gender
        st.cell(row=row, column=4).value = student.id_number
        st.cell(row=row, column=5).value = student.phone_number
        st.cell(row=row, column=6).value = student.middle_school
        st.cell(row=row, column=7).value = student.subject
        st.cell(row=row, column=8).value = student.pwd
        st.cell(row=row, column=9).value = student.room
        st.cell(row=row, column=10).value = student.seat
        st.cell(row=row, column=11).value = student.exam_id

        # 写入成绩
        if student.score != 'O':
            st.cell(row=row, column=12).value = student.score

        # 写入缺考状态
        if student.miss:
            st.cell(row=row, column=13).value = '缺考'

        # 日期格式化为字符串再写入
        dt = datetime.strftime(student.datetime_added, '%Y-%m-%d %H:%M:%S')
        st.cell(row=row, column=14).value = dt

        # 下一行
        row += 1

    # 添加边框
    add_border(st, start_row=2, start_column=1, end_row=row - 1, end_column=14)

    # 设置合适的列宽
    width_dict = {'A': 6, 'B': 9, 'C': 6, 'D': 20, 'E': 12, 'F': 24, 'G': 12,
                  'H': 12, 'I': 8, 'J': 8, 'K': 15, 'L': 7, 'M': 9, 'N': 20}
    set_width_dict(st, width_dict)

    # 输出
    return write_out(wb)


def score_manage(request, task_id):
    """成绩管理主页"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象及所有考生对象
    task = Task.objects.get(id=task_id)
    students = Student.objects.filter(task_belong=task).order_by('exam_id')

    # 生成标题
    title = '{}{}中考美术加试成绩管理'.format(settings.USER_NAME, task.year)

    # 生成提示
    tip = '成绩说明：A(合格)，D(不合格)，O(未设置)'

    context = {
        'task': task,
        'title': title,
        'students': students,
        'tip': tip,
        'form': FileUploadForm(),
        'err': '',
        'have_button': False,
        'err_id': -1,
    }
    return render_ht(request, 'zzbm/score_manage.html', context)


def get_score(request, student_id, pwd):
    """查询某个考生的成绩"""
    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 暂时不允许普通用户访问
    if not student.task_belong.can_que:
        if request.user.username not in DT.managers:
            return HttpResponse('成绩未出，请耐心等待')

    # 检验匹配
    if pwd != student.pwd:
        raise Http404

    # 标题
    title = '{}{}年美术后备生测试成绩'.format(settings.USER_NAME, student.task_belong.year)

    context = {'student': student, 'title': title, 'score': DT.scores_2025[student.score]}
    return render_ht(request, 'zzbm/student_score.html', context)


def download_temp(request, task_id):
    """下载成绩模板"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 定制标题
    title = '{}{}中考美术加试成绩模板'.format(settings.USER_NAME, task.year)

    # 定制说明内容
    tip_msg = '说明：请勿更改表格格式，仅在E列填写考生成绩即可\n成绩只可填写‘A’，‘B’，‘C’或‘D’'

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入标题，设置标题样式，水平居中
    st['A1'].value = title
    st['A1'].font = Font(size=14, bold=True)
    st['A1'].alignment = Alignment(horizontal='center')

    # 写入提示，行高2倍，垂直居中，自动换行
    st['A2'].value = tip_msg
    st.row_dimensions[2].height = 30
    st['A2'].alignment = Alignment(vertical='center', wrap_text=True)

    # 前两行合并单元格
    st.merge_cells(range_string='A1:H1')
    st.merge_cells(range_string='A2:H2')

    # 写入表头
    st['A3'].value = '考场号'
    st['B3'].value = '座位号'
    st['C3'].value = '准考证号'
    st['D3'].value = '考生姓名'
    st['E3'].value = '身份证号'
    st['F3'].value = '考试科目'
    st['G3'].value = '成绩'
    st['H3'].value = '备注'

    # 设置表头格式
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        st[c + '3'].font = Font(bold=True)

    # 初始行
    row = 4

    # 循环遍历所有考生写入信息
    for student in Student.objects.filter(task_belong=task).order_by('exam_id'):
        st.cell(row=row, column=1).value = student.room
        st.cell(row=row, column=2).value = student.seat
        st.cell(row=row, column=3).value = student.exam_id
        st.cell(row=row, column=4).value = student.name
        st.cell(row=row, column=5).value = student.id_number
        st.cell(row=row, column=6).value = student.subject

        # 当成绩不为O时写入
        if student.score != 'O':
            st.cell(row=row, column=7).value = student.score

        # 缺考备注
        if student.miss:
            st.cell(row=row, column=8).value = '缺考'

        # 下一行
        row += 1

    # 为E列添加下拉选项
    options = ['A', 'B', 'C', 'D']
    validation = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                                allow_blank=True)
    for r in range(4, row):
        validation.add('E' + str(r))
    st.add_data_validation(validation)

    # 调整列宽
    width_dict = {'A': 10, 'B': 10, 'C': 18, 'D': 12, 'E': 20, 'F': 18, 'G': 10, 'H': 8}
    set_width_dict(st, width_dict)

    return write_out(wb, fn='score_template.xlsx')


def write_score(request, task_id):
    """根据上传的文件赋予成绩"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出对应的任务对象及所有考生对象
    task = Task.objects.get(id=task_id)
    students = Student.objects.filter(task_belong=task).order_by('exam_id')

    # 生成标题
    title = '{}{}中考美术加试成绩管理'.format(settings.USER_NAME, task.year)

    # 生成提示
    tip = '成绩说明：A(合格)，D(不合格)，O(未设置)'

    # 获取表单提交内容
    form = FileUploadForm(request.POST, request.FILES)

    if form.is_valid():
        # 临时保存文件
        upload_file = form.cleaned_data['file']
        file_path = 'media/temp/' + upload_file.name

        # 保存到指定路径
        with open(file_path, 'wb') as f:
            for chunk in upload_file.chunks():
                f.write(chunk)

        # 处理上传的文件
        try:
            wb = load_workbook(file_path)
            st = wb.active
        except InvalidFileException:
            # 删除临时文件
            os.remove(file_path)

            # 错误提示信息
            return render_ht(request, 'zzbm/score_manage.html', context={
                'task': task,
                'title': title,
                'students': students,
                'tip': tip,
                'form': FileUploadForm(),
                'err': '文件格式必须为xlsx',
                'have_button': False,
                'err_id': -1,
            })

        # 删除临时文件
        os.remove(file_path)

        # 成功数和失败数
        success, fail = 0, 0

        # 错误提示填充
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # 从第2行开始逐行读取
        for row in range(2, st.max_row + 1):
            # 读取准考证号
            exam_id = st.cell(row=row, column=3).value

            # 根据准考证号匹配考生
            sts = students.filter(exam_id=exam_id)

            # 一个准考证号对应多个考生给出提示
            if len(sts) > 1:
                st.cell(row=row, column=9).value = '准考证号{}对应超过一个考生'.format(exam_id)
                st.cell(row=row, column=9).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            try:
                student = sts[0]
            except IndexError:
                # 写出错误提示
                st.cell(row=row, column=9).value = '准考证号{}不存在'.format(exam_id)
                st.cell(row=row, column=9).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # # 已进行过成绩赋值，给出提示
            # if student.score != 'O':
            #     st.cell(row=row, column=9).value = '已给过成绩，为{}'.format(student.score)
            #     st.cell(row=row, column=9).fill = yellow_fill
            #
            #     # 失败数加一
            #     fail += 1
            #     continue

            # # 判断考生姓名是否匹配
            # name = st.cell(row=row, column=4).value
            # if name != student.name:
            #     # 写出错误提示
            #     st.cell(row=row, column=7).value = '准考证号与姓名不匹配'
            #     st.cell(row=row, column=7).fill = yellow_fill
            #
            #     # 失败数加一
            #     fail += 1
            #     continue

            # 读取赋予的成绩
            score = st.cell(row=row, column=7).value

            # 缺考直接跳过
            if st.cell(row=row, column=8).value == '缺考':
                student.miss = True
                student.save()
                continue

            # 成绩格式不合要求
            if score not in ['A', 'B', 'C', 'D', 'O']:
                # 写出错误提示
                st.cell(row=row, column=9).value = '成绩值只能为A、B、C、D、O其中之一'
                st.cell(row=row, column=9).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 通过所有检验，可进行成绩赋值
            student.score = score
            student.miss = False
            student.save()

            # 成功数加一
            success += 1

        # 生成提示信息
        tip_msg = '共上传{}条数据，成功{}条，失败{}条'.format(success + fail, success, fail)

        if fail > 0:
            # 存在处理失败数据，保存失败文件
            st['I1'].value = '错误提示'
            st['I1'].fill = yellow_fill
            st['I1'].font = Font(bold=True)
            set_width_dict(st, {'I': 40})

            # 设定错误文件名
            n = 0

            while True:
                # 生成文件名
                if n > 0:
                    filename = 'error({}).xlsx'.format(n)
                else:
                    filename = 'error.xlsx'

                # 判断是否已有文件名
                if filename in os.listdir(os.path.join('media', 'temp')):
                    # 下一个号
                    n += 1
                else:
                    # 可用文件名
                    wb.save(os.path.join('media', 'temp', filename))
                    break

            # 可下载错误文件
            have_button = True
            err_id = n

            # 创建并启动延时删除文件线程
            remove_thread = threading.Thread(target=remove_lay, args=(
                os.path.join('media', 'temp', filename), 60))
            remove_thread.start()
        else:
            # 无须下载错误文件
            have_button = False
            err_id = -1

        # 返回提示
        return render_ht(request, 'zzbm/score_manage.html', context={
            'task': task,
            'title': title,
            'students': students,
            'tip': tip,
            'form': FileUploadForm(),
            'err': tip_msg,
            'have_button': have_button,
            'err_id': err_id,
        })


def download_error(request, error_num):
    """下载错误文件并然后删除"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.operators:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 连接文件名及文件路径
    if error_num > 0:
        filename = 'error({}).xlsx'.format(error_num)
    else:
        filename = 'error.xlsx'
    file_path = os.path.join('media', 'temp', filename)

    # 下载
    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f, content_type='application/vnd.ms-excel')

            # 设置文件名称
            response['Content-Disposition'] = 'attachment; filename=%s' % filename
    except FileNotFoundError:
        return HttpResponse('没有错误文件可以下载！')

    # 输出
    return response


def change_give(request, task_id):
    """开启/关闭报名自动分配考号"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出对应的任务对象
    task = Task.objects.get(id=task_id)

    # 更改状态
    task.auto_give = not task.auto_give
    task.save()

    # 重定向至任务管理页（当前页）
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def find_pwd(request, task_id):
    """找回报名序号"""
    # 取出任务对象
    task = Task.objects.get(id=task_id)

    if request.method == 'POST':
        # 读取前端数据
        sfz = request.POST.get('sfz', '')
        em = request.POST.get('em', '')

        try:
            # 锁定对象
            st_ob = Student.objects.filter(task_belong=task, id_number=sfz, add_method=1)[0]
        except IndexError:
            # 未找到对象
            return render_ht(request, 'zzbm/find_pwd.html', context={
                'task': task, 'err': '未查询到身份证号{}的考生'.format(sfz)})
        else:
            # 匹配邮箱
            if em != st_ob.email:
                return render_ht(request, 'zzbm/find_pwd.html', context={
                    'task': task, 'err': '邮箱不正确'})

            # 更新口令
            st_ob.update_pwd()
            st_ob.save()

            # 登录邮箱
            sm_ob = smtplib.SMTP()
            if os.name == 'nt':
                sm_ob.connect('smtp.qq.com', 25)
            else:
                sm_ob.connect('smtp.qq.com', 587)
            sm_ob.login('459206164', 'swekqzftsgphbibe')

            # 邮件标题
            mail_title = '{}美术加试报名序号重置'.format(settings.USER_NAME)

            # 邮件内容
            args = [st_ob.id_number, st_ob.pwd]
            mail_text = '尊敬的考生{}：\n你的报名序号已重置为{}，请牢记！'.format(*args)

            # 邮件对象
            message = MIMEText(mail_text, 'plain', 'utf-8')
            message['from'] = '459206164@qq.com'
            message['to'] = em
            message['subject'] = mail_title

            # 发送邮件
            try:
                sm_ob.sendmail('459206164@qq.com', em, message.as_string())
            except:
                return render_ht(request, 'zzbm/find_pwd.html', context={
                    'task': task, 'err': '邮件发送失败'})
            else:
                # 成功提示
                context = {'task': task, 'em': em}
                return render_ht(request, 'zzbm/tip_new_pwd.html', context)

    context = {'task': task, 'err': ''}
    return render_ht(request, 'zzbm/find_pwd.html', context)


def multi_export(request, task_id):
    """批量导入考生主页"""
    # 禁止操作员以外的用户访问此页
    if request.user.username not in DT.operators:
        raise Http404

    # 取出相应的任务对象
    task = Task.objects.get(id=task_id)

    context = {
        'task': task,
        'form': StudentsUploadForm(),
        'err': '',
        'have_button': False,
        'err_id': -1,
    }
    return render_ht(request, 'zzbm/multi_export.html', context)


def multi_temp(request):
    """下载批量导入考生模板"""
    # 禁止操作员以外的用户访问此页
    if request.user.username not in DT.operators:
        raise Http404

    # 生成示例号段
    if request.user.username in DT.managers:
        year = datetime.now().year
        num_head = str(year) + '01'
    else:
        num_head = request.user.username

    # 打开文件，定位工作表
    wb = Workbook()
    st = wb.active

    # 写入标题，设置标题样式，水平居中
    st['A1'].value = '批量导入考生模板'
    st['A1'].font = Font(size=14, bold=True)
    st['A1'].alignment = Alignment(horizontal='center')

    # 写入提示行
    st['A2'].value = '请勿更改表格格式，请删除多余准考证号后再上传，所有数据均为必填'
    st.row_dimensions[2].height = 30

    # 前两行合并单元格
    st.merge_cells(range_string='A1:G1')
    st.merge_cells(range_string='A2:G2')

    # 写入首行，包含所有要导入的信息种类
    st['A3'].value = '准考证号'
    st['B3'].value = '姓名'
    st['C3'].value = '性别'
    st['D3'].value = '身份证号'
    st['E3'].value = '手机号'
    st['F3'].value = '初中毕业学校'
    st['G3'].value = '选考科目'

    # 设置表头格式
    for c in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        st[c + '2'].font = Font(bold=True)

    # 初始行
    row = 4

    for i in range(20):
        # 生成准考证号
        ex = num_head + DT.str_three(i)

        # 写入表格
        st['A' + str(row)].value = ex

        # 下一行
        row += 1

    # 为C列添加下拉选项
    options1 = ['男', '女']
    validation1 = DataValidation(type='list', formula1='"' + ','.join(options1) + '"',
                                 allow_blank=True)
    for r in range(4, row):
        validation1.add('C' + str(r))
    st.add_data_validation(validation1)

    # 为G列添加下拉选项
    options2 = DT.get_subject_list()
    validation2 = DataValidation(type='list', formula1='"' + ','.join(options2) + '"',
                                 allow_blank=True)
    for r in range(4, row):
        validation2.add('G' + str(r))
    st.add_data_validation(validation2)

    # 列宽统一调整为15
    width_dict = {'A': 12, 'B': 12, 'C': 6, 'D': 20, 'E': 12, 'F': 24, 'G': 15}
    set_width_dict(st, width_dict)

    # 输出
    return write_out(wb, fn='student_template.xlsx')


def write_students(request, task_id):
    """模板批量导入考生"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.operators:
        raise Http404

    # 禁止非POST方法访问此页
    if request.method != 'POST':
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 取出对应的任务对象
    task = Task.objects.get(id=task_id)

    # print('get hear!')

    # 获取表单提交内容
    form = StudentsUploadForm(request.POST, request.FILES)

    if form.is_valid():
        # 临时保存文件
        upload_file = form.cleaned_data['file']
        file_path = 'media/temp/' + upload_file.name

        # 保存到指定路径
        with open(file_path, 'wb') as f:
            for chunk in upload_file.chunks():
                f.write(chunk)

        # 处理上传的文件
        try:
            wb = load_workbook(file_path)
            st = wb.active
        except InvalidFileException:
            # 删除临时文件
            os.remove(file_path)

            # 错误提示信息
            return render_ht(request, 'zzbm/multi_export.html', context={
                'task': task,
                'form': FileUploadForm(),
                'err': '文件格式必须为xlsx',
                'have_button': False,
                'err_id': -1,
            })

        # 删除临时文件
        os.remove(file_path)

        # 成功数和失败数
        success, fail = 0, 0

        # 错误提示填充
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # 从第4行开始逐行读取
        for row in range(4, st.max_row + 1):
            # 读取准考证号
            exam_id = str(st.cell(row=row, column=1).value)

            # 默认准考证号格式合法
            lea = True

            # 操作员权限
            if request.user.username in DT.managers:
                power = DT.num_head()
            else:
                power = [request.user.username]

            if request.user.username != 'zz106dyc':
                # 准考证号须为9位
                if len(exam_id) != 9:
                    lea = False

                # 准考证号前六位须为操作员用户名
                if exam_id[:6] not in power:
                    lea = False
            else:
                # 管理员特权，可分配任意格式准考证号
                if len(exam_id) != 9 and len(exam_id) != 12:
                    lea = False

                head7 = str(datetime.now().year) + '106'
                if exam_id[:6] not in power and exam_id[:7] != head7:
                    lea = False

            if not lea:
                st.cell(row=row, column=8).value = '准考证号格式错误'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            if exam_id in all_exam_id(task):
                st.cell(row=row, column=8).value = '准考证号重复'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            if len(exam_id) == 9:
                method = 2
            elif len(exam_id) == 12:
                method = 1
            else:
                st.cell(row=row, column=8).value = '准考证号格式错误'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 读取姓名
            name = st.cell(row=row, column=2).value

            if name is None:
                st.cell(row=row, column=8).value = '姓名不能为空'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            if len(name) < 1 or len(name) > 5:
                st.cell(row=row, column=8).value = '姓名长度须在1~5之间'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 读取性别
            gender = st.cell(row=row, column=3).value

            if gender not in ['男', '女']:
                st.cell(row=row, column=8).value = '性别只能是"男"或"女"'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 读取身份证号
            id_number = str(st.cell(row=row, column=4).value)

            if not id_number:
                # 身份证号暂时缺失，先记作‘0’，待完善
                id_number = '0'
            else:
                if not correct_id(id_number):
                    st.cell(row=row, column=8).value = '身份证号格式非法'
                    st.cell(row=row, column=8).fill = yellow_fill

                    # 失败数加一
                    fail += 1
                    continue

                if len(id_number) == 18:
                    if {0: '女', 1: '男'}[int(id_number[16]) % 2] != gender:
                        st.cell(row=row, column=8).value = '身份证号与性别不匹配'
                        st.cell(row=row, column=8).fill = yellow_fill

                        # 失败数加一
                        fail += 1
                        continue

                if id_number in all_id(task, add_method=method):
                    st.cell(row=row, column=8).value = '身份证号{}的考生重复报名'.format(id_number)
                    st.cell(row=row, column=8).fill = yellow_fill

                    # 失败数加一
                    fail += 1
                    continue

            # 读取手机号
            phone_number = str(st.cell(row=row, column=5).value)

            if len(phone_number) != 11 or phone_number[0] != '1':
                st.cell(row=row, column=8).value = '手机号格式错误'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 读取初中毕业学校
            middle_school = st.cell(row=row, column=6).value

            if middle_school is None:
                st.cell(row=row, column=8).value = '初中毕业学校未填'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            if len(middle_school) < 1 or len(middle_school) > 20:
                st.cell(row=row, column=8).value = '初中毕业学校长度须在1~20之间'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 读取选考科目
            subject = st.cell(row=row, column=7).value

            if subject not in DT.get_subject_list():
                st.cell(row=row, column=8).value = '选考科目只能是“素描或创意画”、”书法或国画“二选一'
                st.cell(row=row, column=8).fill = yellow_fill

                # 失败数加一
                fail += 1
                continue

            # 生成考生对象
            student = Student()
            student.exam_id = exam_id
            student.name = name
            student.gender = gender
            student.card_type = 1
            student.id_number = id_number
            student.phone_number = phone_number
            student.middle_school = middle_school
            student.subject = subject

            # 考生对象与任务对象相关联
            student.task_belong = task
            student.num = task.added + 1
            task.added += 1

            # 保存
            student.add_method = method
            student.save()
            task.save()

            # 成功数加1
            success += 1

        # 生成提示信息
        tip_msg = '共上传{}条数据，成功{}条，失败{}条'.format(success + fail, success, fail)

        if fail > 0:
            # 存在处理失败数据，保存失败文件
            st['H3'].value = '错误提示'
            st['H3'].fill = yellow_fill
            st['H3'].font = Font(bold=True)
            set_width_dict(st, {'F': 50})

            # 设定错误文件名
            n = 0

            while True:
                # 生成文件名
                if n > 0:
                    filename = 'error({}).xlsx'.format(n)
                else:
                    filename = 'error.xlsx'

                # 判断是否已有文件名
                if filename in os.listdir(os.path.join('media', 'temp')):
                    # 下一个号
                    n += 1
                else:
                    # 可用文件名
                    wb.save(os.path.join('media', 'temp', filename))
                    break

            # 可下载错误文件
            have_button = True
            err_id = n

            # 创建并启动延时删除文件线程
            remove_thread = threading.Thread(target=remove_lay, args=(
                os.path.join('media', 'temp', filename), 60))
            remove_thread.start()
        else:
            # 无须下载错误文件
            have_button = False
            err_id = -1

        # 返回提示
        return render_ht(request, 'zzbm/multi_export.html', context={
            'task': task,
            'form': FileUploadForm(),
            'err': tip_msg,
            'have_button': have_button,
            'err_id': err_id,
        })


def get_score_input(request, task_id):
    """直接查询成绩的页面"""
    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 生成标题
    title = '查询{}{}年美术后备生测试成绩'.format(settings.USER_NAME, task.year)

    if request.method == 'POST':
        # 对POST提交的数据作出处理，前端读取输入信息
        # exam_id = request.POST.get('exam-id', '')
        id_number = request.POST.get('id-number', '')

        # 抓取考生对象
        students = Student.objects.filter(task_belong=task, id_number=id_number)

        if not students:
            # 考生不存在
            return render_ht(request, 'zzbm/get_score_input.html', context={
                'task': task,
                'title': title,
                'err': '身份证号{}的考生不存在'.format(id_number),
            })

        # # 默认不匹配
        # match = False
        #
        # for student in students:
        #     if exam_id == student.exam_id:
        #         # 匹配成功
        #         match = True
        #         break
        #
        # if not match:
        #     # 提示不匹配
        #     return render(request, 'zzbm/get_score_input.html', context={
        #         'task': task,
        #         'title': title,
        #         'err': '身份证号与准考证号不匹配',
        #     })

        # 默认初始成绩及考生对象
        score_int = -1
        ob_out = None

        for ob in students:
            if DT.score_int[ob.score] > score_int:
                score_int = DT.score_int[ob.score]
                ob_out = ob

        if score_int < 0:
            # 成绩为O，暂不予显示
            return render_ht(request, 'zzbm/get_score_input.html', context={
                'task': task,
                'title': title,
                'err': '未查询到你的成绩',
            })
        else:
            # 显示成绩
            return HttpResponseRedirect(reverse('zzbm:get_score', args=[
                ob_out.id, ob_out.pwd]))

    # 未提交数据，加载新的查询页面
    context = {'task': task, 'title': title, 'err': ''}
    return render_ht(request, 'zzbm/get_score_input.html', context)


def get_in(request, task_id):
    """核对信息验证"""
    # 取出任务对象
    task = Task.objects.get(id=task_id)

    if request.method == 'POST':
        # 对POST提交的数据作出处理，读取输入框内容
        exam_id = request.POST.get('ex-id', '')
        phone_number = request.POST.get('phone', '')

        # 验证准考证号位数
        if len(exam_id) != 9:
            return render_ht(request, 'zzbm/get_in.html', context={
                'task': task, 'err': '准考证号格式错误'})

        # 尝试取出考生对象
        try:
            student = Student.objects.filter(task_belong=task, exam_id=exam_id)[0]
        except IndexError:
            return render_ht(request, 'zzbm/get_in.html', context={
                'task': task, 'err': '准考证号不存在'})
        else:
            # 检验手机号是否匹配
            if phone_number != student.phone_number:
                return render_ht(request, 'zzbm/get_in.html', context={
                    'task': task, 'err': '手机号不匹配'})

            # 通过检验，导向核对页面
            return HttpResponseRedirect(reverse('zzbm:check', args=[student.id]))
    else:
        # 加载新的验证页面
        context = {'task': task, 'err': ''}
        return render_ht(request, 'zzbm/get_in.html', context)


def check(request, student_id):
    """核对信息"""
    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 直接报名的考生不允许核对
    if len(student.exam_id) != 9:
        raise Http404

    context = {'student': student}
    return render_ht(request, 'zzbm/check.html', context)


def complete_id(request, student_id):
    """完善身份证号"""
    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 身份证号已完善者不可访问此页
    if student.id_number != '0':
        raise Http404

    if request.method == 'POST':
        # 对POST提交的数据作出处理，读取输入内容
        id_number = request.POST.get('sfz', '')

        # 验证身份证号
        if not correct_id(id_number):
            return render_ht(request, 'zzbm/complete_id.html', context={
                'student': student, 'err': '身份证号格式非法'})

        # 检查重复
        if id_number in all_id(student.task_belong, add_method=2):
            return render_ht(request, 'zzbm/complete_id.html', context={
                'student': student, 'err': '身份证号{}的考生已报名'.format(id_number)})

        # 检查身份证号与性别匹配
        if {0: '女', 1: '男'}[int(id_number[16]) % 2] != student.gender:
            return render_ht(request, 'zzbm/complete_id.html', context={
                'student': student, 'err': '身份证号与性别不匹配'})

        # 通过验证，保存数据
        student.id_number = id_number
        student.save()

        # 导向核对页
        return HttpResponseRedirect(reverse('zzbm:check', args=[student.id]))
    else:
        # 加载空白的身份证号输入框
        context = {'student': student, 'err': ''}
        return render_ht(request, 'zzbm/complete_id.html', context)


def see_multi(request, task_id):
    """查看模板批量导入的考生"""
    # 设置访问权限
    if request.user.username not in DT.operators:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 初始化考生列表
    sts = []

    for student in Student.objects.filter(task_belong=task, add_method=2).order_by('exam_id'):
        if request.user.username in DT.managers:
            # 管理员可查看、操作所有
            sts.append(student)
        else:
            if request.user.username == student.exam_id[:6]:
                # 可查看、操作自己导入的
                sts.append(student)

    context = {'task': task, 'sts': tuple(sts)}
    return render_ht(request, 'zzbm/see_multi.html', context)


def change_info(request, student_id):
    """修改已导入考生信息"""
    # 设置访问权限
    if request.user.username not in DT.operators:
        raise Http404

    # 取出要修改的考生对象
    student = Student.objects.get(id=student_id)

    # 生成标题
    title = '修改准考证号{}的考生信息'.format(student.exam_id)

    if request.method != 'POST':
        # 未提交数据，用考生原有信息填充表单
        form = ChangeInfoForm(instance=student)
    else:
        # 对POST提交的数据作出处理
        form = ChangeInfoForm(instance=student, data=request.POST)

        if form.is_valid():
            # 依据表单内容修改信息
            to_change = form.save(commit=False)

            # 验证身份证号格式
            if not correct_id(to_change.id_number):
                return render_ht(request, 'zzbm/change_info', context={
                    'student': student,
                    'title': title,
                    'form': form,
                    'err': '身份证号格式非法！',
                })

            # 验证身份证号与性别是否匹配
            if len(to_change.id_number) == 18:
                if {0: '女', 1: '男'}[int(to_change.id_number[16]) % 2] != to_change.gender:
                    return render_ht(request, 'zzbm/change_info.html', context={
                        'student': student,
                        'title': title,
                        'form': form,
                        'err': '身份证号与性别不匹配！',
                    })

            # 防止重复报名
            if to_change.id_number in all_id(student.task_belong, add_method=2, exc=[student]):
                return render_ht(request, 'zzbm/change_info.html', context={
                    'student': student,
                    'title': title,
                    'form': form,
                    'err': '身份证号{}已报名！'.format(to_change.id_number),
                })

            # 判断手机号格式是否有误
            if len(to_change.phone_number) != 11 or to_change.phone_number[0] != '1':
                return render_ht(request, 'zzbm/chang_info.html', context={
                    'student': student,
                    'title': title,
                    'form': form,
                    'err': '手机号格式错误！',
                })

            # 通过检验，修改考生信息
            student.name = to_change.name
            student.gender = to_change.gender
            student.id_number = to_change.id_number
            student.phone_number = to_change.phone_number
            student.middle_school = to_change.middle_school
            student.subject = to_change.subject
            student.save()

            # 重定向至考生查看页
            return HttpResponseRedirect(reverse('zzbm:see_multi',
                                                args=[student.task_belong.id]))

    context = {'student': student, 'title': title, 'form': form, 'err': ''}
    return render_ht(request, 'zzbm/change_info.html', context)


def reset_task_num(request, task_id):
    """重置任务考场分配情况"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 重设参数
    task.mr1 = 0
    task.mr2 = 50
    task.turn1 = 0
    task.turn2 = 0
    # task.max_len = 0

    # 保存对象
    task.save()

    # 重定向至任务管理页
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def edit_student(request, student_id):
    """修改考生信息"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出考生对象及所属任务对象
    student = Student.objects.get(id=student_id)
    task = student.task_belong

    if request.method != 'POST':
        # 未提交数据，用当前考生填充表单
        form = ChangePutForm(instance=student)
    else:
        # 对POST提交的数据作出处理
        form = ChangePutForm(instance=student, data=request.POST, files=request.FILES)

        if form.is_valid():
            editor = form.save(commit=False)

            # 判断身份证号是否合法
            if not correct_id(editor.id_number):
                return render_ht(request, 'zzbm/edit_student.html', context={
                    'task': task,
                    'student': student,
                    'form': form,
                    'err': '身份证号格式非法！',
                })

            # 判断性别与身份证号是否匹配
            if len(editor.id_number) == 18:
                if {0: '女', 1: '男'}[int(editor.id_number[16]) % 2] != editor.gender:
                    return render_ht(request, 'zzbm/edit_student.html', context={
                        'task': task,
                        'student': student,
                        'form': form,
                        'err': '身份证号与性别不匹配，请检查是否输入有误！',
                    })

            # 防止重复报名
            if editor.id_number in all_id(task, add_method=1, exc=[student]):
                return render_ht(request, 'zzbm/edit_student.html', context={
                    'task': task,
                    'student': student,
                    'form': form,
                    'err': '身份证号{}的考生已报名，请勿重复报名'.format(editor.id_number),
                })

            # 判断手机号格式是否有误
            if len(editor.phone_number) != 11 or editor.phone_number[0] != '1':
                return render_ht(request, 'zzbm/edit_student.html', context={
                    'task': task,
                    'student': student,
                    'form': form,
                    'err': '手机号格式有误！',
                })

            # 判断邮箱格式是否有误
            if editor.email:
                if '@' not in editor.email or '.com' not in editor.email:
                    return render_ht(request, 'zzbm/edit_student.html', context={
                        'task': task,
                        'student': student,
                        'form': form,
                        'err': '邮箱格式有误！',
                    })

            # # 一寸照片必须上传
            # if not editor.photo:
            #     return render(request, 'zzbm/edit_student.html', context={
            #         'task': task,
            #         'student': student,
            #         'form': form,
            #         'err': '请上传个人一寸照片',
            #     })
            #
            # # 一寸照片大小不能超过1MB
            # if editor.photo.size > 1024 * 1024:
            #     return render(request, 'zzbm/edit_student.html', context={
            #         'task': task,
            #         'student': student,
            #         'form': form,
            #         'err': '上传一寸照片大小不能超过1MB',
            #     })

            # # 准考证号格式要符合规范
            # head7 = str(datetime.now().year) + '106'
            # if editor.exam_id != '未分配':
            #     if len(editor.exam_id) != 12 or editor.exam_id[:7] != head7:
            #         return render(request, 'zzbm/edit_student.html', context={
            #             'task': task,
            #             'student': student,
            #             'form': form,
            #             'err': '准考证号格式不符合规范',
            #         })

            # 根据表单填写内容修改考生对象字段
            student.name = editor.name
            student.gender = editor.gender
            student.card_type = editor.card_type
            student.id_number = editor.id_number
            student.phone_number = editor.phone_number
            student.middle_school = editor.middle_school
            student.subject = editor.subject
            student.photo = editor.photo
            student.exam_id = editor.exam_id
            student.save()

            # 重定向至考生查看页
            return HttpResponseRedirect(reverse('zzbm:student', args=[student_id, student.pwd]))

    context = {'task': task, 'student': student, 'form': form, 'err': ''}
    return render_ht(request, 'zzbm/edit_student.html', context)


def change_miss(request, student_id):
    """标记缺考或取消标记缺考"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 更改缺考状态
    student.miss = not student.miss
    student.save()

    # 重定向至考生列表页
    return HttpResponseRedirect(reverse('zzbm:list_put', args=[student.task_belong.id, student_id]))


def export_score(request, task_id, pub):
    """导出成绩单"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '准考证号'
    st['B1'].value = '姓名'
    st['C1'].value = '身份证号'
    st['D1'].value = '考试科目'
    st['E1'].value = '成绩'

    # 初始行
    row = 2

    for ob in Student.objects.filter(task_belong=task):
        if ob.score != 'O':
            st.cell(row=row, column=1).value = ob.exam_id
            st.cell(row=row, column=2).value = ob.name
            st.cell(row=row, column=3).value = ob.id_number
            st.cell(row=row, column=4).value = ob.subject
            if pub:
                st.cell(row=row, column=5).value = DT.scores_2025[ob.score]
            else:
                st.cell(row=row, column=5).value = ob.score

            row += 1

    return write_out(wb)


def none_score(request, task_id):
    """导出无成绩考生"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '准考证号'
    st['B1'].value = '姓名'
    st['C1'].value = '身份证号'
    st['D1'].value = '初中毕业学校'

    # 初始行
    row = 2

    for ob in Student.objects.filter(task_belong=task, add_method=2):
        if ob.score == 'O':
            st.cell(row=row, column=1).value = ob.exam_id
            st.cell(row=row, column=2).value = ob.name
            st.cell(row=row, column=3).value = ob.id_number
            st.cell(row=row, column=4).value = ob.middle_school

            row += 1

    return write_out(wb)


def de_repeat(request, task_id, pub):
    """导出去重成绩单"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 建立身份证号与考生对象对应字典
    id_st = {}
    for ob in Student.objects.filter(task_belong=task):
        if ob.score != 'O':
            id_st.setdefault(ob.id_number, ob)

            # 比较成绩，取大者
            if DT.score_int[ob.score] > DT.score_int[id_st[ob.id_number].score]:
                id_st[ob.id_number] = ob

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '准考证号'
    st['B1'].value = '姓名'
    st['C1'].value = '身份证号'
    st['D1'].value = '考试科目'
    st['E1'].value = '成绩'

    # 初始行
    row = 2

    for student in id_st.values():
        st.cell(row=row, column=1).value = student.exam_id
        st.cell(row=row, column=2).value = student.name
        st.cell(row=row, column=3).value = student.id_number
        st.cell(row=row, column=4).value = student.subject
        if pub:
            st.cell(row=row, column=5).value = DT.scores_2025[student.score]
        else:
            st.cell(row=row, column=5).value = student.score

        row += 1

    return write_out(wb)


def show_repeat(request, task_id):
    """导出重复考生"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 建立字典
    multi_dict = {}
    for student in Student.objects.filter(task_belong=task):
        if student.score != 'O':
            multi_dict.setdefault(student.id_number, [])
            multi_dict[student.id_number].append(student)

    # 打开新的工作表
    wb = Workbook()
    st = wb.active

    # 写入表头
    st['A1'].value = '准考证号'
    st['B1'].value = '姓名'
    st['C1'].value = '身份证号'
    st['D1'].value = '考试科目'
    st['E1'].value = '成绩'

    # 初始行
    row = 2

    for obs in multi_dict.values():
        if len(obs) > 1:
            for ob in obs:
                st.cell(row=row, column=1).value = ob.exam_id
                st.cell(row=row, column=2).value = ob.name
                st.cell(row=row, column=3).value = ob.id_number
                st.cell(row=row, column=4).value = ob.subject
                st.cell(row=row, column=5).value = ob.score

                row += 1

    return write_out(wb)


def que_ctrl(request, task_id):
    """查成绩开关"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 更改状态
    task.can_que = not task.can_que
    task.save()

    # 重定向至任务管理页
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def change_start_turn(request, task_id, target):
    """根据已结束的场次数更改start_turn属性"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出任务对象
    task = Task.objects.get(id=task_id)

    # 更改状态
    if task.start_turn < target:
        task.start_turn = target
    elif task.start_turn == target:
        task.start_turn -= 1
    else:
        # 不允许的操作
        raise Http404

    # 更改两科已分配轮数
    task.turn1 = task.start_turn
    task.turn2 = task.start_turn

    # 重置已分配考场数
    task.mr1 = task.get_mr(str(task.start_turn))[0]
    task.mr2 = task.get_mr(str(task.start_turn))[1]

    # 保存任务
    task.save()

    # 重定向至任务管理页
    return HttpResponseRedirect(reverse('zzbm:task_manage', args=[task_id]))


def get_year_room():
    """获取所有任务年份及对应考场号，以字典形式返回"""
    # 初始化存放字典
    yrd = {}
    for task in Task.objects.all():
        # 初始化考场存放列表
        room_list = []

        # 循环遍历所有考生、加入考场号
        for student in Student.objects.filter(task_belong=task):
            if student.room != '未分配':
                room_list.append(student.room)

        # 去重、排序
        room_list = list(set(room_list))
        room_list.sort()

        # 加入字典
        yrd[task.year] = room_list

    return yrd


def load_miss_main(request):
    """登记缺考考场选择"""
    # # 禁止非管理员用户访问此页
    # if request.user.username not in DT.managers:
    #     raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 加考场-密码载存放对象
    with open('media/room-pwd.json') as f:
        rpds = f.read()
        if rpds:
            rpd = json.loads(rpds)
        else:
            rpd = {}

    # 基础数据
    yrd = get_year_room()
    year_options = tuple(yrd.keys())
    yrds = json.dumps(yrd)

    if request.method != 'POST':
        # 未提交数据，加载新的选择页面
        context = {'year_options': year_options, 'yrds': yrds, 'err': '',
                   'tt': '{}中招美术加试缺考考生登记'.format(settings.USER_NAME)}
        return render_ht(request, 'zzbm/load_miss_main.html', context)
    else:
        # 前端读取年份、考场号、密码
        year = request.POST.get('year', '')
        room = request.POST.get('room', '')
        pwd = request.POST.get('pwd', '')

        # 获取正确的密码
        key = year + '-' + room
        try:
            correct_pwd = rpd[key]
        except KeyError:
            # 默认密码为106106
            correct_pwd = '106106'

        # 验证密码
        if pwd != correct_pwd:
            return render_ht(request, 'zzbm/load_miss_main.html', context={
                'year_options': year_options,
                'yrds': yrds,
                'err': '密码错误！',
            })

        # 获取年份-考场所有考生
        task = Task.objects.get(year=year)
        students = Student.objects.filter(task_belong=task, room=room)

        # 制定标题
        title = '{}{}{}考场考生'.format(year, DT.ei_mid, room)

        # 前端显示
        context = {'title': title, 'students': students, 'task': task, 'st_id': '0'}
        return render_ht(request, 'zzbm/list_put_small.html', context)


def change_miss_1(request, student_id):
    """标记缺考或取消标记缺考"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 取出考生对象
    student = Student.objects.get(id=student_id)

    # 更改缺考状态
    student.miss = not student.miss
    student.save()

    # 重新制定标题，确定考生
    title = '{}{}{}考场考生'.format(student.task_belong.year, DT.ei_mid, student.room)
    students = Student.objects.filter(task_belong=student.task_belong, room=student.room)

    # 前端显示
    context = {'title': title, 'students': students, 'task': student.task_belong, 'st_id': str(student_id)}
    return render_ht(request, 'zzbm/list_put_small.html', context)


def room_pwd_manage(request, task_id):
    """考场密码管理"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 获取任务对象及年份、考场列表
    task = Task.objects.get(id=task_id)
    year = task.year
    room_list = get_year_room()[year]

    # 前端显示
    items = []

    # 加载存放对象
    with open('media/room-pwd.json') as f:
        rpds = f.read()
        if rpds:
            rpd = json.loads(rpds)
        else:
            rpd = {}

    # 筛选年份、考场号
    for room in room_list:
        key = year + '-' + room
        try:
            pwd_val = rpd[key]
        except KeyError:
            pwd_val = ''

        # 加入显示对象
        items.append((room, pwd_val))

    # 制定标题
    title = '{}年中招美术加试考场密码管理'.format(year)

    context = {'task': task, 'title': title, 'items': tuple(items)}
    return render_ht(request, 'zzbm/room_pwd_manage.html', context)


def room_pwd_update(request, task_id):
    """考场密码更新"""
    # 禁止非管理员用户访问此页
    if request.user.username not in DT.managers:
        raise Http404

    # 确保服务器上进入正确的目录，可正常运行
    if os.name != 'nt':
        os.chdir('/root/dormitory_manager/dm')

    # 获取任务对象及年份、考场列表
    task = Task.objects.get(id=task_id)
    year = task.year
    room_list = get_year_room()[year]

    # 加载存放对象
    with open('media/room-pwd.json') as f:
        rpds = f.read()
        if rpds:
            rpd = json.loads(rpds)
        else:
            rpd = {}

    for room in room_list:
        key = year + '-' + room

        # 生成新密码
        new_pwd = ''
        for i in range(6):
            new_pwd += str(random.randint(0, 9))

        # 更新存储
        rpd[key] = new_pwd

    # 重新存储
    with open('media/room-pwd.json', 'w') as fo:
        fo.write(json.dumps(rpd))

    # 重定向至考场密码管理页
    return HttpResponseRedirect(reverse('zzbm:room_pwd_manage', args=[task_id]))
